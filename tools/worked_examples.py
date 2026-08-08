#!/usr/bin/env python3
"""Fuller worked examples, and the column widths to read them in.

Every row-based template shipped with exactly one example row. That is enough to
show the expected format and nothing else: the value stream map's chart was a
single 2-minute bar and its Process Cycle Efficiency read 100%, which is the
precise opposite of the lesson. The FMEA showed one pair of bars. A chart with
one bar teaches nothing and looks broken.

These examples are sized to make the point of each tool visible the moment the
file opens — the VSM's touch time is 26 minutes against 2,400 minutes of waiting,
so PCE lands near 1% where it belongs. Every example row keeps the green fill, so
it stays obvious what to delete.

Called by tools/patch_workbooks.py. Idempotent: a cell is only counted as
changed when its value actually differs.
"""
from __future__ import annotations

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

EX = PatternFill("solid", fgColor="FFECFAEF")      # "this is an example" green


def _rows(ws, first_row: int, rows: list[list], widths: dict | None = None) -> int:
    """Write example rows in the green fill, leaving formula columns alone."""
    changed = 0
    for w, val in (widths or {}).items():
        if ws.column_dimensions[w].width != val:
            ws.column_dimensions[w].width = val
            changed += 1
    for i, row in enumerate(rows):
        r = first_row + i
        for j, val in enumerate(row, start=1):
            # None marks a calculated column. "" reads back as None once the
            # file has been saved, so writing it would churn on every run.
            if val is None or val == "":
                continue
            c = ws.cell(row=r, column=j)
            if c.__class__.__name__ == "MergedCell":
                continue
            if c.value != val:
                c.value = val
                changed += 1
            c.fill = EX
    return changed


# --------------------------------------------------------------------------


def value_stream_map(wb) -> int:
    """A billing dispute end to end. Touch time 26 min, waiting 2,400 min."""
    ws = wb["Value stream"]
    # The shipped widths were offset by several columns: the "#" column was 32
    # wide and "Wait BEFORE this step" was 4, so its header wrapped one letter
    # per line.
    widths = {"A": 5, "B": 34, "C": 15, "D": 22, "E": 13,
              "F": 15, "G": 13, "H": 15, "I": 11, "J": 34}
    rows = [
        [1, "Customer submits the form", "Customer", "Help centre", 2, 0, 0.92,
         "Value-add", "No", "Form has 3 fields nobody reads"],
        [2, "Sits in the unassigned queue", "—", "Zendesk", 0, 240, 1.00,
         "Non-value-add", "No", "Overnight arrivals wait until the morning shift"],
        [3, "Tier 1 triage and tagging", "Tier 1", "Zendesk", 4, 0, 0.78,
         "Business-value-add", "No", "22% get the wrong contact reason, which breaks routing"],
        [4, "Waits for the daily Billing Ops sweep", "—", "Zendesk", 0, 480, 1.00,
         "Non-value-add", "No", "Billing Ops pull the queue once a day at 10:00"],
        [5, "Billing Ops checks the posting", "Billing Ops", "Billing console + warehouse",
         11, 0, 0.85, "Value-add", "Yes", "15% bounce back to Tier 1 for missing evidence"],
        [6, "Waits for the nightly adjustment batch", "—", "Billing platform", 0, 720, 1.00,
         "Non-value-add", "No", "Cut-off is 18:00; miss it and you wait another day"],
        [7, "Adjustment posts, agent verifies", "Billing Ops", "Billing console", 6, 0, 0.94,
         "Value-add", "No", ""],
        [8, "Waits for the customer to confirm", "Customer", "Email", 0, 960, 1.00,
         "Non-value-add", "No", "Nothing chases the customer; this is the longest single wait"],
        [9, "Agent closes and confirms", "Tier 1", "Zendesk", 3, 0, 0.97,
         "Value-add", "No", ""],
    ]
    return _rows(ws, 10, rows, widths)


def fmea(wb) -> int:
    """Six failure modes, so the RPN chart ranks something."""
    ws = wb["FMEA"]
    rows = [
        ["Agent closes a billing ticket", "Adjustment has not posted yet",
         "Customer discovers it themselves and contacts us again", 7,
         "System closes the ticket before the posting webhook confirms", 8, "None",
         "Only detected if the customer complains again", 9, None,
         "Block Resolved status until the posting webhook confirms", "Billing Ops",
         "2026-09-30", "Deferred-close rule shipped", 7, 2, 3, None],
        ["Tier 1 tags the contact reason", "Wrong reason selected",
         "Routing sends it to the wrong queue; customer repeats themselves", 6,
         "40 codes in a flat list, no search", 9, "Annual training",
         "QA audit catches it weeks later on 2% of contacts", 8, None,
         "Cut the list to 12 codes and predict from the first message", "Support Ops",
         # Re-scored only for what has SHIPPED. This row's control is built and
         # not live, so it keeps its original 6 / 9 / 8 and earns no credit yet
         # — the same discipline the "In backlog" row below already followed.
         # Scoring it at 6 / 3 / 4 while pending put 19.6 points into the
         # headline risk reduction, taking it from 57.7% to 77.3% on the
         # strength of a control no customer had met. That is the exact defect
         # this pack's benefit case exists to teach.
         "2026-10-15", "Code list redesigned, pending release — not re-scored until live",
         6, 9, 8, None],
        ["Billing Ops verifies evidence", "Evidence missing, ticket bounced back",
         "Adds a full day to the resolution", 5,
         "Tier 1 has no checklist for what Billing needs", 7, "Wiki page nobody opens",
         "Detected immediately on bounce-back", 3, None,
         "Required-fields form before hand-off", "Support Ops", "2026-09-15",
         "Form live since August", 5, 2, 3, None],
        ["Nightly adjustment batch runs", "Batch fails silently",
         "Every adjustment that day is late; nobody knows until customers call", 8,
         "No alerting on batch exit code", 4, "None",
         "Discovered next morning by chance", 9, None,
         "Alert on non-zero exit to the on-call channel", "Platform Eng", "2026-09-01",
         "Alerting deployed", 8, 2, 2, None],
        ["Agent applies a goodwill credit", "Credit exceeds the agent's authority",
         "Margin leakage; caught in month-end reconciliation", 4,
         "Limit is documented but not enforced in the tool", 6, "SOP states the limit",
         "Month-end reconciliation, up to 30 days later", 7, None,
         "Enforce the limit in the credit form", "Billing Ops", "2026-11-30",
         "In backlog", 4, 6, 7, None],
        ["Customer confirms resolution", "No confirmation ever arrives",
         "Ticket auto-closes while the customer still has the problem", 6,
         "Auto-close at 96 hours with no reminder", 7, "None",
         "Only visible as a reopen days later", 8, None,
         "Send a reminder at 48 hours before auto-close", "Support Ops", "2026-10-01",
         "Reminder shipped", 6, 3, 4, None],
    ]
    return _rows(ws, 11, rows)


def xy_matrix(wb) -> int:
    """Eight candidate causes, so the ranking has something to rank."""
    ws = wb["X-Y matrix"]
    rows = [
        ["Ticket closed before the adjustment posts", "Process", 9, 3, 6, 6, 9, None, None],
        ["Contact reason mis-tagged at triage", "Process", 6, 9, 8, 4, 6, None, None],
        ["Billing Ops sweep runs once a day", "Method", 8, 4, 9, 7, 5, None, None],
        ["Evidence requirements not stated to Tier 1", "Method", 5, 6, 7, 3, 4, None, None],
        ["Nightly batch has no failure alerting", "Machine", 7, 2, 8, 9, 3, None, None],
        ["New-hire tenure under 90 days", "People", 6, 7, 4, 5, 8, None, None],
        ["Credit authority not enforced in the tool", "Machine", 3, 5, 3, 8, 4, None, None],
        ["Auto-close at 96 hours with no reminder", "Process", 8, 3, 5, 6, 7, None, None],
    ]
    return _rows(ws, 15, rows)


def solution_selection(wb) -> int:
    """Six options, so the weighted ranking is a ranking."""
    ws = wb["Solution selection"]
    rows = [
        ["Ticket closed before adjustment posts", "Block Resolved until the webhook confirms",
         5, 5, 3, 4, 3, 4, 5, None, None, "2 Design it out", "Yes"],
        ["Contact reason mis-tagged", "Cut to 12 codes and predict from first message",
         4, 5, 2, 3, 2, 3, 5, None, None, "2 Design it out", "Yes"],
        ["Billing sweep runs once a day", "Move to continuous pull with a WIP cap",
         5, 4, 2, 3, 2, 2, 4, None, None, "3 Mistake-proof it", "Yes"],
        ["Batch fails silently", "Alert on non-zero exit to on-call",
         3, 5, 5, 5, 5, 5, 5, None, None, "4 Automatic detection", "Yes"],
        ["Evidence missing at hand-off", "Required-fields form before hand-off",
         3, 4, 4, 5, 4, 4, 4, None, None, "3 Mistake-proof it", "Yes"],
        ["Contact reason mis-tagged", "Retrain agents on the existing code list",
         2, 1, 5, 5, 4, 5, 2, None, None, "5 Standard work + training", "No"],
    ]
    return _rows(ws, 14, rows)


def hypothesis_log(wb) -> int:
    """Five tests, so the p-value chart has a line to draw."""
    ws = wb["Test log"]
    rows = [
        [1, "Did the deferred-close rule cut reopens?", "2-proportion test", "Proportion",
         "Yes", "Contact reason, tenure", 0.41, 7420, 7180, 0.0001, "-4.9 pts",
         "-6.1 to -3.7", "13.1 pts", None,
         "Reopens fell 4.9 points — real, and short of the 13.1 points registered as the "
         "smallest change that clears the floor. Significant is not the same as bookable",
         "Rejected", "2026-06-14", "M. Berenji"],
        [2, "Did the shorter code list improve routing accuracy?", "2-proportion test",
         "Proportion", "Yes", "Queue, tenure", 0.22, 4100, 4260, 0.0031, "+6.2 pts",
         "+2.1 to +10.3", "3.0 pts", None, "Routing accuracy up 6.2 points",
         "Accepted", "2026-07-02", "M. Berenji"],
        [3, "Is handle time different across the three sites?", "One-way ANOVA",
         "Continuous", "Yes", "Contact reason", 0.68, 5200, 5150, 0.2140, "0.4 min",
         "-0.3 to 1.1", "1.0 min", None,
         "No detectable site difference once contact mix is held constant",
         "Rejected", "2026-07-19", "A. Okafor"],
        [4, "Did the 48-hour reminder reduce auto-closes?", "2-proportion test",
         "Proportion", "Yes", "Contact reason", 0.55, 3900, 3880, 0.0009, "-3.1 pts",
         "-4.8 to -1.4", "2.0 pts", None, "Auto-closes down 3.1 points",
         "Accepted", "2026-08-05", "A. Okafor"],
        [5, "Does tenure predict QA score?", "Simple regression", "Continuous",
         "Yes", "Queue", 0.31, 840, 0, 0.0400, "R-sq 0.09",
         "slope 0.02 to 0.31", "R-sq 0.25", None,
         "Significant but explains only 9% — statistically real, operationally thin",
         "Noted", "2026-08-21", "M. Berenji"],
    ]
    return _rows(ws, 11, rows)


def control_plan(wb) -> int:
    """Five controls across the hierarchy, so the durability chart has a shape."""
    ws = wb["Control plan"]
    rows = [
        ["7-day reopen rate, billing", "<= 8.0%", "Warehouse query, def v3", "warehouse.tickets",
         "Laney p' (attribute, large n)", 0.086, 0.104, 0.068, "Weekly", "Monday ops review",
         "Billing Support Mgr", "Any point above the UCL", "Pull 20 reopens same day and classify",
         "Two consecutive weeks above", "Ops Director + Champion", "2 Design it out",
         "1 — system blocks it", "SOP-BIL-014 v3, QA item 7"],
        ["Routing accuracy", ">= 90%", "QA audit sample, 200/week", "QA tool",
         "Laney p' (attribute, large n)", 0.91, 0.94, 0.88, "Weekly", "Monday ops review",
         "Support Ops Mgr", "Below the LCL, or 8 points under the centre line",
         "Re-audit 50 tickets and review the code list", "Two weeks below target",
         "Ops Director", "2 Design it out", "2 — tool warns", "SOP-SUP-021 v2"],
        ["Nightly batch completion", "100%", "Batch exit code", "Platform monitoring",
         "t-chart (days between failures)", 0, 0, 0, "Daily", "Automated alert",
         "Platform Eng on-call", "Any non-zero exit code",
         "On-call investigates before 08:00; Billing Ops notified", "Two failures in a week",
         "Eng Manager", "4 Automatic detection", "1 — system blocks it", "RUN-BATCH-002"],
        ["Median resolution time, billing", "<= 4.0 h", "Warehouse query, def v3",
         "warehouse.tickets", "I-MR (daily aggregate)", 2.6, 4.1, 1.1, "Daily",
         "Daily stand-up", "Billing Support Mgr", "Any point above the UCL",
         "Check queue depth and staffing for that day", "Three days in a row above",
         "Ops Director", "3 Mistake-proof it", "3 — checklist step", "SOP-BIL-014 v3"],
        ["Goodwill credit within authority", "100%", "Month-end reconciliation",
         "Finance ledger", "g-chart (contacts between breaches)", 0, 0, 0, "Monthly",
         "Month-end close", "Finance Business Partner", "Any credit above the limit",
         "Recover or approve retrospectively; coach the agent", "Two in one month",
         "Finance Director", "5 Standard work + training", "5 — nothing stops them",
         "SOP-BIL-009 v1"],
    ]
    return _rows(ws, 10, rows)


EXAMPLES = {
    "10-value-stream-map.xlsx": value_stream_map,
    "11-cause-effect-xy-matrix.xlsx": xy_matrix,
    "12-fmea.xlsx": fmea,
    "13-hypothesis-test-log.xlsx": hypothesis_log,
    "15-solution-selection-matrix.xlsx": solution_selection,
    "17-control-plan.xlsx": control_plan,
}


def add_examples(wb, wbname: str) -> int:
    fn = EXAMPLES.get(wbname)
    return fn(wb) if fn else 0
