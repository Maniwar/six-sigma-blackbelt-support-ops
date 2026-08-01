#!/usr/bin/env python3
"""Draw a workbook's charts into the on-page preview, as inline SVG.

The preview used to render cells and nothing else: `tools/preview.py` walked
rows and columns, so all 35 native Excel charts were invisible to anyone who
previewed a template rather than downloading it. Every chart fix this repo has
made — axis ranges, reference lines, anchors moved clear of the data — landed
in a file nobody could see without Excel installed.

Two things make this harder than it sounds, and both are why it is done here
rather than through openpyxl:

  * **openpyxl's reader keeps only the first sub-chart of a combo.** Five
    templates are bar+line — the Pareto's cumulative curve, the 80% rule, the
    VSM's touch-time line. Read them back through openpyxl and the line is
    simply gone. So this parses `xl/charts/chartN.xml` straight out of the zip,
    which is the same XML Excel reads and cannot silently drop anything.
  * **openpyxl stores no formula results.** Most series point at calculated
    cells, so the numbers have to come from somewhere. LibreOffice recalculates
    the workbook and writes the cached values back; `values_for()` reads those.
    It runs at build time only — the SVG is baked into previews.json, so
    nothing at runtime depends on LibreOffice being installed.

Nothing here is theme-aware on purpose: the preview modal is a white sheet
imitating Excel, and the SVG has to sit on it without looking pasted in.
"""
from __future__ import annotations

import math
import re
import shutil
import subprocess
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

C = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"
A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
PR = "{http://schemas.openxmlformats.org/package/2006/relationships}"
XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"

INK = "#151b24"
MUTED = "#5b6675"
GRID = "#e6eaf1"
AXIS = "#b9c1cd"
PALETTE = ["#1F4E79", "#B45309", "#1d6f42", "#C0392B", "#6b46c1", "#0e7490"]

FONT = "-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,Arial,sans-serif"
SVG_CSS = (
    f".ct{{font:700 14px {FONT};fill:{INK}}}"
    f".cl{{font:11.5px {FONT};fill:{MUTED}}}"
    f".cat{{font:600 11.5px {FONT};fill:{MUTED}}}"
    f".cv{{font:700 10.5px {FONT};fill:{INK}}}"
    f".cg{{stroke:{GRID};stroke-width:1}}"
    f".ca{{stroke:{AXIS};stroke-width:1.2}}"
    f".cz{{stroke:#8b95a3;stroke-width:1.4}}"
)


# ------------------------------------------------------------------ parsing


def _txt(node) -> str:
    """Every a:t run under this node, joined — titles arrive split by run."""
    if node is None:
        return ""
    return "".join(t.text or "" for t in node.iter(f"{A}t")).strip()


def _val(node, tag, attr="val", default=None):
    el = node.find(tag) if node is not None else None
    return el.get(attr) if el is not None else default


def _ref(holder):
    """The A1 range a cat/val/xVal/yVal points at, whatever flavour of ref."""
    if holder is None:
        return None
    for kind in ("numRef", "strRef", "multiLvlStrRef"):
        el = holder.find(f"{C}{kind}")
        if el is not None:
            f = el.find(f"{C}f")
            if f is not None and f.text:
                return f.text
    return None


def _line_of(sp):
    """(colour, dash, width) off an spPr's a:ln, or (None, None, None)."""
    if sp is None:
        return None, None, None
    ln = sp.find(f"{A}ln")
    if ln is None:
        return None, None, None
    clr = ln.find(f"{A}solidFill/{A}srgbClr")
    dash = ln.find(f"{A}prstDash")
    w = ln.get("w")
    return (("#" + clr.get("val")) if clr is not None else None,
            dash.get("val") if dash is not None else None,
            (int(w) / 12700.0) if w else None)


def _fill_of(sp):
    if sp is None:
        return None
    clr = sp.find(f"{A}solidFill/{A}srgbClr")
    return ("#" + clr.get("val")) if clr is not None else None


def _shows_values(node) -> bool:
    d = node.find(f"{C}dLbls") if node is not None else None
    return d is not None and _val(d, f"{C}showVal") == "1"


def _series(ser, kind: str, idx: int) -> dict:
    sp = ser.find(f"{C}spPr")
    colour, dash, width = _line_of(sp)
    name = _txt(ser.find(f"{C}tx"))
    if not name:
        v = ser.find(f"{C}tx/{C}v")
        name = (v.text or "").strip() if v is not None else ""
    if not name:
        cache = ser.find(f"{C}tx/{C}strRef/{C}strCache/{C}pt/{C}v")
        name = (cache.text or "").strip() if cache is not None else ""
    marker = _val(ser.find(f"{C}marker"), f"{C}symbol") or ("circle" if kind == "scatter" else "none")
    # Per-point colours. Several charts colour each bar to carry meaning — the
    # Kano categories, the control hierarchy, the components of variation —
    # and rendering them all one colour throws that meaning away.
    points = {}
    for dp in ser.findall(f"{C}dPt"):
        at = _val(dp, f"{C}idx")
        fill = _fill_of(dp.find(f"{C}spPr"))
        if at is not None and fill:
            points[int(at)] = fill
    return {
        "points": points,
        "kind": kind,
        "name": name,
        "fill": _fill_of(sp) or colour or PALETTE[idx % len(PALETTE)],
        "line": colour or PALETTE[idx % len(PALETTE)],
        "dash": dash,
        "width": width or 2.0,
        "marker": marker,
        "cat": _ref(ser.find(f"{C}cat")) or _ref(ser.find(f"{C}xVal")),
        "val": _ref(ser.find(f"{C}val")) or _ref(ser.find(f"{C}yVal")),
        "labels": _shows_values(ser),
    }


def _axes(plot) -> dict:
    """axId -> {title, fmt, min, max, kind, skip} for every axis in the plot."""
    out = {}
    for tag, kind in ((f"{C}valAx", "val"), (f"{C}catAx", "cat"), (f"{C}dateAx", "cat")):
        for ax in plot.findall(tag):
            scaling = ax.find(f"{C}scaling")
            out[_val(ax, f"{C}axId")] = {
                "kind": kind,
                "title": _txt(ax.find(f"{C}title")),
                "fmt": _val(ax, f"{C}numFmt", "formatCode"),
                "min": _val(scaling, f"{C}min"),
                "max": _val(scaling, f"{C}max"),
                "skip": int(_val(ax, f"{C}tickLblSkip") or 1),
                "deleted": _val(ax, f"{C}delete") == "1",
                "crosses": _val(ax, f"{C}crosses"),
            }
    return out


def parse_chart(xml: bytes) -> dict:
    """One chartN.xml as a spec the renderer can draw without touching XML."""
    root = ET.fromstring(xml)
    plot = root.find(f".//{C}plotArea")
    axes = _axes(plot)
    groups, idx = [], 0
    for sub in plot:
        tag = sub.tag.replace(C, "")
        if not tag.endswith("Chart"):
            continue
        kind = tag[:-5]                                   # bar / line / scatter
        sers = []
        for ser in sub.findall(f"{C}ser"):
            s = _series(ser, kind, idx)
            s["labels"] = s["labels"] or _shows_values(sub)
            sers.append(s)
            idx += 1
        ax_ids = [e.get("val") for e in sub.findall(f"{C}axId")]
        groups.append({
            "kind": kind,
            "series": sers,
            "dir": _val(sub, f"{C}barDir") or "col",
            "grouping": _val(sub, f"{C}grouping") or "clustered",
            "gap": int(_val(sub, f"{C}gapWidth") or 150),
            "overlap": int(_val(sub, f"{C}overlap") or 0),
            # the second axId is the value axis this group is measured against;
            # a combo puts its line group on a different one (the Pareto's
            # cumulative % against the counts), and that is the only way to know
            "val_ax": ax_ids[1] if len(ax_ids) > 1 else None,
            "cat_ax": ax_ids[0] if ax_ids else None,
        })
    return {
        "title": _txt(root.find(f".//{C}chart/{C}title")),
        "groups": groups,
        "axes": axes,
        "legend": _val(root, f".//{C}legend/{C}legendPos"),
        "has_legend": root.find(f".//{C}legend") is not None,
    }


def charts_by_sheet(path: Path) -> dict[str, list[dict]]:
    """Every chart in the workbook, keyed by the sheet it is anchored to.

    Walks workbook.xml -> sheet rels -> drawing -> drawing rels -> chartN.xml,
    which is how Excel finds them and therefore the only mapping that cannot
    disagree with the file.
    """
    z = zipfile.ZipFile(path)
    names = set(z.namelist())

    def rels(part: str) -> dict[str, str]:
        p = str(Path(part).parent / "_rels" / (Path(part).name + ".rels"))
        if p not in names:
            return {}
        base = Path(part).parent
        out = {}
        for rel in ET.fromstring(z.read(p)):
            tgt = rel.get("Target", "")
            if tgt.startswith("/"):
                out[rel.get("Id")] = tgt.lstrip("/")
            else:
                out[rel.get("Id")] = str(Path(*Path((base / tgt)).parts).as_posix())
        return out

    wb_rels = rels("xl/workbook.xml")
    found: dict[str, list[dict]] = {}
    for sh in ET.fromstring(z.read("xl/workbook.xml")).iter(
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
        target = wb_rels.get(sh.get(f"{R}id"))
        if not target:
            continue
        part = target if target.startswith("xl/") else "xl/" + target
        sheet_rels = rels(part)
        drawings = [v for v in sheet_rels.values() if "drawings/drawing" in v]
        out: list[dict] = []
        for d in drawings:
            d = d if d.startswith("xl/") else "xl/" + d
            if d not in names:
                continue
            drel = rels(d)
            # anchor order is the order they sit on the sheet, top to bottom
            for anchor in ET.fromstring(z.read(d)):
                for gf in anchor.iter(f"{C}chart"):
                    cp = drel.get(gf.get(f"{R}id"))
                    if not cp:
                        continue
                    cp = cp if cp.startswith("xl/") else "xl/" + cp
                    if cp in names:
                        out.append(parse_chart(z.read(cp)))
        if out:
            found[sh.get("name")] = out
    return found


# ------------------------------------------------------------------- values


CELL = re.compile(r"^(?:'([^']+)'|([^!']+))!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")


def _col(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch) - 64
    return n


def resolve(ref: str, cells: dict) -> list:
    """A range like 'Pareto'!$I$5:$I$14 as a flat list of cached values."""
    if not ref:
        return []
    m = CELL.match(ref.strip())
    if not m:
        return []
    sheet = m.group(1) or m.group(2)
    c1, r1 = _col(m.group(3)), int(m.group(4))
    c2, r2 = (_col(m.group(5)), int(m.group(6))) if m.group(5) else (c1, r1)
    out = []
    for r in range(min(r1, r2), max(r1, r2) + 1):
        for c in range(min(c1, c2), max(c1, c2) + 1):
            out.append(cells.get((sheet, r, c)))
    return out


SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"


def values_for(path: Path) -> dict:
    """{(sheet, row, col): value} with every formula already worked out.

    openpyxl writes no cached results, so a freshly built workbook has literally
    no numbers in it as far as a reader is concerned. LibreOffice recalculates
    on load and writes the cache back on save, which is the only way to get the
    real series values without reimplementing Excel.
    """
    from openpyxl import load_workbook

    soffice = SOFFICE if Path(SOFFICE).exists() else shutil.which("soffice")
    if not soffice:
        raise RuntimeError(
            "LibreOffice is needed to work out the chart values for the previews. "
            "Install it, or leave tools/previews.json as committed.")
    with tempfile.TemporaryDirectory() as td:
        src = Path(td) / path.name
        shutil.copyfile(path, src)
        subprocess.run(
            [soffice, "--headless", "--norestore", "--convert-to",
             "xlsx:Calc MS Excel 2007 XML", "--outdir", str(Path(td) / "out"), str(src)],
            check=True, capture_output=True, timeout=180)
        done = Path(td) / "out" / path.name
        wb = load_workbook(done, data_only=True)
        cells = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cells[(ws.title, cell.row, cell.column)] = cell.value
        return cells


# ----------------------------------------------------------------- rendering


def _num(v):
    """The value as a float, or None for text, blanks and #N/A gaps."""
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def fmt_num(v: float, code: str | None) -> str:
    code = (code or "").lower()
    if "%" in code:
        digits = 1 if ".0" in code else 0
        return f"{v * 100:.{digits}f}%"
    if "$" in code:
        return f"${v:,.0f}"
    if abs(v) >= 1000:
        return f"{v:,.0f}"
    if v == int(v):
        return str(int(v))
    return f"{v:,.2f}".rstrip("0").rstrip(".")


def nice_scale(lo: float, hi: float, want: int = 5) -> tuple[float, float, float]:
    """A tick step that lands on 1/2/2.5/5 × 10^n, and bounds snapped to it.

    A raw min/max gives axis labels like 396.4 and a plot squashed into the top
    fifth — the exact complaint the visual QA layer exists to catch.
    """
    if hi <= lo:
        hi = lo + (abs(lo) * 0.1 or 1.0)
    raw = (hi - lo) / max(1, want)
    mag = 10 ** math.floor(math.log10(raw)) if raw > 0 else 1.0
    step = 10 * mag
    for mult in (1, 2, 2.5, 5, 10):
        if mult * mag >= raw:
            step = mult * mag
            break
    return step * math.floor(lo / step), step * math.ceil(hi / step), step


def _esc(s: str) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _wrap(text: str, per_line: int) -> list[str]:
    words, lines, cur = str(text).split(), [], ""
    for w in words:
        if cur and len(cur) + 1 + len(w) > per_line:
            lines.append(cur)
            cur = w
        else:
            cur = (cur + " " + w).strip()
    if cur:
        lines.append(cur)
    return lines


def render(spec: dict, cells: dict, width: int = 900) -> str | None:
    """One chart spec as standalone SVG, or None if it has nothing to plot."""
    groups = spec["groups"]
    axes = spec["axes"]

    # ---- pull the numbers, and drop any series that turns out to be empty
    plot_groups = []
    for g in groups:
        sers = []
        for s in g["series"]:
            ys = [_num(v) for v in resolve(s["val"], cells)]
            if not any(y is not None for y in ys):
                continue
            xs = [_num(v) for v in resolve(s["cat"], cells)] if g["kind"] == "scatter" else None
            sers.append({**s, "ys": ys, "xs": xs})
        if sers:
            plot_groups.append({**g, "series": sers})
    if not plot_groups:
        return None

    scatter = plot_groups[0]["kind"] == "scatter"

    # ---- categories: the longest cat ref anywhere wins, so a reference line
    #      series with its own short range cannot shorten the axis
    cats: list[str] = []
    if not scatter:
        best = []
        for g in plot_groups:
            for s in g["series"]:
                raw = resolve(s["cat"], cells)
                if len(raw) > len(best):
                    best = raw
        for v in best:
            if v is None:
                cats.append("")
            elif isinstance(v, float) and v == int(v):
                cats.append(str(int(v)))
            else:
                cats.append(str(v))
        n = max(len(s["ys"]) for g in plot_groups for s in g["series"])
        if not cats:
            cats = [str(i + 1) for i in range(n)]
        cats = (cats + [""] * n)[:n]
        # A slot with neither a label nor a value is spare room in the template,
        # not a category — the Pareto reserves ten rows and the example fills
        # five, and drawing the empty five leaves the plot looking half-broken.
        # A labelled-but-empty slot (the ROI model's Year 4-10) is kept: there
        # the axis is telling you the model runs further than the example does.
        while cats and not cats[-1] and all(
                len(s["ys"]) < len(cats) or s["ys"][len(cats) - 1] is None
                for g in plot_groups for s in g["series"]):
            cats.pop()
        n = len(cats)
        if not n:
            return None

    horizontal = (not scatter and plot_groups[0]["kind"] == "bar"
                  and plot_groups[0]["dir"] == "bar")

    # ---- which value axis each group is measured against
    prim = [g for g in plot_groups if g["val_ax"] == plot_groups[0]["val_ax"]]
    sec = [g for g in plot_groups if g["val_ax"] != plot_groups[0]["val_ax"]]
    ax_prim = axes.get(plot_groups[0]["val_ax"], {})
    ax_sec = axes.get(sec[0]["val_ax"], {}) if sec else {}
    ax_cat = axes.get(plot_groups[0]["cat_ax"], {})

    def bounds(gs, ax):
        vals = [y for g in gs for s in g["series"] for y in s["ys"] if y is not None]
        if not vals:
            return 0.0, 1.0, 1.0
        stacked = any(g["grouping"] == "stacked" for g in gs)
        if stacked:
            n = max(len(s["ys"]) for g in gs for s in g["series"])
            tot = [sum((s["ys"][i] or 0) for g in gs for s in g["series"]
                       if i < len(s["ys"])) for i in range(n)]
            vals = vals + tot
        lo, hi = min(vals), max(vals)
        lo = 0.0 if lo >= 0 and lo <= hi * 0.55 else lo
        hi = 0.0 if hi <= 0 and hi >= lo * 0.55 else hi
        lo, hi, step = nice_scale(lo, hi)
        if ax.get("min") is not None:
            lo = float(ax["min"])
        if ax.get("max") is not None:
            hi = float(ax["max"])
            step = (hi - lo) / 5 if hi > lo else step
        return lo, hi, step

    lo1, hi1, st1 = bounds(prim, ax_prim)
    lo2, hi2, st2 = bounds(sec, ax_sec) if sec else (0, 1, 1)

    # ---- geometry. Everything that needs room asks for it before we lay out,
    #      and the three things that live under the plot — category labels, the
    #      axis title, the legend — get their own bands. Overlapping them is
    #      exactly the "elements being covered up" defect the visual QA caught
    #      on the sheets, and it is no more acceptable here.
    title = spec["title"]
    title_lines = _wrap(title, 72) if title else []
    # A series whose every bar carries its own colour has no single swatch to
    # show, and drawing one in the series colour claims a colour no bar uses.
    # The categories are labelled on the axis, so the entry adds nothing.
    for g in plot_groups:
        for s in g["series"]:
            if s["kind"] == "bar" and s["points"] and len(s["points"]) >= len(
                    [y for y in s["ys"] if y is not None]):
                s["name"] = ""
    legend_names = [s["name"] for g in plot_groups for s in g["series"] if s["name"]]
    show_legend = spec["has_legend"] and len(legend_names) > 1
    legend_lines = _wrap(" · ".join(legend_names), 84) if show_legend else []

    ylab = [fmt_num(lo1 + i * st1, ax_prim.get("fmt"))
            for i in range(int(round((hi1 - lo1) / st1)) + 1)] if st1 else ["0"]
    pad_l = 14 + max(28, 7.0 * max(len(t) for t in ylab))
    if ax_prim.get("title"):
        pad_l += 16
    # A horizontal chart's last value label is centred on the axis end, so half
    # of it hangs past the plot — "400,000" came out as "400,00(".
    pad_r = 34 if horizontal else 18
    if sec:
        ylab2 = [fmt_num(lo2 + i * st2, ax_sec.get("fmt"))
                 for i in range(int(round((hi2 - lo2) / st2)) + 1)] if st2 else ["0"]
        pad_r = 16 + max(28, 7.0 * max(len(t) for t in ylab2)) + (16 if ax_sec.get("title") else 0)

    top = 16 + 20 * len(title_lines)
    CAT_CAP = 26
    rotate = False
    # The workbook's own tickLblSkip is sized for the range the chart reserves
    # (the VSM books 19 rows), not for the seven the example fills, so trusting
    # it here would drop two labels in three off a chart with ample room. The
    # preview knows exactly how many categories it is about to draw and how
    # wide they are, so it works the skip out from that instead.
    if not scatter and cats:
        est_band = max(1.0, (width - pad_l - pad_r) / len(cats))
        longest0 = min(CAT_CAP, max((len(c) for c in cats), default=4))
        need = 15.0 if longest0 > 7 else 6.2 * longest0 + 8
        skip = max(1, math.ceil(need / est_band))
    else:
        skip = 1
    if horizontal:
        pad_l = 14 + max(60, min(190, 6.6 * max((len(c) for c in cats), default=6)))
        cat_h = 26
    elif scatter:
        cat_h = 26
    else:
        longest = min(CAT_CAP, max((len(c) for i, c in enumerate(cats)
                                    if i % skip == 0), default=4))
        rotate = longest > 7
        # A -38° label anchored at its end runs up and to the LEFT of its tick,
        # so the first category overhangs the value axis by cos(38°) × its
        # width. Without paying for that here it is clipped by the viewBox,
        # which is how the Pareto shipped reading "stment not posted at c".
        if rotate:
            cat_h = 14 + math.sin(math.radians(38)) * 6.2 * longest
            pad_l = max(pad_l, min(165, math.cos(math.radians(38)) * 6.2 * longest + 8))
        else:
            cat_h = 26
    axt_h = 18 if ax_cat.get("title") else 0
    leg_h = (20 * len(legend_lines) + 8) if legend_lines else 0
    bottom = cat_h + axt_h + leg_h

    height = max(300, int(top + bottom + (26 * len(cats) if horizontal else 240)))
    height = min(height, 780)
    pw = width - pad_l - pad_r
    ph = height - top - bottom
    if pw < 80 or ph < 80:
        return None

    o: list[str] = []

    def sy(v, lo, hi):
        return top + ph - (v - lo) / (hi - lo) * ph if hi > lo else top + ph

    def sx_val(v, lo, hi):
        return pad_l + (v - lo) / (hi - lo) * pw if hi > lo else pad_l

    # ---- title
    for i, line in enumerate(title_lines):
        o.append(f'<text x="{pad_l}" y="{18 + i * 20}" class="ct">{_esc(line)}</text>')

    # ---- gridlines and the value scale
    n_ticks = int(round((hi1 - lo1) / st1)) if st1 else 1
    for i in range(n_ticks + 1):
        v = lo1 + i * st1
        if horizontal:
            x = sx_val(v, lo1, hi1)
            o.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{top + ph}" class="cg"/>')
            o.append(f'<text x="{x:.1f}" y="{top + ph + 16}" class="cl" text-anchor="middle">'
                     f'{_esc(fmt_num(v, ax_prim.get("fmt")))}</text>')
        else:
            y = sy(v, lo1, hi1)
            o.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{pad_l + pw}" y2="{y:.1f}" class="cg"/>')
            o.append(f'<text x="{pad_l - 8}" y="{y + 4:.1f}" class="cl" text-anchor="end">'
                     f'{_esc(fmt_num(v, ax_prim.get("fmt")))}</text>')
    if sec:
        n2 = int(round((hi2 - lo2) / st2)) if st2 else 1
        for i in range(n2 + 1):
            v = lo2 + i * st2
            y = sy(v, lo2, hi2)
            o.append(f'<text x="{pad_l + pw + 8}" y="{y + 4:.1f}" class="cl">'
                     f'{_esc(fmt_num(v, ax_sec.get("fmt")))}</text>')

    # ---- axis frame
    o.append(f'<line x1="{pad_l}" y1="{top}" x2="{pad_l}" y2="{top + ph}" class="ca"/>')
    o.append(f'<line x1="{pad_l}" y1="{top + ph}" x2="{pad_l + pw}" y2="{top + ph}" class="ca"/>')
    if sec:
        o.append(f'<line x1="{pad_l + pw}" y1="{top}" x2="{pad_l + pw}" y2="{top + ph}" class="ca"/>')
    if lo1 < 0 < hi1:
        if horizontal:
            zx = sx_val(0, lo1, hi1)
            o.append(f'<line x1="{zx:.1f}" y1="{top}" x2="{zx:.1f}" y2="{top + ph}" class="cz"/>')
        else:
            zy = sy(0, lo1, hi1)
            o.append(f'<line x1="{pad_l}" y1="{zy:.1f}" x2="{pad_l + pw}" y2="{zy:.1f}" class="cz"/>')

    # ---- axis titles
    if ax_prim.get("title"):
        cx, cy = pad_l - (pad_l - 16), top + ph / 2
        o.append(f'<text transform="translate(14,{cy:.1f}) rotate(-90)" class="cat" '
                 f'text-anchor="middle">{_esc(ax_prim["title"])}</text>')
    if sec and ax_sec.get("title"):
        o.append(f'<text transform="translate({width - 6},{top + ph / 2:.1f}) rotate(-90)" '
                 f'class="cat" text-anchor="middle">{_esc(ax_sec["title"])}</text>')

    # ---- categories
    n = len(cats) if not scatter else 0
    band = pw / n if n else pw
    if not scatter and not horizontal:
        base = top + ph
        for i, c in enumerate(cats):
            if not c or i % skip:
                continue
            x = pad_l + band * (i + 0.5)
            if rotate:
                o.append(f'<text transform="translate({x:.1f},{base + 13}) rotate(-38)" '
                         f'class="cl" text-anchor="end">{_esc(c[:CAT_CAP])}</text>')
            else:
                o.append(f'<text x="{x:.1f}" y="{base + 17}" class="cl" '
                         f'text-anchor="middle">{_esc(c[:14])}</text>')
    elif horizontal:
        for i, c in enumerate(cats):
            if not c:
                continue
            y = top + (ph / n) * (i + 0.5) + 4
            o.append(f'<text x="{pad_l - 8}" y="{y:.1f}" class="cl" text-anchor="end">'
                     f'{_esc(c[:30])}</text>')
    if ax_cat.get("title"):
        o.append(f'<text x="{pad_l + pw / 2:.1f}" y="{top + ph + cat_h + 12:.1f}" '
                 f'class="cat" text-anchor="middle">{_esc(ax_cat["title"])}</text>')

    # ---- the series themselves
    bar_groups = [g for g in plot_groups if g["kind"] == "bar"]
    n_bar = sum(len(g["series"]) for g in bar_groups
                if g["grouping"] != "stacked") or 1
    slot = 0
    for g in plot_groups:
        lo, hi = (lo2, hi2) if (sec and g in sec) else (lo1, hi1)
        if g["kind"] == "bar":
            gapf = min(0.6, g["gap"] / 100.0 / 2)
            usable = band * (1 - gapf)
            stacked = g["grouping"] == "stacked"
            run = [0.0] * (n + 1)
            for s in g["series"]:
                w = usable if stacked else usable / n_bar
                for i, y in enumerate(s["ys"][:n]):
                    if y is None:
                        continue
                    if horizontal:
                        bh = (ph / n) * (1 - gapf) / (1 if stacked else n_bar)
                        y0 = top + (ph / n) * i + (ph / n) * gapf / 2 + (0 if stacked else bh * slot)
                        # The baseline is zero clamped into the axis, not
                        # min(0, y) — with an all-negative scale that made every
                        # bar start at its own value and draw zero pixels wide,
                        # which is how the DOE effects chart came out blank.
                        x0 = sx_val(max(lo, min(0, hi)), lo, hi)
                        x1 = sx_val(y, lo, hi)
                        o.append(f'<rect x="{min(x0, x1):.1f}" y="{y0:.1f}" '
                                 f'width="{abs(x1 - x0):.1f}" height="{bh:.1f}" '
                                 f'fill="{s["points"].get(i, s["fill"])}" rx="1"/>')
                        if s["labels"]:
                            lx, anc = ((max(x0, x1) + 5, "start") if y >= 0
                                       else (min(x0, x1) - 5, "end"))
                            o.append(f'<text x="{lx:.1f}" y="{y0 + bh / 2 + 4:.1f}" '
                                     f'class="cv" text-anchor="{anc}">'
                                     f'{_esc(fmt_num(y, ax_prim.get("fmt")))}</text>')
                    else:
                        x0 = pad_l + band * i + band * gapf / 2 + (0 if stacked else w * slot)
                        yb = sy(run[i] if stacked else max(lo, min(0, hi)), lo, hi)
                        yt = sy((run[i] + y) if stacked else y, lo, hi)
                        o.append(f'<rect x="{x0:.1f}" y="{min(yb, yt):.1f}" width="{w:.1f}" '
                                 f'height="{max(1.0, abs(yt - yb)):.1f}" '
                                 f'fill="{s["points"].get(i, s["fill"])}" rx="1"/>')
                        if s["labels"] and not stacked:
                            o.append(f'<text x="{x0 + w / 2:.1f}" y="{min(yb, yt) - 5:.1f}" '
                                     f'class="cv" text-anchor="middle">'
                                     f'{_esc(fmt_num(y, ax_prim.get("fmt")))}</text>')
                        if stacked:
                            run[i] += y
                if not stacked:
                    slot += 1
        elif g["kind"] == "line":
            for s in g["series"]:
                # a None is a real gap — NA() is how these workbooks stop a
                # series being drawn across a period it has no data for
                runs, cur = [], []
                for i, y in enumerate(s["ys"][:n]):
                    if y is None:
                        if cur:
                            runs.append(cur)
                        cur = []
                        continue
                    # On a horizontal bar chart the categories run down the
                    # side, so a combo line has to be transposed with them —
                    # the DOE's noise floor is a vertical rule, not a
                    # horizontal one laid across the top bar.
                    if horizontal:
                        cur.append((sx_val(y, lo, hi), top + (ph / n) * (i + 0.5)))
                        continue
                    cur.append((pad_l + band * (i + 0.5), sy(y, lo, hi)))
                if cur:
                    runs.append(cur)
                dash = ' stroke-dasharray="7 5"' if s["dash"] in ("dash", "sysDash", "lgDash") else \
                       (' stroke-dasharray="2 4"' if s["dash"] in ("sysDot", "dot") else "")
                for pts in runs:
                    if len(pts) == 1:
                        x, y = pts[0]
                        o.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{s["line"]}"/>')
                        continue
                    d = " ".join(f"{'M' if i == 0 else 'L'}{x:.1f},{y:.1f}"
                                 for i, (x, y) in enumerate(pts))
                    o.append(f'<path d="{d}" fill="none" stroke="{s["line"]}" '
                             f'stroke-width="{s["width"]:.1f}" stroke-linejoin="round" '
                             f'stroke-linecap="round"{dash}/>')
                if s["marker"] not in ("none", None):
                    for pts in runs:
                        for x, y in pts:
                            o.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" '
                                     f'fill="#fff" stroke="{s["line"]}" stroke-width="1.6"/>')
        elif g["kind"] == "scatter":
            xs_all = [x for s in g["series"] for x in (s["xs"] or []) if x is not None]
            xlo, xhi, xst = nice_scale(min(xs_all), max(xs_all)) if xs_all else (0, 1, 1)
            for i in range(int(round((xhi - xlo) / xst)) + 1):
                v = xlo + i * xst
                x = sx_val(v, xlo, xhi)
                o.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{top + ph}" class="cg"/>')
                o.append(f'<text x="{x:.1f}" y="{top + ph + 16}" class="cl" '
                         f'text-anchor="middle">{_esc(fmt_num(v, ax_cat.get("fmt")))}</text>')
            for s in g["series"]:
                pts = [(sx_val(x, xlo, xhi), sy(y, lo, hi))
                       for x, y in zip(s["xs"] or [], s["ys"])
                       if x is not None and y is not None]
                if s["marker"] == "none":
                    # A scatter series with no marker is a reference line drawn
                    # as two endpoints — the stakeholder map's influence and
                    # support thresholds. Plotting them as dots put four stray
                    # red points in the middle of the quadrants and drew no
                    # quadrants at all.
                    dash = ' stroke-dasharray="7 5"' if s["dash"] in (
                        "dash", "sysDash", "lgDash") else ""
                    d = " ".join(f"{'M' if i == 0 else 'L'}{x:.1f},{y:.1f}"
                                 for i, (x, y) in enumerate(sorted(pts)))
                    if len(pts) > 1:
                        o.append(f'<path d="{d}" fill="none" stroke="{s["line"]}" '
                                 f'stroke-width="{s["width"]:.1f}"{dash}/>')
                    continue
                for x, y in pts:
                    o.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" '
                             f'r="4.5" fill="{s["fill"]}" fill-opacity=".78" '
                             f'stroke="{s["line"]}" stroke-width="1"/>')

    # ---- legend, as swatch + name so a reference line reads as a reference line
    if legend_lines:
        y = top + ph + cat_h + axt_h + 16
        x = pad_l
        for g in plot_groups:
            for s in g["series"]:
                if not s["name"]:
                    continue
                w = 13 + 6.6 * len(s["name"]) + 16
                if x + w > width - 4 and x > pad_l:
                    x = pad_l
                    y += 18
                if s["kind"] == "bar":
                    o.append(f'<rect x="{x:.1f}" y="{y - 8}" width="11" height="11" '
                             f'fill="{s["fill"]}" rx="2"/>')
                else:
                    dash = ' stroke-dasharray="5 4"' if s["dash"] in ("dash", "sysDash", "lgDash") else ""
                    o.append(f'<line x1="{x:.1f}" y1="{y - 3}" x2="{x + 12:.1f}" y2="{y - 3}" '
                             f'stroke="{s["line"]}" stroke-width="2.6"{dash}/>')
                o.append(f'<text x="{x + 17:.1f}" y="{y + 1}" class="cl">{_esc(s["name"])}</text>')
                x += w

    body = "".join(o)
    # The style rides inside the SVG rather than in the page stylesheet, so the
    # chart still looks like itself when the preview is saved as a standalone
    # page or dropped into anything else that does not carry our CSS.
    return (f'<svg class="xchart" xmlns="http://www.w3.org/2000/svg" '
            f'viewBox="0 0 {width} {height}" role="img" '
            f'aria-label="{_esc(title or "chart")}" '
            f'preserveAspectRatio="xMidYMid meet"><style>{SVG_CSS}</style>{body}</svg>')


def sheet_charts_html(sheet: str, by_sheet: dict, cells: dict) -> str:
    """Every chart on one sheet, wrapped so the preview can lay them out."""
    out = []
    for spec in by_sheet.get(sheet, []):
        svg = render(spec, cells)
        if svg:
            out.append(f'<figure class="xchartbox">{svg}</figure>')
    if not out:
        return ""
    return ('<div class="xcharts"><div class="xchartshead">This sheet&rsquo;s chart, '
            'drawn from the cells above &mdash; live in the .xlsx, so it moves when '
            'you change one</div>' + "".join(out) + "</div>")


# --------------------------------------------------------------------- cache


ROOT = Path(__file__).resolve().parent.parent
CACHE = ROOT / "tools" / "chartcache.json"


def _digest(path: Path) -> str:
    """The workbook's bytes AND this renderer's, so editing either misses.

    Keying on the workbook alone would serve a stale chart for every template
    the moment the drawing code changed — the cache would quietly outvote the
    fix. Hashing the module source in means any edit here invalidates the lot.
    """
    import hashlib
    h = hashlib.sha256(path.read_bytes())
    h.update(Path(__file__).read_bytes())
    return h.hexdigest()[:32]


def charts_html(path: Path) -> dict[str, str]:
    """{sheet name: chart markup} for one workbook, keyed off its bytes.

    Rendering needs LibreOffice to work the formulas out, which is a heavy
    dependency to put in the path of every HTML sync. The cache is committed and
    keyed on the workbook's own digest, so a machine without LibreOffice can
    still rebuild the page, and a workbook that did not change is never
    recalculated. Change the workbook and the key misses, which is the point.
    """
    import json
    key = _digest(path)
    cache = {}
    if CACHE.exists():
        cache = json.loads(CACHE.read_text(encoding="utf-8"))
    if key in cache:
        return cache[key]
    by_sheet = charts_by_sheet(path)
    out: dict[str, str] = {}
    if by_sheet:
        cells = values_for(path)
        for sheet in by_sheet:
            html = sheet_charts_html(sheet, by_sheet, cells)
            if html:
                out[sheet] = html
    cache[key] = out
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=0, sort_keys=True),
                     encoding="utf-8")
    return out


def prune_cache(paths) -> int:
    """Drop entries for workbooks that no longer exist at that digest."""
    import json
    if not CACHE.exists():
        return 0
    cache = json.loads(CACHE.read_text(encoding="utf-8"))
    live = {_digest(p) for p in paths}
    dead = [k for k in cache if k not in live]
    for k in dead:
        del cache[k]
    if dead:
        CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=0, sort_keys=True),
                         encoding="utf-8")
    return len(dead)
