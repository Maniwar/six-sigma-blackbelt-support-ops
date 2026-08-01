#!/usr/bin/env python3
"""Propagate templates/ into the single-file HTML, then into docs/.

Every workbook exists in four places: templates/*.xlsx, the base64 copy inside
the HTML, the preview table's formula tooltips, and docs/index.html. Only the
first is edited by hand; this script derives the other three, so a formula fix
cannot land in one copy and miss the rest.

What it rewrites inside `const TPLS = {...}`:
  * b64      - re-read from templates/*.xlsx
  * content  - re-read from templates/*.md
  * preview  - each <td>'s title tooltip is set from the formula in the matching
               workbook cell, and text cells are re-read from the workbook.
               Cell position is recovered by accumulating colspan across each
               row, which reproduces the original generator exactly (verified
               against all 233 formulas before any change was made).

    python3 tools/patch_workbooks.py
    python3 tools/sync_html.py
    python3 tools/verify.py

Requires: openpyxl
"""
from __future__ import annotations

import base64
import html as H
import json
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import chartsvg                                                  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
HTML = ROOT / "six-sigma-blackbelt-support-ops.html"
DOCS = ROOT / "docs" / "index.html"
TEMPLATES = ROOT / "templates"

# A sheet's cells end at its `</table>`, NOT at `</table></div>` — charts now
# sit between those two tags, and a pattern that reached past the table walked
# the next sheet's rows against this sheet's cells. tools/verify.py imports this
# and made the same walk, so both were wrong in the same way.
RE_SHEET = re.compile(
    r'(<div class="xsheet[^"]*" data-xsheet="(\d+)"><table class="xgrid">)(.*?)(</table>)', re.S
)
# Charts injected on a previous run, so a re-sync replaces them rather than
# stacking a second copy underneath the first.
# `</figure></div>` occurs exactly once per block — between two figures the
# markup reads `</figure><figure`, so the non-greedy match cannot stop early.
RE_CHARTS = re.compile(r'<div class="xcharts">.*?</figure></div>', re.S)
# The end of a sheet is its table plus whatever explanatory note follows it,
# then the closing div. Matching `</table></div>` alone silently stopped
# matching the moment a note was added after the table, and the charts vanished
# from those previews rather than moving.
RE_SHEET_END = re.compile(r"(</table>(?:<p class=\"xscaf\">.*?</p>)?)</div>", re.S)

# Previews that tools/build_templates.py generates from the workbook. For these
# the generated markup IS the preview; for the older workbooks there is no
# generator and the shipped markup is patched in place instead.
_PREVIEWS = ROOT / "tools" / "previews.json"
GENERATED: dict[str, str] = (
    json.loads(_PREVIEWS.read_text(encoding="utf-8")) if _PREVIEWS.exists() else {})
RE_ROW = re.compile(r"(<tr>)(.*?)(</tr>)", re.S)
RE_TD = re.compile(r"<td([^>]*)>(.*?)</td>", re.S)

# Cells whose rendered text the workbook cannot supply on its own: openpyxl
# stores no cached results, so any *numeric* cell added or changed after the
# previews were generated needs its display value stated here.
TEXT_OVERRIDES: dict[tuple[str, str, str], str] = {
    ("13-hypothesis-test-log.xlsx", "Test log", "B8"): "0.05",
}

# Classes for cells that did not exist when the previews were generated.
CLASS_OVERRIDES: dict[tuple[str, str, str], str] = {
    ("13-hypothesis-test-log.xlsx", "Test log", "B8"): "x-fill",
    ("13-hypothesis-test-log.xlsx", "Test log", "A8"): "x-b",
}


def extract_tpls(src: str) -> tuple[int, int, dict]:
    """Return (start, end, parsed) for the `const TPLS = {...}` object literal."""
    i = src.index("const TPLS=")
    start = src.index("{", i)
    depth = 0
    in_str = False
    esc = False
    for j in range(start, len(src)):
        ch = src[j]
        if esc:
            esc = False
            continue
        if ch == "\\" and in_str:
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return start, j + 1, json.loads(src[start : j + 1])
    raise SystemExit("could not find the end of const TPLS")


def _set_attr(attrs: str, name: str, value: str) -> str:
    """Replace an attribute's value in place, or append it, preserving order."""
    pat = re.compile(r'(\s%s=")([^"]*)(")' % name)
    if pat.search(attrs):
        return pat.sub(lambda m: m.group(1) + value + m.group(3), attrs, count=1)
    return attrs + ' %s="%s"' % (name, value)


def generated_preview(wbpath: Path) -> str | None:
    """Build this workbook's preview from the workbook, values and all.

    Same call the built pack has always used. Returns None if it cannot be
    produced, so the caller falls back to editing the existing markup rather
    than shipping a blank sheet.
    """
    from preview import shown_from, workbook_html
    import chartsvg
    try:
        wb = load_workbook(wbpath)
        try:
            shown = shown_from(wb, chartsvg.values_for(wbpath))
        except Exception:                                        # noqa: BLE001
            shown = {}                    # no recalculation available
        return workbook_html(wb, shown)
    except Exception as exc:                                     # noqa: BLE001
        print(f"    (could not generate preview for {wbpath.name}: {exc})")
        return None


def rebuild_preview(preview: str, wbpath: Path, stats: dict) -> str:
    wb = load_workbook(wbpath)
    fname = wbpath.name

    def do_sheet(m: re.Match) -> str:
        idx = int(m.group(2))
        ws = wb.worksheets[idx]
        row_no = [0]
        # Columns claimed by a rowspan started in an earlier row. Without this
        # the walk shifts left on every row under a vertical merge, and the
        # fishbone's bones are vertical merges.
        held: dict[int, set[int]] = {}

        def do_row(rm: re.Match) -> str:
            row_no[0] += 1
            r = row_no[0]
            col = [1]

            def do_td(tm: re.Match) -> str:
                attrs, inner = tm.group(1), tm.group(2)
                cs = re.search(r'colspan="(\d+)"', attrs)
                rs = re.search(r'rowspan="(\d+)"', attrs)
                nc = int(cs.group(1)) if cs else 1
                nr = int(rs.group(1)) if rs else 1
                while col[0] in held.get(r, ()):
                    col[0] += 1
                c = col[0]
                col[0] += nc
                if nr > 1:
                    for rr in range(r + 1, r + nr):
                        held.setdefault(rr, set()).update(range(c, c + nc))
                cell = ws.cell(row=r, column=c)
                val = cell.value
                key = (fname, ws.title, cell.coordinate)

                if isinstance(val, str) and val.startswith("="):
                    esc_f = H.escape(val, quote=True)
                    if 'title="' not in attrs:
                        stats["title_added"] += 1
                        cls = re.search(r'class="([^"]*)"', attrs)
                        newcls = (cls.group(1) + " x-f") if cls else "x-calc x-f"
                        attrs = _set_attr(attrs, "class", newcls)
                    elif H.unescape(re.search(r'title="([^"]*)"', attrs).group(1)) != val:
                        stats["title_changed"] += 1
                    attrs = _set_attr(attrs, "title", esc_f)
                elif 'title="' in attrs:
                    # This cell used to hold a formula and no longer does. The
                    # tooltip has to go, or it survives every future sync and
                    # describes a cell somewhere else entirely.
                    stats["title_changed"] += 1
                    attrs = re.sub(r'\s*title="[^"]*"', "", attrs)
                    attrs = _set_attr(attrs, "class",
                                      (re.search(r'class="([^"]*)"', attrs).group(1)
                                       if re.search(r'class="([^"]*)"', attrs) else "").replace("x-f", "").strip())
                    if isinstance(val, str):
                        inner = H.escape(val, quote=True)

                if isinstance(val, str) and not val.startswith("="):
                    want = H.escape(val, quote=True)
                    if want != inner:
                        stats["text_changed"] += 1
                        inner = want
                elif key in TEXT_OVERRIDES:
                    want = TEXT_OVERRIDES[key]
                    if want != inner:
                        stats["text_changed"] += 1
                        inner = want

                if key in CLASS_OVERRIDES:
                    attrs = _set_attr(attrs, "class", CLASS_OVERRIDES[key])
                return "<td%s>%s</td>" % (attrs, inner)

            return rm.group(1) + RE_TD.sub(do_td, rm.group(2)) + rm.group(3)

        return m.group(1) + RE_ROW.sub(do_row, m.group(3)) + m.group(4)

    return RE_SHEET.sub(do_sheet, preview)


def inject_charts(preview: str, wbpath: Path, stats: dict) -> str:
    """Put each sheet's charts under its table.

    The preview rendered cells and nothing else, so every native Excel chart in
    the pack — and every fix made to one — was invisible to anyone who opened
    the preview instead of downloading the file. They go in as inline SVG, drawn
    from the same cells the table above shows.
    """
    by_sheet = chartsvg.charts_html(wbpath)
    if not by_sheet:
        return preview
    order = load_workbook(wbpath, read_only=True).sheetnames
    seen = [0]

    def repl(m: re.Match) -> str:
        i = seen[0]
        seen[0] += 1
        html = by_sheet.get(order[i], "") if i < len(order) else ""
        stats["charts"] += html.count('<svg class="xchart"')
        return m.group(1) + html + "</div>"

    return RE_SHEET_END.sub(repl, preview)


def main() -> int:
    src = HTML.read_text(encoding="utf-8")
    start, end, tpls = extract_tpls(src)
    stats = {"title_changed": 0, "title_added": 0, "text_changed": 0, "b64": 0,
             "md": 0, "charts": 0, "regenerated": 0}

    for slug, entry in tpls.items():
        path = TEMPLATES / entry["file"]
        if not path.exists():
            raise SystemExit(f"missing template: {path}")
        if entry.get("ext") == "xlsx":
            new_b64 = base64.b64encode(path.read_bytes()).decode("ascii")
            if new_b64 != entry.get("b64"):
                entry["b64"] = new_b64
                stats["b64"] += 1
            # Charts come off FIRST. rebuild_preview walks to `</table></div>`
            # to find the end of a sheet, and a chart block sitting between
            # those two tags makes that walk run on into the next sheet.
            bare = RE_CHARTS.sub("", entry["preview"])
            if entry["file"] in GENERATED:
                # tools/build_templates.py regenerates these previews from the
                # workbook itself on every build — and until now nothing carried
                # the result into the page, so a preview could only ever be
                # updated in the cells rebuild_preview happens to rewrite. Any
                # new row, column or computed value was written to
                # previews.json and then quietly dropped. Take it wholesale.
                if bare != GENERATED[entry["file"]]:
                    stats["regenerated"] += 1
                bare = GENERATED[entry["file"]]
            else:
                # The other nineteen carried hand-written markup that
                # rebuild_preview could only ever edit in place: it walks the
                # <td>s that are already there, so it can correct a value but
                # cannot add a row. Thirty-one pieces of prose existed in a
                # workbook and appeared nowhere in its preview — including all
                # six definitions of the hierarchy of controls and nineteen of
                # the notes just written into the calculator pack. Generating
                # the preview from the workbook, exactly as the built pack has
                # always done, makes the two agree by construction instead of
                # by whoever remembered to update both.
                bare = generated_preview(path) or rebuild_preview(bare, path, stats)
            entry["preview"] = inject_charts(bare, path, stats)
        else:
            text = path.read_text(encoding="utf-8")
            if text != entry.get("content"):
                entry["content"] = text
                stats["md"] += 1

    payload = json.dumps(tpls, ensure_ascii=False)
    if "</script" in payload.lower():
        raise SystemExit("template content contains </script> and would break the page")

    out = src[:start] + payload + src[end:]
    HTML.write_text(out, encoding="utf-8")
    shutil.copyfile(HTML, DOCS)

    print(f"  workbooks re-embedded : {stats['b64']}")
    print(f"  markdown re-embedded  : {stats['md']}")
    print(f"  tooltips rewritten    : {stats['title_changed']} (added {stats['title_added']})")
    print(f"  preview text updated  : {stats['text_changed']}")
    print(f"  previews regenerated  : {stats['regenerated']}")
    print(f"  charts drawn in preview: {stats['charts']}"
          f" (cache pruned {chartsvg.prune_cache(sorted(TEMPLATES.glob('*.xlsx')))})")
    print(f"  docs/index.html synced ({DOCS.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
