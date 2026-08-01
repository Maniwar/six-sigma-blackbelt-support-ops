#!/usr/bin/env python3
"""Charts for the eight original workbooks, which shipped with none.

Every one of these files computed the right answer and then showed it as a
column of numbers. An FMEA whose RPNs you cannot see ranked, a value stream map
that never shows you that 98% of the lead time is waiting, an ROI tab with no
breakeven line — the arithmetic was there and the point of it was not.

Each spec writes a small block of chart-feeding cells (blue, labelled, off to
the side of the working area) and binds a native Excel chart to it. Nothing is
hard-coded: change a yellow cell and the chart moves.

Called by tools/patch_workbooks.py. Idempotent — the cell writes only count as
changes when a value actually differs, so repeated runs do not churn the
embedded base64 in the HTML.
"""
from __future__ import annotations

import pathlib

from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.packaging.custom import StringProperty
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CALC = PatternFill("solid", fgColor="FFF2F7FF")
IN = PatternFill("solid", fgColor="FFFFF9E3")   # "you fill this in"
F_MICRO = Font(italic=True, size=9, color="FF6B7280")
F_B = Font(bold=True, size=10)

BLUE, GREEN, RED, PURPLE, AMBER = "1F4E79", "3F8F5A", "C0392B", "6B4FA0", "B45309"


class Sheet:
    """Writes chart-feeding cells and counts only the ones that really changed."""

    def __init__(self, ws):
        self.ws = ws
        self.changed = 0

    def put(self, row, col, value, fmt=None, bold=False, note=False):
        c = self.ws.cell(row=row, column=col)
        if c.__class__.__name__ == "MergedCell":
            owner = next((str(m) for m in self.ws.merged_cells.ranges
                          if (row, col) in [(r, cc) for r, cc in
                                            [(rr, ccc) for rr in range(m.min_row, m.max_row + 1)
                                             for ccc in range(m.min_col, m.max_col + 1)]]), "?")
            raise ValueError(
                f"{self.ws.title!r} row {row} col {col} is inside merged range {owner} — "
                "pick a free block for the chart data")
        if c.value != value:
            c.value = value
            self.changed += 1
        if fmt:
            c.number_format = fmt
        if bold:
            c.font = F_B
        if note:
            c.font = F_MICRO
        elif not bold and isinstance(value, str) and value.startswith("="):
            c.fill = CALC
        return c

    def label(self, row, col, text):
        c = self.put(row, col, text, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        return c


def _bar(ws, title, cats, vals, anchor, colours=None, horizontal=False,
         fmt=None, height=7.5, width=15, labels=True, gap=60, name=None):
    ch = BarChart()
    ch.type = "bar" if horizontal else "col"
    ch.style = 10
    ch.title = title
    ch.legend = None
    ch.height, ch.width = height, width
    ch.gapWidth = gap
    ch.add_data(vals, titles_from_data=False)
    ch.set_categories(cats)
    if name and ch.series:
        # An unnamed series shows as "Series1" the moment anything turns the
        # legend on — which laying a reference line over it does.
        ch.series[0].tx = SeriesLabel(v=name)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    if fmt:
        ch.y_axis.numFmt = fmt
    if labels:
        ch.dataLabels = DataLabelList()
        ch.dataLabels.showVal = True
        ch.dataLabels.numFmt = fmt or "#,##0"
        # without these the label reads "YOU; Column B; 50,667" — the category
        # is already on the axis and the series name is meaningless here
        ch.dataLabels.showCatName = False
        ch.dataLabels.showSerName = False
        ch.dataLabels.showLegendKey = False
        ch.dataLabels.showPercent = False
        ch.dataLabels.showBubbleSize = False
    if colours:
        # per-bar colour, so "you" reads differently from the benchmark bars
        from openpyxl.chart.marker import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties
        ch.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(solidFill=col))
            for i, col in enumerate(colours)]
    elif ch.series:
        ch.series[0].graphicalProperties.solidFill = BLUE
        # OOXML's invertIfNegative defaults on, and with no negative fill
        # defined the renderer draws negative bars as nothing at all. The DOE
        # effects chart is entirely negative values, so it was drawing an empty
        # plot area with a correctly scaled axis beside it.
    for ser in ch.series:
        ser.invertIfNegative = False
    ws.add_chart(ch, anchor)
    return ch


def ranked_block(sh, first, last, sort_col, carry, dest, label, fmts=None):
    """Sort a scoring table by value, in formulas, with no macro and no re-typing.

    A chart titled "the tall bars are where to look" that plots rows in the
    order somebody typed them makes the reader do the ranking the chart was
    supposed to do. Worse on an FMEA, where the whole method is "work the
    highest RPN first" — burying it in row four defeats the tool.

    Writes a tie-broken rank beside the source table, then resolves each
    position with INDEX/MATCH into a block the chart reads. Re-ranks itself the
    moment a score changes.

    sort_col : column holding the value to sort by, descending
    carry    : columns to carry across, in chart order (first is the category)
    dest     : first column of the ranked block
    """
    L = get_column_letter
    rank_col = dest - 1
    sc = L(sort_col)
    for r in range(first, last + 1):
        c = sh.ws.cell(row=r, column=rank_col)
        c.value = (f'=IF({sc}{r}="","",RANK({sc}{r},${sc}${first}:${sc}${last},0)'
                   f'+COUNTIF(${sc}${first}:{sc}{r},{sc}{r})-1)')
        c.number_format = "0"
        c.font = F_MICRO
    sh.put(first - 1, rank_col, "rank", note=True)
    sh.put(first - 1, dest, "#", bold=True)
    for i, src in enumerate(carry):
        sh.put(first - 1, dest + 1 + i, label[i] if i < len(label) else "", bold=True)
    rk = L(rank_col)
    for i, r in enumerate(range(first, last + 1)):
        sh.put(r, dest, i + 1, fmt="0")
        for j, src in enumerate(carry):
            col = dest + 1 + j
            sh.put(r, col,
                   f'=IFERROR(INDEX(${L(src)}${first}:${L(src)}${last},'
                   f'MATCH({L(dest)}{r},${rk}${first}:${rk}${last},0)),"")',
                   fmt=(fmts or {}).get(j))
    return {"cats": Reference(sh.ws, min_col=dest + 1, min_row=first, max_row=last),
            "vals": [Reference(sh.ws, min_col=dest + 2 + j, min_row=first, max_row=last)
                     for j in range(len(carry) - 1)]}


def _overlay(ch, ws, ref, name, colour=RED):
    """Lay a reference line over a bar chart.

    A ranked bar chart of a column you already have is worth about ten seconds
    of your own time. The line is what makes it worth downloading: it brings the
    standard — the mean, the action threshold, the 80% mark — to the data.
    """
    ln = LineChart()
    sr = Series(ref, title=name)
    sr.graphicalProperties.line.solidFill = colour
    sr.graphicalProperties.line.width = 18000
    sr.graphicalProperties.line.dashStyle = "dash"
    sr.marker = Marker(symbol="none")
    sr.smooth = False
    ln.series.append(sr)
    ch += ln
    # _bar suppresses the legend for a single series; with a reference line
    # there are two, and an unlabelled dashed line is a mystery
    from openpyxl.chart.legend import Legend
    ch.legend = Legend()
    ch.legend.position = "b"
    return ch


def _grouped(ws, title, cats, series, anchor, fmt=None, height=8, width=17):
    """Two bars per category — before/after, gross/realised."""
    ch = BarChart()
    ch.type = "col"
    ch.style = 10
    ch.title = title
    ch.height, ch.width = height, width
    ch.gapWidth = 60
    ch.overlap = -10
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    for ref, name, colour in series:
        s = Series(ref, title=name)
        s.graphicalProperties.solidFill = colour
        s.invertIfNegative = False       # negative bars render as nothing otherwise
        ch.series.append(s)
    ch.set_categories(cats)
    if fmt:
        ch.y_axis.numFmt = fmt
    ws.add_chart(ch, anchor)
    return ch


def _line(ws, title, cats, series, anchor, fmt=None, height=8, width=17, y_title=None):
    ch = LineChart()
    ch.title = title
    ch.style = 2
    ch.height, ch.width = height, width
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    if y_title:
        ch.y_axis.title = y_title
    for ref, name, colour, dashed, marker in series:
        s = Series(ref, title=name)
        s.graphicalProperties.line.solidFill = colour
        s.graphicalProperties.line.width = 16000 if dashed else 28000
        if dashed:
            s.graphicalProperties.line.dashStyle = "dash"
        s.marker = Marker(symbol="circle", size=6) if marker else Marker(symbol="none")
        if marker:
            s.marker.graphicalProperties.solidFill = colour
        s.smooth = False
        ch.series.append(s)
    ch.set_categories(cats)
    if fmt:
        ch.y_axis.numFmt = fmt
    ws.add_chart(ch, anchor)
    return ch


# ---------------------------------------------------------------- 05


def data_collection_plan(wb) -> int:
    """How the required sample size explodes as you ask for a tighter answer."""
    name = next((s for s in wb.sheetnames if "sample size" in s.lower()), None)
    if not name:
        return 0
    sh = Sheet(wb[name])
    # This block used to sit at row 20. chart_specs only ever writes cells and
    # patch_workbooks loads the workbook from disk, so moving it left the old
    # copy behind as a ghost table. Clear the ground we abandoned.
    for r in range(20, 33):
        for c in range(1, 4):
            cell = sh.ws.cell(row=r, column=c)
            if cell.__class__.__name__ != "MergedCell" and cell.value is not None:
                cell.value = None
                sh.changed += 1
    r0 = 34                      # below every merged note row on this tab
    sh.label(r0, 1, "How sample size moves with the precision you ask for")
    sh.put(r0 + 1, 1, "Worst-case proportion (p=0.5), 95% confidence. Halving the "
                      "margin of error costs you four times the sample.", note=True)
    sh.put(r0 + 2, 1, "Margin of error", bold=True)
    sh.put(r0 + 2, 2, "Sample needed", bold=True)
    for i, moe in enumerate([0.10, 0.075, 0.05, 0.04, 0.03, 0.025, 0.02, 0.015, 0.01]):
        r = r0 + 3 + i
        sh.put(r, 1, moe, fmt="0.0%")
        sh.put(r, 2, f"=CEILING(1.96^2*0.25/A{r}^2,1)", fmt="#,##0")
    sh.put(r0 + 2, 3, "At 5% margin", bold=True)
    for i in range(9):
        r = r0 + 3 + i
        sh.put(r, 3, "=CEILING(1.96^2*0.25/0.05^2,1)", fmt="#,##0")
    _line(sh.ws, "Sample size against margin of error",
          Reference(sh.ws, min_col=1, min_row=r0 + 3, max_row=r0 + 11),
          [(Reference(sh.ws, min_col=2, min_row=r0 + 3, max_row=r0 + 11),
            "Sample needed", BLUE, False, True),
           (Reference(sh.ws, min_col=3, min_row=r0 + 3, max_row=r0 + 11),
            "The usual choice: 5% margin", RED, True, False)],
          f"E{r0}", fmt="#,##0", y_title="Contacts to sample")
    return sh.changed


# ---------------------------------------------------------------- 10


def value_stream_map(wb) -> int:
    """The whole point of a VSM: touch time is a rounding error next to the wait."""
    ws = wb["Value stream"]
    LAST = 28                    # the sheet's own totals sum E10:E28
    cats = Reference(ws, min_col=2, min_row=10, max_row=LAST)
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "stacked"
    ch.overlap = 100
    ch.style = 10
    ch.title = "Where the time actually goes, step by step"
    ch.height, ch.width = 9, 22
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.y_axis.title = "Minutes"
    for col, nm, colour in ((5, "Touch time", GREEN), (6, "Waiting", RED)):
        s = Series(Reference(ws, min_col=col, min_row=10, max_row=LAST), title=nm)
        s.graphicalProperties.solidFill = colour
        s.invertIfNegative = False
        ch.series.append(s)
    ch.set_categories(cats)
    sh = Sheet(ws)
    sh.put(9, 12, "Average wait", bold=True)
    for r in range(10, LAST + 1):
        sh.put(r, 12, f'=IF(B{r}="","",AVERAGE($F$10:$F${LAST}))', fmt="#,##0")
    _overlay(ch, ws, Reference(ws, min_col=12, min_row=10, max_row=LAST),
             "Average wait per step")
    ws.add_chart(ch, "N9")
    return sh.changed


# ---------------------------------------------------------------- 11


def xy_matrix(wb) -> int:
    """Which candidate causes actually drive the CTQs."""
    ws = wb["X-Y matrix"]
    sh = Sheet(ws)
    blk = ranked_block(sh, 15, 36, sort_col=8, carry=[1, 8], dest=12,
                       label=["Candidate cause", "Weighted total"],
                       fmts={1: "#,##0"})
    for r in range(15, 37):
        sh.put(r, 15, f'=IF(N{r}="","",AVERAGE($N$15:$N$36))', fmt="#,##0")
    sh.put(14, 15, "Mean", bold=True)
    ch = _bar(ws, "Weighted total by candidate cause — ranked, tallest first",
              blk["cats"], blk["vals"][0],
              "Q14", horizontal=True, height=11, width=17, labels=False,
              name="Weighted total")
    _overlay(ch, ws, Reference(ws, min_col=15, min_row=15, max_row=36), "Mean score")
    return sh.changed


# ---------------------------------------------------------------- 12


def fmea(wb) -> int:
    """RPN before and after the action — the only view that shows whether it worked."""
    ws = wb["FMEA"]
    sh = Sheet(ws)
    # The method is "work the highest RPN first". Plotting in entry order buries
    # the worst failure mode wherever it happened to be typed.
    blk = ranked_block(sh, 11, 36, sort_col=10, carry=[2, 10, 18], dest=21,
                       label=["Failure mode", "RPN now", "RPN after"],
                       fmts={1: "#,##0", 2: "#,##0"})
    # The chart drew its action line at 100 while the sheet's own summary
    # counted ">=200 (act now)". Two thresholds on one screen, and a reviewer
    # who spots that stops trusting the rest of the file. One live input drives
    # both now, and the summary label reads it rather than restating it.
    sh.put(45, 1, "Act above this RPN", bold=True)
    thr = sh.ws.cell(row=45, column=4)
    if thr.value != 200:
        thr.value = 200
        sh.changed += 1
    thr.fill = IN
    thr.number_format = "#,##0"
    sh.put(45, 5, "The conventional line is 100-200. Whatever you set here drives "
                  "the chart line and the count above.", note=True)
    for r in range(11, 37):
        sh.put(r, 25, f'=IF(W{r}="","",$D$45)', fmt="#,##0")
    sh.put(10, 25, "Act above", bold=True)
    # keep the summary honest against the same cell
    lab = sh.ws.cell(row=40, column=1)
    if lab.value != "Above the action threshold (act now)":
        lab.value = "Above the action threshold (act now)"
        sh.changed += 1
    cnt = sh.ws.cell(row=40, column=4)
    want = '=COUNTIF(J11:J35,">="&$D$45)'
    if cnt.value != want:
        cnt.value = want
        sh.changed += 1
    ch = _grouped(ws, "Risk priority number — ranked, and did the action reduce it?",
                  blk["cats"],
                  [(blk["vals"][0], "RPN now", RED),
                   (blk["vals"][1], "RPN after", GREEN)],
                  "AA10", height=10, width=22)
    _overlay(ch, ws, Reference(ws, min_col=25, min_row=11, max_row=36),
             "Act above this threshold", AMBER)
    return sh.changed


# ---------------------------------------------------------------- 13


def hypothesis_log(wb) -> int:
    """Every p-value against the line you agreed to judge them by."""
    ws = wb["Test log"]
    sh = Sheet(ws)
    LAST = 35                    # row 36 is the merged note banner
    # Two bugs in one line. It hardcoded 0.05 while the sheet has a validated
    # alpha input at B8, so changing alpha moved every verdict and not the line
    # you judge them against. And it keyed on the p-value column, so the
    # standard stopped dead at the last completed test — absent across four
    # fifths of the axis. Key on the row number instead.
    for r in range(11, LAST + 1):
        sh.put(r, 20, f'=IF(A{r}="","",$B$8)', fmt="0.000")
    sh.put(10, 20, "alpha", bold=True)
    _line(ws, "p-value by test, against alpha = 0.05",
          Reference(ws, min_col=1, min_row=11, max_row=LAST),
          [(Reference(ws, min_col=10, min_row=11, max_row=LAST), "p-value", BLUE, False, True),
           (Reference(ws, min_col=20, min_row=11, max_row=LAST), "alpha (0.05)", RED, True, False)],
          "T10", fmt="0.000", height=9, width=20, y_title="p-value")
    return sh.changed


# ---------------------------------------------------------------- 15


def solution_selection(wb) -> int:
    ws = wb["Solution selection"]
    sh = Sheet(ws)
    blk = ranked_block(sh, 14, 33, sort_col=10, carry=[2, 10], dest=16,
                       label=["Solution", "Weighted score"],
                       fmts={1: "#,##0"})
    for r in range(14, 34):
        sh.put(r, 19, f'=IF(R{r}="","",AVERAGE($R$14:$R$33))', fmt="#,##0")
    sh.put(13, 19, "Mean", bold=True)
    ch = _bar(ws, "Weighted score by solution — ranked, best first",
              blk["cats"], blk["vals"][0],
              "U13", horizontal=True, height=11, width=17, labels=False,
              name="Weighted score")
    _overlay(ch, ws, Reference(ws, min_col=19, min_row=14, max_row=33), "Mean score")
    return sh.changed


# ---------------------------------------------------------------- 17


def control_plan(wb) -> int:
    """Are your controls designed in, or are they 'remind people to be careful'?"""
    ws = wb["Control plan"]
    sh = Sheet(ws)
    r0 = 37                      # the sheet's notes run to row 34 and are merged
    sh.label(r0, 1, "Controls by level — how durable is this control plan?")
    sh.put(r0 + 1, 1, "A plan weighted towards levels 5 and 6 decays the moment attrition "
                      "or a busy week arrives. Levels 1-3 survive without anyone remembering.",
           note=True)
    levels = ["1 Eliminate the demand", "2 Design it out", "3 Mistake-proof it",
              "4 Automatic detection", "5 Standard work + training", "6 Reminder or awareness"]
    for i, lvl in enumerate(levels):
        r = r0 + 2 + i
        sh.put(r, 1, lvl)
        sh.put(r, 2, f'=COUNTIF($P$10:$P$27,LEFT(A{r},1)&"*")', fmt="0")
    _bar(sh.ws, "Controls by level — lower is more durable",
         Reference(sh.ws, min_col=1, min_row=r0 + 2, max_row=r0 + 7),
         Reference(sh.ws, min_col=2, min_row=r0 + 2, max_row=r0 + 7),
         f"E{r0}", horizontal=True, height=8, width=17,
         # levels 1-3 survive attrition; 4-6 rely on someone remembering
         colours=[GREEN, GREEN, GREEN, AMBER, RED, RED])
    return sh.changed


# ---------------------------------------------------------------- 19


def calculators(wb) -> int:
    changed = 0

    # 1 — where you sit on the scale everyone quotes
    sh = Sheet(wb["1 Sigma level"])
    sh.label(19, 1, "Your DPMO against the standard scale")
    for i, (lab, dpmo) in enumerate([("2 sigma", 308538), ("3 sigma", 66807),
                                     ("4 sigma", 6210), ("5 sigma", 233), ("6 sigma", 3.4)]):
        sh.put(20 + i, 1, lab)
        sh.put(20 + i, 2, dpmo, fmt="#,##0")
    sh.put(25, 1, "YOU", bold=True)
    sh.put(25, 2, "=IFERROR(B12,\"\")", fmt="#,##0")
    _bar(sh.ws, "DPMO — you against the scale", Reference(sh.ws, min_col=1, min_row=20, max_row=25),
         Reference(sh.ws, min_col=2, min_row=20, max_row=25), "D4",
         colours=[BLUE, BLUE, BLUE, BLUE, BLUE, AMBER], horizontal=True, height=8, width=16)
    changed += sh.changed

    # 2 — kappa against the thresholds people actually use
    sh = Sheet(wb["2 QA agreement (kappa)"])
    sh.label(16, 1, "Your kappa against the usual bars")
    for i, (lab, v) in enumerate([("Unusable below", 0.60), ("Acceptable", 0.80),
                                  ("Good", 0.90)]):
        sh.put(17 + i, 1, lab)
        sh.put(17 + i, 2, v, fmt="0.00")
    sh.put(20, 1, "YOUR KAPPA", bold=True)
    sh.put(20, 2, "=IFERROR(B13,\"\")", fmt="0.00")
    kch = _bar(sh.ws, "Cohen's kappa against the usual thresholds",
               Reference(sh.ws, min_col=1, min_row=17, max_row=20),
               Reference(sh.ws, min_col=2, min_row=17, max_row=20), "D4",
               colours=[RED, AMBER, GREEN, BLUE], fmt="0.00", height=8, width=16)
    # kappa runs 0 to 1 by definition; letting it auto-scale to 0.55-0.95 makes
    # a 0.72 look like it is nearly at the top of the scale
    kch.y_axis.scaling.min, kch.y_axis.scaling.max = 0, 1
    changed += sh.changed

    # 3 — the capability picture: your distribution against the SLA
    sh = Sheet(wb["3 SLA capability"])
    sh.label(19, 1, "Your resolution-time distribution against the SLA limit")
    sh.put(20, 1, "Hours", bold=True)
    sh.put(20, 2, "How often", bold=True)
    sh.put(20, 3, "SLA limit", bold=True)
    for i in range(41):
        r = 21 + i
        sh.put(r, 1, f"=MAX(0,$B$6-4*$B$7)+({i}/40)*(8*$B$7)", fmt="0.00")
        sh.put(r, 2, f"=IFERROR(NORMDIST(A{r},$B$6,$B$7,FALSE),\"\")", fmt="0.0000")
        sh.put(r, 3, f"=IF(ABS(A{r}-$B$5)<=(8*$B$7)/80,MAX($B$21:$B$61),0)", fmt="0.0000")
    _line(sh.ws, "Where you sit against the SLA — area past the red line is a breach",
          Reference(sh.ws, min_col=1, min_row=21, max_row=61),
          [(Reference(sh.ws, min_col=2, min_row=21, max_row=61), "Your process", BLUE, False, False),
           (Reference(sh.ws, min_col=3, min_row=21, max_row=61), "SLA limit", RED, False, False)],
          "E4", height=9, width=18, y_title="Relative frequency")
    changed += sh.changed

    # 4 — Little's Law, as a picture
    sh = Sheet(wb["4 Backlog and lead time"])
    sh.label(16, 1, "Lead time: today against your target")
    sh.put(17, 1, "Today")
    sh.put(17, 2, "=IFERROR(B9,\"\")", fmt="#,##0.00")
    sh.put(18, 1, "Target")
    sh.put(18, 2, "=B7", fmt="#,##0.00")
    _bar(sh.ws, "Lead time in days — today against target",
         Reference(sh.ws, min_col=1, min_row=17, max_row=18),
         Reference(sh.ws, min_col=2, min_row=17, max_row=18), "D4",
         colours=[RED, GREEN], fmt="#,##0.00", height=7, width=14)
    sh.label(20, 1, "Backlog: now against the cap that delivers your target")
    sh.put(21, 1, "Open now")
    sh.put(21, 2, "=B5", fmt="#,##0")
    sh.put(22, 1, "Cap needed")
    sh.put(22, 2, "=IFERROR(B11,\"\")", fmt="#,##0")
    _bar(sh.ws, "Open tickets against the cap",
         Reference(sh.ws, min_col=1, min_row=21, max_row=22),
         Reference(sh.ws, min_col=2, min_row=21, max_row=22), "D20",
         colours=[RED, GREEN], height=7, width=14)
    changed += sh.changed

    # 5 — PCE: the number that shocks people
    sh = Sheet(wb["5 Process efficiency"])
    sh.label(14, 1, "Where the elapsed time went")
    sh.put(15, 1, "Actual work")
    sh.put(15, 2, "=B5", fmt="#,##0")
    sh.put(16, 1, "Waiting")
    sh.put(16, 2, "=IFERROR(B9,\"\")", fmt="#,##0")
    _bar(sh.ws, "Minutes of work against minutes of waiting",
         Reference(sh.ws, min_col=1, min_row=15, max_row=16),
         Reference(sh.ws, min_col=2, min_row=15, max_row=16), "D4",
         colours=[GREEN, RED], height=8, width=15)
    changed += sh.changed

    # 6 — where the headcount goes
    sh = Sheet(wb["6 Staffing"])
    sh.label(15, 1, "What the headcount is actually made of")
    sh.put(16, 1, "Offered load")
    sh.put(16, 2, "=IFERROR(B10,\"\")", fmt="#,##0.0")
    sh.put(17, 1, "+ occupancy gap")
    sh.put(17, 2, "=IFERROR(B11-B10,\"\")", fmt="#,##0.0")
    sh.put(18, 1, "+ shrinkage")
    sh.put(18, 2, "=IFERROR(B12,\"\")", fmt="#,##0.0")
    sh.put(19, 1, "PAID FTE", bold=True)
    sh.put(19, 2, "=IFERROR(B13,\"\")", fmt="#,##0.0")
    _bar(sh.ws, "From offered load to paid FTE — where does the headcount actually go?",
         Reference(sh.ws, min_col=1, min_row=16, max_row=19),
         Reference(sh.ws, min_col=2, min_row=16, max_row=19), "D4",
         colours=[BLUE, AMBER, RED, GREEN], fmt="#,##0.0", height=8, width=16)
    changed += sh.changed

    # 7 and 8 — gross against what you can actually book
    for tab, gross, real in (("7 Benefit — AHT", 14, 15), ("8 Benefit — avoided contacts", 13, 14)):
        sh = Sheet(wb[tab])
        r0 = 19
        sh.label(r0, 1, "Gross against realised")
        sh.put(r0 + 1, 1, "Gross annual value")
        sh.put(r0 + 1, 2, f"=IFERROR(B{gross},\"\")", fmt='"$"#,##0')
        sh.put(r0 + 2, 1, "Realised benefit")
        sh.put(r0 + 2, 2, f"=IFERROR(B{real},\"\")", fmt='"$"#,##0')
        _bar(sh.ws, "Gross value against what Finance will actually book",
             Reference(sh.ws, min_col=1, min_row=r0 + 1, max_row=r0 + 2),
             Reference(sh.ws, min_col=2, min_row=r0 + 1, max_row=r0 + 2), f"D{r0 - 15}",
             colours=[AMBER, GREEN], fmt='"$"#,##0', height=8, width=16)
        changed += sh.changed

    # 9 — the breakeven chart Finance asks for
    sh = Sheet(wb["9 ROI and payback"])
    sh.label(19, 1, "Cumulative position in today's money")
    sh.put(20, 1, "Year", bold=True)
    sh.put(20, 2, "Cumulative", bold=True)
    sh.put(20, 3, "Breakeven", bold=True)
    sh.put(21, 1, "Year 0")
    sh.put(21, 2, "=-B5", fmt='"$"#,##0')
    sh.put(21, 3, 0, fmt='"$"#,##0')
    for i in range(1, 4):
        r = 21 + i
        sh.put(r, 1, f"Year {i}")
        sh.put(r, 2, f"=B{r - 1}+B{9 + i}", fmt='"$"#,##0')
        sh.put(r, 3, 0, fmt='"$"#,##0')
    _line(sh.ws, "Cumulative discounted position — it crosses zero at payback",
          Reference(sh.ws, min_col=1, min_row=21, max_row=24),
          [(Reference(sh.ws, min_col=2, min_row=21, max_row=24), "Cumulative", BLUE, False, True),
           (Reference(sh.ws, min_col=3, min_row=21, max_row=24), "Breakeven", RED, True, False)],
          "E4", fmt='"$"#,##0', height=9, width=18)
    changed += sh.changed
    return changed


SPECS = {
    "05-data-collection-plan.xlsx": data_collection_plan,
    "10-value-stream-map.xlsx": value_stream_map,
    "11-cause-effect-xy-matrix.xlsx": xy_matrix,
    "12-fmea.xlsx": fmea,
    "13-hypothesis-test-log.xlsx": hypothesis_log,
    "15-solution-selection-matrix.xlsx": solution_selection,
    "17-control-plan.xlsx": control_plan,
    "19-black-belt-calculators.xlsx": calculators,
}


def _fingerprint(wb) -> str:
    """What the charts currently are, in enough detail to notice an edit.

    Counting charts is not enough: correcting a series range from B11:B30 to
    B11:B36 leaves the count identical, so a count-based check would decide
    nothing had changed and never write the fix to disk.
    """
    parts = []
    for ws in wb.worksheets:
        for ch in getattr(ws, "_charts", []):
            title = ""
            rich = getattr(getattr(getattr(ch, "title", None), "tx", None), "rich", None)
            if rich is not None:
                title = "".join(r.t or "" for p in rich.p for r in (p.r or []))
            refs = []
            for ser in getattr(ch, "series", []) or []:
                for holder in ("val", "cat", "xVal", "yVal"):
                    h = getattr(ser, holder, None)
                    if h is None:
                        continue
                    for kind in ("numRef", "strRef"):
                        rr = getattr(h, kind, None)
                        if rr is not None and rr.f:
                            refs.append(f"{holder}:{rr.f}")
            # deliberately not the anchor: it is a plain string when we set it
            # and an anchor object once the file has been read back, so it would
            # never compare equal and every run would re-save
            parts.append(f"{ws.title}|{ch.__class__.__name__}|{title}|{';'.join(refs)}")
    return "\n".join(parts)


def _spec_version() -> str:
    """A short hash of this file.

    The reference fingerprint catches a chart being added, removed or
    re-pointed. It cannot see a styling change — turning off the category name
    in a data label leaves every reference identical — so a fix to how charts
    LOOK was computed, found to match, and thrown away without ever reaching
    disk. Stamping the spec version into the workbook closes that hole: edit
    this file and every workbook it builds is rewritten once, then left alone.
    """
    import hashlib
    return hashlib.sha256(pathlib.Path(__file__).read_bytes()).hexdigest()[:12]


def _stamp(wb) -> str:
    for prop in getattr(wb.custom_doc_props, "props", []):
        if prop.name == "chartspec":
            return str(prop.value or "")
    return ""


def add_charts(wb, wbname: str) -> int:
    """Rebuild this workbook's charts from the spec. Returns the change count."""
    spec = SPECS.get(wbname)
    if spec is None:
        return 0
    version = _spec_version()
    stale = _stamp(wb) != version
    before = _fingerprint(wb)
    for ws in wb.worksheets:
        ws._charts = []
    changed = spec(wb)
    if stale:
        props = wb.custom_doc_props
        for prop in list(getattr(props, "props", [])):
            if prop.name == "chartspec":
                props.props.remove(prop)
        props.append(StringProperty(name="chartspec", value=version))
    # A workbook whose charts were added, removed or re-pointed must be saved
    # even if not a single cell value moved.
    return changed + (1 if stale or _fingerprint(wb) != before else 0)
