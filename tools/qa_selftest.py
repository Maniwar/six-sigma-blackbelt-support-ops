#!/usr/bin/env python3
"""Mutation testing for the QA harness itself.

Every quality check in this repo was written after a defect got past the
previous ones, and each one has passed cleanly ever since. That is either
because the defect class is gone, or because the check cannot fire. Those two
states look identical from a green run, and only one of them is good news.

So this reintroduces each shipped defect into a real workbook and asserts the
audit catches it. A mutant that survives is a decorative check: it is reported
as a failure here, loudly, because a check nobody can trip is worse than no
check at all — it buys confidence it has not earned.

    python3 tools/qa_selftest.py

Exit status is non-zero if any mutant survived.
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
import qa_citations as C                                          # noqa: E402
import qa_templates as Q                                          # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "templates"

HDR = "FF333C49"


# ------------------------------------------------------------------ mutants
# Each returns a description of the damage it did, or None if it could not find
# anything to damage in this workbook (in which case the mutant is skipped, not
# passed — an unapplied mutant proves nothing).


def _first_block(wb):
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in Q.GUIDE_TABS):
            continue
        blocks = Q._blocks(ws)
        for hrow, first, last, cols in blocks:
            if last - first + 1 >= 3:
                return ws, hrow, first, last, cols
    return None


def m_empty_column(wb):
    """A column the example declares and then never fills in.

    24-doe-design-matrix shipped with a Notes header over eight blank rows.
    """
    found = _first_block(wb)
    if not found:
        return None
    ws, hrow, first, last, cols = found
    for col, head in sorted(cols.items(), reverse=True):
        cells = [ws.cell(row=r, column=col) for r in range(first, last + 1)]
        if not any(c.value not in (None, "") for c in cells):
            continue
        if any(c.fill and c.fill.patternType
               and str(c.fill.fgColor.rgb) == Q.FILL_INPUT for c in cells):
            continue                        # a yellow column may be blank
        for c in cells:
            if (c.row, c.column) in Q.merged_shadow(ws):
                continue
            try:
                c.value = None
            except AttributeError:          # merged shadow
                return None
        return f"emptied {ws.title!r} column {head!r} across {last - first + 1} rows"
    return None


def m_opaque_header(wb):
    """A header that names nothing, with the legend taken away.

    The DOE design matrix read 'A B C AB AC BC' with no key anywhere.
    """
    found = _first_block(wb)
    if not found:
        return None
    ws, hrow, first, last, cols = found
    shadow = Q.merged_shadow(ws)
    for r in (hrow - 1, hrow + 1):
        if r < 1:
            continue
        for c in range(1, max(cols) + 1):
            cell = ws.cell(row=r, column=c)
            if (r, c) in shadow:
                continue
            if isinstance(cell.value, str) and len(cell.value) > 25:
                cell.value = None
    target = max(cols)
    ws.cell(row=hrow, column=target).value = "XZ"
    return f"renamed a {ws.title!r} header to 'XZ' and removed the legend"


def m_boilerplate(wb):
    """One sentence pasted onto every row of a block.

    The DOE effects table said the same thing on all six effect rows.
    """
    found = _first_block(wb)
    if not found:
        return None
    ws, hrow, first, last, cols = found
    shadow = Q.merged_shadow(ws)
    col = max(cols)
    n = 0
    for r in range(first, min(last, first + 4) + 1):
        if (r, col) in shadow:
            continue
        ws.cell(row=r, column=col).value = (
            "Bigger absolute value = bigger effect. A large interaction "
            "means the two factors cannot be set independently.")
        n += 1
    if n < 3:
        return None
    return f"pasted one identical note onto {n} rows of {ws.title!r}"


def m_ghost_block(wb):
    """A moved helper block that left its old copy behind."""
    for ws in wb.worksheets:
        shadow = Q.merged_shadow(ws)
        for row in ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and len(c.value) > 24):
                    continue
                if c.value.startswith("=") or not (c.font and c.font.bold):
                    continue
                r = ws.max_row + 4
                d = ws.cell(row=r, column=c.column)
                if (r, c.column) in shadow:
                    continue
                d.value = c.value
                d.font = Font(bold=True, size=11)
                return f"left a duplicate heading on {ws.title!r} at row {r}"
    return None


def m_naked_input(wb):
    """A yellow cell with nothing anywhere near it saying what to type."""
    for ws in wb.worksheets:
        shadow = Q.merged_shadow(ws)
        r = ws.max_row + 3
        c = ws.cell(row=r, column=2)
        if (r, 2) in shadow:
            continue
        c.value = 0
        c.fill = PatternFill("solid", fgColor=Q.FILL_INPUT)
        return f"added an unexplained yellow input at {ws.title}!B{r}"
    return None


def m_empty_series(wb):
    """A chart still pointing at a block a later edit stopped feeding."""
    for ws in wb.worksheets:
        for ch in getattr(ws, "_charts", []):
            for ser in ch.series or []:
                vref, _ = Q.series_refs(ser)
                p = Q.parse_ref(vref or "")
                if not p:
                    continue
                sheet, c1, r1, c2, r2 = p
                src = wb[sheet]
                shadow = Q.merged_shadow(src)
                blanked = 0
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        if (r, c) in shadow:
                            continue
                        src.cell(row=r, column=c).value = None
                        blanked += 1
                if blanked:
                    return f"blanked the {blanked} cells behind a chart on {ws.title!r}"
    return None


def m_untitled_chart(wb):
    """A chart with its title taken off — a decoration, not an answer."""
    for ws in wb.worksheets:
        for ch in getattr(ws, "_charts", []):
            if Q.chart_title_text(ch).strip():
                ch.title = None
                return f"removed the title from a chart on {ws.title!r}"
    return None


def m_paint_over_the_inputs(wb):
    """Turn a sheet all-green AND take away the sentence that explains it.

    Not hypothetical: this shipped. Repainting the worked examples green walked
    every pre-filled row of a block, took 825 cells with it, and left sheets
    with no input cell at all while the instructions still said to overwrite the
    yellow column — a colour that by then existed nowhere on them.

    Both halves are needed, because an all-green sheet is legitimate when the
    workbook says the green is what you replace. The defect is the mismatch, so
    the mutant creates the mismatch: paint the inputs over, then delete the
    instruction that would have made it coherent.
    """
    import re as _re

    green = PatternFill("solid", fgColor="FFECFAEF")
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in Q.GUIDE_TABS):
            continue
        shadow = Q.merged_shadow(ws)
        hit = [c for row in ws.iter_rows() for c in row
               if (c.row, c.column) not in shadow and _fill(c) == Q.FILL_INPUT]
        if not hit:
            continue
        for c in hit:
            c.fill = green
        # The instruction lives across two cells — "Green cells" in one, "Replace
        # it with your own data" in the next — so deleting whole cells that match
        # on their own removes neither half. Take the word out instead.
        for w in wb.worksheets:
            for row in w.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and "green" in c.value.lower():
                        c.value = _re.sub("green", "shaded", c.value, flags=_re.I)
        return (f"repainted all {len(hit)} input cell(s) on {ws.title!r} as example "
                "data and removed the sentence that says to overwrite the green")
    return None


def _fill(c) -> str:
    try:
        return str(c.fill.fgColor.rgb) if c.fill and c.fill.patternType else ""
    except Exception:                                            # noqa: BLE001
        return ""


def m_green_formula(wb):
    """Paint a computed cell as a worked example — "replace this" on a formula."""
    green = PatternFill("solid", fgColor="FFECFAEF")
    for ws in wb.worksheets:
        shadow = Q.merged_shadow(ws)
        for row in ws.iter_rows():
            for c in row:
                if (c.row, c.column) in shadow:
                    continue
                if isinstance(c.value, str) and c.value.startswith("=") \
                        and _fill(c) == Q.FILL_CALC:
                    c.fill = green
                    return f"painted the formula at {ws.title}!{c.coordinate} as an example"
    return None


def m_bare_row(wb):
    """A name and a number, and nothing telling the reader what it is."""
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in Q.GUIDE_TABS):
            continue
        shadow = Q.merged_shadow(ws)
        r = ws.max_row + 3
        if (r, 1) in shadow or (r, 2) in shadow:
            continue
        ws.cell(row=r, column=1, value="DPO")
        c = ws.cell(row=r, column=2, value=0.0507)
        c.number_format = "0.0000"
        c.fill = PatternFill("solid", fgColor=Q.FILL_CALC)
        return f"added an unexplained label:value row at {ws.title}!A{r}"
    return None


def m_strip_row_note(wb):
    """Take the explanation off a CALCULATED row that already has one.

    This is the one that matters. m_bare_row only proves the check notices a row
    invented for it; this proves it defends the notes actually in the workbook,
    and it targets a blue cell on purpose — GUIDED reads yellow only, so if
    ROWLABEL ever stops firing nothing else in the harness covers the row.
    """
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in Q.GUIDE_TABS):
            continue
        shadow = Q.merged_shadow(ws)
        for r in range(1, ws.max_row + 1):
            lab, val = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2)
            if not (isinstance(lab, str) and lab.strip()) or val.value in (None, ""):
                continue
            try:
                rgb = str(val.fill.fgColor.rgb) if val.fill and val.fill.patternType else ""
            except Exception:                                    # noqa: BLE001
                rgb = ""
            if rgb != Q.FILL_CALC or Q._speaks_for_itself(val):
                continue
            notes = [c for c in Q.ROWLABEL_COLS
                     if (r, c) not in shadow
                     and isinstance(ws.cell(row=r, column=c).value, str)
                     and len(ws.cell(row=r, column=c).value) >= 12]
            if not notes:
                continue
            for c in notes:
                ws.cell(row=r, column=c).value = None
            return f"blanked the note beside {ws.title}!A{r} ({lab.strip()[:30]!r})"
    return None


MUTANTS = [
    ("empty declared column", m_empty_column, "EXAMPLE"),
    ("opaque header, no legend", m_opaque_header, "EXAMPLE"),
    ("boilerplate note on every row", m_boilerplate, "EXAMPLE"),
    ("ghost block left behind", m_ghost_block, "GUIDED"),
    ("unexplained yellow input", m_naked_input, "GUIDED"),
    ("chart wired to an emptied block", m_empty_series, "CHARTS"),
    ("chart with no title", m_untitled_chart, "CHARTS"),
    ("entry grid repainted as example data", m_paint_over_the_inputs, "GUIDED"),
    ("formula cell painted as a worked example", m_green_formula, "GUIDED"),
    ("label and number, no explanation", m_bare_row, "ROWLABEL"),
    ("explanation stripped off a calculated row", m_strip_row_note, "ROWLABEL"),
]


# --------------------------------------------------------------------- run


# ------------------------------------------------------- markdown mutants
# The MARKDOWN layer is the newest, and it found 97 real defects on its first
# run — which is exactly the profile of a check that will pass forever
# afterwards and never be questioned again. These two put it back in front of
# the defects it was written for.


def md_empty_column(text: str):
    """Blank out one column of one table, on every row."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if not Q.RE_MD_ROW.match(line) or i + 1 >= len(lines):
            continue
        if not Q.RE_MD_RULE.match(lines[i + 1]):
            continue
        head = [c.strip() for c in Q.RE_MD_ROW.match(line).group(1).split("|")]
        if len(head) < 2 or any(h.strip(" *").lower() in ("signature", "signed")
                                for h in head):
            continue
        j, n = i + 2, 0
        while j < len(lines) and Q.RE_MD_ROW.match(lines[j]) \
                and not Q.RE_MD_RULE.match(lines[j]):
            cells = [c for c in Q.RE_MD_ROW.match(lines[j]).group(1).split("|")]
            if len(cells) == len(head):
                cells[-1] = "  "
                lines[j] = "|" + "|".join(cells) + "|"
                n += 1
            j += 1
        if n:
            return "\n".join(lines), f"emptied column {head[-1]!r} on {n} row(s)"
    return None, None


def md_strip_method(text: str):
    """Take away every statement of how the arithmetic is done."""
    if not Q.RE_ARITH.search(text) or not Q.RE_METHOD.search(text):
        return None, None
    out = Q.RE_METHOD.sub("something", text)
    return out, "removed every formula, method and workbook pointer"


def md_break_shares(text: str):
    """Nudge one share so a column totalled at 100% no longer adds up."""
    import re as _re
    IS_TOTAL = ("total", "all ", "check")
    for head, body, line in Q.md_tables(text):
        for col in range(len(head)):
            if col >= len(head):
                continue
            pct = _re.compile(r"\*?\*?(-?\d+(?:\.\d+)?)\s*%")
            total = None
            vals = []
            for row, _ln in body:
                if col >= len(row):
                    continue
                m = pct.match(row[col].strip())
                if not m:
                    continue
                if row[0].strip(" *").lower().startswith(IS_TOTAL):
                    total = float(m.group(1))
                else:
                    vals.append((row, float(m.group(1))))
            # Only mutate a column the check actually audits — one totalled at
            # 100% — and only a row it actually sums. Mutating the total row, or
            # a column totalled at 14.2%, changes nothing the check claims to
            # cover, and a survivor there says nothing about the check.
            if total is None or abs(total - 100.0) > 0.5 or len(vals) < 3:
                continue
            vals = [(r, v) for r, v in vals]
            # Edit the raw line by number. Rebuilding it from md_tables' cells
            # cannot work — those are stripped, so the reconstructed row never
            # matches the spacing in the file and the mutant silently no-ops.
            row, lineno = vals[0][0], None                       # noqa: F841
            for cells, ln in body:
                if cells is row:
                    lineno = ln
                    break
            if lineno is None:
                continue
            lines = text.splitlines()
            raw = lines[lineno - 1]
            parts = raw.split("|")
            # a leading "|" makes parts[0] empty, so cell N is parts[N+1]
            target = col + 1
            if target >= len(parts):
                continue
            m = _re.search(r"(-?\d+(?:\.\d+)?)", parts[target])
            if not m:
                continue
            parts[target] = parts[target].replace(
                m.group(1), str(float(m.group(1)) + 9), 1)
            lines[lineno - 1] = "|".join(parts)
            return "\n".join(lines), f"added 9 points to one share in {head[col]!r}"
    return None, None


def md_bury_in_comment(text: str):
    """Hide a line of guidance inside an HTML comment, where nobody reads it.

    Four real rules shipped this way — a markdown viewer hides the comment and
    the page escaped it into visible body text, so the same sentence was both
    invisible in the file and printed raw, angle brackets and all.
    """
    for i, line in enumerate(text.split("\n")):
        s = line.strip()
        if s.startswith("*") and s.endswith("*") and 25 <= len(s) <= 120:
            lines = text.split("\n")
            lines[i] = f"<!-- {s.strip('*')} -->"
            return "\n".join(lines), f"buried {s[:34]!r} in an HTML comment"
    return None, None


MD_MUTANTS = [("markdown: empty declared column", md_empty_column),
              ("markdown: arithmetic with no method", md_strip_method),
              ("markdown: shares that do not add up", md_break_shares),
              ("markdown: guidance buried in an HTML comment", md_bury_in_comment)]


def run_markdown() -> tuple[int, int, list[str]]:
    killed = applied = 0
    survivors = []
    with tempfile.TemporaryDirectory() as td:
        for doc in sorted(TEMPLATES.glob("*.md")):
            text = doc.read_text(encoding="utf-8")
            Q.fails.clear()
            Q.audit_markdown(doc)
            baseline = set(Q.fails)
            for name, mutate in MD_MUTANTS:
                out, what = mutate(text)
                if out is None:
                    continue
                tmp = Path(td) / doc.name
                tmp.write_text(out, encoding="utf-8")
                applied += 1
                Q.fails.clear()
                Q.audit_markdown(tmp)
                if [f for f in set(Q.fails) - baseline if "[MARKDOWN]" in f]:
                    killed += 1
                else:
                    survivors.append(f"    SURVIVED [MARKDOWN] {name} on {doc.name} — {what}")
    return killed, applied, survivors


# ------------------------------------------------- control-chart mutants
# The two ways a control chart's worked example goes wrong, both of which this
# repo shipped: data so quiet the rule never fires, and a baseline that is
# itself out of control so every limit is computed from an unstable process.


def cc_flatten(wb):
    """Put the monitoring window back inside the limits — the state four of
    the seven charts shipped in, where the Signal column is blank on all 24
    rows and the rule has never once been evaluated."""
    ws = wb["I-MR"]
    for row, v in zip(range(34, 38), (442, 404, 416, 428)):
        ws[f"B{row}"] = v
    return "I-MR monitoring window back inside the limits"


def cc_unstable_baseline(wb):
    """Push a BASELINE point outside the limits, so the sheet computes its own
    limits from a process it shows is not in control."""
    ws = wb["Xbar-R"]
    for col in "BCDEF":
        ws[f"{col}20"] = 505
    return "Xbar-R baseline point 7 forced out of control"


CC_MUTANTS = [("control chart: worked data never signals", cc_flatten),
              ("control chart: baseline is out of control", cc_unstable_baseline)]


def run_control() -> tuple[int, int, list[str]]:
    import importlib

    V = importlib.import_module("verify")
    book = TEMPLATES / "27-control-charts.xlsx"
    killed = applied = 0
    survivors = []
    with tempfile.TemporaryDirectory() as td:
        def audit(path: Path) -> set[str]:
            V.FAILURES.clear()
            V.PASSES[0] = 0
            V.audit_control_signals(V._engine(path).calculate(), book.name)
            return set(V.FAILURES)

        # Same FILENAME, different directory. `formulas` keys every cell by the
        # workbook's file name, so a copy called base_27-... resolves nothing at
        # all — every lookup returns None, the audit fails identically before and
        # after, and both mutants survive against a check that was working. That
        # cost a round trip and is exactly the false negative a mutation suite
        # is supposed to catch, so it is written down rather than just fixed.
        def stage(sub: str) -> Path:
            d = Path(td) / sub
            d.mkdir()
            dst = d / book.name
            shutil.copyfile(book, dst)
            return dst

        baseline = audit(stage("base"))
        for i, (name, mutate) in enumerate(CC_MUTANTS):
            tmp = stage(f"m{i}")
            wb = load_workbook(tmp)
            what = mutate(wb)
            if not what:
                continue
            wb.save(tmp)
            applied += 1
            if audit(tmp) - baseline:
                killed += 1
            else:
                survivors.append(f"    SURVIVED [CONTROL] {name} — {what}")
    V.FAILURES.clear()
    V.PASSES[0] = 0
    return killed, applied, survivors


# ------------------------------------------------------ guidance mutants
# md_guidance.py holds a second copy of the pack's worked example, and a second
# copy of a number is a number that will drift. It had: it still carried the
# benefit chain that was withdrawn for being causally impossible, and it
# rewrote every document's preamble with a bare "14.2% reopen rate" — the one
# sentence those documents now exist to warn against. Nothing in the build ran
# it, so nothing noticed for a release.
#
# These mutate the module rather than a file, so each one saves and restores.


def gm_drift_figure(MG):
    """A worked-example number the pack does not state anywhere."""
    for name, vals in MG.EXAMPLE.items():
        for label, v in vals.items():
            if any(c.isdigit() for c in v):
                vals[label] = "999,777 units at $888.55"
                return f"{name} {label!r} offers a figure nothing states"
    return None


def gm_dead_row(MG):
    """A key that addresses a row no template has — dead weight carrying a
    number, which is how the stale ones hid."""
    for name, vals in MG.EXAMPLE.items():
        if vals:
            k = next(iter(vals))
            vals["Row that no template has"] = vals.pop(k)
            return f"{name} {k!r} renamed to a row that does not exist"
    return None


def gm_unpopulated_preamble(MG):
    """The preamble that says 14.2% without saying 14.2% of WHAT."""
    orig = MG.block
    MG.block = lambda n: orig(n).replace("OD-BIL-004-ADJ", "the billing queue")
    return "preamble no longer names the population it measures"


GUIDE_MUTANTS = [("guidance: a figure the pack does not state", gm_drift_figure),
                 ("guidance: a row no template has", gm_dead_row),
                 ("guidance: preamble drops the population", gm_unpopulated_preamble)]


def run_guidance() -> tuple[int, int, list[str]]:
    import copy
    import importlib

    V = importlib.import_module("verify")
    MG = importlib.import_module("md_guidance")

    def fails() -> list[str]:
        V.FAILURES.clear()
        V.PASSES[0] = 0
        V.test_guidance()
        return list(V.FAILURES)

    killed = applied = 0
    survivors = []
    baseline = fails()
    for name, mutate in GUIDE_MUTANTS:
        saved, saved_block = copy.deepcopy(MG.EXAMPLE), MG.block
        try:
            what = mutate(MG)
            if not what:
                continue
            applied += 1
            if set(fails()) - set(baseline):
                killed += 1
            else:
                survivors.append(f"    SURVIVED [GUIDANCE] {name} — {what}")
        finally:
            MG.EXAMPLE.clear()
            MG.EXAMPLE.update(saved)
            MG.block = saved_block
    V.FAILURES.clear()
    V.PASSES[0] = 0
    return killed, applied, survivors


# ------------------------------------------------------- numeric mutants
# The #N/A exemption used to be a hand-written list of column spans, and two
# workbooks that use the same NA() idiom were never added to it — so the gate
# ran red on 37 deliberate cells for months and everyone learned to read past
# it. It is now a property of the formula: a deliberate gap CALLS NA(), an
# accidental one falls out of a lookup that missed.
#
# That rule is only worth having if it can tell the two apart, so the first
# mutant leaves the cell reporting the very same #N/A and changes nothing but
# the REASON. If the check survives that, it is reading the value and calling
# it a design decision, which is where the old list ended up.
#
# audit_numeric recalculates a whole workbook, far too slow for the per-mutant
# loop above, so this runs only on the four workbooks that use NA() at all.
NA_BOOKS = ["05-data-collection-plan", "19-black-belt-calculators",
            "25-pareto-and-distribution", "27-control-charts"]
MISS = '=MATCH("no such row",$A$1:$A$2,0)'


def nm_reason_swap(wb):
    """A gap that is #N/A because a lookup missed, not because it was written.

    Every NA() in the book, not the first one found: `=IF(I5="",NA(),...)` on a
    row that HAS data never takes its NA() branch, so swapping that one changes
    no value and proves nothing.
    """
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "NA()" in c.value.replace(" ", ""):
                    c.value = c.value.replace("NA()", MISS.lstrip("="))
                    n += 1
    return f"{n} deliberate gap(s) now come from a missed lookup" if n else None


def nm_hidden_break(wb):
    """A broken lookup somewhere ordinary — the base check, not the exemption."""
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        ws.cell(row=ws.max_row + 3, column=60).value = MISS
        return f"broken lookup dropped into {ws.title}!BH{ws.max_row}"
    return None


NUM_MUTANTS = [("numeric: #N/A for the wrong reason", nm_reason_swap),
               ("numeric: a lookup that misses", nm_hidden_break)]


def run_numeric() -> tuple[int, int, list[str]]:
    killed = applied = 0
    survivors = []
    with tempfile.TemporaryDirectory() as td:
        for stem in NA_BOOKS:
            book = TEMPLATES / f"{stem}.xlsx"
            base = Path(td) / ("base_" + book.name)
            shutil.copyfile(book, base)
            load_workbook(base).save(base)
            Q.fails.clear()
            Q.audit_numeric(base)
            baseline = set(Q.fails)
            for name, mutate in NUM_MUTANTS:
                safe = "".join(ch if ch.isalnum() else "_" for ch in name)
                tmp = Path(td) / (safe + "_" + book.name)
                shutil.copyfile(book, tmp)
                wb = load_workbook(tmp)
                what = mutate(wb)
                if not what:
                    continue
                wb.save(tmp)
                applied += 1
                Q.fails.clear()
                Q.audit_numeric(tmp)
                if [f for f in set(Q.fails) - baseline if "[NUMERIC]" in f]:
                    killed += 1
                else:
                    survivors.append(
                        f"    SURVIVED [NUMERIC] {name} on {book.name} — {what}")
    return killed, applied, survivors


# ------------------------------------------------------ citation mutants
# This layer shipped DEAD. tools/qa_citations.py was written, wired into
# verify.py, and never called by main() — so for one whole release the pack's
# 163 cross-references were resolved by nothing at all, and the harness said
# "PASSED" with the same confidence it says everything else. That is the exact
# failure a mutation suite exists to catch, and it caught it only because these
# mutants were written afterwards. Anything not mutated here can die the same
# way without anybody noticing.
#
# Citations are cross-document, so unlike the markdown mutants these cannot run
# against one file in isolation: the whole templates directory is copied, one
# document is edited, and the resolver is pointed at the copy.


def _cite_runs(text: str, want_figures: bool):
    """Citation runs in this document, optionally only the ones whose clause
    demands a figure (the ones the resolution check can actually judge)."""
    out = []
    for i, line in enumerate(text.split("\n"), start=1):
        masked = C.RE_RUN.sub(lambda x: " " * len(x.group(0)), line)
        for m in C.RE_RUN.finditer(line):
            if not m.group(1).endswith(".md"):
                continue
            want = C.figures(C.sentence_around(masked, m.start()))
            if want or not want_figures:
                out.append((i, m, want))
    return out


def _retarget(text: str, i: int, m, line_no) -> str:
    lines = text.split("\n")
    lines[i - 1] = (lines[i - 1][:m.start()] + f"{m.group(1)}:{line_no}"
                    + lines[i - 1][m.end():])
    return "\n".join(lines)


def cite_drift(text: str):
    """A citation that no longer lands on its row.

    The pack cites by line number and every edit shifts them — one pass added
    seventeen lines to the charter and silently broke references in four other
    files. The mutant moves a reference onto a line that carries none of the
    figures its sentence claims, and stays clear of the +/-2 window the check
    forgives as an off-by-a-line.
    """
    for i, m, want in _cite_runs(text, want_figures=True):
        tgt = TEMPLATES / m.group(1)
        if not tgt.exists():
            continue
        body = tgt.read_text(encoding="utf-8").split("\n")
        for j, ln in enumerate(body, start=1):
            if not ln.strip():
                continue
            near = "\n".join(body[max(0, j - 3):j + 2])
            if want & {C.norm(x.group(0)) for x in C.RE_FIG.finditer(near)}:
                continue
            return _retarget(text, i, m, j), f"moved {m.group(0)} to :{j}"
    return None, None


def cite_past_eof(text: str):
    """A citation into a document that has since been shortened."""
    for i, m, _ in _cite_runs(text, want_figures=False):
        if (TEMPLATES / m.group(1)).exists():
            return _retarget(text, i, m, 99999), f"sent {m.group(0)} past the end"
    return None, None


def cite_blank(text: str):
    """A citation onto a line with nothing on it — what a deleted row leaves."""
    for i, m, _ in _cite_runs(text, want_figures=False):
        tgt = TEMPLATES / m.group(1)
        if not tgt.exists():
            continue
        body = tgt.read_text(encoding="utf-8").split("\n")
        for j, ln in enumerate(body, start=1):
            if not ln.strip():
                return _retarget(text, i, m, j), f"pointed {m.group(0)} at a blank line"
    return None, None


def cite_paren_drift(text: str):
    """Drift a citation that sits INSIDE brackets, annotating a figure that
    sits outside them — `966 adjustments in the baseline month (09-...md:106)`.

    This is the shape the check was blind to. "(" read as a clause boundary, so
    a bracket holding nothing but the citation produced an empty clause, an
    empty clause claims no figure, and the reference was resolved for existence
    only. 131 of the pack's 157 citations are written that way, so the figure
    test was running on 17% of them and reporting that every citation resolves.
    """
    for i, m, want in _cite_runs(text, want_figures=True):
        line = text.split("\n")[i - 1]
        if m.start() == 0 or line[m.start() - 1] != "(":
            continue
        tgt = TEMPLATES / m.group(1)
        if not tgt.exists():
            continue
        body = tgt.read_text(encoding="utf-8").split("\n")
        for j, ln in enumerate(body, start=1):
            if not ln.strip():
                continue
            near = "\n".join(body[max(0, j - 3):j + 2])
            if want & {C.norm(x.group(0)) for x in C.RE_FIG.finditer(near)}:
                continue
            return _retarget(text, i, m, j), f"parenthetical {m.group(0)} moved to :{j}"
    return None, None


CITE_MUTANTS = [("citation: parenthetical drifted off its row", cite_paren_drift),
                ("citation: drifted off its row", cite_drift),
                ("citation: past the end of the file", cite_past_eof),
                ("citation: onto a blank line", cite_blank)]


def _cite_fails(work: Path) -> set[str]:
    C.fails.clear()
    C.warns.clear()
    C.checked[0] = 0
    cache: dict = {}
    for doc in sorted(work.glob("*.md")):
        C.check_file(doc, cache)
    return set(C.fails)


def run_citations() -> tuple[int, int, list[str]]:
    killed = applied = 0
    survivors = []
    with tempfile.TemporaryDirectory() as td:
        work = Path(td) / "templates"
        shutil.copytree(TEMPLATES, work)
        real = C.TEMPLATES
        C.TEMPLATES = work
        try:
            baseline = _cite_fails(work)
            for doc in sorted(TEMPLATES.glob("*.md")):
                text = doc.read_text(encoding="utf-8")
                for name, mutate in CITE_MUTANTS:
                    out, what = mutate(text)
                    if out is None:
                        continue
                    (work / doc.name).write_text(out, encoding="utf-8")
                    applied += 1
                    if _cite_fails(work) - baseline:
                        killed += 1
                    else:
                        survivors.append(
                            f"    SURVIVED [CITATION] {name} on {doc.name} — {what}")
                    (work / doc.name).write_text(text, encoding="utf-8")
        finally:
            C.TEMPLATES = real
    return killed, applied, survivors


# -------------------------------------------------------- visual mutants
# The visual layer this replaces rendered PNGs and told a human to look at
# them. It made no claim, so it could not fail, so a green run from it meant
# nothing at all. The replacement asserts — and an assertion is only worth
# having if it can be shown to fire.


# ------------------------------------------------------ Word-export mutants
# The visual mutants below cover the PREVIEW svg only. The Word-export checks —
# the ones that caught the worst defect this project has had — had no mutation
# coverage at all, which is how they came to be listed for several sessions as
# "two surviving mutants" that do not exist in this file.
#
# What they caught: the business case draws four charts and the Word copy
# carried their titles, labels and numbers with a blank column down the middle
# where every bar should have been. And, found while fixing that, negative bars
# a fifth of their true length on screen and in every email — a bar crossing
# zero is drawn inside half a split cell and its width was written as a share of
# the whole axis. On the benefit bridge that drew the realisation haircut, a
# fifth of the gross, at a thirtieth of the bar beside it.
#
# Driving a real browser costs about three seconds, so the document and its
# geometry are captured ONCE and the mutants work on copies.


def wm_short_negative_bar(doc, geo):
    """The negative-bar bug exactly: a bar that crosses zero drawn at a
    fraction of its true length, while every label and number stays right."""
    import qa_visual as V
    for rows in geo:
        for r in rows:
            # Genuinely negative, not merely money. Written loosely the first
            # time, this matched the first positive bar instead and still
            # killed — a mutant that passes for the wrong reason proves the
            # check fires on SOMETHING, which is not what its name claims.
            if V._money(r["value"]) < 0 and r["w"] > 20:
                r["w"] = r["w"] / 5.0
                return doc, geo, (f"{r['label']!r} ({r['value']}) redrawn at a fifth "
                                  f"of its length")
    return None, None, None


def wm_zero_line_split(doc, geo):
    """Positive and negative bars sit in different cells, so a common zero line
    is not something the markup gives you for free."""
    for rows in geo:
        if len(rows) > 2:
            rows[1]["x"] = rows[1]["x"] + 40
            return doc, geo, f"{rows[1]['label']!r} shifted off the zero line"
    return None, None, None


def wm_missing_chart(doc, geo):
    """A chart that never reaches the document — the original defect."""
    return doc, geo[:-1], "one chart dropped before it reached the page"


def wm_drop_word_table(doc, geo):
    """A chart that reaches the screen and not the .docx."""
    import xml.etree.ElementTree as ET

    import qa_visual as V
    root = ET.fromstring(doc)
    body = root.find(f"{V.W}body")
    for t in body.findall(f"{V.W}tbl"):
        if t.find(f"{V.W}tblPr/{V.W}tblBorders") is None:
            body.remove(t)
            return ET.tostring(root, encoding="unicode"), geo, "one chart table cut from the .docx"
    return None, None, None


WORD_MUTANTS = [("word: negative bar drawn short", wm_short_negative_bar),
                ("word: bars do not share a zero line", wm_zero_line_split),
                ("word: a chart never drawn", wm_missing_chart),
                ("word: a chart missing from the .docx", wm_drop_word_table)]


skipped: list[str] = []


def run_word() -> tuple[int, int, list[str]]:
    import copy

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import qa_visual as V
    got = V.word_document()
    if got is None:
        # No browser, so there is no document to mutate. Reported as a layer
        # that did not run, never as zero mutants: the caller's `if a:` guard
        # printed nothing at all for a zero-attempt layer, so the suite went on
        # to say "Every check still has teeth" with this whole layer absent.
        skipped.append("(Word export): " + (V.browser_missing() or "no document"))
        return 0, 0, []
    doc, geo = got
    real = V.word_document
    killed = applied = 0
    survivors = []
    try:
        V.fails.clear()
        V.passes[0] = 0
        V.word_document = lambda: (doc, copy.deepcopy(geo))
        V.audit_word_charts()
        baseline = set(V.fails)
        for name, mutate in WORD_MUTANTS:
            d2, g2, what = mutate(doc, copy.deepcopy(geo))
            if what is None:
                continue
            applied += 1
            V.fails.clear()
            V.passes[0] = 0
            V.word_document = lambda d=d2, g=g2: (d, g)
            V.audit_word_charts()
            if set(V.fails) - baseline:
                killed += 1
            else:
                survivors.append(f"    SURVIVED [WORD] {name} — {what}")
    finally:
        V.word_document = real
        V.fails.clear()
        V.passes[0] = 0
    return killed, applied, survivors


def run_visual() -> tuple[int, int, list[str]]:
    import re as _re
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import qa_visual as V
    from sync_html import extract_tpls
    src = (ROOT / "six-sigma-blackbelt-support-ops.html").read_text(encoding="utf-8")
    _, _, tpls = extract_tpls(src)
    svg = next((s for e in tpls.values() if e.get("ext") == "xlsx"
                for s in _re.findall(r'<svg class="xchart".*?</svg>', e["preview"], _re.S)),
               None)
    if svg is None:
        return 0, 0, []
    muts = [
        ("visual: label outside the frame",
         lambda s: _re.sub(r'<text x="[\d.]+"', '<text x="-260"', s, count=1)),
        ("visual: chart that plots nothing",
         lambda s: _re.sub(r"<(rect|circle|path)\b[^>]*/?>", "", s)),
        # Two labels at the same coordinates. The first attempt at this mutant
        # slid a right-edge axis label 90px left, which lands in empty space and
        # collides with nothing — it survived, and the check looked decorative
        # when it was the mutant that was wrong. Duplicating a label in place
        # guarantees the overlap the check exists to find.
        ("visual: two labels written over each other",
         lambda s: (lambda m: s.replace(m.group(0), m.group(0) + m.group(1)
                                        + "Average wait per step" + m.group(3), 1)
                    if m else s)(
             _re.search(r'(<text x="[\d.]+" y="[\d.]+" class="cl">)([^<]{6,})(</text>)', s))),
        # Label grey lightened until it stops being readable. 4.5:1 is the WCAG
        # AA bar for body text; #cfd4dc on white is about 1.5:1, which looks
        # fine in a thumbnail and disappears on a real screen.
        ("visual: axis labels too faint to read",
         lambda s: s.replace("fill:#5b6675", "fill:#cfd4dc")),
    ]
    contrast_mutants = {"visual: axis labels too faint to read"}
    killed = 0
    survivors = []
    for name, mutate in muts:
        V.fails.clear()
        V.passes[0] = 0
        if name in contrast_mutants:
            V.audit_contrast("mutant.xlsx", name, mutate(svg))
        else:
            V.audit_svg("mutant.xlsx", name, mutate(svg))
        if V.fails:
            killed += 1
        else:
            survivors.append(f"    SURVIVED [VISUAL] {name}")
    V.fails.clear()
    return killed, len(muts), survivors


def run_properties() -> tuple[int, int, list[str]]:
    """The property suite went from five permanent failures to none in one
    sitting. That is either five real fixes or five checks quietly declawed, and
    the two look identical from a green run — so pin an axis to the shipped
    example's magnitude and require the harness to say so.
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import qa_properties as P
    book = TEMPLATES / "25-pareto-and-distribution.xlsx"
    if not book.exists():
        return 0, 0, []
    killed = applied = 0
    survivors = []
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / book.name
        shutil.copyfile(book, tmp)
        wb = load_workbook(tmp)
        pinned = False
        for ws in wb.worksheets:
            for ch in getattr(ws, "_charts", []):
                ch.y_axis.scaling.min, ch.y_axis.scaling.max = 0, 500
                pinned = True
        if pinned:
            wb.save(tmp)
            applied = 1
            P.fails.clear()
            P.passes[0] = 0
            try:
                P.test_axis_survives_rescale(tmp)
            except Exception:                                    # noqa: BLE001
                pass
            if P.fails:
                killed = 1
            else:
                survivors.append(
                    "    SURVIVED [PROPERTY] axis pinned to the example's magnitude")
            P.fails.clear()
    return killed, applied, survivors


def audit_to_set(path: Path) -> set[str]:
    Q.fails.clear()
    Q.warns.clear()
    wb = load_workbook(path)
    Q.audit_charts(path, wb)
    Q.audit_guided(path, wb)
    Q.audit_example(path, wb)
    Q.audit_rowlabel(path, wb)
    Q.audit_formula_colour(path, wb)
    return set(Q.fails)


def run(book: Path) -> tuple[int, int, list[str]]:
    """Apply every mutant to this workbook; return (killed, applied, survivors)."""
    killed = applied = 0
    survivors = []
    with tempfile.TemporaryDirectory() as td:
        # Round-tripping through openpyxl loses a little (combo sub-charts, for
        # one), so the baseline is the SAVED copy, not the shipped file. Only a
        # fail that is new against that baseline counts as a kill.
        base = Path(td) / ("base_" + book.name)
        shutil.copyfile(book, base)
        load_workbook(base).save(base)
        baseline = audit_to_set(base)
        for name, mutate, layer in MUTANTS:
            tmp = Path(td) / (name.replace(" ", "_") + "_" + book.name)
            shutil.copyfile(book, tmp)
            wb = load_workbook(tmp)
            try:
                what = mutate(wb)
            except Exception as exc:                             # noqa: BLE001
                # A mutant that raises is evidence of nothing, and it used to
                # print a note and move on. `m_paint_over_the_inputs` died on a
                # NameError against sixteen of the twenty-one workbooks and the
                # suite still reported every mutant killed, because a mutant
                # that never applied is not counted as applied. Same shape as
                # the citation check that was never called: silent absence
                # reading as success. It fails the run now.
                what = None
                survivors.append(f"    ERRORED  [{layer}] {name} on {book.name} "
                                 f"— {type(exc).__name__}: {exc}")
            if not what:
                continue
            applied += 1
            wb.save(tmp)
            new = audit_to_set(tmp) - baseline
            hit = [f for f in new if f"[{layer}]" in f]
            if hit:
                killed += 1
            else:
                survivors.append(f"    SURVIVED [{layer}] {name} — {what}")
    return killed, applied, survivors


# ------------------------------------------------- baseline-series mutants
# The three checks doing the most work in this repo had no mutation coverage,
# which is the failure this whole file exists to catch, sitting in the file
# that catches it. test_baseline_series_recomputes carries seven assertions
# over the baseline's whole distribution; test_worked_example_figures_are_the_packs
# guards the page's narrative against the templates; test_nothing_retired_survives
# is the general registry every past correction is now pinned to. All three
# were mutation-tested by hand when written and by nothing afterwards, so a
# later edit could quietly hollow any of them out while this file went on
# printing "Every check still has teeth".
#
# The containment bug is in here on purpose. The first version of the series
# check tested `value in row` rather than anchoring at the start, so a row
# reading "-1.56, a left skew - restated from the 1.56" passed it while
# telling the reader the opposite sign. An adversarial pass found it. If the
# check is ever loosened back, bl_skew_restated fires.

def bl_sd(text: str):
    if "| Standard deviation | *1.8 points" not in text:
        return None, None
    return (text.replace("| Standard deviation | *1.8 points",
                         "| Standard deviation | *1.4 points"),
            "the standard deviation put back below the floor its own maximum forces")


def bl_rate(text: str):
    if "18.9, 14.0, 12.6" not in text:
        return None, None
    return (text.replace("18.9, 14.0, 12.6", "16.9, 14.0, 12.6"),
            "one weekly rate altered, so the series no longer gives the stated SD")


def bl_skew_restated(text: str):
    """The exact evasion an adversarial pass found in the first version."""
    if "| Skewness | *1.56 — the moment" not in text:
        return None, None
    return (text.replace("| Skewness | *1.56 — the moment",
                         "| Skewness | *−1.56, a left skew — restated from the 1.56 — the moment"),
            "skewness sign flipped while still containing the right digits")


def bl_percentiles(text: str):
    if "| p10 / p50 / p90 / p95 | *12.4% / 14.0% / 14.9% / 16.7%" not in text:
        return None, None
    return (text.replace("| p10 / p50 / p90 / p95 | *12.4% / 14.0% / 14.9% / 16.7%",
                         "| p10 / p50 / p90 / p95 | *11.0% / 13.0% / 18.0% / 18.9%, superseding "
                         "12.4% / 14.0% / 14.9% / 16.7%"),
            "percentiles replaced, with the right set quoted after them as superseded")


def bl_ordinary_count(text: str):
    if "and five of the twelve weekly rates" not in text:
        return None, None
    return (text.replace("and five of the twelve weekly rates",
                         "and six of the twelve weekly rates"),
            "the count of weeks outside the ordinary limits overstated by one")


def bl_ad(text: str):
    if "Anderson-Darling A² = 0.81 on" not in text:
        return None, None
    return (text.replace("Anderson-Darling A² = 0.81 on", "Anderson-Darling A² = 0.93 on"),
            "the Anderson-Darling statistic no longer the one the series gives")


def bl_ppu(text: str):
    if "Ppu −1.15 — (8.0% − 14.2%)" not in text:
        return None, None
    return (text.replace("Ppu −1.15 — (8.0% − 14.2%)", "Ppu −1.48 — (8.0% − 14.2%)"),
            "Ppu no longer following from the mean and SD beside it")


def bl_series_removed(text: str):
    old = "1.8 points, on the twelve weekly rates this section is cut on: 14.9,"
    if old not in text:
        return None, None
    return (text.replace(old, "1.8 points. 14.9,"),
            "the series unpublished, so every statistic goes back to an assertion")


def bl_ucl(text: str):
    if "| UCL / LCL | *17.2% / 11.2%* |" not in text:
        return None, None
    return (text.replace("| UCL / LCL | *17.2% / 11.2%* |", "| UCL / LCL | *14.5% / 11.2%* |"),
            "limits narrowed until five weeks signal, against section 2's claim of one")


BASELINE_MUTANTS = [
    ("baseline: SD below its arithmetic floor", bl_sd),
    ("baseline: a weekly rate altered", bl_rate),
    ("baseline: skewness sign flipped, digits kept", bl_skew_restated),
    ("baseline: percentiles replaced, right set quoted as superseded", bl_percentiles),
    ("baseline: ordinary-limit count overstated", bl_ordinary_count),
    ("baseline: Anderson-Darling detached from the series", bl_ad),
    ("baseline: Ppu detached from the mean and SD", bl_ppu),
    ("baseline: the series unpublished", bl_series_removed),
    ("baseline: limits narrowed past the special cause", bl_ucl),
]


def pg_retired_claim(text: str):
    """A retired claim back in the page's authored prose, wrapped as the page
    wraps it — which is what defeated the matcher before scan() normalised
    whitespace."""
    anchor = "<p><strong>Baseline.</strong> Laney p\u2032 chart over the twelve signed weeks"
    if anchor not in text:
        return None, None
    return (text.replace(anchor,
                         "<p><strong>Baseline.</strong> Laney p\u2032 chart over 52 weeks showed\n"
                         "  a stable process, over the twelve signed weeks"),
            "a retired claim reintroduced across a line wrap")


def pg_figure_drift(text: str):
    if "<strong>$21,294</strong>" not in text:
        return None, None
    return (text.replace("<strong>$21,294</strong>", "<strong>$23,881</strong>"),
            "the walkthrough's realised benefit moved to a figure no template states")


def pg_volume_drift(text: str):
    if "on 61,400 resolved billing" not in text:
        return None, None
    return (text.replace("on 61,400 resolved billing", "on 74,300 resolved billing"),
            "the walkthrough's record count moved away from the baseline's")


def pg_chain_avoidable(text: str):
    if "represents 719 avoidable" not in text:
        return None, None
    return (text.replace("represents 719 avoidable", "represents 1,646 avoidable"),
            "avoidable reopens taken off the whole queue instead of the in-scope population")


def pg_chain_gross(text: str):
    if "= $25,051 gross" not in text:
        return None, None
    return (text.replace("= $25,051 gross", "= $31,343 gross"),
            "gross benefit no longer the reduction at the stated unit cost")


def pg_chain_lesson(text: str):
    """The one that used to survive: the sentence exists in the embedded
    template payload too, so searching the whole file found it there no matter
    what the case study said."""
    old = "so <strong>the project was not bookable as chartered</strong>"
    if old not in text:
        return None, None
    return (text.replace(old, "so <strong>the project was bookable as chartered</strong>"),
            "the lesson sentence deleted from the page while the payload still carries it")


def pg_unit_cost(text: str):
    old = "$38.60 it costs to serve a billing contact twice"
    if old not in text:
        return None, None
    return (text.replace(old, "$6.80 it costs to serve a billing contact twice"),
            "an avoided reopen priced at cost-to-serve, the defect the example teaches against")


PAGE_MUTANTS = [
    ("page: retired claim reintroduced, line-wrapped", pg_retired_claim),
    ("page: walkthrough benefit drifted from the pack", pg_figure_drift),
    ("page: walkthrough record count drifted from the pack", pg_volume_drift),
]

# Run against test_case_study_chain_closes ALONE. These are arithmetic breaks,
# and the drift check kills most of them as a side effect -- a figure that no
# longer follows from the chain is usually also a figure the pack never states.
# Accepting either kill proved nothing about the chain check: deleting one of
# its assertions left every mutant still dying, to the other check.
# Each of these breaks ONE link and leaves the rest of the chain arithmetically
# intact, so it can only be killed by the assertion guarding that link. Without
# them, deleting an assertion left every mutant still dying to a neighbour.

def pg_realised_only(text: str):
    """Moves the realised figure AND the headline that quotes it, so the only
    broken link is realised = gross x factor. Moving realised alone also trips
    the headline assertion, and then deleting this one changes nothing."""
    if "<strong>$21,294</strong>" not in text:
        return None, None
    return (text.replace("<strong>$21,294</strong>", "<strong>$22,900</strong>")
                .replace('<div class="num">$21.3k</div>', '<div class="num">$22.9k</div>'),
            "realised benefit no longer the gross at the stated realisation factor")


def pg_headline_only(text: str):
    if '<div class="num">$21.3k</div>' not in text:
        return None, None
    return (text.replace('<div class="num">$21.3k</div>', '<div class="num">$24.8k</div>'),
            "the headline stat no longer the benefit the page derives below it")


def pg_reduction_only(text: str):
    """Moves the reduction AND the two figures downstream of it, so only the
    'reduction follows from the in-scope population' link is left broken."""
    if "649 reopens avoided annualized" not in text or "= $25,051 gross" not in text:
        return None, None
    out = (text.replace("649 reopens avoided annualized", "700 reopens avoided annualized")
               .replace("= $25,051 gross", "= $27,020 gross")
               .replace("<strong>$21,294</strong>", "<strong>$22,967</strong>")
               .replace('<div class="num">$21.3k</div>', '<div class="num">$23.0k</div>'))
    return out, "reduction no longer 5.6 points of the in-scope population, chain otherwise consistent"

CHAIN_MUTANTS = [
    ("chain: avoidable reopens off the wrong population", pg_chain_avoidable),
    ("chain: gross benefit detached from the reduction", pg_chain_gross),
    ("chain: the lesson sentence deleted", pg_chain_lesson),
    ("chain: a reopen priced at cost-to-serve", pg_unit_cost),
    ("chain: realised detached from gross x factor", pg_realised_only),
    ("chain: headline stat detached from the derivation", pg_headline_only),
    ("chain: reduction off the in-scope population", pg_reduction_only),
]


def run_baseline() -> tuple[int, int, list[str]]:
    """Mutate the artefacts on disk, in a copy, and require each check to fire.

    These checks read files rather than in-memory fixtures, so the whole tree
    is copied and verify's module constants are pointed at the copy. Anything
    less and a surviving mutant would be left in the real repo.
    """
    import importlib

    V = importlib.import_module("verify")
    killed = applied = 0
    survivors: list[str] = []

    with tempfile.TemporaryDirectory() as td:
        work = Path(td)
        shutil.copytree(TEMPLATES, work / "templates")
        page = ROOT / "six-sigma-blackbelt-support-ops.html"
        shutil.copy2(page, work / page.name)
        real_t, real_h = V.TEMPLATES, V.HTML
        V.TEMPLATES, V.HTML = work / "templates", work / page.name

        def fails(fn) -> set[str]:
            V.FAILURES.clear()
            V.PASSES[0] = 0
            fn()
            return set(V.FAILURES)

        try:
            doc = work / "templates" / "09-baseline-document.md"
            clean = doc.read_text(encoding="utf-8")
            base = fails(V.test_baseline_series_recomputes)
            for name, mutate in BASELINE_MUTANTS:
                out, what = mutate(clean)
                if out is None:
                    continue
                doc.write_text(out, encoding="utf-8")
                applied += 1
                if fails(V.test_baseline_series_recomputes) - base:
                    killed += 1
                else:
                    survivors.append(f"    SURVIVED [BASELINE] {name} — {what}")
                doc.write_text(clean, encoding="utf-8")

            html = work / page.name
            clean_h = html.read_text(encoding="utf-8")
            base_w = fails(V.test_worked_example_figures_are_the_packs)
            base_r = fails(V.test_nothing_retired_survives)
            base_c = fails(V.test_case_study_chain_closes)
            for name, mutate in PAGE_MUTANTS:
                out, what = mutate(clean_h)
                if out is None:
                    continue
                html.write_text(out, encoding="utf-8")
                applied += 1
                caught = ((fails(V.test_worked_example_figures_are_the_packs) - base_w)
                          or (fails(V.test_nothing_retired_survives) - base_r))
                if caught:
                    killed += 1
                else:
                    survivors.append(f"    SURVIVED [PAGE] {name} — {what}")
                html.write_text(clean_h, encoding="utf-8")

            for name, mutate in CHAIN_MUTANTS:
                out, what = mutate(clean_h)
                if out is None:
                    continue
                html.write_text(out, encoding="utf-8")
                applied += 1
                if fails(V.test_case_study_chain_closes) - base_c:
                    killed += 1
                else:
                    survivors.append(f"    SURVIVED [CHAIN] {name} — {what}")
                html.write_text(clean_h, encoding="utf-8")
        finally:
            V.TEMPLATES, V.HTML = real_t, real_h
            V.FAILURES.clear()
            V.PASSES[0] = 0
    return killed, applied, survivors


def main() -> int:
    only = [a for a in sys.argv[1:] if not a.startswith("-")]
    books = sorted(TEMPLATES.glob("*.xlsx"))
    if only:
        books = [b for b in books if any(o in b.name for o in only)]
    print(f"Mutation testing the QA harness against {len(books)} workbook(s).\n"
          "Every mutant reintroduces a defect this repo has actually shipped.\n")
    total_k = total_a = 0
    survivors: list[str] = []
    for book in books:
        k, a, s = run(book)
        total_k += k
        total_a += a
        survivors += s
        flag = "" if k == a else "   <-- a check did not fire"
        print(f"  {book.name:36s} {k}/{a} mutants killed{flag}")
    if not only:
        k, a, s = run_markdown()
        total_k += k
        total_a += a
        survivors += s
        docs = len(sorted(TEMPLATES.glob("*.md")))
        print(f"  {'(markdown templates)':36s} {k}/{a} mutants killed"
              f"{'' if k == a else '   <-- a check did not fire'}   across {docs} docs")
        k, a, s = run_control()
        total_k += k
        total_a += a
        survivors += s
        print(f"  {'(control charts)':36s} {k}/{a} mutants killed"
              f"{'' if k == a else '   <-- a check did not fire'}")
        k, a, s = run_guidance()
        total_k += k
        total_a += a
        survivors += s
        print(f"  {'(template filler)':36s} {k}/{a} mutants killed"
              f"{'' if k == a else '   <-- a check did not fire'}")
        k, a, s = run_numeric()
        total_k += k
        total_a += a
        survivors += s
        print(f"  {'(recalculated values)':36s} {k}/{a} mutants killed"
              f"{'' if k == a else '   <-- a check did not fire'}")
        k, a, s = run_citations()
        total_k += k
        total_a += a
        survivors += s
        print(f"  {'(cross-references)':36s} {k}/{a} mutants killed"
              f"{'' if k == a else '   <-- a check did not fire'}")
        k, a, s = run_word()
        total_k += k
        total_a += a
        survivors += s
        if a:
            print(f"  {'(Word export)':36s} {k}/{a} mutants killed"
                  f"{'' if k == a else '   <-- a check did not fire'}")
        else:
            print(f"  {'(Word export)':36s} NOT RUN")
        k, a, s = run_visual()
        total_k += k
        total_a += a
        survivors += s
        print(f"  {'(preview visuals)':36s} {k}/{a} mutants killed"
              f"{'' if k == a else '   <-- a check did not fire'}")
        k, a, s = run_properties()
        total_k += k
        total_a += a
        survivors += s
        if a:
            print(f"  {'(input properties)':36s} {k}/{a} mutants killed"
                  f"{'' if k == a else '   <-- a check did not fire'}")
        k, a, s = run_baseline()
        total_k += k
        total_a += a
        survivors += s
        print(f"  {'(baseline series + page prose)':36s} {k}/{a} mutants killed"
              f"{'' if k == a else '   <-- a check did not fire'}")
    print(f"\n  {total_k}/{total_a} mutants killed across {len(books)} workbooks")

    # README quotes this number as a claim about the pack. verify.py cannot
    # cheaply know it -- the total is mutants x workbooks and only falls out of
    # a real run -- so the check lives where the number does. It went stale at
    # 309 while the suite reached 328, in the paragraph that argues a check you
    # cannot trip is worse than no check at all.
    readme = ROOT / "README.md"
    if readme.exists():
        want = f"{total_a} mutants, all killed"
        if want not in readme.read_text(encoding="utf-8"):
            # If a layer did not run, this run applied fewer mutants than a full
            # one and the README is right while this is wrong. Saying "expected
            # 324" without that caveat proposes a fix that would bless the
            # missing layer — the exact defect this suite exists to catch, in
            # its own failure message.
            why = (f" — but {len(skipped)} layer(s) did not run this time, so this "
                   f"run applied fewer mutants than a full one. Fix the layer, not "
                   f"the README; see the NOT RUN lines below") if skipped else ""
            survivors.append(
                f"    README.md quotes a mutant count that is not this one — "
                f"expected {want!r} for the {total_a} mutants this run applied{why}")
    if survivors:
        print(f"\n{len(survivors)} SURVIVING MUTANT(S) — these checks cannot fail:")
        print("\n".join(sorted(set(survivors))))
        return 1
    if skipped:
        # "Every check still has teeth" is a claim about checks that ran. A
        # layer that never ran has no teeth to report on, and saying it anyway
        # is the exact failure this suite exists to catch.
        print(f"\n{len(skipped)} LAYER(S) NOT RUN — their checks were not tested:")
        print("\n".join("  - " + n for n in skipped))
        return 1
    print("\nEvery check still has teeth.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
