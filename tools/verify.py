#!/usr/bin/env python3
"""Test suite for the workbooks and the single-file HTML.

Three layers:

  STRUCTURE  every formula is checked for the defect class that produced the
             worst bug in this repo - a formula reading a cell that sits inside
             a merged range but is not its top-left anchor, and is therefore
             always empty. The SLA verdict did exactly that and returned
             "NOT CAPABLE" for every possible input.

  NUMERIC    the fixed calculators are recalculated with a real formula engine
             against inputs chosen to expose the original bugs. Needs the
             optional `formulas` package; skipped with a warning if absent.

  SYNC       the four copies of every template agree: templates/*.xlsx, the
             base64 blob in the HTML, the preview tooltips, and docs/index.html.

    python3 tools/verify.py            # everything
    python3 tools/verify.py --fast     # skip the numeric layer

Requires: openpyxl  (+ formulas, optional)
"""
from __future__ import annotations

import base64
import html as H
import json
import re
import shutil
import sys
import tempfile
import warnings
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "tools"))
from sync_html import RE_ROW, RE_SHEET, RE_TD, extract_tpls  # noqa: E402

HTML = ROOT / "six-sigma-blackbelt-support-ops.html"
DOCS = ROOT / "docs" / "index.html"
TEMPLATES = ROOT / "templates"
CALC = "19-black-belt-calculators.xlsx"

FAILURES: list[str] = []
PASSES = [0]


def check(cond: bool, label: str, detail: str = "") -> None:
    if cond:
        PASSES[0] += 1
    else:
        FAILURES.append(f"{label}" + (f"\n      {detail}" if detail else ""))


def approx(a, b, tol=1e-6) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol * max(1.0, abs(float(b)))
    except (TypeError, ValueError):
        return False


# ---------------------------------------------------------------- STRUCTURE
CELL_REF = re.compile(r"(?<![A-Za-z0-9_!$])(\$?)([A-Z]{1,3})(\$?)([0-9]{1,7})(?![0-9(])")


def covered_non_anchor(ws) -> set[str]:
    """Cells inside a merged range that are not its top-left anchor."""
    dead = set()
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(rng))
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    dead.add(f"{get_column_letter(c)}{r}")
    return dead


RE_BOK = re.compile(r"<h5>BOK mapping</h5>\s*<p>(.*?)</p>", re.S)
RE_ASQ_CODE = re.compile(r"\b((?:I|II|III|IV|V|VI|VII|VIII|IX)(?:\.[A-Z])?(?:\.\d+)?)\b")
RE_IASSC_CODE = re.compile(r"\b([1-5]\.[1-5])\b")



SIGNAL_COL = {"I-MR": ("I", 13), "Laney p-prime": ("I", 13),
              "Laney u-prime": ("I", 13), "Xbar-R": ("L", 13),
              "EWMA": ("I", 15), "CUSUM": ("I", 15),
              "t and g (rare events)": ("H", 13)}


def audit_control_signals(sol, name: str) -> None:
    """Every control chart must trip its own rule in its own worked data, and
    must not trip it inside the baseline window.

    Four of the seven shipped a stable run that never crossed a limit, so the
    Signal column was blank down all 24 rows. A control chart is a signal
    detector; one whose example never signals shows the reader the shape of the
    chart and nothing about reading it, and — the part a harness can act on —
    leaves the rule formula never once evaluated against a true case. It could
    have been broken from the day it was written and every other check here
    would still have passed.

    The second half matters more. Xbar-R signalled three times INSIDE its
    baseline window, so every limit on that sheet was computed from a process
    the sheet itself showed was out of control. That is the error the method
    most needs a reader not to make, shipped as the worked example.

    Split out of the test so the mutation suite can drive this exact code
    rather than a copy of it that could agree with a bug.
    """
    for sheet, (col, hdr) in SIGNAL_COL.items():
        fired = [r - hdr for r in range(hdr + 1, hdr + 25)
                 if isinstance(_read(sol, name, sheet, f"{col}{r}"), str)
                 and _read(sol, name, sheet, f"{col}{r}").strip()]
        check(bool(fired),
              f"{sheet}: the worked data trips the chart's own rule",
              "the Signal column is blank on all 24 points — the rule has never "
              "been evaluated against a true case")
        window = _read(sol, name, sheet, "B3")
        if isinstance(window, (int, float)):     # sheets with a frozen baseline
            inside = [p for p in fired if p <= int(window)]
            check(not inside,
                  f"{sheet}: the baseline window is in control",
                  f"points {inside} signal inside the {int(window)}-point baseline, so "
                  f"the limits are computed from a process this sheet shows is not stable")


# Jargon headers that still carry no key line. Each one is a column a reader
# meets cold, in a downloaded file where nothing is clickable. The list only
# ever shrinks: a NEW uncovered term fails the build, and covering one keeps it
# passing, so the work is visible without the gate rotting into a report.
UNGLOSSED: set[str] = set()
RE_JARGON = re.compile(
    r"kappa|sigma|r²|adjusted|std|t-stat|p-value|vif|auc|logit|odds|coeff|intercept|"
    r"slope|residual|fitted|ucl|lcl|centre line|dpmo|dpo|rty|pce|erlang|occupancy|"
    r"shrinkage|percentile|median|variance|component|ndc|appraiser|trial|effect|"
    r"degrees of freedom|confidence|rank|rpn|sev|occ|det|weight|score|u-bar|p-bar|"
    r"moving range|breakeven|npv|payback|roi|discount|cumulative|share|contribution|"
    r"study variation|lead time|touch time|unit of analysis", re.I)


def test_legend_matches_sheets() -> None:
    """Every workbook's colour key names exactly the fills it uses.

    Both directions are defects and both shipped. The Kanban board and the
    Erlang table explained green cells they do not have; the control charts
    and the Gage R&R study used all three and explained one. It reads the
    SAVED workbook, because the first attempt at this derived the key at the
    end of the builder — before the polish pass that adds fills — and so
    reported every workbook clean while two of them shipped wrong.
    """
    from openpyxl import load_workbook
    import xlpolish as X

    for path in sorted(TEMPLATES.glob("*.xlsx")):
        wb = load_workbook(path)
        used = X.fills_used(wb)
        has = {n for n, c, _ in X.LEGEND_KEY if c in used}
        txt = "\n".join(str(c.value) for ws in wb.worksheets
                        for row in ws.iter_rows() for c in row
                        if isinstance(c.value, str))
        says = {n for n, _, _ in X.LEGEND_KEY if n in txt}
        check(says == has,
              f"{path.name}: the colour key matches the sheets",
              f"key says {sorted(says)}, workbook uses {sorted(has)}")

        # A key headed "HOW TO USE THIS SHEET" is answerable for that sheet,
        # not for the workbook. Checking only the workbook lets a sheet with
        # no worked example sit under a key that promises one, because some
        # other tab has green cells and satisfies the total.
        names = [n for n, _, _ in X.LEGEND_KEY]
        for ws in wb.worksheets:
            block = {v.strip() for r in range(1, min(ws.max_row, 14) + 1)
                     for v in [ws.cell(row=r, column=1).value]
                     if isinstance(v, str) and v.strip() in names}
            if not block:
                continue
            here = {n for n, c, _ in X.LEGEND_KEY if c in X.sheet_fills(ws)}
            check(block == here,
                  f"{path.name} / {ws.title}: the sheet's own key matches it",
                  f"key says {sorted(block)}, sheet uses {sorted(here)}")


def test_chart_notes() -> None:
    """Every chart says how to read it and what would worry you.

    A chart drawn from live cells that nobody can interpret is decoration, which
    is the same defect as a check that cannot fail. The titles carried a hint at
    best; a reader who does not already know what a residual plot is could not
    act on "a shapeless cloud is the only shape you can use".
    """
    sys.path.insert(0, str(ROOT / "tools"))
    import chart_notes
    import chartsvg
    from xlpolish import _chart_title as xlpolish_title

    missing = []
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        for _sheet, specs in (chartsvg.charts_by_sheet(path) or {}).items():
            for spec in specs:
                t = (spec.get("title") or "").strip()
                if not t:
                    continue
                if t not in chart_notes.NOTES:
                    missing.append(f"{path.name}: {t!r} (no note)")
                elif t not in chart_notes.ACTIONS:
                    missing.append(f"{path.name}: {t!r} (note but no action)")
    check(not missing, "every chart says how to read it AND what to do about it",
          f"{len(missing)} without one: {missing[:3]}")

    # The note has to be in the FILE, not only the preview. This pack is
    # downloaded; a caption that lives in the page is no use in Excel. It also
    # catches the anchor bug that made this land in the nine patched workbooks
    # and none of the thirteen generated ones while the build stayed green.
    from openpyxl import load_workbook as _lw
    absent = []
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        wb = _lw(path)
        want = sum(1 for ws in wb.worksheets for ch in getattr(ws, "_charts", [])
                   if xlpolish_title(ch) in chart_notes.NOTES)
        found = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("HOW TO READ IT."))
        if found < want:
            absent.append(f"{path.name} {found}/{want}")
    check(not absent, "the chart notes are in the workbooks, not only the preview",
          f"missing in: {absent[:4]}")

    # Nothing the build writes into a cell may stack. The key line under each
    # header is rewritten on every run, and when its strip pattern stopped
    # matching its own output the key was appended instead of replaced —
    # four copies of the same paragraph on eleven sheets, and every gate green,
    # because nothing compared a cell with itself.
    stacked = []
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        for ws in _lw(path).worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if isinstance(v, str) and (v.count("Key:") > 1
                                               or v.count("HOW TO READ IT.") > 1):
                        stacked.append(f"{path.name} {ws.title}!{c.coordinate}")
    check(not stacked, "no cell carries the same generated block twice",
          f"{len(stacked)} stacked: {stacked[:4]}")


def test_idempotent() -> None:
    """Running the finishing pass twice must produce the same bytes.

    test_deterministic checks that the timestamps are frozen, which is a
    PRECONDITION of a meaningful diff and not the thing itself — its own
    docstring says it avoids building twice because that would double the run.
    So a pass that APPENDS on every run still writes frozen-timestamp bytes and
    sails through: the key line under each header stopped matching its own
    output, stacked to four copies on eleven sheets, and every gate stayed
    green because nothing ever ran the pass twice and compared.

    This runs polish_workbook twice over a copy of each workbook and diffs the
    result. It is cheap — no rebuild, no recalculation — and it covers the half
    of the pack that accumulates: the workbooks patched in place rather than
    generated from nothing, which is where both of this release's stacking and
    anchor bugs landed.
    """
    import shutil
    import tempfile

    from openpyxl import load_workbook as _lw

    sys.path.insert(0, str(ROOT / "tools"))
    from xlpolish import polish_workbook, save_workbook

    drifted = []
    with tempfile.TemporaryDirectory() as td:
        for path in sorted(TEMPLATES.glob("*.xlsx")):
            a, b = Path(td) / "a.xlsx", Path(td) / "b.xlsx"
            shutil.copyfile(path, a)
            wb = _lw(a)
            polish_workbook(wb)
            save_workbook(wb, a)
            shutil.copyfile(a, b)
            wb = _lw(b)
            polish_workbook(wb)
            save_workbook(wb, b)
            # CONTENT, not bytes. openpyxl's own round trip is not byte-stable
            # — load-and-save re-emits xl/styles.xml differently on three of
            # these workbooks whatever pass is run, including one that only
            # writes document properties. Comparing bytes measures openpyxl;
            # comparing cells measures us, and it is us that stacked four
            # copies of a key line onto eleven sheets.
            wa, wbk = _lw(a), _lw(b)
            for s1, s2 in zip(wa.worksheets, wbk.worksheets):
                for r1, r2 in zip(s1.iter_rows(), s2.iter_rows()):
                    for c1, c2 in zip(r1, r2):
                        if c1.value != c2.value:
                            drifted.append(f"{path.name} {s1.title}!{c1.coordinate}")
                            break
                if {str(m) for m in s1.merged_cells.ranges} != \
                        {str(m) for m in s2.merged_cells.ranges}:
                    drifted.append(f"{path.name} {s1.title} merges")
    check(not drifted,
          "the finishing pass is idempotent — a second run writes no new content",
          f"{len(drifted)} cell(s)/sheet(s) move on a second pass: {drifted[:4]}")


# Generated blocks that are too dense to read comfortably. The threshold is
# what a reader can take, not what the pack currently does: 350 characters
# without a line break, or 1,500 in one cell. Choosing a looser number so the
# present state passes is how a gate becomes decoration, and this release has
# spent most of its commits undoing exactly that.
#
# Empty. Anything that trips this now is new. The set only shrinks — a NEW dense block fails
# the build immediately, and one that gets shortened has to leave the list, so
# it cannot rot in either direction.
DENSE_OK: set[str] = set()   # empty, and it must stay that way
MAX_LINE, MAX_BLOCK = 350, 1500


def test_readable_blocks() -> None:
    """Nothing the build generates may arrive as a wall of text.

    The user found four stacked copies of a key line by looking at the page,
    and before that a seven-gloss key run together with separators into 1,400
    characters of unbroken prose. Every gate was green for both. Nothing here
    measured what a reader actually meets — the checks read structure, so a
    correct, complete, unreadable block passed every one of them.
    """
    from openpyxl import load_workbook as _lw
    marks = ("Key:", "HOW TO READ IT.")
    fresh = []
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        for ws in _lw(path).worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if not isinstance(v, str) or not any(m in v for m in marks):
                        continue
                    # The defect was never paragraph LENGTH — a 400-character
                    # note in a wrapped cell reads fine, and Excel wraps it.
                    # It was seven separate definitions run together with "·"
                    # into one undifferentiated line. So the line rule applies
                    # only where a block holds SEVERAL entries: those must be
                    # separated. A single paragraph is judged on total size.
                    entries = v.count(" = ")
                    runon = (entries >= 2
                             and max(len(seg) for seg in v.split("\n")) > MAX_LINE)
                    if runon or len(v) > MAX_BLOCK:
                        where = f"{path.name} {ws.title}!{c.coordinate}"
                        if where not in DENSE_OK:
                            fresh.append(where)
    check(not fresh, "no new wall of generated text",
          f"{fresh[:4]} — break it into lines, shorten it, or move it off the "
          f"reader's path; add to DENSE_OK only with a reason")
    print(f"           {len(DENSE_OK)} generated block(s) still too dense to read easily")


def test_glossary_coverage() -> None:
    """Every jargon column header should carry a plain-English key.

    On the page every acronym is clickable. In a downloaded workbook nothing is,
    so a reader meets "Widest trial spread" or "% Study variation" with no way to
    find out what it holds. xlpolish writes a key line under any header row whose
    labels it can explain — it just could not explain many of them: 49 jargon
    headers across twelve workbooks had no key at all, and 232 headers of all
    kinds had none.

    The gap is tracked as a set rather than a count, so a header added tomorrow
    without a gloss fails even while the backlog is being worked through, and
    the backlog itself cannot quietly grow.
    """
    from openpyxl import load_workbook as _lw

    sys.path.insert(0, str(ROOT / "tools"))
    import xlpolish

    # A dict literal takes the LAST value for a repeated key, silently. The page
    # glossary shipped fifteen terms defined twice for exactly that reason, and
    # it was not consistently the better definition that survived; adding a
    # richer RPN entry here re-created the same defect within one commit.
    import re as _re
    blk = (ROOT / "tools" / "xlpolish.py").read_text(encoding="utf-8")
    blk = blk[blk.index("GLOSSES = {"):blk.index("_GL_LOW")]
    keys = _re.findall(r'^\s{4}"((?:[^"\\]|\\.)+)":', blk, _re.M)
    dupes = sorted({k for k in keys if keys.count(k) > 1})
    check(not dupes, "no workbook gloss is defined twice",
          f"silently overwritten: {dupes}")

    covered = xlpolish.glossed          # ONE matcher, not a second copy of it

    bare = set()
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        wb = _lw(path)
        for ws in wb.worksheets:
            if ws.title.lower().startswith(("how to use", "read me", "legend")):
                continue
            for row in ws.iter_rows():
                for c in row:
                    try:
                        dark = (c.fill and c.fill.patternType
                                and str(c.fill.fgColor.rgb) == "FF333C49")
                    except Exception:                            # noqa: BLE001
                        dark = False
                    if not (dark and isinstance(c.value, str) and c.value.strip()):
                        continue
                    h = c.value.strip()
                    if h.startswith("=") or len(h) > 60:
                        continue
                    if RE_JARGON.search(h) and not covered(h):
                        bare.add(h)
    new = sorted(bare - UNGLOSSED)
    check(not new, "no new jargon header ships without a plain-English key",
          f"{new} — add a gloss to xlpolish.GLOSSES, or to UNGLOSSED with a reason")
    done = sorted(UNGLOSSED - bare)
    check(not done, "the unglossed backlog list has no stale entries",
          f"these now have a key and can leave UNGLOSSED: {done}")
    print(f"           {len(bare)} jargon header(s) still without a key "
          f"(49 at the start of this pass)")


def test_case_study() -> None:
    """The curriculum's worked project says its figures are internally
    consistent. That has to be true of the one a reader would quote.

    It was true of every step except the last. Define closes: 480,000 contacts
    x 6.2 points is 29,800 avoidable, x $6.80 is $203,000. Results closes as
    far as the volume: 14.2% to 8.6% is 5.6 points, x 480,000 is 26,900. Then
    the benefit was stated as $190,400, which is $7.08 a contact — against the
    $6.80 the same page states in Define, and above the gross the page's own
    numbers produce. The headline stat repeated it. A benefit a reader cannot
    reproduce from the numbers printed beside it is the exact defect this
    example exists to teach against, sitting in the example.
    """
    src = HTML.read_text(encoding="utf-8")

    def num(pattern: str):
        m = re.search(pattern, src)
        return float(m.group(1).replace(",", "")) if m else None

    vol = num(r"across ([\d,]+) contacts")
    avoidable = num(r"represents ([\d,]+) avoidable contacts")
    cost = num(r"\$([\d.]+) = \$[\d,]+ gross")
    cut = num(r"Volume reduction ([\d,]+) contacts annualized")
    gross = num(r"= \$([\d,]+) gross")
    realised = num(r"benefit\s*</?\w*>?\s*\$([\d,]+)")
    factor = num(r"gross, &times; ([\d.]+) realisation")
    check(None not in (vol, avoidable, cost, cut, gross, realised, factor),
          "the case study's benefit chain is still readable from the page",
          f"vol={vol} avoidable={avoidable} cost={cost} cut={cut} "
          f"gross={gross} realised={realised} factor={factor}")
    if None in (vol, avoidable, cost, cut, gross, realised, factor):
        return
    check(abs(vol * 0.062 - avoidable) < 100,
          "case study: the avoidable-contact count follows from volume x the gap",
          f"{vol:,.0f} x 6.2 points = {vol * 0.062:,.0f}, page says {avoidable:,.0f}")
    check(abs(vol * 0.056 - cut) < 100,
          "case study: the volume reduction follows from the improvement achieved",
          f"{vol:,.0f} x 5.6 points = {vol * 0.056:,.0f}, page says {cut:,.0f}")
    check(abs(cut * cost - gross) < 100,
          "case study: gross benefit is the volume reduction at the stated unit cost",
          f"{cut:,.0f} x ${cost} = {cut * cost:,.0f}, page says {gross:,.0f}")
    check(abs(gross * factor - realised) < 100,
          "case study: the realised benefit is the gross at the stated factor",
          f"{gross:,.0f} x {factor} = {gross * factor:,.0f}, page says {realised:,.0f}")
    head = re.search(r'<div class="num">\$(\d+)k</div><div class="lbl">Validated annu', src)
    check(head and abs(int(head.group(1)) * 1000 - realised) < 1000,
          "case study: the headline stat matches the benefit the page derives",
          f"headline ${head.group(1) if head else '?'}k against ${realised:,.0f}")


def test_guidance() -> None:
    """tools/md_guidance.py holds a SECOND copy of the pack's worked example.

    It fills the blank fields of a markdown template with the billing-adjustment
    case, which means it carries its own copy of every figure the pack states —
    and a second copy of a number is a number that will drift. It had. It still
    held the benefit chain that was withdrawn for being causally impossible, so
    regenerating an emptied template re-injected 61,400 tickets, 3,807 reopens
    avoided and $22,005 realized, none of which survive anywhere else.

    Worse, it rewrote each document's "how to use this" block unconditionally,
    and its copy of the preamble said "a 14.2% 7-day reopen rate" with no
    population named — the exact defect the reconciliation removed, since the
    whole queue and the in-scope adjustments both run 14.2% and multiplying one
    by the other is how the case died. One run of the tool would have put that
    sentence back into all eleven documents.

    Nothing in the build invokes it, which is precisely why it rotted unseen.
    So the build checks it instead: every figure the tool would write has to be
    one the templates already carry, and every row it addresses has to exist.
    """
    sys.path.insert(0, str(ROOT / "tools"))
    import md_guidance
    import qa_citations

    def figs(text: str) -> set:
        out = set()
        for m in qa_citations.RE_FIG.finditer(text or ""):
            n = qa_citations.norm(m.group(0))
            if len(n.replace(".", "")) >= 3 or "$" in m.group(0):
                out.add(n)
        return out

    # The preamble is the dangerous string: it is rewritten on every run, so a
    # stale one cannot be protected by "only fills blanks".
    preamble = md_guidance.block("01-project-charter")
    check("OD-BIL-004-ADJ" in preamble,
          "the guidance preamble names the population it measures",
          "it says '14.2% reopen rate' with no population — the whole queue and "
          "the in-scope adjustments both run 14.2% and they are different quantities")

    dead, drift = [], []
    for name, values in md_guidance.EXAMPLE.items():
        path = TEMPLATES / f"{name}.md"
        if not path.exists():
            dead += [f"{name} (no such template)"] if values else []
            continue
        rows: dict = {}
        for line in path.read_text(encoding="utf-8").split("\n"):
            if line.startswith("|"):
                cells = line.split("|")
                if len(cells) >= 3:
                    rows.setdefault(cells[1].strip(), []).append(" | ".join(cells[2:-1]))
        for label, val in values.items():
            if label not in rows:
                dead.append(f"{name}: {label!r}")
                continue
            want = figs(val)
            if want and not (want & set().union(*(figs(r) for r in rows[label]))):
                drift.append(f"{name}: {label!r} offers {sorted(want)[:2]}, "
                             f"the document has none of them")
    check(not drift,
          f"every figure md_guidance would write is one the pack states "
          f"({sum(len(v) for v in md_guidance.EXAMPLE.values())} values)",
          "; ".join(drift[:5]))
    check(not dead, "every row md_guidance addresses still exists",
          "; ".join(dead[:6]))


def test_citations() -> None:
    """A cross-reference has to land on what it claims.

    The templates cite each other by LINE NUMBER, which is not a stable
    address: editing any document shifts every pointer into it from every
    sibling. One pass added seventeen lines to the charter and silently broke
    references in four other files. tools/qa_citations.py resolves them.
    """
    import qa_citations

    qa_citations.fails.clear()
    qa_citations.warns.clear()
    qa_citations.checked[0] = 0
    cache: dict = {}
    for path in sorted(qa_citations.TEMPLATES.glob("*.md")):
        qa_citations.check_file(path, cache)
    check(qa_citations.checked[0] > 100,
          f"the citation check walked the pack ({qa_citations.checked[0]} references)",
          "it found almost nothing to resolve — the reference format has changed")
    for w in qa_citations.warns:
        print(f"           warn {w}")
    check(not qa_citations.fails,
          "every cross-reference resolves to a line that carries the figure",
          "; ".join(qa_citations.fails[:4]))


def test_bok() -> None:
    """Every body-of-knowledge section the page cites has to exist.

    The page asserts certification coverage on all 23 modules, and an assertion
    nobody checks is a claim rather than a mapping. It cited six ASQ sections
    that exist in neither the 2015 nor the 2022 body of knowledge — I.C, I.D,
    I.E, III.E, VII.D, VIII.E — and several that exist but mean something else,
    each with a parenthetical contradicting the real section title: "IV.B (VOC)"
    where IV.B is the business case, "V.E (MSA)" where V.E is probability, "VI.D
    (correlation, linear/multiple regression)" where VI.D is gap, root cause and
    waste analysis.

    tools/bok.py holds both outlines as data, transcribed from the two documents
    the page's own reference list links to, so the claim and the check cannot
    drift apart.
    """
    sys.path.insert(0, str(ROOT / "tools"))
    from bok import ALL_ASQ, IASSC as IASSC_BOK, MODULE_MAP, render
    src = HTML.read_text(encoding="utf-8")

    # The line on the page is rendered from MODULE_MAP, so drift between the two
    # means somebody typed a mapping by hand again — which is the habit that put
    # six nonexistent sections on the page in the first place.
    # Counted, because this loop silently stopped covering 25 of the 26 modules
    # once they gained id attributes: the split pattern was the literal
    # `<details class="mod">`, which after that matched nothing, so the whole
    # file arrived as one block, the first module id won, and every other
    # module's line went unchecked. A loop that can quietly examine one item
    # instead of twenty-six is the same failure as a check that cannot fail.
    seen = 0
    for block in re.split(r'(?=<details class="mod"[ >])', src):
        m = re.search(r'<span class="mid">(M\d+)</span>', block)
        if not m:
            continue
        seen += 1
        want = render(m.group(1))
        got = re.search(r"<h5>BOK mapping</h5>\s*<p>(.*?)</p>", block, re.S)
        check(want is not None, f"{m.group(1)} is declared in MODULE_MAP",
              "the module is on the page but tools/bok.py does not map it")
        if want is None:
            continue
        check(got is not None and got.group(1).strip() == want,
              f"{m.group(1)} BOK line matches tools/bok.py",
              f"page says {got.group(1).strip()[:60]!r} but MODULE_MAP renders "
              f"{want[:60]!r} — run tools/apply_bok.py" if got else
              "the module carries no BOK mapping at all")

    check(seen == len(MODULE_MAP), "the BOK check walked every module",
          f"it examined {seen} of {len(MODULE_MAP)} — the module-block split has "
          "stopped matching, so most modules are going unchecked")

    # The headline coverage claim, checked against what the modules actually
    # map to. It read "the full ASQ CSSBB and IASSC bodies of knowledge" in six
    # places — the lede, three meta descriptions and two JSON-LD blocks — while
    # three ASQ sections had no module at all. A number in the claim is worth
    # more than the word "full" precisely because a number can be wrong.
    from bok import coverage as _coverage
    asq_cov, iassc_cov = _coverage()
    a_have = sum(1 for v in asq_cov.values() if v)
    i_have = sum(1 for v in iassc_cov.values() if v)
    full = a_have == len(asq_cov)
    claim = ("the complete ASQ CSSBB" if full
             else f"{a_have} of the {len(asq_cov)} ASQ CSSBB sections")
    check(claim in src, "the coverage claim matches the modules",
          f"the page should say {claim!r}; the word 'complete' is only allowed "
          f"while it is true, and {a_have} of {len(asq_cov)} sections are covered")
    check(full or "complete ASQ" not in src,
          "the page claims complete ASQ coverage only while it has it",
          f"{a_have} of {len(asq_cov)} ASQ sections are covered")
    check(i_have == len(iassc_cov) or "complete IASSC" not in src,
          "the page claims complete IASSC coverage only while it has it",
          f"{i_have} of {len(iassc_cov)} IASSC sections are covered")

    # Every chip in the coverage table is a link into the module it names. One
    # module ships expanded, as <details class="mod" open>, and the first pass
    # gave anchors only to the ones written <details class="mod"> exactly — so
    # that module's chips pointed at nothing.
    mod_ids = set(re.findall(r'<details class="mod" id="(M\d+)"', src))
    targets = set(re.findall(r'class="pill" href="#(M\d+)"', src))
    check(not targets - mod_ids, "every coverage-table link lands on a module",
          f"{sorted(targets - mod_ids)} are linked but have no anchor")
    check(len(mod_ids) == len(MODULE_MAP),
          "every module carries an anchor to link at",
          f"{len(mod_ids)} anchors for {len(MODULE_MAP)} modules")

    mappings = RE_BOK.findall(src)
    check(len(mappings) >= 20, "every module declares a BOK mapping",
          f"only {len(mappings)} found — the curriculum claims full coverage")
    for text in mappings:
        plain = re.sub(r"<[^>]+>", "", text)
        asq_part, _, iassc_part = plain.partition("IASSC")
        for code in set(RE_ASQ_CODE.findall(asq_part)):
            check(code in ALL_ASQ, f"ASQ {code} is a real BOK section",
                  f"cited in {plain.strip()[:70]!r} but no such section exists in the "
                  "ASQ CSSBB body of knowledge")
        for code in set(RE_IASSC_CODE.findall(iassc_part)):
            check(code in IASSC_BOK, f"IASSC {code} is a real BOK section",
                  f"cited in {plain.strip()[:70]!r} but no such section exists")


def test_export_charts() -> None:
    """The email / standalone export has to carry the charts and the legend.

    It walked the preview's tables and never its charts, so a workbook whose
    whole point is a control chart exported as a wall of numbers — and the
    closing note still described only yellow and blue, having never learned
    that green now means the worked example you replace.

    Checked at the source, because the export is built in the browser: the
    behaviour itself was confirmed by opening it and counting two SVGs in the
    Pareto's export body where there had been none.
    """
    src = HTML.read_text(encoding="utf-8")
    i = src.find("function tplEmailHTML")
    body = src[i:src.find("\n}", i)] if i >= 0 else ""
    check(bool(body), "the email export builder is present")
    # What this used to assert: that the string "svg.xchart" appeared in the
    # function's source. A check on the code, not on its output — and it passed
    # while the Pareto's Word export opened as a paragraph of raw CSS followed
    # by its axis labels one per line, because inline SVG that a client cannot
    # draw is stripped down to its TEXT CHILDREN rather than removed.
    # qa_visual.audit_template_export runs the real function in a real browser
    # and reads what comes out. This one now only asserts the shape the output
    # check depends on.
    check("svg.xchart" in body and "image/svg+xml" in body,
          "the email export turns each chart into a data-URI image",
          "an <img> cannot leak text into a client that will not draw it; "
          "inline SVG degrades to its own <style> block and labels")
    for colour in ("Green cells", "Yellow cells", "Blue cells"):
        check(colour in body, f"the export legend explains {colour.lower()}",
              "the note tells the reader what a colour means; leaving one out "
              "is how the pack ended up with a legend nobody could trust")


def test_deterministic() -> None:
    """Every shipped workbook carries frozen timestamps, so a diff means something.

    An .xlsx records the time in three independent places — the created and
    modified properties in docProps/core.xml, and a modification time on every
    zip entry — and openpyxl writes the clock into all three. So a rebuild that
    changed nothing still produced twenty modified binary files plus the page
    that embeds them, and `git diff` could never answer "did that build actually
    do anything?". It hid a real one: a chart was re-anchored and a note's font
    rewritten on every single run, growing the style table by one entry per
    build, for as long as this repo has existed.

    Checked here rather than by building twice, which would double every run.
    """
    import zipfile
    from xlpolish import ZIP_EPOCH
    stamp = "2026-01-01T00:00:00Z"
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        z = zipfile.ZipFile(path)
        core = z.read("docProps/core.xml").decode("utf-8", "replace")
        for field in ("created", "modified"):
            got = re.search(rf"<dcterms:{field}[^>]*>([^<]*)<", core)
            check(bool(got) and got.group(1) == stamp,
                  f"{path.name} docProps {field} frozen",
                  f"is {got.group(1) if got else '(absent)'}, not {stamp} — the "
                  "bytes move on every build and a diff stops meaning anything")
        moving = [i.filename for i in z.infolist() if i.date_time != ZIP_EPOCH]
        check(not moving, f"{path.name} zip entry timestamps frozen",
              f"{len(moving)} entry/entries carry a build-time clock: {moving[:3]}")


def test_structure() -> None:
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        wb = load_workbook(path)
        for ws in wb.worksheets:
            dead = covered_non_anchor(ws)
            if not dead:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    # ignore quoted string literals inside the formula
                    bare = re.sub(r'"[^"]*"', '""', v)
                    for m in CELL_REF.finditer(bare):
                        ref = f"{m.group(2)}{m.group(4)}"
                        if ref in dead:
                            check(
                                False,
                                f"{path.name} [{ws.title}] {cell.coordinate} reads {ref}, "
                                f"which is swallowed by a merged range and is always empty",
                                v[:120],
                            )
            PASSES[0] += 1


# ------------------------------------------------------------------ NUMERIC
def _engine(path: Path):
    warnings.filterwarnings("ignore")
    import formulas

    return formulas.ExcelModel().loads(str(path)).finish()


def _read(sol, fname: str, sheet: str, cell: str):
    key = f"'[{fname.upper()}]{sheet.upper()}'!{cell}"
    for k, v in sol.items():
        if k.upper() == key:
            try:
                return v.value[0, 0]
            except Exception:
                return v
    return "<missing>"


def recalc(scenarios: list[tuple[str, str, object]], sheets_cells):
    """Copy the calculator workbook, apply inputs, recalculate, read outputs."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / CALC
        shutil.copyfile(TEMPLATES / CALC, tmp)
        if scenarios:
            wb = load_workbook(tmp)
            for sheet, cell, val in scenarios:
                wb[sheet][cell] = val
            wb.save(tmp)
        sol = _engine(tmp).calculate()
        return {(s, c): _read(sol, CALC, s, c) for s, c in sheets_cells}


def test_numeric() -> None:
    K = "2 QA agreement (kappa)"
    S = "3 SLA capability"
    R = "9 ROI and payback"

    # --- kappa verdict must follow kappa (B13), not chance agreement (B12).
    # Perfect agreement: the old formula read Pe = 0.50 and said MARGINAL.
    out = recalc([(K, "B5", 35), (K, "B6", 0), (K, "B7", 0), (K, "B8", 35)], [(K, "B13"), (K, "B14")])
    check(approx(out[(K, "B13")], 1.0), "kappa: perfect agreement gives kappa = 1.0", repr(out[(K, "B13")]))
    check(str(out[(K, "B14")]).startswith("EXCELLENT"),
          "kappa verdict: perfect agreement reads EXCELLENT (was MARGINAL)", repr(out[(K, "B14")]))

    # Worse than chance: the old formula read Pe = 0.918 and said EXCELLENT.
    out = recalc([(K, "B5", 64), (K, "B6", 3), (K, "B7", 3), (K, "B8", 0)], [(K, "B13"), (K, "B14")])
    check(float(out[(K, "B13")]) < 0, "kappa: worse-than-chance is negative", repr(out[(K, "B13")]))
    check(str(out[(K, "B14")]).startswith("UNACCEPTABLE"),
          "kappa verdict: worse-than-chance reads UNACCEPTABLE (was EXCELLENT)", repr(out[(K, "B14")]))

    # --- SLA verdict must follow Ppu (B13). The old formula read B15, an empty
    #     merged cell, so every input returned NOT CAPABLE.
    out = recalc([(S, "B5", 8), (S, "B6", 2.6), (S, "B7", 1.1)], [(S, "B13"), (S, "B17")])
    check(approx(out[(S, "B13")], (8 - 2.6) / (3 * 1.1)), "SLA: Ppu arithmetic", repr(out[(S, "B13")]))
    check(out[(S, "B17")] == "CAPABLE",
          "SLA verdict: a capable process reads CAPABLE (was always NOT CAPABLE)", repr(out[(S, "B17")]))

    # Ppu = (6.5 - 2.6) / (3 x 1.1) = 1.18, i.e. inside the 1.00-1.33 band.
    out = recalc([(S, "B5", 6.5), (S, "B6", 2.6), (S, "B7", 1.1)], [(S, "B13"), (S, "B17")])
    check(1.0 <= float(out[(S, "B13")]) < 1.33, "SLA: test input lands in the marginal band",
          repr(out[(S, "B13")]))
    check(out[(S, "B17")].startswith("MARGINAL"), "SLA verdict: middle band reads MARGINAL", repr(out[(S, "B17")]))

    out = recalc([], [(S, "B17")])  # shipped example, Ppu 0.42
    check(out[(S, "B17")].startswith("NOT CAPABLE"),
          "SLA verdict: shipped example still reads NOT CAPABLE", repr(out[(S, "B17")]))

    # --- The workbook must not state two answers to its own question. The ROI
    #     tab's realised benefit is a yellow input, and rightly so — a generic
    #     calculator has to accept a benefit from anywhere. But its WORKED
    #     EXAMPLE was $156,672 while tab 8 computed $172,012.80 from the inputs
    #     shipped beside it, so one workbook disagreed with itself by $15,340.80
    #     with nothing to say which was meant. The example now carries tab 8's
    #     own output, and this keeps the two tied as either changes.
    out = recalc([], [(R, "B6"), ("8 Benefit — avoided contacts", "B14")])
    roi_in = float(out[(R, "B6")])
    tab8 = float(out[("8 Benefit — avoided contacts", "B14")])
    check(approx(roi_in, tab8, 1e-4),
          "ROI: the worked example equals the benefit the workbook computes for it",
          f"tab 9 states {roi_in:,.2f}, tab 8 computes {tab8:,.2f} — "
          f"{abs(roi_in - tab8):,.2f} apart")

    # --- NPV must honour any number of years, matching the HTML card's loop.
    for years in (1, 3, 5, 10):
        out = recalc([(R, "B5", 193000), (R, "B6", 156672), (R, "B7", years), (R, "B8", 0.1)], [(R, "B15")])
        want = -193000 + sum(156672 / (1.1 ** y) for y in range(1, years + 1))
        check(approx(out[(R, "B15")], want, 1e-6),
              f"ROI: NPV correct at {years} year(s)", f"got {out[(R, 'B15')]!r} want {want:,.2f}")

    out = recalc([(R, "B5", 193000), (R, "B6", 156672), (R, "B7", 3), (R, "B8", 0)], [(R, "B15")])
    check(approx(out[(R, "B15")], -193000 + 156672 * 3), "ROI: NPV handles a zero discount rate",
          repr(out[(R, "B15")]))

    # --- zero defects used to make NORMSINV(1) a #NUM!.
    out = recalc([("1 Sigma level", "B7", 0)], [("1 Sigma level", "B15")])
    check(not isinstance(out[("1 Sigma level", "B15")], float) or out[("1 Sigma level", "B15")] == out[("1 Sigma level", "B15")],
          "sigma: zero defects does not raise #NUM!", repr(out[("1 Sigma level", "B15")]))


def test_numeric_other() -> None:
    """Hypothesis log and VSM, which live in their own workbooks."""
    # Hypothesis log: a blank practical threshold must not be read as "matters".
    src = TEMPLATES / "13-hypothesis-test-log.xlsx"
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        ws = wb["Test log"]
        ws["M11"] = None  # clear the practical threshold, keep p and effect
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        got = _read(sol, src.name, "Test log", "N11")
        check(str(got).strip() in ("", "0"),
              "hypothesis log: blank practical threshold leaves the verdict blank "
              "(was 'YES - real and matters')", repr(got))

    # Durability counters must survive a free-text paste in the hierarchy column
    # instead of turning the whole summary into #VALUE!, and must still count.
    src = TEMPLATES / "15-solution-selection-matrix.xlsx"
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        ws = wb["Solution selection"]
        rows = [("a", "1 Eliminate demand", "Yes"), ("b", "3 Guide it", "Yes"),
                ("c", "2 Design it out", "Yes"), ("d", "6 Train and remind", "Yes"),
                ("e", "not a level at all", "Yes"), ("f", "1 Eliminate demand", "No")]
        for i, (name, lvl, sel) in enumerate(rows, start=14):
            ws[f"B{i}"], ws[f"L{i}"], ws[f"M{i}"] = name, lvl, sel
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        e37 = _read(sol, src.name, "Solution selection", "E37")
        e38 = _read(sol, src.name, "Solution selection", "E38")
        check(approx(e37, 3), "solution matrix: durable count survives a junk paste", f"got {e37!r} want 3")
        check(approx(e38, 1), "solution matrix: decaying count survives a junk paste", f"got {e38!r} want 1")

    src = TEMPLATES / "17-control-plan.xlsx"
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        ws = wb["Control plan"]
        # the sheet now ships with worked example rows; this fixture is about
        # the formula, so clear the data region before planting it
        for r in range(10, 28):
            ws[f"A{r}"], ws[f"P{r}"] = None, None
        for i, (name, lvl) in enumerate([("m1", "2 Design it out"), ("m2", "3 Guide it"),
                                         ("m3", "5 Standardise it"), ("m4", "garbage")], start=10):
            ws[f"A{i}"], ws[f"P{i}"] = name, lvl
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        e30 = _read(sol, src.name, "Control plan", "E30")
        e31 = _read(sol, src.name, "Control plan", "E31")
        e32 = _read(sol, src.name, "Control plan", "E32")
        check(approx(e30, 2), "control plan: durable count survives a junk paste", f"got {e30!r} want 2")
        check(approx(e31, 1), "control plan: decaying count survives a junk paste", f"got {e31!r} want 1")
        check(approx(e32, 0.5), "control plan: durable share", f"got {e32!r} want 0.5")

    # Kano: the published evaluation table, cell by cell. Two of these were
    # transposed — Like/Dislike returned "Delighter" and Expect-it/Dislike
    # returned "Performance", so the two most common answers in a support
    # survey each got the other's investment advice.
    src = TEMPLATES / "23-kano-analysis.xlsx"
    KANO = [
        ("Like", "Dislike", "Performance"),        # one-dimensional
        ("Like", "Live with it", "Delighter"),
        ("Like", "Neutral", "Delighter"),
        ("Like", "Expect it", "Delighter"),
        ("Like", "Like", "Questionable"),
        ("Expect it", "Dislike", "Must-have"),
        ("Neutral", "Dislike", "Must-have"),
        ("Live with it", "Dislike", "Must-have"),
        ("Neutral", "Neutral", "Indifferent"),
        ("Expect it", "Neutral", "Indifferent"),
        ("Dislike", "Neutral", "Reverse"),
        ("Neutral", "Like", "Reverse"),
        ("Dislike", "Dislike", "Questionable"),
    ]
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        ws = wb["Kano analysis"]
        for i, (fun, dys, _) in enumerate(KANO):
            ws[f"B{7 + i}"], ws[f"C{7 + i}"] = fun, dys
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        for i, (fun, dys, want) in enumerate(KANO):
            got = _read(sol, src.name, "Kano analysis", f"D{7 + i}")
            check(got == want, f"Kano: {fun} / {dys} classifies as {want}", f"got {got!r}")

    # Pareto: it is sorted by definition. The cumulative used to be a running
    # sum down the rows, which is only a Pareto if the user happens to type
    # their categories in descending order. Type them ascending — the worst
    # case — and the ranked block must still come out ranked.
    src = TEMPLATES / "25-pareto-and-distribution.xlsx"
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        ws = wb["Pareto"]
        ascending = [("Refund timing", 74), ("Proration misunderstood", 96),
                     ("Duplicate charge", 151), ("Wrong plan applied", 233),
                     ("Adjustment not posted at closure", 412)]
        for i, (nm, n) in enumerate(ascending):
            ws.cell(row=5 + i, column=1, value=nm)
            ws.cell(row=5 + i, column=2, value=n)
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        counts = [_read(sol, src.name, "Pareto", f"I{5 + i}") for i in range(5)]
        try:
            nums = [float(c) for c in counts]
        except (TypeError, ValueError):
            nums = []
        check(nums == sorted(nums, reverse=True),
              "Pareto ranks itself even when the categories are typed ascending",
              f"chart plots {nums}")
        check(approx(_read(sol, src.name, "Pareto", "J7"), 0.8240165631469979, 1e-3),
              "Pareto cumulative follows the ranking, not the typing order",
              repr(_read(sol, src.name, "Pareto", "J7")))
        check(approx(_read(sol, src.name, "Pareto", "C29"), 0.8240165631469979, 1e-3),
              "'share explained by the top three' means the top three by rank",
              repr(_read(sol, src.name, "Pareto", "C29")))

    # Every chart that claims a ranking must actually rank. These plotted rows
    # in the order somebody typed them, which makes the reader do the ranking
    # the chart was supposed to do — and on an FMEA, whose whole method is
    # "work the highest RPN first", buries the worst failure mode wherever it
    # happened to be entered.
    for book, sheet, cat_col, val_col, first, n in [
            ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "M", "N", 15, 8),
            ("15-solution-selection-matrix.xlsx", "Solution selection", "Q", "R", 14, 6),
            ("12-fmea.xlsx", "FMEA", "V", "W", 11, 6)]:
        src = TEMPLATES / book
        sol = _engine(src).calculate()
        vals = []
        for i in range(n):
            v = _read(sol, book, sheet, f"{val_col}{first + i}")
            try:
                vals.append(float(v))
            except (TypeError, ValueError):
                pass
        check(len(vals) == n and vals == sorted(vals, reverse=True),
              f"{book}: the chart plots its ranking in rank order",
              f"got {vals}")

    # A reference line that contradicts the sheet it sits on destroys trust in
    # the whole file. The FMEA drew its action line at 100 while its own summary
    # counted ">=200 (act now)"; the hypothesis log hardcoded alpha 0.05 while
    # the sheet has a validated alpha input, and keyed the line on the p-value
    # column so it stopped at the last completed test.
    src = TEMPLATES / "12-fmea.xlsx"
    sol = _engine(src).calculate()
    thr = _read(sol, src.name, "FMEA", "D45")
    line = _read(sol, src.name, "FMEA", "Y11")
    check(approx(thr, line), "FMEA: the chart's action line is the sheet's action threshold",
          f"threshold={thr!r} line={line!r}")
    check(approx(_read(sol, src.name, "FMEA", "D40"), 4),
          "FMEA: the count above the threshold follows that same cell",
          repr(_read(sol, src.name, "FMEA", "D40")))

    src = TEMPLATES / "13-hypothesis-test-log.xlsx"
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        wb["Test log"]["B8"] = 0.01          # the user tightens alpha
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        drawn = []
        for r in range(11, 36):
            v = _read(sol, src.name, "Test log", f"T{r}")
            if isinstance(v, (int, float)):
                drawn.append(float(v))
        check(drawn and all(abs(v - 0.01) < 1e-9 for v in drawn),
              "the alpha line follows the alpha the sheet validates",
              f"values {sorted(set(drawn))[:3]}")
        check(len(drawn) >= 20,
              "the alpha line spans the log, not just the rows already filled in",
              f"drawn on {len(drawn)} of 25 rows")

    # A chart must reconcile to the sheet it sits on.
    src = TEMPLATES / "27-control-charts.xlsx"
    sol = _engine(src).calculate()
    lo = float(_read(sol, src.name, "t and g (rare events)", "K14"))
    hi = float(_read(sol, src.name, "t and g (rare events)", "J14"))
    cl = float(_read(sol, src.name, "t and g (rare events)", "L14"))
    check(lo < cl < hi, "t-chart: the centre line sits inside its own limits",
          f"limits {lo:.1f}..{hi:.1f}, centre {cl:.1f}")
    # the limits are built on the transformed scale, so the centre of that
    # construction is the back-transformed mean, not the mean of the raw gaps
    check(not approx(cl, _read(sol, src.name, "t and g (rare events)", "B6"), 1e-3),
          "t-chart: the centre line is the back-transformed centre, not the arithmetic mean")
    ucl = _read(sol, src.name, "t and g (rare events)", "M14")
    check(approx(ucl, _read(sol, src.name, "t and g (rare events)", "B11")),
          "g-chart: the UCL the sheet computes is the UCL the chart plots", repr(ucl))

    # Kano's chart has to account for every class its own classifier returns
    src = TEMPLATES / "23-kano-analysis.xlsx"
    sol = _engine(src).calculate()
    counts = [float(_read(sol, src.name, "Kano analysis", f"D{28 + i}")) for i in range(6)]
    check(sum(counts) == 8,
          "Kano: the summary accounts for every classified row",
          f"classes {counts} sum to {sum(counts)}, example has 8 rows")
    check(counts[4] == 1, "Kano: a Reverse classification is counted, not computed and dropped",
          f"Reverse count {counts[4]}")

    # the control plan's own durability score and its chart must count the
    # same rows, or a control appears on one and not the other
    src = TEMPLATES / "17-control-plan.xlsx"
    sol = _engine(src).calculate()
    total = _read(sol, src.name, "Control plan", "E29")
    feed = sum(float(_read(sol, src.name, "Control plan", f"B{r}")) for r in range(39, 45))
    check(approx(total, feed),
          "control plan: the durability score and the chart count the same rows",
          f"summary {total!r} vs chart feed {feed}")

    # Control limits must come from a frozen baseline, not the whole series.
    # Recomputing over every point you have is how a drifting process drags its
    # own limits along and never signals — which the picker tab warns about in
    # bold while all five estimated-limit charts did exactly that.
    src = TEMPLATES / "27-control-charts.xlsx"
    sol = _engine(src).calculate()
    vals = [408, 415, 402, 431, 419, 396, 424, 410, 438, 405, 417, 429,
            401, 422, 413, 407, 435, 398, 420, 411, 449, 471, 478, 466]
    base = int(_read(sol, src.name, "I-MR", "B3"))
    check(base < len(vals),
          "the baseline window is shorter than the series, not the whole thing",
          f"baseline {base} of {len(vals)}")
    check(approx(_read(sol, src.name, "I-MR", "B6"), sum(vals[:base]) / base),
          "I-MR centre line comes from the baseline window only",
          f"got {_read(sol, src.name, 'I-MR', 'B6')!r}")
    mrs = [abs(vals[i] - vals[i - 1]) for i in range(1, len(vals))]
    check(approx(_read(sol, src.name, "I-MR", "B7"), sum(mrs[:base - 1]) / (base - 1)),
          "I-MR average moving range comes from the baseline window only")
    for sheet, cell in (("Laney p-prime", "B5"), ("Laney u-prime", "B5"),
                        ("Xbar-R", "B6"), ("Xbar-R", "B7")):
        v = _read(sol, src.name, sheet, cell)
        check(isinstance(v, (int, float)),
              f"{sheet}!{cell} still computes with a frozen baseline", repr(v))

    # --- Every control chart must trip its own rule in its own worked data,
    #     and must not trip it inside the baseline window.
    #
    #     Four of the seven shipped a stable run that never crossed a limit, so
    #     the Signal column was blank down all 24 rows. A control chart is a
    #     signal detector; one whose example never signals shows the reader the
    #     shape of the chart and nothing about reading it, and — the part a
    #     harness can act on — leaves the rule formula never once evaluated
    #     against a true case. It could have been broken from the day it was
    #     written and every check here would still have passed.
    #
    #     The second half matters more. Xbar-R signalled three times INSIDE its
    #     baseline window, so every limit on that sheet was computed from a
    #     process the sheet itself showed was out of control. That is the error
    #     the method most needs a reader not to make, shipped as the example.
    audit_control_signals(sol, src.name)

    # The baseline divisor must be what SUMIF actually summed, not the window
    # the user asked for. Dividing 15 pasted points by a baseline of 20 put the
    # centre line 25% low and every limit with it, silently — a defect the
    # freeze introduced and the picker tab itself anticipates by warning that
    # 20-25 points is the working minimum.
    src = TEMPLATES / "27-control-charts.xlsx"
    vals = [408, 415, 402, 431, 419, 396, 424, 410, 438, 405, 417, 429,
            401, 422, 413, 407, 435, 398, 420, 411, 449, 471, 478, 466]
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        ws = wb["I-MR"]
        for i in range(15, 24):                 # leave only 15 points
            ws.cell(row=14 + i, column=2).value = None
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        check(approx(_read(sol, src.name, "I-MR", "B6"), sum(vals[:15]) / 15),
              "the baseline divides by the points that exist, not the window requested",
              repr(_read(sol, src.name, "I-MR", "B6")))
        warn_txt = _read(sol, src.name, "I-MR", "J3")
        check(isinstance(warn_txt, str) and "only 15" in warn_txt,
              "the sheet says so when the baseline window exceeds the data",
              repr(warn_txt))

    # The breakeven chart must agree with the NPV cell on the same screen. It
    # was hardwired to four points while "Years to model" is validated 1-10, so
    # a five-year model showed $196,620 on the chart and $400,918 in the cell
    # labelled NET PRESENT VALUE.
    for yrs in (3, 5, 10):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td) / CALC
            shutil.copyfile(TEMPLATES / CALC, tmp)
            wb = load_workbook(tmp)
            wb["9 ROI and payback"]["B7"] = yrs
            wb.save(tmp)
            sol = _engine(tmp).calculate()
            last = None
            for r in range(22, 33):
                v = _read(sol, CALC, "9 ROI and payback", f"B{r}")
                if isinstance(v, (int, float)):
                    last = v
            npv = _read(sol, CALC, "9 ROI and payback", "B15")
            check(approx(last, npv, 1e-4),
                  f"ROI: the chart's last point equals NPV at {yrs} years",
                  f"chart {last!r} vs NPV {npv!r}")

    # VSM: "% of lead time" must divide by lead time, not by waiting time.
    src = TEMPLATES / "10-value-stream-map.xlsx"
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / src.name
        shutil.copyfile(src, tmp)
        wb = load_workbook(tmp)
        ws = wb["Value stream"]
        # same here: wipe the shipped example so the arithmetic is the only
        # thing under test
        for r in range(10, 29):
            for col in "ABCDEFGHIJ":
                ws[f"{col}{r}"] = None
        ws["E10"], ws["F10"] = 10, 90       # touch 10, wait 90 -> lead 100
        ws["B44"], ws["B45"] = 60, 30
        wb.save(tmp)
        sol = _engine(tmp).calculate()
        lead = _read(sol, src.name, "Value stream", "E33")
        c44 = _read(sol, src.name, "Value stream", "C44")
        check(approx(lead, 100), "VSM: lead time = touch + wait", repr(lead))
        check(approx(c44, 0.60), "VSM: waiting state is a share of lead time, not of waiting time",
              f"got {c44!r}, want 0.60 (0.667 would mean it still divides by waiting time)")


# --------------------------------------------------------------------- SYNC
def test_sync() -> None:
    src = HTML.read_text(encoding="utf-8")
    _, _, tpls = extract_tpls(src)
    test_tool_links(src, tpls)
    # A title and a description are set with textContent, so an HTML entity in
    # one is shown to the reader literally. The modal read "Control Charts
    # &mdash; all seven types" for as long as the entry has existed, because the
    # text was copied out of the card markup where the entity was correct.
    ent = [f"{s}.{f}" for s, e in tpls.items() for f in ("title", "desc")
           if isinstance(e.get(f), str) and re.search(r"&(\w+|#\d+);", e[f])]
    check(not ent, "no template title or description carries a raw HTML entity",
          ", ".join(ent[:4]))
    check(len(tpls) == 33, f"33 templates registered (found {len(tpls)})")

    exts = [e.get("ext") for e in tpls.values()]
    check(exts.count("xlsx") == 22, f"22 Excel workbooks (found {exts.count('xlsx')})")

    # The page counts itself out loud, in the schema block, the meta description
    # and the download button. Those numbers had been stale by two since the
    # regression and Gage R&R workbooks landed: the registry said 30 and 19, the
    # prose said 28 and 17, and nothing compared them.
    said = re.findall(r"(\d+) downloadable templates including (\d+) Excel workbooks", src)
    check(said and all((int(a), int(b)) == (len(tpls), exts.count("xlsx")) for a, b in said),
          f"the page's own template count matches the registry ({len(tpls)}/"
          f"{exts.count('xlsx')})", f"the prose says {sorted(set(said))}")
    # The check above greps ONE exact phrase, so the page went on saying
    # "Twenty-eight working files ... Seventeen are real Excel workbooks" in the
    # templates section — spelled out, therefore invisible to a \d+ pattern —
    # four lines above a button reading "Download all 32 templates". A stat card
    # said 17 as well. Every place the page counts itself is checked now,
    # whichever way the number is written.
    WORD = {"nine": 9, "ten": 10, "eleven": 11, "twelve": 12, "seventeen": 17,
            "nineteen": 19, "twenty": 20, "twenty-one": 21, "twenty-eight": 28,
            "thirty": 30, "thirty-two": 32, "thirty-three": 33, "twenty-two": 22,
            "forty-two": 42}

    def spoken(tok: str):
        return int(tok) if tok.isdigit() else WORD.get(tok.lower())

    for pattern, want, what in (
            (r"([A-Za-z-]+|\d+) working files", len(tpls), "working files"),
            (r"([A-Za-z-]+|\d+) are real Excel\s+workbooks", exts.count("xlsx"),
             "real Excel workbooks"),
            (r'<div class="num">(\d+)</div><div class="lbl">Excel workbooks</div>',
             exts.count("xlsx"), "the Excel-workbooks stat card"),
            (r'<div class="num">(\d+)</div><div class="lbl">Documents</div>',
             exts.count("md"), "the Documents stat card")):
        found = re.findall(pattern, src)
        got = [spoken(t) for t in found]
        check(found and all(g == want for g in got),
              f"the page's count of {what} matches the registry ({want})",
              f"it says {found} — {[g for g in got if g != want]} is wrong")

    # The glossary is one JSON literal plus eight Object.assign blocks, and a key
    # written twice does not raise anything — the later one silently wins. Fifteen
    # terms had been defined twice, so fifteen definitions were written, shipped
    # and unreachable, and it was not consistently the better one that survived:
    # the Tukey entry that won gave a Minitab menu path, and the one it replaced
    # explained that comparing five queues is ten tests and produces a false
    # positive about 40% of the time.
    terms = gloss_keys(src)
    twice = sorted({t for t in terms if terms.count(t) > 1})
    check(not twice, "no glossary term is defined twice",
          f"silently overwritten: {twice}")
    said = re.findall(r"(\d+) plain-English definitions", src)
    check(said and all(int(n) == len(set(terms)) for n in said),
          f"the page's definition count matches the glossary ({len(set(terms))})",
          f"the prose says {sorted(set(said))}")

    btn = re.findall(r"Download all (\d+) templates", src)
    check(btn and all(int(n) == len(tpls) for n in btn),
          "the download-all button counts the templates it downloads",
          f"the button says {sorted(set(btn))}")
    check(exts.count("md") == 11, f"11 Markdown templates (found {exts.count('md')})")

    for slug, entry in tpls.items():
        path = TEMPLATES / entry["file"]
        check(path.exists(), f"{entry['file']} exists on disk")
        if not path.exists():
            continue
        if entry.get("ext") == "xlsx":
            want = base64.b64encode(path.read_bytes()).decode("ascii")
            check(entry.get("b64") == want, f"{entry['file']}: embedded base64 matches the file on disk")
            _check_preview(entry, path)
            _check_preview_charts(entry, path)
            _check_preview_generated(entry, path)
            _check_preview_fidelity(entry, path)
        else:
            check(entry.get("content") == path.read_text(encoding="utf-8"),
                  f"{entry['file']}: embedded markdown matches the file on disk")

    check(HTML.read_bytes() == DOCS.read_bytes(), "docs/index.html is identical to the root HTML")

    # Version: the meta tag and the sidebar badge must agree, so a deploy can be
    # identified without opening the page. The release notes themselves live in
    # CHANGELOG.md — a reader who has just opened the document should not have to
    # scroll past its maintenance history — but the newest entry there has to
    # match what the page claims to be.
    meta = re.search(r'<meta name="app-version" content="([0-9.]+)"', src)
    badge = re.search(r"Customer Support Operations &middot; v([0-9.]+)", src)
    check(bool(meta), "an <meta name=\"app-version\"> tag is present")
    check(bool(badge), "the sidebar shows a version")
    check(not re.search(r'<div class="t">(New )?[Ii]n v[0-9.]+ &mdash;', src),
          "release notes are not in the product — they belong in CHANGELOG.md")
    changelog = ROOT / "CHANGELOG.md"
    check(changelog.exists(), "CHANGELOG.md exists")
    if changelog.exists():
        notes = re.findall(r"^## (?:New in )?v?([0-9.]+)", changelog.read_text(encoding="utf-8"), re.M)
        check(bool(notes), "CHANGELOG.md has at least one release entry")
        if meta and badge and notes:
            newest = max(notes, key=lambda v: [int(p) for p in v.split(".")])
            check(meta.group(1) == badge.group(1) == newest,
                  "meta tag, sidebar badge and CHANGELOG agree on the version",
                  f"meta={meta.group(1)} sidebar={badge.group(1)} changelog={newest}")
    # The glossary explainer must outrank every dialog. It renders inside the
    # template preview, the export dialog and the download menu, and at a lower
    # z-index it opens behind them where nobody can read it.
    zs = {}
    for zm in re.finditer(r"z-index\s*:\s*(\d+)", src):
        a, b = src.rfind("}", 0, zm.start()), src.rfind("{", 0, zm.start())
        zs[src[a + 1:b].strip().split()[-1]] = int(zm.group(1))
    pop = zs.get("#pop", 0)
    over = {k: v for k, v in zs.items() if k != "#pop" and v >= pop and k != ".skiplink"}
    check(pop > 0 and not over,
          "the glossary explainer sits above every dialog",
          f"#pop={pop} but these are at or above it: {over}")

    # Every template card must be a direct child of .tplgrid. Cards nested one
    # level deeper still render, but the CSS grid only lays out its own
    # children, so the whole section collapses into a single narrow column.
    g0 = src.index('<div class="tplgrid">')
    depth, j = 0, g0
    while j < len(src):
        if src.startswith("<div", j):
            depth += 1
        elif src.startswith("</div>", j):
            depth -= 1
            if depth == 0:
                break
        j += 1
    grid = src[g0:j]
    total_cards = src.count('class="tplc"')
    check(grid.count('class="tplc"') == total_cards,
          "every template card is a direct child of the grid",
          f"{grid.count('class=\"tplc\"')} of {total_cards} inside .tplgrid — the rest are nested")
    check(total_cards == len(tpls),
          f"a card for every registered template ({len(tpls)})", f"cards={total_cards}")

    # Nested explainers. An explainer that leans on three more undefined terms
    # is not an explanation, so #pop is glossed too — which needs a self-link
    # guard, a back trail, and one code path for pointer and keyboard. Enter
    # used to navigate without recording the trail, so the back button never
    # appeared for keyboard users.
    # Was: assert the function is declared. qa_visual.audit_glossary dispatches
    # a real click and a real Enter on a real .gl and compares what opens —
    # which is the claim. Cutting the keydown call site takes the keyboard
    # result from 907 characters of popover to 0 while the click is unaffected.
    # "openTerm(g);" with the semicolon, so the definition itself is not counted
    check(src.count("openTerm(g);") == 2,
          "both the click and the keydown handler go through openTerm()",
          f"found {src.count('openTerm(g);')} call sites")
    # Was a source string. qa_visual.audit_glossary opens a real popover and
    # counts the links in it.
    # Was the literal guard line. Now counted in the DOM: the harness picks
    # the term whose explainer mentions the most OTHER terms, so the popover
    # has links to be wrong about — without that, "no self-links" holds
    # trivially and the check cannot fail.
    check("n.id==='pop'&&!POP_OPEN" in src,
          "#pop is glossed only when showPop asks, not on document-wide passes")
    check(src.count("POP_TRAIL.length = 0") >= 2,
          "the back trail is cleared when the explainer closes and on a fresh term")

    # Case. The matcher is deliberately case-sensitive — an "i" flag would let
    # short acronyms swallow ordinary words (IT, OR, US, AND). But ordinary-word
    # terms are routinely written lower-case mid-sentence, and those matched
    # nothing at all: "gemba" appears more often in this document than "Gemba"
    # did, and only the capitalised form was ever clickable.
    # Both of these asserted an exact line of JavaScript. qa_visual.audit_glossary
    # runs the matcher over real prose and counts what it links instead — and the
    # difference is not academic. The claim attached to the first one named the
    # acronym regex as what "keeps IT/OR/US safe"; the regex and the
    # `t !== t.toUpperCase()` beside it are each sufficient alone, so removing
    # either changes nothing and both look dead until you cut them together.
    # A source check cannot tell redundancy from rot. A mutation can.
    for term in ("Gemba", "Kaizen", "Poka-yoke"):
        check(f'"{term}"' in src, f"{term} is still a glossary entry")

    # The eleven markdown templates shipped as well-structured blank forms with
    # nothing telling a first-time user what a good answer looks like, while
    # every workbook opened with a how-to tab and a worked example row. They
    # carry the same now.
    for md in sorted(TEMPLATES.glob("*.md")):
        body = md.read_text(encoding="utf-8")
        check("## How to use this" in body,
              f"{md.name}: has a how-to block")
        check("**The mistake this prevents.**" in body,
              f"{md.name}: names the mistake it prevents")
        check(body.count("*") > 6,
              f"{md.name}: carries a worked example", "no italic example entries found")

    # Discoverability. A single-file page with no metadata is invisible to both
    # search engines and the AI assistants people increasingly ask first.
    import json as _json
    for tag in ('name="description"', 'rel="canonical"', 'property="og:title"',
                'name="twitter:card"', 'name="robots"', 'application/ld+json'):
        check(tag in src, f"page carries {tag}")
    m = re.search(r'<script type="application/ld\+json">(.*?)</script>', src, re.S)
    check(bool(m), "structured data is present")
    if m:
        try:
            data = _json.loads(m.group(1))
            types = [x.get("@type") for x in data.get("@graph", [])]
        except Exception:                                        # noqa: BLE001
            types = []
        check("Course" in types, "structured data declares a Course", str(types))
        check("FAQPage" in types, "structured data carries an FAQ", str(types))
        faq = next((x for x in data.get("@graph", []) if x.get("@type") == "FAQPage"), {})
        check(len(faq.get("mainEntity", [])) >= 5,
              "the FAQ answers at least five real questions")
    check((ROOT / "docs" / "robots.txt").exists(), "docs/robots.txt exists")
    check((ROOT / "docs" / "sitemap.xml").exists(), "docs/sitemap.xml exists")

    for dead in ("parseCSV", "renderCSV"):
        check(dead not in src, f"dead {dead}() removed")
    check("function esc2(" in src, "esc2() retained (renderMD depends on it)")
    check(src.count("-year net<") == 0 and "'-year NPV</th>" in src,
          "wizard sensitivity column is labelled NPV, not 'net'")


def test_tool_links(src: str, tpls: dict) -> None:
    """A tool must link to a template that is about that tool.

    "Linear and multiple regression" and "Logistic regression" both pointed at
    the root-cause evidence pack — a narrative document with nowhere to write a
    coefficient. It was reported by a reader, because nothing here compared the
    two halves: the tool library and the template pack were authored
    separately and joined by a hand-typed slug that no check ever read.
    """
    from openpyxl import load_workbook as _lw

    def corpus(slug: str) -> str:
        e = tpls.get(slug)
        if not e:
            return ""
        bits = [e.get("title", ""), e.get("desc", ""), e.get("content", "")]
        if e.get("ext") == "xlsx":
            try:
                wb = _lw(TEMPLATES / e["file"], read_only=True)
                for ws in wb.worksheets:
                    bits.append(ws.title)
                    for row in ws.iter_rows():
                        bits += [c.value for c in row if isinstance(c.value, str)]
            except Exception:                                    # noqa: BLE001
                pass
        return " ".join(b for b in bits if b).lower()

    STOP = {"and", "the", "of", "for", "test", "tests", "analysis", "chart",
            "charts", "study", "plan", "models", "model"}
    seg = src[src.index('id="toolList"'):]
    bad = []
    # Bound each tool at its own </details>. Letting the last one run to the end
    # of the file swallowed the whole template grid and produced thirty
    # spurious hits the first time this was written.
    for block in seg.split('<details class="tool"')[1:]:
        block = block.split("</details>")[0]
        m = re.search(r'<span class="tn">([^<]*)</span>', block)
        if not m:
            continue
        name = m.group(1)
        words = [w for w in re.findall(r"[A-Za-z][A-Za-z'’-]{3,}", name.lower())
                 if w not in STOP]
        if not words:
            continue
        for slug in sorted(set(re.findall(r'data-(?:tpl|tplchip|dl)="([0-9][^"]*)"', block))):
            text = corpus(slug)
            if text and not any(w in text for w in words):
                bad.append(f"{name!r} -> {slug}, which never mentions {words}")
    check(not bad, "every tool links to a template that mentions what the tool is",
          "\n      ".join(bad[:4]))


def _check_preview_fidelity(entry: dict, path: Path) -> None:
    """Everything written in the workbook has to appear in its preview.

    Nineteen previews were hand-written markup that the sync could only edit in
    place — it rewrites the <td>s already present, so it could correct a value
    but never add a row. Anything added to those workbooks afterwards existed in
    the download and appeared nowhere on the page: thirty-one pieces of prose,
    including all six definitions of the hierarchy of controls and nineteen of
    the notes written into the calculator pack. Nothing compared the two, so
    nothing said so.

    Scaffolding columns are excluded because the preview summarises them on
    purpose — that is the one place where differing is the correct behaviour.
    """
    import html as _html
    sys.path.insert(0, str(ROOT / "tools"))
    from preview import scaffold_from
    shown = re.sub(r"\s+", " ", _html.unescape(
        re.sub(r"<[^>]+>", " ", entry.get("preview", ""))))
    missing = []
    for ws in load_workbook(path).worksheets:
        scaf = scaffold_from(ws)
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and len(v) >= 40 and not v.startswith("=")):
                    continue
                if scaf and cell.column >= scaf:
                    continue
                if re.sub(r"\s+", " ", v)[:50] not in shown:
                    missing.append(f"{ws.title}!{cell.coordinate}")
    check(not missing, f"{path.name} preview shows what the workbook says",
          f"{len(missing)} piece(s) of prose are in the workbook and nowhere in "
          f"the preview: {missing[:4]}")


def _check_preview_generated(entry: dict, path: Path) -> None:
    """A generated preview in the page must be the one the builder generated.

    tools/build_templates.py rebuilds these from the workbook on every run and
    writes them to previews.json — and for a long time nothing carried them into
    the page, so any preview change that was not a plain text edit was written
    to disk and silently dropped. Nothing compared the two, so nothing said so.
    """
    from sync_html import GENERATED, RE_CHARTS
    want = GENERATED.get(path.name)
    if want is None:
        return
    got = RE_CHARTS.sub("", entry["preview"])
    check(got == want,
          f"{path.name}: the page carries the preview the builder generated",
          f"differs by {abs(len(got) - len(want))} chars — run tools/sync_html.py")


def _check_preview_charts(entry: dict, path: Path) -> None:
    """Every chart in the workbook has to be visible in the preview.

    The preview rendered cells only, so 35 charts were downloadable but not
    viewable, and nobody noticed for as long as the previews existed — no check
    compared the two. Counting the workbook's charts against the preview's
    <svg> tags is the smallest thing that could have caught it, so it is what
    runs now, on every workbook, every build.
    """
    import zipfile
    with zipfile.ZipFile(path) as z:
        want = len([n for n in z.namelist()
                    if re.match(r"xl/charts/chart\d+\.xml$", n)])
    got = entry["preview"].count('<svg class="xchart"')
    check(got == want,
          f"{path.name}: all {want} charts are drawn in the preview",
          f"workbook has {want}, preview shows {got} — run tools/sync_html.py")


def _check_preview(entry: dict, path: Path) -> None:
    wb = load_workbook(path)
    bad = []
    seen = 0
    for m in RE_SHEET.finditer(entry["preview"]):
        ws = wb.worksheets[int(m.group(2))]
        held: dict[int, set[int]] = {}      # columns claimed by an earlier rowspan
        for r, rm in enumerate(RE_ROW.finditer(m.group(3)), start=1):
            col = 1
            for tm in RE_TD.finditer(rm.group(2)):
                attrs = tm.group(1)
                cs = re.search(r'colspan="(\d+)"', attrs)
                rs = re.search(r'rowspan="(\d+)"', attrs)
                ti = re.search(r'title="([^"]*)"', attrs)
                nc = int(cs.group(1)) if cs else 1
                nr = int(rs.group(1)) if rs else 1
                while col in held.get(r, ()):
                    col += 1
                cell = ws.cell(row=r, column=col)
                if nr > 1:
                    for rr in range(r + 1, r + nr):
                        held.setdefault(rr, set()).update(range(col, col + nc))
                col += nc
                wf = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
                pf = H.unescape(ti.group(1)) if ti else None
                if pf or wf:
                    seen += 1
                    if pf != wf:
                        bad.append(f"[{ws.title}] {cell.coordinate}: preview={pf!r} workbook={wf!r}")
    check(not bad, f"{path.name}: all {seen} preview tooltips match the workbook", "\n      ".join(bad[:4]))


# ------------------------------------------------------------------- EXPORT
# The business case is generated in the browser: an email-safe HTML rendering
# and a real .xlsx with live formulas and native charts. These tests pull the
# shipped JavaScript straight out of the HTML, run it under node, and check the
# workbook it produces actually recalculates to the numbers the page showed.

JS_START = "/* ============================================================ xlsx writer"
# Stop before the docx writer: everything after it needs a live DOM, which node
# has no business providing. The docx package is validated in the browser
# instead, by parsing its own ZIP with DOMParser.
JS_END = "/* ============================================================ docx writer"

# Each problem archetype emits a different benefit model, so the row the gross
# value lands on shifts. Every branch gets generated and recalculated.
GROSS_ROW = {"rate": 5, "volume": 4, "aht": 5, "shrink": 6,
             "deflect": 5, "attrition": 5, "copq": 5, "churn": 5}

HARNESS = r"""
const fs=require('fs');
global.fm=function(n,d){ if(n===undefined||n===null||!isFinite(n)) return '—';
  d=(d===undefined)?0:d;
  return Number(n).toLocaleString('en-US',{minimumFractionDigits:d,maximumFractionDigits:d}); };
eval(fs.readFileSync(process.argv[2],'utf8'));
const outDir=process.argv[3];
const base={vol:480000,cpc:6.8,rate:14.0,hr:38,occ:82,aht:420,ahtsave:14,agents:120,shrink:32,
  shrinktgt:28,target:8.0,harvest:'reduce',realz:80,
  deflrate:35,cpcnew:0.35,attr:38,attrtgt:30,replcost:9500,incidentcost:85,ltv:1400,
  bbmonths:9,bbcost:120000,training:35000,eng:60000,tooling:8000,years:3,disc:10};
const kinds=[
 {kind:'rate',   n:'Reduce rework and reopens',  metric:'Reopen rate',        gross:480000*0.06*6.8},
 {kind:'volume', n:'Eliminate a contact driver', metric:'Contacts per year',
  V:{rate:40000,target:8000},                                                 gross:(40000-8000)*6.8},
 {kind:'aht',    n:'Reduce handle time',         metric:'Average handle time',gross:480000*14/3600/0.82*38},
 {kind:'shrink', n:'Recover shrinkage capacity', metric:'Shrinkage %',        gross:120*1760*0.04*38},
 {kind:'deflect',  n:'Deflect to self-service',  metric:'Share deflected',
  gross:480000*0.35*(6.8-0.35)},
 {kind:'attrition',n:'Reduce agent attrition',   metric:'Annual attrition',
  gross:120*0.08*9500},
 {kind:'copq',     n:'Cut the cost of poor quality', metric:'Incident rate',
  gross:480000*0.06*85},
 {kind:'churn',    n:'Protect revenue at risk',  metric:'Churn rate',
  gross:480000*0.06*1400}
];
const meta={};
for(const k of kinds){
  const V=Object.assign({},base,k.V||{});
  const gross=k.gross, real=gross*0.8, inv=120000*0.75+35000+60000+8000;
  let npv=-inv; for(let y=1;y<=3;y++) npv+=real/Math.pow(1.1,y);
  const m={gross,real,inv,npv,pb:inv/real,roi:(real*3-inv)/inv,fte:1,realz:0.8,
    detail:[['Improvement','a → b','6.0 pts']]};
  const a={n:k.n,d:'desc',metric:k.metric,kind:k.kind};
  const ctx={m,V,S:{arch:k.kind},a};
  fs.writeFileSync(outDir+'/'+k.kind+'.xlsx', Buffer.from(bizXlsx(ctx)));
  meta[k.kind]={gross,real,inv,npv,html:bizHTML(ctx).length};
}
fs.writeFileSync(outDir+'/meta.json', JSON.stringify(meta));
"""


def test_export() -> None:
    import subprocess

    src = HTML.read_text(encoding="utf-8")
    check("business-case.md" not in src, "markdown business case replaced by HTML/Excel")
    check("  function doc(){" not in src, "old markdown generator removed")
    for needed in ("var XLSX = (function(){", "var DOCX = (function(){",
                   "function bizXlsx(", "function bizHTML(",
                   "function openExport(", "function tplEmailHTML(", 'id="expCopy"',
                   "function showFmtMenu(", "function dlTemplateAs(",
                   # the template modal's button is created at runtime, not in the markup
                   "b.id = 'tplEmail'", "fullCalcOnLoad",
                   "wordprocessingml.document.main+xml"):
        check(needed in src, f"export code present: {needed}")

    # Download must offer a format rather than pushing a .md at everyone.
    # (Markdown is still *available* from the menu - it just isn't the default.)
    check("""if(t.ext==='xlsx'){
    dlBlob(t.file, b64ToBlob(t.b64,""" not in src,
          "old markdown-by-default download path removed")
    # Was the literal ternary. Now read off the FILENAME that would reach disk:
    # the harness stubs the download primitives and drives dlTemplate for one
    # workbook and one document, expecting a .xlsx and a .docx. Forcing the
    # branch to 'docx' fails it — and a workbook exported as Word loses every
    # formula, which is the whole reason to download one.
    for fmt in ("'docx'", "'html'"):
        check(fmt in src, f"format menu offers {fmt}")
    # Markdown is a developer format; this audience gets Word and HTML only.
    check("'text/markdown'" not in src, "Markdown is no longer offered as a download")
    check("['md','Markdown'" not in src, "Markdown removed from the format menu")

    if JS_START not in src or JS_END not in src:
        check(False, "export JS block markers found in the HTML")
        return
    js = src[src.index(JS_START):src.index(JS_END)]

    if not shutil.which("node"):
        print("           node not found - skipping the generated-workbook test")
        return

    with tempfile.TemporaryDirectory() as td:
        d = Path(td)
        (d / "bundle.js").write_text(js, encoding="utf-8")
        (d / "run.js").write_text(HARNESS, encoding="utf-8")
        r = subprocess.run(["node", str(d / "run.js"), str(d / "bundle.js"), str(d)],
                           capture_output=True, text=True)
        if r.returncode != 0:
            check(False, "business case workbook generates without error", r.stderr.strip()[:400])
            return
        check(True, "business case workbook generates under node for every archetype")
        meta = json.loads((d / "meta.json").read_text())

        import zipfile
        import xml.etree.ElementTree as ET

        try:
            import formulas  # noqa: F401
            can_calc = True
        except ImportError:
            can_calc = False
            print("           formulas not installed - skipping workbook recalculation")

        for kind, e in meta.items():
            path = d / (kind + ".xlsx")
            check(e["html"] > 8000, f"{kind}: business case HTML renders ({e['html']} chars)")

            z = zipfile.ZipFile(path)
            names = set(z.namelist())
            broken = []
            for n in names:
                if n.endswith((".xml", ".rels")):
                    try:
                        ET.fromstring(z.read(n))
                    except Exception as ex:
                        broken.append(f"{n}: {ex}")
            check(not broken, f"{kind}: every xlsx part is well-formed XML", "; ".join(broken[:2]))
            charts = [n for n in names if "/charts/chart" in n]
            check(len(charts) == 4, f"{kind}: workbook carries all 4 charts (found {len(charts)})")
            ct = ET.fromstring(z.read("[Content_Types].xml"))
            ovr = {o.get("PartName") for o in ct if o.tag.endswith("Override")}
            missing = [n for n in names if not n.endswith(".rels")
                       and n != "[Content_Types].xml" and "/" + n not in ovr]
            check(not missing, f"{kind}: every xlsx part has a content type", ", ".join(missing[:3]))

            wb = load_workbook(path)
            check(wb.sheetnames == ["Business case", "Inputs", "Benefit model",
                                    "Financials", "Sensitivity"],
                  f"{kind}: workbook sheet layout", str(wb.sheetnames))

            if not can_calc:
                continue
            sol = _engine(path).calculate()
            gr = GROSS_ROW[kind]
            for label, sheet, cell, want in [
                ("gross annual value", "Benefit model", f"B{gr}", e["gross"]),
                ("realised annual benefit", "Benefit model", f"B{gr + 2}", e["real"]),
                ("total investment", "Financials", "B7", e["inv"]),
                ("net present value", "Financials", "B13", e["npv"]),
                ("cover sheet cross-reference", "Business case", "B9", e["real"]),
                ("base-case sensitivity NPV", "Sensitivity", "D5", e["npv"]),
            ]:
                got = _read(sol, path.name, sheet, cell)
                check(approx(got, want, 1e-6), f"{kind}: workbook recalculates the {label}",
                      f"got {got!r} want {want}")


# --------------------------------------------------------------------- A11Y
def test_a11y() -> None:
    """Guard the accessibility and compatibility fixes from silently regressing."""
    src = HTML.read_text(encoding="utf-8")

    # A regex lookbehind throws a SyntaxError at construction on Safari < 16.4,
    # which would abort the whole script block - taking the formula cards, the
    # wizard, the template previews and every download with it.
    check("(?<!" not in src,
          "no regex lookbehind (it aborts the entire script on Safari < 16.4)")
    check("(^|[^\\\\w-])(" in src, "glossary regex uses the portable leading-capture form")

    # The glossary terms carry tabindex, so they must be operable from a keyboard.
    check("if(e.key !== 'Enter' && e.key !== ' ' && e.key !== 'Spacebar') return;" in src,
          "glossary terms respond to Enter and Space, not just a mouse click")
    check("s.setAttribute('role','button')" in src, "glossary terms expose a button role")
    check("s.setAttribute('aria-label'" in src, "glossary terms carry an accessible name")
    check("back.focus()" in src, "closing the explainer returns focus to the term that opened it")

    # Both modals should announce themselves.
    check(src.count('aria-modal="true"') + src.count("m.setAttribute('aria-modal','true')") >= 2,
          "both dialogs are marked aria-modal")
    check('<div id="tplPrev" role="dialog"' in src, "template preview is a labelled dialog")

    check('class="skiplink"' in src, "a skip-to-content link exists for keyboard users")
    check("prefers-reduced-motion" in src, "the page honours prefers-reduced-motion")
    check("behavior:'smooth'" not in src,
          "scripted scrolls go through SCROLL_BEHAVIOR rather than hardcoding smooth")
    check(":focus-visible" in src, "a visible focus ring is defined")


JARGON = [
    # Terms that were used in the prose with no explainer behind them. If any of
    # these stops resolving, the page is talking jargon at people again.
    "Six Sigma", "Lean", "DMADV", "PDCA", "TQM", "QA", "ASQ", "IASSC", "CSSBB", "BOK",
    "SOP", "KB", "WFM", "CRM", "ACD", "IVR", "ETL", "ROI", "KPI", "SME", "DOE", "EVOP",
    "RACI", "SCAMPER", "TRIZ", "ADKAR", "QFD", "Non-parametric", "Parametric",
    "Transformation", "Normality", "Skew", "Poisson", "Binomial", "Exponential",
    "Hypothesis test", "Null hypothesis", "Type I error", "Type II error", "Sample size",
    "Statistical significance", "Welch", "Wilcoxon", "Tukey", "Levene", "Post-hoc",
    "Bonferroni", "Independence", "Clustering", "Poisson regression", "Negative binomial",
    "Offset", "Control chart", "Control limits", "Special cause", "Common cause",
    "Rational subgrouping", "I-MR", "Moving range", "p-chart", "u-chart", "g-chart",
    "t-chart", "Nelson rules", "Capability", "Cpu", "Pp", "USL", "LSL", "Z-score",
    "Attribute agreement", "Repeatability", "Reproducibility", "Bias", "Linearity",
    "Mix shift", "Operational definition", "Value stream", "Cycle time", "Lead time",
    "Touch time", "Queue time", "Kanban", "Backlog", "Deflection", "Tier 2",
    "Disposition code", "After-call work", "Concurrency", "Intraday", "Async",
    "Schedule adherence", "Forecast accuracy", "Queue discipline", "Swivel-chair",
    "Discount rate", "Soft savings", "Loaded cost", "Charter", "Control plan", "Kano",
    "Pugh", "Multiple regression", "Residual", "Multicollinearity", "Holdout",
    "Histogram", "Boxplot", "Run chart", "Pareto chart", "Percentile",
    "Standard deviation", "Variance", "Changeover", "Pull system", "Bottleneck",
]



def gloss_keys(src: str) -> list[str]:
    """Every term slot in the glossary, duplicates included.

    Walks each object literal by brace depth rather than by regex, because the
    definitions themselves are full of braces, quotes and apostrophes.
    """
    out: list[str] = []
    for spot in [m.end() - 1 for m in
                 re.finditer(r"const GLOSS=\{|Object\.assign\(GLOSS,\s*\{", src)]:
        depth, quote, i, start = 0, None, spot, spot + 1
        while i < len(src):
            ch = src[i]
            if quote:
                if ch == "\\":
                    i += 2
                    continue
                if ch == quote:
                    quote = None
            elif src[i:i + 2] == "/*":
                i = src.index("*/", i) + 2
                continue
            elif ch in "\"'":
                quote = ch
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    break
            elif ch == ":" and depth == 1:
                m = re.search(r'(?:"([^"]+)"|\'([^\']+)\'|([A-Za-z_$][\w$ .&/-]*))\s*$',
                              src[start:i].rstrip())
                if m:
                    out.append(next(g for g in m.groups() if g is not None).strip())
            i += 1
    return out


def test_glossary() -> None:
    """No term should be used in the prose without an explainer behind it."""
    src = HTML.read_text(encoding="utf-8")
    gl = src[src.index("const GLOSS="):src.index("/* ================================================================ helpers */")]
    missing = [t for t in JARGON if f'"{t}":{{' not in gl and f"'{t}'" not in gl and f'"{t}"' not in gl]
    check(not missing, f"every audited term has a glossary entry ({len(JARGON)} checked)",
          "undefined: " + ", ".join(missing[:8]))

    # The A-Z index used to be hand-written and drifted 148 entries behind GLOSS.
    check("listEl.innerHTML = Object.keys(GLOSS)" in src,
          "the glossary index is generated from GLOSS, not hand-maintained")
    # Anything rendered after load starts with no links unless it is re-glossed.
    check(src.count("window.reGloss(") >= 5,
          f"dynamic surfaces are re-glossed (found {src.count('window.reGloss(')} call sites)")
    check("<pre id=\"tplBody\">" not in src,
          "template preview is not a <pre> (PRE is skipped, so it could never gloss)")
    check("t + 's?'" in src, "acronym plurals (SLAs, CTQs) resolve to the singular entry")
    # Test-selector results must not be dead ends.
    check("function resultLinks(" in src and "var R_TOOL={" in src,
          "statistical test results link to a tool and a template")

    # A one-character alias links every stray capital letter on the page; "Z"
    # for Z-score did exactly that. The runtime guard drops anything shorter
    # than two characters and de-duplicates aliases added by separate batches.
    check("if(a.length<2) return;" in src,
          "single-character glossary aliases are rejected at runtime")
    check("seen[a.toLowerCase()]=1" in src, "duplicate aliases are collapsed")

    # NOAUTO suppresses a term from being linked inline. It had grown to 21
    # entries and was quietly hiding t-test, ANOVA, Chi-square, Kappa, Gemba, A3
    # and Tollgate from the prose entirely. Keep it small and deliberate.
    noauto = re.search(r"var NOAUTO = \{([^}]*)\}", src)
    check(bool(noauto), "NOAUTO list is present")
    if noauto:
        n = len(re.findall(r"'[^']+':1", noauto.group(1)))
        check(n <= 10, f"NOAUTO stays short ({n} terms) so technical terms remain clickable")
        for t in ("'t-test'", "'ANOVA'", "'Chi-square'", "'Kappa'", "'Tollgate'", "'Gemba'"):
            check(t not in noauto.group(1), f"{t} is not suppressed from the prose")
    # Ordinary verbs must not be registered as synonyms for Lean vocabulary.
    for bad in ('"Pull","Push"', '"Paired data","Matched"'):
        check(bad not in src, f"no common-verb alias: {bad}")
    check('"P&L":{' in src, "P&L is defined")


def test_toollib() -> None:
    """The tool library's navigation aids, and that nothing dangles."""
    src = HTML.read_text(encoding="utf-8")

    names = re.findall(r'<span class="tn">([^<]*)</span>', src)
    check(len(names) == 52, f"52 tools present (found {len(names)})")

    # Every tool the picker recommends must exist under exactly that name, or the
    # picker renders a gap and the user hits a dead end.
    # PHKEY exists in both tool modules, so anchor the end of the block *after*
    # PICK starts or the slice runs backwards and silently matches nothing.
    p0 = src.index("var PICK=[")
    p1 = src.index("var PHKEY=", p0)
    block = src[p0:p1]
    # Scan both quote styles in one left-to-right pass. A single-quote-only regex
    # desynchronises on "Levene's test (...)" - the apostrophe inside a
    # double-quoted string reads as an opening quote and shifts every match after it.
    picked = set()
    for arr in re.findall(r"tools:\[(.*?)\]", block, re.S):
        for a, b in re.findall(r"'([^']*)'|\"([^\"]*)\"", arr):
            picked.add(a or b)
    missing = [n for n in names if n not in picked]
    check(not missing, f"every tool is reachable from the picker",
          "unreachable: " + ", ".join(missing[:5]))

    # Tool -> calculator links must point at formula cards that exist.
    calc_ids = set(re.findall(r'\{id:"([a-z]+)",group:', src))
    linked = set(re.findall(r"^\s*([a-z]+):\['([a-z',]+)'\],?$", src, re.M))
    referenced = set()
    blk = src[src.index("var TOOL_CALC={"):src.index("var CALC_NAME=")]
    for grp in re.findall(r"\[([^\]]+)\]", blk):
        referenced |= {x.strip().strip("'") for x in grp.split(",")}
    check(referenced and referenced <= calc_ids,
          "every tool->calculator link points at a real formula card",
          f"dangling: {sorted(referenced - calc_ids)}")
    # Was the literal assignment. qa_visual.audit_glossary counts elements with
    # an fml- id in the built DOM: 14 today, 0 if the assignment goes.

    for needed in ("function slugify(", "'tgroup p-'", "id=\"tExpand\"", "id=\"tReset\"",
                   "'tempty'", "mark.thit", "id=\"toolPick\"", "openFromHash",
                   "idx.className='tindex'", "className='fmlback'", "window.__gotoTool"):
        check(needed in src, f"tool library affordance present: {needed}")

    # Every formula card should name the tool that explains its method, so the
    # link runs both ways instead of stranding the reader in the arithmetic.
    card_ids = set(re.findall(r'\{id:"([a-z]+)",group:', src))
    linked_back = set()
    blk = src[src.index("var TOOL_CALC={"):src.index("var CALC_NAME=")]
    for grp in re.findall(r"\[([^\]]+)\]", blk):
        linked_back |= {x.strip().strip("'") for x in grp.split(",")}
    orphans = sorted(card_ids - linked_back)
    check(not orphans, "every formula card links back to a tool",
          f"no backlink for: {orphans}")

    # Finance is told they can check the maths in the workbook, so every chart
    # drawn in the HTML case must have a counterpart bound to cells in the Excel.
    html_charts = re.findall(r"emBar\(\{title:'([^']*)'", src)
    xl_charts = re.findall(r"type:'(?:col|bar|line)', title:'([^']*)'", src)
    check(len(html_charts) == len(xl_charts) and len(html_charts) == 4,
          f"every on-screen chart has an Excel counterpart "
          f"({len(html_charts)} in HTML, {len(xl_charts)} in Excel)",
          f"HTML={html_charts} XL={xl_charts}")


# --------------------------------------------------------------------- main
def main() -> int:
    fast = "--fast" in sys.argv
    print("BOK        every cited certification section exists")
    test_bok()
    print("EXPORT-CH  the email export carries charts and the legend")
    test_export_charts()
    print("BUILD      workbooks are byte-reproducible")
    test_deterministic()
    print("READABLE   nothing the build writes is a wall of text")
    test_readable_blocks()
    print("IDEMPOTENT the finishing pass changes nothing on a second run")
    test_idempotent()
    print("STRUCTURE  merged-cell reference audit")
    test_structure()
    print("SYNC       four-way template consistency")
    test_sync()
    print("A11Y       keyboard, dialogs and browser compatibility")
    test_a11y()
    print("TOOLS      library navigation, picker and calculator links")
    test_toollib()
    print("GLOSSARY   jargon coverage and dynamic linking")
    test_glossary()
    print("CITATIONS  every file.md:NN lands on a line that carries the figure")
    test_citations()
    print("CHARTS+    every chart says how to read it")
    test_legend_matches_sheets()
    test_chart_notes()
    print("GLOSSARY+  every jargon column header carries a plain-English key")
    test_glossary_coverage()
    print("CASE       the worked project's figures reproduce from each other")
    test_case_study()
    print("GUIDANCE   the template filler's numbers match the pack's")
    test_guidance()
    print("EXPORT     business case HTML + live-formula workbook")
    test_export()
    if fast:
        print("NUMERIC    skipped (--fast)")
    else:
        try:
            import formulas  # noqa: F401
        except ImportError:
            print("NUMERIC    SKIPPED - pip install formulas to run the recalculation tests")
        else:
            print("NUMERIC    recalculating fixed formulas")
            test_numeric()
            test_numeric_other()

    print()
    if FAILURES:
        print(f"FAILED  {len(FAILURES)} check(s), {PASSES[0]} passed\n")
        for f in FAILURES:
            print("  x " + f)
        return 1
    print(f"PASSED  all {PASSES[0]} checks")
    return 0


if __name__ == "__main__":
    sys.exit(main())
