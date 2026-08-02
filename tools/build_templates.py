#!/usr/bin/env python3
"""Build the templates that were missing, and their previews.

Eleven tools had no template at all, and several more pointed at one that did
not fit — the fishbone and the 5 Whys both linked to the X-Y matrix, which is a
scoring grid, not a cause diagram. These fill those gaps.

Every workbook follows the same contract as the originals: a "How to use this"
tab, yellow cells you fill in, blue cells that calculate, one worked example
row, and a note on every input saying where the number comes from.

    python3 tools/build_templates.py     # writes templates/*.xlsx + previews.json
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.comments import Comment

sys.path.insert(0, str(Path(__file__).resolve().parent))
import chartsvg  # noqa: E402
from preview import shown_from, workbook_html  # noqa: E402
from xlpolish import polish_workbook, save_workbook  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "templates"

TITLE = PatternFill("solid", fgColor="FF151B24")
BAND = PatternFill("solid", fgColor="FFEEF1F6")
HDR = PatternFill("solid", fgColor="FF333C49")
IN = PatternFill("solid", fgColor="FFFFF9E3")
CALC = PatternFill("solid", fgColor="FFF2F7FF")
EX = PatternFill("solid", fgColor="FFECFAEF")
THIN = Border(bottom=Side("thin", color="FFD8DEE7"))

F_TITLE = Font(bold=True, size=15, color="FFFFFFFF")
F_HDR = Font(bold=True, size=10, color="FFFFFFFF")
F_BAND = Font(bold=True, size=11)
F_B = Font(bold=True, size=11)
F_NOTE = Font(italic=True, size=9, color="FF6B7280")
F_CALC = Font(size=11, color="FF1D4ED8")

SHOWN: dict = {}          # (sheet, cell) -> what the preview should display


def note(ws, row, col, text):
    """Attach 'where this number comes from' to an input cell.

    Inputs inside a table are explained once by their column header. A
    standalone input has no header, so it carries its own note — which is what
    the page means when it says every yellow cell tells you where to look.
    """
    c = ws.cell(row=row, column=col)
    c.comment = Comment(text, "Template")
    c.comment.width = 320
    c.comment.height = 110
    return c


def title(ws, text, sub, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.cell(row=1, column=1, value=text).font = F_TITLE
    for c in range(1, width + 1):
        ws.cell(row=1, column=c).fill = TITLE
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.cell(row=2, column=1, value=sub).font = F_NOTE
    ws.sheet_view.showGridLines = False


def band(ws, row, text, width):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    ws.cell(row=row, column=1, value=text).font = F_BAND
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = BAND


def header(ws, row, labels):
    for i, l in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=l)
        c.fill, c.font = HDR, F_HDR
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30


def note_cell(ws, row, col, text, width):
    """A wrapped explanation spanning the rest of the row."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=width)
    c = ws.cell(row=row, column=col, value=text)
    c.font = F_NOTE
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = max(15, 15 * (1 + len(text) // 95))
    return c


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def howto(wb, lines):
    ws = wb.create_sheet("How to use this", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 108
    for i, (bold, text) in enumerate(lines, start=2):
        c = ws.cell(row=i, column=2, value=text)
        c.font = F_B if bold else Font(size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if not bold:
            ws.row_dimensions[i].height = max(15, 15 * (len(text) // 100 + 1))
    return ws



def overlay(ch, ws, ref, name, colour="C0392B"):
    """Lay a dashed reference line over a bar chart — the standard, on the data."""
    from openpyxl.chart.legend import Legend
    ln = LineChart()
    sr = Series(ref, title=name)
    sr.graphicalProperties.line.solidFill = colour
    sr.graphicalProperties.line.width = 18000
    sr.graphicalProperties.line.dashStyle = "dash"
    sr.marker = Marker(symbol="none")
    sr.smooth = False
    ln.series.append(sr)
    ch += ln
    ch.legend = Legend()
    ch.legend.position = "b"
    return ch


def bar(ws, title_, cat_ref, val_ref, anchor, horizontal=False, pct=False, series=None,
        colours=None, name=None):
    """A chart bound to the cells, so it moves when the numbers do.

    Series titles are set here rather than read from a header cell, because the
    header row is usually inside a merged band and merged cells are read-only.
    """
    ch = BarChart()
    ch.type = "bar" if horizontal else "col"
    ch.style = 10
    ch.title = title_
    ch.legend = None
    ch.height, ch.width = 7.5, 15
    ch.gapWidth = 60
    ch.add_data(val_ref, titles_from_data=False)
    ch.set_categories(cat_ref)
    if name and ch.series:
        ch.series[0].tx = SeriesLabel(v=name)   # else the legend reads "Series1"
    if ch.series and colours:
        from openpyxl.chart.marker import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties
        ch.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(solidFill=c))
            for i, c in enumerate(colours)]
    elif ch.series:
        ch.series[0].graphicalProperties.solidFill = "1F4E79"
        ch.series[0].graphicalProperties.line.solidFill = "1F4E79"
    # OOXML's invertIfNegative defaults on, and with no negative fill defined
    # the renderer draws negative bars as nothing at all. This has to apply on
    # both paths above — the per-point-colour branch used to skip it.
    for ser in ch.series:
        ser.invertIfNegative = False
    if series:
        ch.series[0].tx = None
    ch.y_axis.majorGridlines = None if horizontal else ch.y_axis.majorGridlines
    if pct:
        ch.y_axis.numFmt = "0%"
    ws.add_chart(ch, anchor)
    return ch


def mark(ws, row, col, kind, note=None):
    c = ws.cell(row=row, column=col)
    if kind == "in":
        c.fill, c.border = IN, THIN
    elif kind == "calc":
        c.fill, c.font, c.border = CALC, F_CALC, THIN
    elif kind == "ex":
        c.fill, c.border = EX, THIN
    return c


LEGEND = [
    (True, "Yellow cells"), (False, "You fill these in."),
    (True, "Blue cells"), (False, "Calculated for you. Do not type over them — they contain formulas."),
    # Plural, and "replace" rather than "delete": on the sheets that ship seeded
    # with a full dataset — the control charts, the Gage R&R study, regression —
    # the worked example IS the entry area. Deleting it leaves an empty chart.
    (True, "Green cells"), (False, "A worked example, so you can see the expected format and the "
                                   "charts say something the moment you open the file. Replace it "
                                   "with your own data when you start."),
]


# ---------------------------------------------------------------- 20 five whys
def five_whys():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("5 Whys tree")
    title(ws, "5 Whys — run it as a tree, not a chain",
          "A single chain gives you one cause and false confidence. Branching gives you a hypothesis set you can test.", 9)
    widths(ws, [10, 30, 30, 30, 30, 30, 12, 14, 32])   # 6 clipped "Instance" to "Instan"
    band(ws, 4, "THE PROBLEM — one specific instance, not a general statement", 9)
    ws.cell(row=5, column=1, value="Instance").font = F_B
    ws.merge_cells("B5:E5"); mark(ws, 5, 2, "in")
    ws.cell(row=5, column=2, value="Billing ticket #48217, closed 12 Mar, reopened 14 Mar")
    note(ws, 5, 2, "One concrete record, with it open in front of you: ticket number, what happened, "
                   "and when. 'Reopens are high' is not an instance and cannot be walked back to a cause.")
    ws.cell(row=5, column=7, value="Date").font = F_B
    mark(ws, 5, 8, "in").value = "2026-03-14"
    note(ws, 5, 8, "The date of the instance, not the date you ran the session. You will want to go "
                   "back to the system state as it was.")
    ws.cell(row=6, column=1, value="Effect").font = F_B
    ws.merge_cells("B6:E6"); mark(ws, 6, 2, "in")
    ws.cell(row=6, column=2, value="Customer had to contact us a second time about the same adjustment")
    note(ws, 6, 2, "What the customer experienced, in their terms. Not the internal symptom and not "
                   "the fix you already have in mind.")
    ws.cell(row=7, column=2, value="Start from a concrete instance with the record in front of you. "
            "At each level ask 'what else could cause this?' before you go deeper.").font = F_NOTE

    # It is called the 5 Whys. Collapsing four and five into one column asks
    # four, and four is usually one short of anything you can act on.
    header(ws, 9, ["#", "Why 1", "Why 2", "Why 3", "Why 4", "Why 5", "Branch?",
                   "Testable with data?", "How you would test it"])
    # The tool is called "a tree, not a chain", so a single-branch example
    # argues against the template. Three branches from the same instance.
    ex = [
        ["1", "The adjustment had not posted when the ticket was closed",
         "Closure is allowed before the posting webhook confirms",
         "The ticket status model has no 'pending adjustment' state",
         "The billing integration was specified as fire-and-forget, with no callback",
         "No one owned the ticket lifecycle across billing and support when it was designed",
         "A", "Yes", "Count reopens where closure preceded the webhook timestamp"],
        ["2", "The agent could not see whether the adjustment had posted",
         "Billing status lives in another system with no view inside the ticket",
         "The integration was scoped to write, never to read",
         "Scope was cut on integration cost, with no one costing the agent's workflow",
         "Design reviews do not require sign-off from anyone who works the queue",
         "B", "Yes", "Sample 50 reopens and check whether the agent had visibility at closure"],
        ["3", "We only count reopens raised for the same reason",
         "The metric was defined by reporting without asking operations",
         "No operational definition was agreed when the project started",
         "The charter template does not require one before the baseline is taken",
         "Metric definitions have no owner anywhere in the organisation",
         "C", "Yes", "Recount reopens on an any-reason definition and compare the two series"],
    ]
    for k, row in enumerate(ex):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=10 + k, column=i, value=v); c.fill = EX; c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
        # size the row to its longest cell: a fixed 30pt clipped the last line
        # of every branch, which is exactly the line that names the root cause
        widest = max(len(str(v)) for v in row[1:6])
        ws.row_dimensions[10 + k].height = max(32, 15 * -(-widest // 25))
    for r in range(13, 26):
        ws.cell(row=r, column=1, value=r - 9).font = F_NOTE
        for c in range(2, 10):
            mark(ws, r, c, "in").alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv); dv.add("H10:H25")

    band(ws, 27, "CHECK BEFORE YOU LEAVE THIS TOOL", 9)
    rows = [("Branches explored", '=COUNTA(B10:B25)', "3"),
            ("Branches you can test with data", '=COUNTIF(H10:H25,"Yes")', "3"),
            ("Chains that reached five levels", '=COUNTA(F10:F25)', "3"),
            ("Chains that stopped at a person",
             '=COUNTIF(F10:F25,"*training*")+COUNTIF(F10:F25,"*follow the process*")', "0")]
    for i, (label, f, shown) in enumerate(rows, start=28):
        ws.cell(row=i, column=1, value=label).font = F_B
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
        c = mark(ws, i, 6, "calc"); c.value = f
        SHOWN[("5 Whys tree", f"F{i}")] = shown
    ws.merge_cells("A33:I33")
    ws.cell(row=33, column=1, value="If your last why is 'not enough training' or 'the agent did not follow the "
            "process', you stopped one level too early. Ask why the process was skippable — that is the cause you "
            "can actually fix.").font = F_NOTE
    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Moving from a symptom to a cause you can act on, without the false confidence a single chain gives you."),
        (True, "The two rules"),
        (False, "1. Start from one concrete instance with the record open. Not 'reopens are high' — ticket #48217."),
        (False, "2. At every level ask 'what else could cause this?' before descending. Two or three branches, not one line."),
        (True, "When to stop"),
        (False, "When you reach something the organisation can change, or when the next why leaves your sphere entirely."),
        (True, "What comes out"),
        (False, "A hypothesis list, not a root cause. It becomes a root cause only when you test it — that is what the "
                "last two columns are for. Carry the testable branches into the X-Y matrix and the hypothesis test log."),
    ])
    return wb, "20-five-whys-tree.xlsx"


# --------------------------------------------------------- 21 cause and effect
def fishbone():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Fishbone")
    title(ws, "Cause & effect (Ishikawa) — support categories, not 6M",
          "Machine / Method / Material / Measurement / Man / Mother Nature was written for a factory. These are the categories a support process actually fails in.", 6)
    widths(ws, [30, 46, 14, 14, 16, 34])
    band(ws, 4, "THE EFFECT — use the problem statement's metric and magnitude", 6)
    ws.merge_cells("B5:F5"); mark(ws, 5, 2, "in")
    ws.cell(row=5, column=1, value="Effect").font = F_B
    ws.cell(row=5, column=2, value="7-day reopen rate on billing tickets is 14.2% against a target of 8%")
    note(ws, 5, 2, "Copy the problem statement's metric and magnitude verbatim — the measured value, "
                   "the target, and the population. A vague effect gives you a fishbone of vague causes.")

    header(ws, 7, ["Category", "Candidate cause", "Likelihood 1-5", "Impact 1-5",
                   "Priority", "Evidence you would need"])
    cats = ["People", "Process", "Systems", "Knowledge", "Routing & demand",
            "Product / upstream", "Measurement"]
    # One example cause left six of the seven branches reading "EMPTY — look
    # again", so the template's own quality gate was firing on its own example.
    ex = [
        ["Process", "Ticket can be closed before the billing adjustment posts", 5, 5,
         "Reopens where closure timestamp precedes the posting webhook"],
        ["Process", "Billing Ops pull the queue once a day, so nothing moves overnight", 4, 4,
         "Queue age distribution by hour of arrival"],
        ["Systems", "The status model has no pending-adjustment state", 5, 5,
         "Count of closures with an adjustment still in flight"],
        ["Systems", "The nightly adjustment batch fails silently", 3, 5,
         "Batch exit codes against the days reopens spiked"],
        ["People", "New hires close on the request, not on the posting", 3, 4,
         "Reopen rate split by agent tenure band"],
        ["Knowledge", "No written definition of 'resolved' for a billing adjustment", 4, 3,
         "Ask ten agents to define it and compare the answers"],
        ["Routing & demand", "Adjustment requests route to general billing, not the adjustments desk", 3, 3,
         "Share of reopens that were transferred at least once"],
        ["Product / upstream", "The billing platform posts asynchronously with no callback", 4, 5,
         "Distribution of the gap between request and posting"],
        ["Measurement", "Reopen rate counts same-reason reopens only, hiding the rest", 3, 4,
         "Recount on an any-reason definition"],
    ]
    for k, row in enumerate(ex):
        rr = 8 + k
        for i, v in enumerate(row, start=1):
            col = i if i < 5 else 6            # column 5 is the calculated priority
            c = ws.cell(row=rr, column=col, value=v); c.fill = EX; c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
        c = mark(ws, rr, 5, "calc"); c.value = f'=IF(COUNT(C{rr}:D{rr})<2,"",C{rr}*D{rr})'
        c.fill = EX
        SHOWN[("Fishbone", f"E{rr}")] = str(row[2] * row[3])
    for rr in range(8 + len(ex), 36):
        for cc in range(1, 7):
            mark(ws, rr, cc, "in").alignment = Alignment(wrap_text=True, vertical="top")
        c = mark(ws, rr, 5, "calc")
        c.value = f'=IF(COUNT(C{rr}:D{rr})<2,"",C{rr}*D{rr})'
        SHOWN[("Fishbone", f"E{rr}")] = ""
    # Key the table so the diagram can pull "the nth cause in this category".
    # INDEX/MATCH on a composite key works in every Excel version; FILTER and
    # TEXTJOIN do not.
    ws.cell(row=7, column=20, value="key (used by the diagram)").font = F_NOTE
    for rr in range(8, 36):
        k = ws.cell(row=rr, column=20)
        k.value = f'=IF(A{rr}="","",A{rr}&"|"&COUNTIF($A$8:A{rr},A{rr}))'
        k.font = F_NOTE
        SHOWN[("Fishbone", f"T{rr}")] = ""
    ws.column_dimensions[get_column_letter(20)].width = 22
    ws.column_dimensions[get_column_letter(20)].hidden = True   # plumbing, not content

    dvc = DataValidation(type="list", formula1='"%s"' % ",".join(cats), allow_blank=True)
    ws.add_data_validation(dvc); dvc.add("A8:A35")
    dvs = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    ws.add_data_validation(dvs); dvs.add("C8:D35")

    band(ws, 37, "BRANCH BALANCE — a thin branch is a blind spot, not an absence of causes", 6)
    for i, cat in enumerate(cats):
        r = 38 + i
        ws.cell(row=r, column=1, value=cat).font = F_B
        c = mark(ws, r, 2, "calc"); c.value = f'=COUNTIF($A$8:$A$35,A{r})'
        n = sum(1 for e in ex if e[0] == cat)
        SHOWN[("Fishbone", f"B{r}")] = str(n)
        c2 = mark(ws, r, 3, "calc")
        c2.value = f'=IF(COUNTIF($A$8:$A$35,A{r})=0,"EMPTY — look again","")'
        SHOWN[("Fishbone", f"C{r}")] = "" if n else "EMPTY — look again"
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for i in range(7):
        r = 38 + i
        c = mark(ws, r, 8, "calc")
        c.value = "=AVERAGE($B$38:$B$44)"
        c.number_format = "#,##0.0"
        SHOWN[("Fishbone", f"H{r}")] = "%.1f" % (len(ex) / 7.0)
    ws.cell(row=37, column=8, value="Mean causes per branch").font = F_NOTE
    fb = bar(ws, "Causes per branch — a thin bar is a blind spot",
             Reference(ws, min_col=1, min_row=38, max_row=44),
             Reference(ws, min_col=2, min_row=38, max_row=44), "J8",
             name="Causes on this branch")
    overlay(fb, ws, Reference(ws, min_col=8, min_row=38, max_row=44),
            "Mean — below this is a branch you are not looking at")
    ws.merge_cells("A46:F46")
    ws.cell(row=46, column=1, value="Support teams over-populate People and under-populate Systems and Knowledge, "
            "because blaming agents is culturally available. An empty Measurement branch is almost always wrong — "
            "if you have not questioned the measurement, you have not finished.").font = F_NOTE

    # ---- the diagram itself -------------------------------------------
    # An Ishikawa is a fishbone-SHAPED diagram: a spine pointing at the effect
    # with angled bones for each category. A table and a bar chart of counts is
    # not one. This draws the real thing, and every cause cell is a formula
    # reading the table, so the picture updates as you type.
    wsd = wb.create_sheet("Fishbone diagram", 1)
    wsd.sheet_view.showGridLines = False
    BONE = Side("medium", color="FF333C49")
    SPINE = Side("thick", color="FF151B24")

    wsd.column_dimensions["A"].width = 2
    for c in range(2, 27):
        wsd.column_dimensions[get_column_letter(c)].width = 7.5
    for c in range(27, 31):
        wsd.column_dimensions[get_column_letter(c)].width = 13

    wsd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=30)
    t = wsd.cell(row=1, column=1, value="Cause & effect — the diagram")
    t.font = F_TITLE
    for c in range(1, 31):
        wsd.cell(row=1, column=c).fill = TITLE
    wsd.row_dimensions[1].height = 26
    wsd.merge_cells(start_row=2, start_column=1, end_row=2, end_column=30)
    wsd.cell(row=2, column=1, value="Every box below is a formula reading the Fishbone tab. Type there; "
             "this redraws itself. Four causes per branch are shown — the tab holds as many as you like.").font = F_NOTE

    TOP = ["People", "Process", "Systems", "Knowledge"]
    BOT = ["Routing & demand", "Product / upstream", "Measurement"]
    CAUSE_ROWS_TOP = [6, 8, 10, 12]
    CAUSE_ROWS_BOT = [18, 20, 22, 24]
    SPINE_ROW = 15

    def cause(row, c1, c2, category, n):
        wsd.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = wsd.cell(row=row, column=c1)
        cell.value = ('=IFERROR(INDEX(Fishbone!$B$8:$B$35,'
                      'MATCH("%s|%d",Fishbone!$T$8:$T$35,0)),"")' % (category, n))
        cell.font = Font(size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.fill = CALC
        wsd.row_dimensions[row].height = 26
        SHOWN[("Fishbone diagram", cell.coordinate)] = ""

    def catbox(row, c1, c2, name):
        wsd.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = wsd.cell(row=row, column=c1, value=name)
        cell.font = F_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(c1, c2 + 1):
            wsd.cell(row=row, column=c).fill = HDR
        wsd.row_dimensions[row].height = 20

    for j, name in enumerate(TOP):
        c = 2 + 6 * j
        catbox(4, c, c + 3, name)
        for i, r in enumerate(CAUSE_ROWS_TOP, start=1):
            cause(r, c, c + 3, name, i)
        # one merged block with a single diagonal is a clean straight bone;
        # a stack of per-cell diagonals renders as a dashed staircase
        wsd.merge_cells(start_row=5, start_column=c + 4, end_row=SPINE_ROW - 1, end_column=c + 5)
        wsd.cell(row=5, column=c + 4).border = Border(diagonal=BONE, diagonalDown=True)

    for j, name in enumerate(BOT):
        c = 5 + 6 * j
        for i, r in enumerate(CAUSE_ROWS_BOT, start=1):
            cause(r, c, c + 3, name, i)
        catbox(26, c, c + 3, name)
        wsd.merge_cells(start_row=SPINE_ROW + 1, start_column=c + 4, end_row=25, end_column=c + 5)
        wsd.cell(row=SPINE_ROW + 1, column=c + 4).border = Border(diagonal=BONE, diagonalUp=True)

    for c in range(2, 27):
        wsd.cell(row=SPINE_ROW, column=c).border = Border(bottom=SPINE)
    wsd.row_dimensions[SPINE_ROW].height = 10

    wsd.merge_cells(start_row=SPINE_ROW - 3, start_column=27, end_row=SPINE_ROW + 3, end_column=30)
    eff = wsd.cell(row=SPINE_ROW - 3, column=27, value="=Fishbone!B5")
    eff.font = Font(bold=True, size=12, color="FFFFFFFF")
    eff.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for r in range(SPINE_ROW - 3, SPINE_ROW + 4):
        for c in range(27, 31):
            wsd.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFC0392B")
    SHOWN[("Fishbone diagram", eff.coordinate)] = \
        "7-day reopen rate on billing tickets is 14.2% against a target of 8%"

    wsd.merge_cells(start_row=28, start_column=1, end_row=28, end_column=30)
    wsd.cell(row=28, column=1, value="A branch with no boxes is where you are not looking, not where there is "
             "nothing to find. An empty Measurement branch is almost always wrong.").font = F_NOTE

    ws2 = wb.create_sheet("Category prompts")
    ws2.sheet_view.showGridLines = False
    widths(ws2, [26, 96])
    title(ws2, "What lives in each branch", "Prompts to force entries into the thin branches.", 2)
    prompts = [
        ("People", "Skill, tenure, authority limits, coaching, workload, incentive conflicts."),
        ("Process", "Steps that can be skipped, missing pending states, approvals, handoffs, queue discipline."),
        ("Systems", "Tooling gaps, swivel-chair work, latency, permissions, integrations that fail silently."),
        ("Knowledge", "Missing or wrong articles, search that does not find them, content nobody owns."),
        ("Routing & demand", "Mis-routing, IVR options, skill-based routing rules, arrival patterns, mix shift."),
        ("Product / upstream", "Defects reaching customers, confusing UX, billing logic, release cadence."),
        ("Measurement", "Disposition codes, QA rubric, operational definitions, instrumentation gaps."),
    ]
    for i, (k, v) in enumerate(prompts, start=4):
        ws2.cell(row=i, column=1, value=k).font = F_B
        c = ws2.cell(row=i, column=2, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[i].height = 28
    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Getting every candidate cause out of the team's heads and onto one page, grouped so the gaps are visible."),
        (True, "How to run it"),
        (False, "Write the effect precisely first. Brainwrite silently for five minutes before anyone speaks — "
                "verbal-first brainstorming is dominated by the most senior voice in the room."),
        (False, "Push each branch down two levels with 'why does that happen?'. Then score likelihood and impact."),
        (True, "The check that matters"),
        (False, "Look at the branch balance table. A thin or empty branch is where you are not looking, not where "
                "there is nothing to find."),
        (True, "What comes out"),
        (False, "Causes on a fishbone are hypotheses, not findings. Nothing leaves this tool without being tested — "
                "carry the highest-priority ones into the X-Y matrix."),
    ])
    return wb, "21-cause-effect-fishbone.xlsx"


# ------------------------------------------------------ 22 stakeholder / RACI
STAKEHOLDERS = [('A. Okafor', 'Billing Ops Manager', 5, -1, 'Proof this will not add work to her team', 'Black Belt'), ('R. Mehta', 'Support Director', 5, 2, 'A monthly number she can take to the exec review', 'Champion'), ('J. Lindqvist', 'Finance Business Partner', 4, 0, 'The benefit accounting policy agreed before baseline', 'Black Belt'), ('S. Duarte', 'WFM Lead', 4, -2, 'Assurance the harvest is scheduled, not assumed', 'Champion'), ('P. Nwosu', 'Platform Engineering Manager', 3, 1, 'A scoped ticket, not a standing request', 'Black Belt'), ('K. Tanaka', 'QA Lead', 3, 2, 'Involvement in the operational definition', 'Black Belt'), ('M. Alvarez', 'Tier 1 Team Lead', 2, 1, 'Her agents consulted before the process changes', 'Process owner'), ('D. Byrne', 'Compliance', 2, 0, 'Sight of anything touching retention or consent', 'Champion')]

RACI_ROWS = [('Sign the benefit and harvest mechanism', 'C', 'A', 'C', 'C', 'R'), ('Agree the operational definition of a defect', 'R', 'I', 'A', 'I', 'C'), ('Approve the pilot design and stopping rule', 'R', 'A', 'C', 'C', 'I'), ('Change the routing rules in production', 'C', 'I', 'A', 'C', 'I'), ('Own the control chart and its reaction plan', 'C', 'I', 'A', 'R', 'I'), ('Release the headcount the project frees', 'I', 'C', 'C', 'R', 'A')]


def stakeholder():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Stakeholders")
    title(ws, "Stakeholder analysis — who can stop this, and who will",
          "Projects rarely fail on the statistics. They fail because someone with influence was never asked.", 7)
    widths(ws, [26, 26, 14, 14, 16, 30, 30])
    header(ws, 4, ["Name", "Role", "Influence 1-5", "Support -2..+2", "Action needed",
                   "What they need from you", "Owner"])
    # One stakeholder is not a stakeholder map, and it left every quadrant of
    # the grid empty but one.
    for k, (nm, role, inf, sup, need, owner) in enumerate(STAKEHOLDERS):
        r = 5 + k
        for i, v in enumerate((nm, role, inf, sup), start=1):
            c = ws.cell(row=r, column=i, value=v); c.fill = EX; c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
        for i, v in ((6, need), (7, owner)):
            c = ws.cell(row=r, column=i, value=v); c.fill = EX; c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
        c = mark(ws, r, 5, "calc"); c.fill = EX
        c.value = (f'=IF(COUNT(C{r}:D{r})<2,"",IF(AND(C{r}>=4,D{r}<0),"ENGAGE NOW",'
                   f'IF(C{r}>=4,"KEEP CLOSE",IF(D{r}<0,"MONITOR","INFORM"))))')
        SHOWN[("Stakeholders", f"E{r}")] = (
            "ENGAGE NOW" if inf >= 4 and sup < 0 else
            "KEEP CLOSE" if inf >= 4 else "MONITOR" if sup < 0 else "INFORM")
        ws.row_dimensions[r].height = 26
    for r in range(5 + len(STAKEHOLDERS), 24):
        for cc in [1, 2, 3, 4, 6, 7]:
            mark(ws, r, cc, "in").alignment = Alignment(wrap_text=True, vertical="top")
        c = mark(ws, r, 5, "calc")
        c.value = (f'=IF(COUNT(C{r}:D{r})<2,"",IF(AND(C{r}>=4,D{r}<0),"ENGAGE NOW",'
                   f'IF(C{r}>=4,"KEEP CLOSE",IF(D{r}<0,"MONITOR","INFORM"))))')
        SHOWN[("Stakeholders", f"E{r}")] = ""
    # The stakeholder map itself: influence against support, so the quadrant a
    # name lands in is visible rather than inferred from two columns of digits.
    sc = ScatterChart()
    sc.title = "Stakeholder map — influence against support"
    sc.style = 13
    sc.height, sc.width = 10, 15
    sc.x_axis.title = "Support:  -2 opposed  \u2192  +2 advocate"
    sc.y_axis.title = "Influence over the outcome"
    sc.x_axis.delete = False
    sc.y_axis.delete = False
    sc.x_axis.scaling.min, sc.x_axis.scaling.max = -2.5, 2.5
    sc.y_axis.scaling.min, sc.y_axis.scaling.max = 0, 6
    # Without an explicit scatterStyle no series gets a line, so the quadrant
    # dividers were written to the file and then not drawn by anything.
    sc.scatterStyle = "lineMarker"
    sc.legend = None
    pts = Series(Reference(ws, min_col=3, min_row=5, max_row=4 + len(STAKEHOLDERS)),
                 xvalues=Reference(ws, min_col=4, min_row=5, max_row=4 + len(STAKEHOLDERS)),
                 title="Stakeholders")
    pts.marker = Marker(symbol="circle", size=9)
    pts.marker.graphicalProperties.solidFill = "1F4E79"
    pts.graphicalProperties.line.noFill = True      # points only, no joining line
    sc.series.append(pts)

    # Without the dividers this is a scatter of two columns you already have.
    # With them it is the map: the top-left quadrant is who can stop you.
    # Two endpoints each, pinned to the axis bounds. Drawing the dividers FROM
    # the stakeholder data made each line span only the range the current data
    # happened to occupy, and follow row order to get there — so the "high
    # influence" line stopped at whichever support scores existed instead of
    # crossing the plot.
    n = len(STAKEHOLDERS)
    for col, vals in ((9, (-2.5, 2.5)), (10, (3.5, 3.5)),      # influence threshold
                      (11, (0, 0)), (12, (0, 6))):             # support threshold
        for k, v in enumerate(vals):
            c = ws.cell(row=5 + k, column=col, value=v)
            c.font = F_NOTE
            SHOWN[("Stakeholders", f"{get_column_letter(col)}{5 + k}")] = str(v)
    ws.cell(row=4, column=9, value="quadrant dividers — the chart reads these").font = F_NOTE
    for col in "IJKL":
        ws.column_dimensions[col].width = 6      # narrow, not hidden: a chart
                                                 # will not plot a hidden cell

    hline = Series(Reference(ws, min_col=10, min_row=5, max_row=6),
                   xvalues=Reference(ws, min_col=9, min_row=5, max_row=6),
                   title="High influence threshold")
    hline.marker = Marker(symbol="none")
    hline.graphicalProperties.line.solidFill = "C0392B"
    hline.graphicalProperties.line.dashStyle = "dash"
    sc.series.append(hline)

    vline = Series(Reference(ws, min_col=12, min_row=5, max_row=6),
                   xvalues=Reference(ws, min_col=11, min_row=5, max_row=6),
                   title="Opposed / supportive threshold")
    vline.marker = Marker(symbol="none")
    vline.graphicalProperties.line.solidFill = "C0392B"
    vline.graphicalProperties.line.dashStyle = "dash"
    sc.series.append(vline)
    sc.legend = Legend()
    sc.legend.position = "b"
    ws.add_chart(sc, "N4")

    dvi = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    ws.add_data_validation(dvi); dvi.add("C5:C23")
    dvs = DataValidation(type="whole", operator="between", formula1=-2, formula2=2, allow_blank=True)
    ws.add_data_validation(dvs); dvs.add("D5:D23")
    band(ws, 25, "SUMMARY", 7)
    for i, (label, f, shown) in enumerate([
            ("Stakeholders mapped", '=COUNTA(A5:A23)', str(len(STAKEHOLDERS))),
            ("High influence and opposed — deal with these first", '=COUNTIF(E5:E23,"ENGAGE NOW")',
             str(sum(1 for x in STAKEHOLDERS if x[2] >= 4 and x[3] < 0))),
            ("High influence and supportive — your sponsors", '=COUNTIF(E5:E23,"KEEP CLOSE")',
             str(sum(1 for x in STAKEHOLDERS if x[2] >= 4 and x[3] >= 0)))], start=26):
        ws.cell(row=i, column=1, value=label).font = F_B
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        mark(ws, i, 5, "calc").value = f
        SHOWN[("Stakeholders", f"E{i}")] = shown

    ws2 = wb.create_sheet("RACI")
    ws2.sheet_view.showGridLines = False
    title(ws2, "RACI — exactly one Accountable per row",
          "Two names in the Accountable column is the most common reason a control plan quietly stops being executed.", 7)
    widths(ws2, [40, 16, 16, 16, 16, 16, 26])
    header(ws2, 4, ["Activity / decision", "Black Belt", "Champion", "Process owner",
                    "WFM", "Finance", "Accountable count"])
    for k, row in enumerate(RACI_ROWS):
        r = 5 + k
        for i, v in enumerate(row, start=1):
            c = ws2.cell(row=r, column=i, value=v); c.fill = EX; c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical="center")
        c = mark(ws2, r, 7, "calc"); c.fill = EX
        c.value = (f'=IF(COUNTIF(B{r}:F{r},"A")=1,"OK",IF(COUNTIF(B{r}:F{r},"A")=0,'
                   f'"NO OWNER","MORE THAN ONE"))')
        SHOWN[("RACI", f"G{r}")] = "OK"
    for r in range(5 + len(RACI_ROWS), 22):
        for cc in range(1, 7):
            mark(ws2, r, cc, "in")
        c = mark(ws2, r, 7, "calc")
        c.value = (f'=IF(COUNTA(A{r})=0,"",IF(COUNTIF(B{r}:F{r},"A")=1,"OK",'
                   f'IF(COUNTIF(B{r}:F{r},"A")=0,"NO OWNER","MORE THAN ONE")))')
        SHOWN[("RACI", f"G{r}")] = ""
    dvr = DataValidation(type="list", formula1='"R,A,C,I"', allow_blank=True)
    ws2.add_data_validation(dvr); dvr.add("B5:F21")
    ws2.merge_cells("A23:G23")
    ws2.cell(row=23, column=1, value="R = does the work · A = owns the outcome (exactly one) · "
             "C = asked before the decision · I = told after it.").font = F_NOTE
    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Finding the person who can stop your project before they do, and settling who owns what while it is cheap."),
        (True, "Stakeholders tab"),
        (False, "Score influence 1-5 and support -2 to +2 honestly. The Action column works out who to deal with first: "
                "high influence and opposed is where projects die."),
        (True, "RACI tab"),
        (False, "The count column flags any row that does not have exactly one Accountable. Fix every flag before you "
                "leave the tollgate — 'we are both accountable' means nobody is."),
    ])
    return wb, "22-stakeholder-and-raci.xlsx"


# ------------------------------------------------------------------- 23 Kano
def kano():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Kano analysis")
    title(ws, "Kano analysis — which requirements actually win anything",
          "Ask each customer two questions per feature: how do you feel if it is present, and how if it is absent.", 6)
    widths(ws, [40, 24, 24, 20, 18, 40])
    ws.merge_cells("A4:F4")
    ws.cell(row=4, column=1, value="Answer scale for both questions: Like · Expect it · Neutral · Live with it · Dislike").font = F_NOTE
    header(ws, 6, ["Feature or requirement", "If present (functional)", "If absent (dysfunctional)",
                   "Category", "Invest?", "What it means for you"])
    scale = "Like,Expect it,Neutral,Live with it,Dislike"
    # Kano's published evaluation table. The previous version had two cells of
    # it transposed: Like/Dislike is One-dimensional (Performance) and
    # Expect-it/Dislike is Must-be — they were returning each other's answer,
    # which inverts the investment advice on the two most common responses.
    CATF = ('=IF(OR(B{r}="",C{r}=""),"",'
            'IF(B{r}="Like",'
            'IF(C{r}="Like","Questionable",IF(C{r}="Dislike","Performance","Delighter")),'
            'IF(B{r}="Dislike",'
            'IF(C{r}="Dislike","Questionable","Reverse"),'
            'IF(C{r}="Dislike","Must-have",'
            'IF(C{r}="Like","Reverse","Indifferent")))))')
    INVF = ('=IF(D{r}="","",IF(D{r}="Must-have","Fix it — no credit for doing it, all the blame for not",'
            'IF(D{r}="Performance","Invest — more is better, and measurable",'
            'IF(D{r}="Delighter","Consider — wins goodwill, costs you nothing if absent",'
            'IF(D{r}="Reverse","Stop doing it — they want the opposite",'
            'IF(D{r}="Questionable","Re-ask — the two answers contradict each other",'
            '"Do not spend here"))))))')
    # One row demonstrated one category and left the other three reading zero,
    # so the chart was a single bar and the summary said nothing.
    EXROWS = [
        ("Resolved without me repeating myself", "Expect it", "Dislike", "Must-have"),
        ("Answered within the time you promised", "Expect it", "Dislike", "Must-have"),
        ("Agent already knows my account history", "Like", "Dislike", "Performance"),
        ("Fewer transfers between teams", "Like", "Dislike", "Performance"),
        ("Proactive notice before I notice the problem", "Like", "Live with it", "Delighter"),
        ("A follow-up note a week later", "Like", "Neutral", "Delighter"),
        ("Choice of chat colour scheme", "Neutral", "Neutral", "Indifferent"),
        ("A survey after every single contact", "Dislike", "Neutral", "Reverse"),
    ]
    for k, (feat, fun, dys, _cat) in enumerate(EXROWS):
        rr = 7 + k
        for i, v in enumerate((feat, fun, dys), start=1):
            c = ws.cell(row=rr, column=i, value=v); c.fill = EX; c.border = THIN
        for col, f in [(4, CATF.format(r=rr)), (5, INVF.format(r=rr))]:
            c = mark(ws, rr, col, "calc"); c.fill = EX; c.value = f
        SHOWN[("Kano analysis", f"D{rr}")] = _cat
        SHOWN[("Kano analysis", f"E{rr}")] = ""
    for col, f, shown in [(4, CATF.format(r=7), "Must-have"),
                          (5, INVF.format(r=7), "Fix it — no credit for doing it, all the blame for not")]:
        c = mark(ws, 7, col, "calc"); c.fill = EX; c.value = f
        SHOWN[("Kano analysis", f"{get_column_letter(col)}7")] = shown
    # "What it means for you" is the column that turns a category into a
    # decision, and it shipped empty on all nineteen rows — the eight worked
    # example rows included, so the reader saw eight demonstrations of the
    # classifier and not one of the conclusion.
    EXWHY = [
        "Every reopen is one of these. This is the project.",
        "Our 14.2% reopen rate breaks this one; fixing it earns no praise, only the absence of complaints.",
        "Worth building: agents currently re-read the case history on every contact.",
        "Two transfers a case today. Directly measurable, and it scales.",
        "Cheap to add on the back of the posting-confirmation fix — send the note when the batch confirms.",
        "Nice, not urgent. Park it until the must-haves hold.",
        "Do not spend a day on this.",
        "Actively harmful — survey fatigue is why our response rate is 17%.",
    ]
    for k, why in enumerate(EXWHY):
        c = mark(ws, 7 + k, 6, "in")
        c.fill = EX
        c.value = why
        c.alignment = Alignment(wrap_text=True, vertical="top")
    note_cell(ws, 26, 1,
              "How to fill the last column. The category is arithmetic; this column is "
              "the decision, and it is the only part of the sheet a sponsor will read. "
              "Write what the category means for THIS operation — name the metric it "
              "moves, or say plainly that it is not worth spending on. A Must-have you "
              "already meet needs one line saying so; a Must-have you fail is your "
              "project. Rows 7 to 14 are worked examples: delete them and write your "
              "own.", 6)
    # Start BELOW the worked example. This loop used to begin at row 8 and paint
    # every row from there as an empty input, which ran after the example was
    # written and repainted seven of its eight rows yellow. The legend says
    # green means "a worked example, delete it when you start" and yellow means
    # "you fill this in", so the sheet was telling the reader to type over rows
    # that were there to be read.
    for r in range(7 + len(EXROWS), 26):
        for cc in [1, 2, 3, 6]:
            mark(ws, r, cc, "in").alignment = Alignment(wrap_text=True, vertical="top")
        for col, f in [(4, CATF.format(r=r)), (5, INVF.format(r=r))]:
            mark(ws, r, col, "calc").value = f
            SHOWN[("Kano analysis", f"{get_column_letter(col)}{r}")] = ""
    dv = DataValidation(type="list", formula1=f'"{scale}"', allow_blank=True)
    ws.add_data_validation(dv); dv.add("B7:C25")
    band(ws, 27, "WHERE YOUR EFFORT SHOULD GO", 6)
    for i, (label, cat, shown) in enumerate([
            ("Must-haves — table stakes", "Must-have", "2"),
            ("Performance — scale with investment", "Performance", "2"),
            ("Delighters — differentiators", "Delighter", "2"),
            ("Indifferent — stop spending here", "Indifferent", "1"),
            ("Reverse — you are actively annoying people", "Reverse", "1"),
            ("Questionable — the answers contradict; re-ask", "Questionable", "0")], start=28):
        ws.cell(row=i, column=1, value=label).font = F_B
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        mark(ws, i, 4, "calc").value = f'=COUNTIF($D$7:$D$25,"{cat}")'
        SHOWN[("Kano analysis", f"D{i}")] = shown
    # The classifier returns six classes; the summary counted four. A feature
    # classified Reverse — you are actively annoying people — was computed and
    # then counted nowhere, and the chart could not reconcile to the table
    # above it. Colour moves with it: CHART-QA says red means a breach, and
    # Reverse is the only breach here, so Must-have takes amber.
    bar(ws, "Where your requirements fall — and which ones are worth money",
        Reference(ws, min_col=1, min_row=28, max_row=33),
        Reference(ws, min_col=4, min_row=28, max_row=33), "G7",
        colours=["B45309", "1F4E79", "3F8F5A", "9AA4B2", "C0392B", "6B4FA0"])
    ws.merge_cells("A35:F35")
    ws.cell(row=35, column=1, value="Speed in support is usually a must-have: being twice as fast wins you nothing "
            "once you are fast enough, while being slow loses you everything. Spending your improvement budget on a "
            "must-have that is already met is the most common way to move a metric and change nothing.").font = F_NOTE
    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Deciding which customer requirements are worth investment, instead of assuming more of everything is better."),
        (True, "How to run it"),
        (False, "For each requirement ask two questions of real customers: how do you feel if it IS present, and how "
                "do you feel if it is NOT. The pair of answers gives you the category automatically."),
        (True, "Reading the categories"),
        (False, "Must-have: absence is punished, presence earns nothing. Performance: satisfaction scales with it. "
                "Delighter: absence is forgiven, presence wins goodwill. Indifferent: stop spending."),
    ])
    return wb, "23-kano-analysis.xlsx"


# --------------------------------------------------------------------- 24 DOE
def doe():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("2^3 design")
    title(ws, "Design of Experiments — 2³ full factorial",
          "Change several things at once in a planned pattern, so you can see interactions one-at-a-time testing can never find.", 9)
    widths(ws, [8, 14, 14, 14, 18, 18, 18, 16, 30])
    band(ws, 4, "YOUR FACTORS — name them and set the two levels you will actually run", 9)
    for i, (f, lo, hi) in enumerate([("Routing rule", "Current", "Skill-based"),
                                     ("Authority limit", "$50", "$250"),
                                     ("Article shown", "No", "Yes")], start=5):
        # chr(63 + i - 4) started at '@': the matrix columns are A, B, C, so a
        # factor labelled '@' pointed at nothing
        ws.cell(row=i, column=1, value=f"Factor {chr(64 + i - 4)}").font = F_B
        mark(ws, i, 2, "in").value = f
        note(ws, i, 2, "Name the factor as the team says it out loud. It must be something you can "
                       "actually set to either level for the whole run — if you cannot control it, "
                       "it is noise to block or stratify, not a factor.")
        note(ws, i, 6, "The level you run today. Keep one arm at current practice so the experiment "
                       "still tells you whether changing anything was worth it.")
        note(ws, i, 8, "The level you are testing. Push it far enough to move the response — a timid "
                       "high level is the most common reason a DOE returns nothing.")
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        ws.cell(row=i, column=5, value="Low (−1)").font = F_NOTE
        mark(ws, i, 6, "in").value = lo
        ws.cell(row=i, column=7, value="High (+1)").font = F_NOTE
        mark(ws, i, 8, "in").value = hi
    # "A B C AB AC BC" with -1 and +1 in the cells tells a reader nothing
    # without scrolling up. The headers now name the factor they belong to and
    # follow whatever the user types in B5:B7.
    header(ws, 9, ["Run", "A", "B", "C", "AB", "AC", "BC", "Response", "Notes"])
    for col, src in ((2, 5), (3, 6), (4, 7)):
        c = ws.cell(row=9, column=col)
        c.value = '=IF($B$%d="","%s","%s: "&$B$%d)' % (
            src, chr(63 + col), chr(63 + col), src)
        SHOWN[("2^3 design", "%s9" % get_column_letter(col))] = {
            2: "A: Routing rule", 3: "B: Authority limit", 4: "C: Article shown"}[col]
    ws.cell(row=8, column=2, value="−1 = the low level you set above · +1 = the high level").font = F_NOTE
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=7)
    h = ws.cell(row=9, column=8)
    h.value = "Response (mean AHT, seconds)"
    ws.cell(row=9, column=9).value = "What you observed on the day"
    design = [(-1, -1, -1), (1, -1, -1), (-1, 1, -1), (1, 1, -1),
              (-1, -1, 1), (1, -1, 1), (-1, 1, 1), (1, 1, 1)]
    resp = [412, 388, 401, 372, 395, 366, 380, 344]
    RUN_NOTES = [
        "Baseline: current routing, $50 authority, no article",
        "Skill-based routing alone",
        "Higher authority alone — fewer transfers to a supervisor",
        "Routing and authority together; the biggest single drop",
        "Article shown alone — small effect on its own",
        "Routing plus article",
        "Authority plus article",
        "All three at the high level. Best result, and the one to confirm.",
    ]
    for i, (a, bq, c) in enumerate(design):
        r = 10 + i
        ws.cell(row=r, column=1, value=i + 1).font = F_B
        n = ws.cell(row=r, column=9, value=RUN_NOTES[i])
        n.font = F_NOTE
        n.alignment = Alignment(wrap_text=True, vertical="center")
        for col, v in zip((2, 3, 4), (a, bq, c)):
            cc = ws.cell(row=r, column=col, value=v); cc.fill = CALC; cc.font = F_CALC; cc.border = THIN
            cc.alignment = Alignment(horizontal="center")
        for col, f in [(5, f"=B{r}*C{r}"), (6, f"=B{r}*D{r}"), (7, f"=C{r}*D{r}")]:
            mark(ws, r, col, "calc").value = f
            SHOWN[("2^3 design", f"{get_column_letter(col)}{r}")] = str(
                {5: a * bq, 6: a * c, 7: bq * c}[col])
        e = mark(ws, r, 8, "in"); e.value = resp[i]
        e.alignment = Alignment(horizontal="center")
        mark(ws, r, 9, "in")
    EFFECT_NOTES = {
        "Effect of A": "Negative means the HIGH level lowers handle time. Biggest single effect "
                       "here: skill-based routing takes about 29 seconds off the average.",
        "Effect of B": "Raising the authority limit from $50 to $250 takes off about 16 seconds — "
                       "fewer calls escalated to a supervisor.",
        "Effect of C": "Showing the article takes off about 22 seconds.",
        "Interaction AB": "Routing and authority together. Small against the main effects, so the "
                          "two can be decided independently.",
        "Interaction AC": "Routing and article together. Also small.",
        "Interaction BC": "Authority and article together. Also small — which is the useful "
                          "finding: none of these has to be rolled out as a package.",
    }
    band(ws, 19, "EFFECTS — how much each factor moves the response", 9)
    eff = [("Effect of A", "B", 2), ("Effect of B", "C", 3), ("Effect of C", "D", 4),
           ("Interaction AB", "E", 5), ("Interaction AC", "F", 6), ("Interaction BC", "G", 7)]
    vals = {}
    for a_, b_, c_ in [(0, 0, 0)]:
        pass
    import statistics
    cols = {2: [d[0] for d in design], 3: [d[1] for d in design], 4: [d[2] for d in design]}
    cols[5] = [d[0] * d[1] for d in design]
    cols[6] = [d[0] * d[2] for d in design]
    cols[7] = [d[1] * d[2] for d in design]
    for label, letter, idx in eff:
        r = 20 + [e[0] for e in eff].index(label)
        ws.cell(row=r, column=1, value=label).font = F_B
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        f = f"=SUMPRODUCT({letter}10:{letter}17,$I$10:$I$17)/4"
        cc = mark(ws, r, 4, "calc"); cc.value = f.replace("$I$", "$H$")
        v = sum(s * y for s, y in zip(cols[idx], resp)) / 4
        SHOWN[("2^3 design", f"D{r}")] = f"{v:.1f}"
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        # The same sentence on all six rows told the reader nothing about their
        # own result. Each row now reads its own number back.
        ws.cell(row=r, column=5, value=EFFECT_NOTES.get(label, "")).font = F_NOTE
    # A..I are merged across the effects block, so the reference column goes
    # to the right of everything the sheet already uses
    for r in range(20, 26):
        c = ws.cell(row=r, column=11)
        c.value = "=-AVERAGE(ABS($D$23),ABS($D$24),ABS($D$25))"
        c.number_format = "#,##0.0"
        c.font = F_NOTE
        SHOWN[("2^3 design", f"K{r}")] = "-2.8"
    ws.cell(row=19, column=11, value="Noise floor").font = F_NOTE
    ws.column_dimensions["K"].width = 11         # visible: charts skip hidden cells
    de = bar(ws, "Effect size — which factor actually moved the response",
             Reference(ws, min_col=1, min_row=20, max_row=25),
             Reference(ws, min_col=4, min_row=20, max_row=25), "M5", horizontal=True,
             name="Effect on the response")
    overlay(de, ws, Reference(ws, min_col=11, min_row=20, max_row=25),
            "Noise floor — the mean interaction, anything smaller is nothing")
    ws.merge_cells("A27:I27")
    ws.cell(row=27, column=1, value="Randomise the run order before you execute, and repeat the whole design if you "
            "can afford it. Eight runs with no replication tells you about size, not about noise.").font = F_NOTE
    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Finding the best combination of settings, and detecting interactions — where two changes together do "
                "something neither does alone. One-factor-at-a-time testing cannot see those at all."),
        (True, "How to use it"),
        (False, "Name three factors and their two levels. Run all eight combinations in randomised order and enter the "
                "response you measured. The effects table then tells you which factor moved it and by how much."),
        (True, "In support"),
        (False, "Factors are things like routing rule, authority limit, whether an article is shown, script version. "
                "The response is handle time, resolution rate or satisfaction."),
        (True, "The trap"),
        (False, "Run the design as designed. Dropping the combinations that look silly destroys the balance the whole "
                "method depends on and the effects become uninterpretable."),
    ])
    return wb, "24-doe-design-matrix.xlsx"


# ------------------------------------------------- 25 pareto and distribution
def pareto():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Pareto")
    title(ws, "Pareto — where the bulk of the problem actually sits",
          "Ranked categories with a running cumulative share, so you can see how much of the problem the top few explain.", 5)
    widths(ws, [40, 12, 11, 9, 34, 3, 7, 40, 12, 13, 11])
    header(ws, 4, ["Category", "Count", "Share", "Rank", "Validated by reading tickets?"])
    data = [("Adjustment not posted at closure", 412), ("Wrong plan applied", 233),
            ("Duplicate charge", 151), ("Proration misunderstood", 96), ("Refund timing", 74)]
    total = sum(d[1] for d in data)

    # A Pareto is sorted by definition. The cumulative used to be a running sum
    # down the rows, which is only a Pareto if the user happens to type their
    # categories in descending order — and if they do not, the 80% line names
    # the wrong vital few, confidently. Type them in any order now: the ranked
    # block to the right sorts itself, and the chart reads that.
    RANKF = ('=IF(B{r}="","",RANK(B{r},$B$5:$B$24,0)+COUNTIF($B$5:B{r},B{r})-1)')
    for i, (name, n) in enumerate(data):
        r = 5 + i
        c = mark(ws, r, 1, "ex" if i == 0 else "in"); c.value = name
        c2 = mark(ws, r, 2, "ex" if i == 0 else "in"); c2.value = n
        sh = mark(ws, r, 3, "calc")
        sh.value = f'=IF(OR(B{r}="",SUM($B$5:$B$24)=0),"",B{r}/SUM($B$5:$B$24))'
        sh.number_format = "0.0%"
        rk = mark(ws, r, 4, "calc"); rk.value = RANKF.format(r=r); rk.number_format = "0"
        if i == 0:
            sh.fill = EX; rk.fill = EX
        SHOWN[("Pareto", f"C{r}")] = f"{n/total*100:.1f}%"
        SHOWN[("Pareto", f"D{r}")] = str(i + 1)
        v = mark(ws, r, 5, "ex" if i == 0 else "in")
        # The one column that decides whether the Pareto is worth anything, and
        # it shipped empty on all twenty rows — a Yes/No dropdown with nothing
        # showing what answering it honestly looks like.
        if i == 0:
            v.value = "Yes"          # the column is a Yes/No dropdown
    for r in range(10, 25):
        for cc in [1, 2, 5]:
            mark(ws, r, cc, "in")
        sh = mark(ws, r, 3, "calc")
        sh.value = f'=IF(OR(B{r}="",SUM($B$5:$B$24)=0),"",B{r}/SUM($B$5:$B$24))'
        sh.number_format = "0.0%"
        rk = mark(ws, r, 4, "calc"); rk.value = RANKF.format(r=r); rk.number_format = "0"
        SHOWN[("Pareto", f"C{r}")] = ""
        SHOWN[("Pareto", f"D{r}")] = ""
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv); dv.add("E5:E24")
    note_cell(ws, 25, 1,
              "How to answer the last column, and why it is the one that matters. "
              "Open 50 tickets at random from your top category and read them. Count "
              "how many are genuinely that category rather than the code an agent "
              "picked under time pressure. Answer Yes only if 40 or more of the 50 "
              "hold up; below that your Pareto is ranking your data-entry habits and "
              "not your problems, and the business case built on it will not survive "
              "contact with the process. Row 5 is the worked example: 50 read, 41 held "
              "up, so Yes.", 5)

    # ---- the ranked block the chart actually plots ----------------------
    ws.cell(row=3, column=7, value="RANKED — this is what the chart plots. "
            "Type above in any order; this sorts itself.").font = F_BAND
    for c in range(7, 12):
        ws.cell(row=3, column=c).fill = BAND
    # not header(), which writes every column in the row and would blank the
    # input block's own headers in A4:E4
    for col, lab in ((7, "#"), (8, "Category"), (9, "Count"),
                     (10, "Cumulative"), (11, "80% line")):
        c = ws.cell(row=4, column=col, value=lab)
        c.fill, c.font = HDR, F_HDR
        c.alignment = Alignment(wrap_text=True, vertical="center")
    run = 0
    for i in range(20):
        r = 5 + i
        ws.cell(row=r, column=7, value=i + 1).font = F_NOTE
        for col, f, fmt in [
            (8, f'=IFERROR(INDEX($A$5:$A$24,MATCH(G{r},$D$5:$D$24,0)),"")', "General"),
            (9, f'=IFERROR(INDEX($B$5:$B$24,MATCH(G{r},$D$5:$D$24,0)),"")', "#,##0"),
            (10, f'=IF(I{r}="",NA(),SUM($I$5:I{r})/SUM($I$5:$I$24))', "0.0%"),
            (11, f'=IF(I{r}="",NA(),0.8)', "0%"),
        ]:
            c = mark(ws, r, col, "calc"); c.value = f; c.number_format = fmt
        if i < len(data):
            run += data[i][1]
            SHOWN[("Pareto", f"H{r}")] = data[i][0]
            SHOWN[("Pareto", f"I{r}")] = f"{data[i][1]:,}"
            SHOWN[("Pareto", f"J{r}")] = f"{run/total*100:.1f}%"
            SHOWN[("Pareto", f"K{r}")] = "80%"
        else:
            for col in "HIJK":
                SHOWN[("Pareto", f"{col}{r}")] = ""
    band(ws, 26, "SUMMARY", 5)
    for i, (label, f, shown) in enumerate([
            ("Total", '=SUM(B5:B24)', f"{total:,}"),
            ("Categories", '=COUNTA(A5:A24)', str(len(data))),
            # the top three by RANK, not the top three rows somebody happened
            # to type first
            ("Share explained by the top three", '=IFERROR(SUM(I5:I7)/SUM(I5:I24),"")',
             f"{sum(d[1] for d in data[:3])/total*100:.1f}%")], start=27):
        ws.cell(row=i, column=1, value=label).font = F_B
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        c = mark(ws, i, 3, "calc"); c.value = f
        if "%" in shown: c.number_format = "0.0%"
        SHOWN[("Pareto", f"C{i}")] = shown
    pc = BarChart(); pc.type = "col"; pc.style = 10
    pc.title = "Pareto — ranked, with cumulative share"
    pc.height, pc.width = 9, 19
    pc.gapWidth = 40
    pc.legend = None
    pc.legend = Legend()
    pc.legend.position = "b"
    pc.add_data(Reference(ws, min_col=9, min_row=5, max_row=14), titles_from_data=False)
    pc.set_categories(Reference(ws, min_col=8, min_row=5, max_row=14))
    for ser in pc.series:
        ser.invertIfNegative = False
        ser.graphicalProperties.solidFill = "1F4E79"
        ser.tx = SeriesLabel(v="Count")
    pc.y_axis.title = "Count"
    line = LineChart()
    line.add_data(Reference(ws, min_col=10, min_row=5, max_row=14), titles_from_data=False)
    line.series[0].tx = SeriesLabel(v="Cumulative share")
    line.add_data(Reference(ws, min_col=11, min_row=5, max_row=14), titles_from_data=False)
    line.series[1].tx = SeriesLabel(v="80% — the vital few are left of here")
    line.series[1].graphicalProperties.line.dashStyle = "dash"
    line.series[1].graphicalProperties.line.solidFill = "C0392B"
    line.series[0].graphicalProperties.line.solidFill = "B45309"
    line.series[0].graphicalProperties.line.width = 24000
    line.y_axis.delete = False
    line.y_axis.axId = 200
    line.y_axis.numFmt = "0%"
    line.y_axis.title = "Cumulative share"
    line.y_axis.crosses = "max"
    pc += line
    pc.height, pc.width = 11, 26
    ws.add_chart(pc, "N3")
    ws.merge_cells("A31:E31")
    ws.cell(row=31, column=1, value="A Pareto inherits the quality of your categories. Disposition codes are chosen "
            "by agents under time pressure, so a Pareto of them ranks your data-entry habits as much as your problems. "
            "Read 50 tickets in the top bar before you size a business case on it — that is what the last column is for.").font = F_NOTE

    ws2 = wb.create_sheet("Shape and spread")
    ws2.sheet_view.showGridLines = False
    title(ws2, "Is this data the shape you assumed?",
          "Normality assessment for support data. Support durations are right-skewed: "
          "the mean flatters you, the median and p90 do not. Read the skew before you "
          "pick a test — a two-sample t-test on data this shaped will mislead you, and "
          "the non-parametric equivalent will not.", 4)
    widths(ws2, [34, 18, 18, 60])
    ws2.cell(row=4, column=1, value="Paste your values into column B").font = F_NOTE
    header(ws2, 5, ["Statistic", "Value", "", "What it tells you"])
    stats = [("Count", '=COUNT($B$20:$B$1000)', "0"),
             ("Mean", '=IFERROR(AVERAGE($B$20:$B$1000),"")', ""),
             ("Median (p50)", '=IFERROR(MEDIAN($B$20:$B$1000),"")', ""),
             ("p90", '=IFERROR(PERCENTILE($B$20:$B$1000,0.9),"")', ""),
             ("Standard deviation", '=IFERROR(STDEV($B$20:$B$1000),"")', ""),
             ("Interquartile range", '=IFERROR(QUARTILE($B$20:$B$1000,3)-QUARTILE($B$20:$B$1000,1),"")', ""),
             ("Mean ÷ median", '=IFERROR(AVERAGE($B$20:$B$1000)/MEDIAN($B$20:$B$1000),"")', "")]
    notes = ["How many values you pasted.",
             "Dragged upward by the long tail. Rarely the number to quote.",
             "The typical experience. Quote this.",
             "What your worst-served customers get. This is what generates complaints.",
             "Assumes symmetry, so it describes skewed data poorly.",
             "The middle half. A better spread measure for this data.",
             "Above about 1.2 means meaningful right skew — use non-parametric tests and quote percentiles."]
    for i, ((label, f, shown), note) in enumerate(zip(stats, notes), start=6):
        ws2.cell(row=i, column=1, value=label).font = F_B
        c = mark(ws2, i, 2, "calc"); c.value = f; c.number_format = "#,##0.00"
        SHOWN[("Shape and spread", f"B{i}")] = shown
        n = ws2.cell(row=i, column=4, value=note); n.font = F_NOTE
        n.alignment = Alignment(wrap_text=True, vertical="top")
    band(ws2, 14, "VERDICT", 4)
    c = mark(ws2, 14, 1, "calc")
    ws2.merge_cells("A15:D15")
    v = ws2.cell(row=15, column=1)
    v.value = ('=IF(COUNT($B$20:$B$1000)=0,"Paste your data into column B below.",'
               'IF(AVERAGE($B$20:$B$1000)/MEDIAN($B$20:$B$1000)>1.2,'
               '"RIGHT-SKEWED — quote the median and p90, and prefer non-parametric tests.",'
               '"ROUGHLY SYMMETRIC — the mean and a parametric test are defensible."))')
    v.fill, v.font = CALC, F_CALC
    SHOWN[("Shape and spread", "A15")] = "Paste your data into column B below."
    ws2.cell(row=18, column=1, value="Your data").font = F_B
    ws2.cell(row=19, column=2, value="Values ↓").font = F_NOTE
    # This tab shipped completely empty: every statistic read blank and the
    # verdict said "paste your data", so the one sheet whose whole purpose is
    # to show you what skewed data looks like showed you nothing at all.
    # 40 handle times with the right tail support data actually has.
    HANDLE = [182, 201, 214, 227, 236, 241, 248, 255, 259, 263,
              268, 272, 277, 281, 286, 292, 298, 305, 311, 318,
              326, 335, 346, 358, 372, 389, 408, 431, 459, 494,
              538, 594, 668, 771, 918, 1140, 1502, 2160, 3480, 5220]
    for i, v in enumerate(HANDLE):
        # All forty are the worked example: they are one distribution, and the
        # whole point of the tab is to show you what a skewed one looks like.
        # Splitting them green-then-yellow would say the first value is a
        # demonstration and the other thirty-nine are yours to invent.
        c = mark(ws2, 20 + i, 2, "ex")
        c.value = v
        c.number_format = "#,##0"
    for r in range(20 + len(HANDLE), 60):
        mark(ws2, r, 2, "in")

    # The lesson made visible: mean, median and p90 side by side. On this data
    # the mean is 43% above the median, which is the whole argument for
    # quoting percentiles instead of averages.
    import statistics as _st
    srt = sorted(HANDLE)
    _mean = sum(HANDLE) / len(HANDLE)
    _med = _st.median(HANDLE)
    _p90 = srt[int(round(0.9 * (len(srt) - 1)))]
    ws2.cell(row=18, column=6, value="What you would quote").font = F_B
    for k, (lab, f, shown) in enumerate([
            ("Mean", '=IFERROR(AVERAGE($B$20:$B$1000),"")', _mean),
            ("Median (p50)", '=IFERROR(MEDIAN($B$20:$B$1000),"")', _med),
            ("p90", '=IFERROR(PERCENTILE($B$20:$B$1000,0.9),"")', _p90)]):
        ws2.cell(row=19 + k, column=6, value=lab).font = F_NOTE
        c = mark(ws2, 19 + k, 7, "calc")
        c.value = f
        c.number_format = "#,##0"
        SHOWN[("Shape and spread", f"G{19 + k}")] = f"{shown:,.0f}"
    bar(ws2, "Mean, median and p90 — quoting the mean hides the tail",
        Reference(ws2, min_col=6, min_row=19, max_row=21),
        Reference(ws2, min_col=7, min_row=19, max_row=21), "F23",
        colours=["9AA4B2", "1F4E79", "C0392B"], name="Seconds")
    ws2.column_dimensions["F"].width = 16
    ws2.column_dimensions["G"].width = 11
    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Two things you should do before any analysis: find out where the bulk of the problem is, and find out "
                "what shape your data actually has."),
        (True, "Pareto tab"),
        (False, "Enter categories and counts. Share and cumulative calculate themselves. The last column is the one "
                "people skip — validate the top category by reading tickets before you build a case on it."),
        (True, "Shape and spread tab"),
        (False, "Paste raw values into column B. The verdict tells you whether your data is skewed enough that the "
                "mean and a t-test would mislead you. Support durations almost always are."),
    ])
    return wb, "25-pareto-and-distribution.xlsx"


# ------------------------------------------------------------ 26 flow and WIP
def flow():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("WIP and lead time")
    title(ws, "Kanban sizing — pick a wait, get the WIP limit",
          "Little's Law is arithmetic, not a theory. If the closing rate is fixed and you want shorter waits, you must cap the work in progress.", 4)
    widths(ws, [40, 18, 18, 64])
    band(ws, 4, "YOUR NUMBERS", 4)
    inputs = [("Open items right now", 4200, "Live count from your ticketing system, same time of day each time."),
              ("Items closed per day", 1100, "Average daily closing rate over a stable period."),
              ("Lead time you want to promise (days)", 2, "What you want customers to wait."),
              ("Agents working this queue", 40, "Used to turn the queue-level cap into a per-agent limit.")]
    for i, (label, v, note) in enumerate(inputs, start=5):
        ws.cell(row=i, column=1, value=label).font = F_B
        mark(ws, i, 2, "in").value = v
        n = ws.cell(row=i, column=4, value=note); n.font = F_NOTE
        n.alignment = Alignment(wrap_text=True, vertical="top")
    band(ws, 10, "WHAT THAT MEANS", 4)
    outs = [("Lead time today (days)", '=IFERROR(B5/B6,"")', "3.82",
             "Backlog ÷ closing rate. Nothing else affects it."),
            ("Lead time today (hours)", '=IFERROR(B5/B6*24,"")', "91.6",
             "The same wait in hours, which is usually the unit your promise and your SLA are written in."),
            ("Backlog cap to hit your target", '=IFERROR(B6*B7,"")', "2,200",
             "This is your queue-level WIP limit."),
            ("Backlog you must remove first", '=IFERROR(MAX(0,B5-B6*B7),"")', "2,000",
             "Until this is cleared the cap cannot hold."),
            ("WIP limit per agent", '=IFERROR(ROUND(B6*B7/B8,0),"")', "55",
             "Enforce this in the tool, not in a briefing. A limit nobody can exceed is a control; a limit in a "
             "document is a suggestion."),
            ("Days to clear the excess at current rate", '=IFERROR(MAX(0,B5-B6*B7)/B6,"")', "1.82",
             "Assumes arrivals stop, which they will not — plan a temporary lift in closing rate instead.")]
    for i, (label, f, shown, note) in enumerate(outs, start=11):
        ws.cell(row=i, column=1, value=label).font = F_B
        c = mark(ws, i, 2, "calc"); c.value = f; c.number_format = "#,##0.00"
        SHOWN[("WIP and lead time", f"B{i}")] = shown
        if note:
            n = ws.cell(row=i, column=4, value=note); n.font = F_NOTE
            n.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = 28
    ws.cell(row=21, column=1, value="Lead time today").font = F_B
    c = mark(ws, 21, 2, "calc"); c.value = "=B11"; c.number_format = "#,##0.00"
    ws.cell(row=22, column=1, value="Lead time promised").font = F_B
    c = mark(ws, 22, 2, "calc"); c.value = "=B7"; c.number_format = "#,##0.00"
    # The chart reads these two rows. Left bare they were a label and a number
    # with nothing to say what the reader is looking at.
    for r, txt in ((21, "What customers actually experience now — the same figure as above, repeated "
                        "here so the chart can draw it against the promise."),
                   (22, "The wait you want to commit to. The gap between the two bars is backlog, not "
                        "effort: closing it means capping the work, not pushing the team harder.")):
        n = ws.cell(row=r, column=4, value=txt); n.font = F_NOTE
        n.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
    SHOWN[("WIP and lead time", "B21")] = "3.82"
    SHOWN[("WIP and lead time", "B22")] = "2.00"
    bar(ws, "Where you are against what you promised (days)",
        Reference(ws, min_col=1, min_row=21, max_row=22),
        Reference(ws, min_col=2, min_row=21, max_row=22), "F5",
        colours=["C0392B", "3F8F5A"])
    ws.merge_cells("A19:D19")
    ws.cell(row=19, column=1, value="There is no third option. If you cannot cap the work and cannot raise the "
            "closing rate, the wait will not fall — and saying so early is more useful than promising otherwise.").font = F_NOTE
    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Turning a lead-time promise into a work-in-progress limit you can actually enforce."),
        (True, "How to use it"),
        (False, "Enter your open count, your closing rate and the wait you want to promise. The cap that follows is "
                "arithmetic — Little's Law — not an opinion."),
        (True, "The part people skip"),
        (False, "You usually have to clear the excess backlog before the cap can hold. Plan that as a separate push "
                "with a start and an end, or the limit gets abandoned in week one."),
    ])
    return wb, "26-kanban-and-wip.xlsx"



# --------------------------------------------------------------------------
# Control charts. The program's chart-selection table names seven chart types
# and, until now, nothing in the repo produced one. This builds all seven off
# your own numbers, with the limits as live formulas so you can see the maths.
# --------------------------------------------------------------------------

def bounds(*seqs, pad=0.10):
    """Axis min/max that frame the data instead of crushing it against the top.

    Left to auto-scale, a control chart of values around 415 with limits at 363
    and 470 renders on a 0-500 axis in most viewers: every point sits in the top
    fifth and the variation you are supposed to be reading is invisible.
    """
    vals = [float(v) for seq in seqs for v in seq
            if isinstance(v, (int, float)) and v == v]
    if not vals:
        return None, None
    lo, hi = min(vals), max(vals)
    span = (hi - lo) or (abs(hi) or 1.0)
    lo -= span * pad
    hi += span * pad
    if min(vals) >= 0 and lo < 0:
        lo = 0.0                       # counts and durations never go negative
    return lo, hi


MARK_COL = 30            # markers live well clear of the plotted area


def markers(ws, sheet, first, last, col, header, formula):
    """A markers-only column: the value where a rule fires, NA() everywhere else.

    NA() is the only thing that reliably leaves a gap. A formula returning ""
    plots as zero, which on any chart whose axis reaches the floor draws a row
    of markers along the bottom.
    """
    ws.cell(row=first - 1, column=col, value=header).font = F_NOTE
    # White text, not hidden. A chart will not plot a hidden cell, so the
    # column has to stay visible — but a reader should not be looking at
    # twenty rows of #N/A, which is what NA() leaves behind by design.
    invisible = Font(color="FFFFFFFF", size=9)
    for r in range(first, last + 1):
        c = ws.cell(row=r, column=col)
        c.value = formula(r)
        c.font = invisible
        SHOWN[(sheet, "%s%d" % (get_column_letter(col), r))] = ""
    ws.column_dimensions[get_column_letter(col)].width = 4
    return Reference(ws, min_col=col, min_row=first, max_row=last)


def spc_chart(ws, title_, cats, series_defs, anchor, y_title="", pct=False,
              height=8, width=19, ylim=None, flags=()):
    """A control chart: the data as marked points, the limits as flat lines."""
    ch = LineChart()
    ch.title = title_
    ch.style = 2
    ch.height, ch.width = height, width
    if y_title:
        ch.y_axis.title = y_title      # setting this to "" writes the text "None"
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    # Deliberately NOT setting explicit bounds any more. OOXML axis bounds are
    # static values — they cannot follow a formula — so framing the axis on the
    # shipped example guarantees the chart is wrong for anyone whose data has a
    # different magnitude. Tripling the inputs put 95 of 97 points outside the
    # frame, on a tab whose How-to says nothing else needs touching.
    #
    # The crushed-axis problem this was solving is real but smaller: it comes
    # from a renderer choosing a zero-anchored range for data that sits well
    # above zero. Every one of these charts plots its control limits as series,
    # so auto-scaling brackets the limits and the data together and lands in
    # very nearly the same place — and, unlike a constant, it stays right when
    # the user pastes their own numbers.
    _ = ylim
    if pct:
        ch.y_axis.numFmt = "0.0%"
    for ref, name, colour in flags:
        # markers only: no line at all, so a fired rule reads as a highlight on
        # the point rather than a second series wandering across the plot
        fl = Series(ref, title=name)
        fl.graphicalProperties.line.noFill = True
        fl.marker = Marker(symbol="circle", size=11)
        fl.marker.graphicalProperties.solidFill = colour
        fl.marker.graphicalProperties.line.solidFill = colour
        fl.smooth = False
        ch.series.append(fl)
    for ref, name, colour, dashed, marker in series_defs:
        sr = Series(ref, title=name)
        lp = sr.graphicalProperties.line
        lp.solidFill = colour
        lp.width = 28000 if not dashed else 16000
        if dashed:
            lp.dashStyle = "dash"
        if marker:
            sr.marker = Marker(symbol="circle", size=6)
            sr.marker.graphicalProperties.solidFill = colour
        else:
            sr.marker = Marker(symbol="none")
        sr.smooth = False
        ch.series.append(sr)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)
    return ch


N_PTS = 24                       # 24 periods is enough to set honest limits
BASE_N = 12                      # baseline window for EWMA/CUSUM: the "before" period


def _mrbar(seq):
    """Average moving range — the short-term sigma estimator every chart here uses."""
    return sum(abs(seq[i] - seq[i - 1]) for i in range(1, len(seq))) / (len(seq) - 1)


def _laney(ns, ks, poisson=False):
    """p-bar (or u-bar), sigma-z and the per-point sigma, exactly as the sheet computes them."""
    pbar = sum(ks) / float(sum(ns))
    sig = [((pbar / n) ** 0.5) if poisson else ((pbar * (1 - pbar) / n) ** 0.5) for n in ns]
    z = [((ks[i] / float(ns[i])) - pbar) / sig[i] for i in range(len(ns))]
    return pbar, sig, z, max(1.0, _mrbar(z) / 1.128)


BASE_PTS = 20        # whole weeks, so day-of-week mix is balanced


def baseline_input(ws, sheet, width, first, last):
    """The window the limits are estimated from — frozen, not the whole series.

    Recomputing limits over every point you have is the single most common way
    SPC is quietly defeated: a process that drifts drags its own limits along
    with it and never signals. The picker tab says so in bold, and until now
    all five estimated-limit charts did exactly what it warns against.

    Defaults to 20 points rather than the full count. Defaulting to the full
    count would preserve the defect silently for everyone who never finds
    this cell.
    """
    # A point index the baseline formulas can filter on. INDEX($B$14:$B$37,$B$3)
    # is the idiomatic Excel way to say "the first N of this range", but it is a
    # range-endpoint construction that several tools cannot parse, including the
    # engine this repo verifies with. SUMIF against an index column is plain,
    # portable and blank-safe.
    ws.cell(row=first - 1, column=20, value="pt").font = F_NOTE
    for i, r in enumerate(range(first, last + 1)):
        c = ws.cell(row=r, column=20, value=i + 1)
        c.font = Font(color="FFFFFFFF", size=9)
        SHOWN[(sheet, "T%d" % r)] = ""
    ws.column_dimensions["T"].width = 3
    ws.cell(row=3, column=1, value="Limits frozen from the first").font = F_B
    c = mark(ws, 3, 2, "in")
    c.value = BASE_PTS
    c.number_format = "0"
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=width)
    ws.cell(row=3, column=3, value="points — the baseline window. Re-baseline when you have "
            "deliberately changed the process, never because the chart is signalling.").font = F_NOTE
    # rows 1-2 are the merged title block and row 3 columns 3..width carry the
    # merged note, so the warning sits just past the note's right edge
    w = ws.cell(row=3, column=width + 1)
    w.value = ('=IF(COUNT($B$%d:$B$%d)=0,"",IF($B$3>COUNT($B$%d:$B$%d),'
               '"only "&COUNT($B$%d:$B$%d)&" points — limits use those",""))'
               % (first, last, first, last, first, last))
    w.font = Font(bold=True, size=9, color="FFC0392B")
    SHOWN[(sheet, "%s3" % get_column_letter(width + 1))] = ""
    SHOWN[(sheet, "B3")] = str(BASE_PTS)
    return "$B$3"


def _spc_stats(ws, rows, width):
    """The stats band every control-chart tab opens with."""
    band(ws, 4, "The maths, calculated from your data — nothing here is hard-coded", width)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 32)
    r = 5
    for label, formula, fmt, note in rows:
        ws.cell(row=r, column=1, value=label).font = F_B
        c = mark(ws, r, 2, "calc")
        c.value = formula
        c.number_format = fmt
        if note:
            # one narrow column cannot hold a sentence; span the rest of the row
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=width)
            n = ws.cell(row=r, column=3, value=note)
            n.font = F_NOTE
            n.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[r].height = 15 * (1 + len(note) // 95)
        r += 1
    return r


# The Xbar-R worked example, which used to be generated by
# `round(base + 22*sin(i*1.7 + j*2.1))` over j = 0..4. That is not five
# observations: j*2.1 reaches 6.3 at j=3, and 2*pi is 6.283, so observation 4
# repeated observation 1 and observation 5 repeated observation 2 in almost
# every subgroup — [437, 437, 404, 437, 437]. The R chart exists to measure
# within-subgroup spread and it was being fed three distinct numbers padded out
# to five, which understates every range on the sheet.
#
# The sine also put THREE of the twenty baseline subgroups outside the limits
# (points 5, 11 and 20), so every limit on the sheet was computed from a process
# the sheet itself showed was not in control. That is the error the method most
# needs a reader to avoid, shipped as the worked example.
#
# Regenerated: 20 baseline subgroups drawn from a stable process, each with five
# distinct observations, every mean inside 429.69 / 386.77 and every range inside
# the R limit of 78.64 — then four subgroups at a shifted level, all four above
# the upper limit with the range chart still in control. X-double-bar 408.23,
# R-bar 37.20. Held as data rather than a formula so the build stays
# byte-reproducible.
XBAR_R_DATA = [
    [446, 410, 407, 399, 400], [428, 399, 370, 401, 432], [414, 421, 423, 416, 391],
    [442, 441, 393, 410, 429], [407, 409, 391, 419, 403], [415, 429, 421, 407, 389],
    [415, 418, 421, 428, 426], [414, 385, 404, 389, 416], [400, 404, 406, 392, 434],
    [419, 399, 433, 405, 376], [394, 399, 398, 380, 414], [397, 401, 424, 403, 429],
    [392, 379, 406, 404, 396], [408, 392, 420, 390, 394], [415, 383, 411, 403, 407],
    [399, 415, 400, 452, 406], [425, 419, 418, 407, 411], [397, 383, 403, 413, 401],
    [420, 414, 391, 417, 383], [424, 416, 402, 386, 436],
    [459, 437, 461, 443, 433], [448, 463, 458, 417, 428],
    [459, 448, 447, 429, 442], [433, 429, 452, 425, 459],
]


def control_charts():
    """27 — every chart type the selection table sends you to.

    The worked data belongs HERE, not in chart_specs.py. That file's specs
    build the charts they feed, so add_charts() clears every chart on a
    workbook before calling its spec; registering this workbook there would
    have deleted all ten of its native charts on the next build. It is also the
    file whose hash stamps every workbook it owns, so even a comment added to
    it rewrites eight unrelated binaries — which is what the byte-reproducible
    build exists to prevent.
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ---- I-MR -----------------------------------------------------------
    ws = wb.create_sheet("I-MR")
    title(ws, "Individuals and Moving Range (I-MR)",
          "One number per day or week — daily AHT, daily backlog, weekly SLA%. "
          "Not for raw per-contact durations: those are heavily skewed and the limits will lie to you.", 9)
    widths(ws, [16, 13, 14, 11, 11, 11, 11, 11, 20])
    baseline_input(ws, "I-MR", 9, 14, 37)
    _spc_stats(ws, [
        ("Points with data", "=COUNT(B14:B37)", "0", "Total points plotted. The limits use only the baseline window above."),
        ("Centre line (mean)", '=IFERROR(SUMIF($T$14:$T$37,"<="&$B$3,$B$14:$B$37)/COUNTIFS($T$14:$T$37,"<="&$B$3,$B$14:$B$37,">"&-9.9E+307),"")', "#,##0.00",
         "The process average over the BASELINE window, not the whole series."),
        ("Average moving range", '=IFERROR(SUMIF($T$15:$T$37,"<="&$B$3,$C$15:$C$37)/COUNTIFS($T$15:$T$37,"<="&$B$3,$C$15:$C$37,">"&-9.9E+307),"")', "#,##0.00", "Mean gap between consecutive points — this is your short-term variation."),
        ("Sigma estimate", "=B7/1.128", "#,##0.00", "MR-bar / 1.128. Uses only point-to-point movement, so a slow drift does not widen the limits."),
        ("Upper control limit", "=B6+2.66*B7", "#,##0.00", "2.66 = 3 / 1.128. Same thing as mean + 3 sigma."),
        ("Lower control limit", "=B6-2.66*B7", "#,##0.00", "If this goes below zero on a count, treat the lower limit as zero."),
        ("Moving-range UCL", "=3.267*B7", "#,##0.00", "A point above this means the process jumped between two periods."),
    ], 9)
    band(ws, 12, "Your data — one row per period, oldest first", 9)
    header(ws, 13, ["Period", "Value", "Moving range", "CL", "UCL", "LCL", "MR CL", "MR UCL", "Signal"])
    # Points 1-20 are the baseline the limits are frozen from, so they stay
    # in control; 21-24 are the monitoring window and step up past the UCL of
    # 466.99 after a release. The moving range never signals, because the LEVEL
    # moved and the spread did not — which is the whole reason there are two
    # charts here rather than one. A worked example that never crosses a limit
    # teaches the shape of the chart and nothing whatever about reading it.
    vals = [408, 415, 402, 431, 419, 396, 424, 410, 438, 405, 417, 429,
            401, 422, 413, 407, 435, 398, 420, 411, 449, 471, 478, 466]
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=1, value="Day %d" % (i + 1))
        c = mark(ws, r, 2, "in" if i else "ex")
        c.value = vals[i]
        c.number_format = "#,##0.00"
        if i:
            f = mark(ws, r, 3, "calc")
            f.value = "=IF(COUNT(B%d:B%d)=2,ABS(B%d-B%d),\"\")" % (r - 1, r, r, r - 1)
            f.number_format = "#,##0.00"
            SHOWN[("I-MR", "C%d" % r)] = "%.2f" % abs(vals[i] - vals[i - 1])
        for col, ref in ((4, "$B$6"), (5, "$B$9"), (6, "$B$10"), (7, "$B$7"), (8, "$B$11")):
            g = mark(ws, r, col, "calc")
            g.value = "=IF(B%d=\"\",\"\",%s)" % (r, ref)
            g.number_format = "#,##0.00"
        sg = mark(ws, r, 9, "calc")
        if i >= 7:
            # the run rule needs eight rows of history; before row 21 the window
            # would reach up into the banded header, which is merged and empty
            sg.value = ("=IF(B{r}=\"\",\"\",IF(OR(B{r}>$B$9,B{r}<$B$10),\"OUT OF CONTROL\","
                        "IF(AND(COUNT(B{p8}:B{r})=8,OR(MIN(B{p8}:B{r})>$B$6,MAX(B{p8}:B{r})<$B$6)),"
                        "\"8 in a row one side — shift\",\"\")))").format(r=r, p8=r - 7)
        else:
            sg.value = ("=IF(B{r}=\"\",\"\",IF(OR(B{r}>$B$9,B{r}<$B$10),"
                        "\"OUT OF CONTROL\",\"\"))").format(r=r)
    base = vals[:BASE_PTS]                 # the window the sheet actually uses
    mu = sum(base) / float(len(base))
    mrb = _mrbar(base)
    SHOWN.update({("I-MR", "B5"): "%d" % N_PTS, ("I-MR", "B6"): "%.2f" % mu,
                  ("I-MR", "B7"): "%.2f" % mrb, ("I-MR", "B8"): "%.2f" % (mrb / 1.128),
                  ("I-MR", "B9"): "%.2f" % (mu + 2.66 * mrb), ("I-MR", "B10"): "%.2f" % (mu - 2.66 * mrb),
                  ("I-MR", "B11"): "%.2f" % (3.267 * mrb)})
    for k, v in (("D", mu), ("E", mu + 2.66 * mrb), ("F", mu - 2.66 * mrb), ("G", mrb), ("H", 3.267 * mrb)):
        for i in range(N_PTS):
            SHOWN[("I-MR", "%s%d" % (k, 14 + i))] = "%.2f" % v
    # The two rules a control chart is actually read by. Deliberately only two:
    # the full Nelson set across ten charts drives the family-wise false-alarm
    # rate to about one signal per chart per fortnight on a perfectly stable
    # process, which is the tampering this workbook warns against.
    out_i = markers(ws, "I-MR", 14, 37, MARK_COL, "out of control",
                    lambda r: f'=IF($B{r}="",NA(),IF(OR($B{r}>$B$9,$B{r}<$B$10),$B{r},NA()))')
    run_i = markers(ws, "I-MR", 14, 37, MARK_COL + 1, "8 in a row one side",
                    lambda r: ('=NA()' if r < 21 else
                               f'=IF(COUNT($B{r-7}:$B{r})<8,NA(),IF(OR(MIN($B{r-7}:$B{r})>$B$6,'
                               f'MAX($B{r-7}:$B{r})<$B$6),$B{r},NA()))'))
    out_mr = markers(ws, "I-MR", 14, 37, MARK_COL + 2, "MR out of control",
                     lambda r: f'=IF($C{r}="",NA(),IF($C{r}>$B$11,$C{r},NA()))')
    cats = Reference(ws, min_col=1, min_row=14, max_row=37)
    spc_chart(ws, "Individuals — is the process stable?", cats, [
        (Reference(ws, min_col=2, min_row=14, max_row=37), "Value", "1F4E79", False, True),
        (Reference(ws, min_col=4, min_row=14, max_row=37), "Centre line", "3F8F5A", False, False),
        (Reference(ws, min_col=5, min_row=14, max_row=37), "UCL", "C0392B", True, False),
        (Reference(ws, min_col=6, min_row=14, max_row=37), "LCL", "C0392B", True, False),
    ], "K4", ylim=bounds(vals, [mu - 2.66 * mrb, mu + 2.66 * mrb]),
        flags=[(out_i, "Out of control", "C0392B"),
               (run_i, "8 in a row one side — shift", "B45309")])
    spc_chart(ws, "Moving range — did it jump between periods?", cats, [
        (Reference(ws, min_col=3, min_row=14, max_row=37), "Moving range", "6B4FA0", False, True),
        (Reference(ws, min_col=7, min_row=14, max_row=37), "MR centre line", "3F8F5A", False, False),
        (Reference(ws, min_col=8, min_row=14, max_row=37), "MR UCL", "C0392B", True, False),
    ], "K22", height=7,
        ylim=bounds([abs(vals[i] - vals[i - 1]) for i in range(1, len(vals))], [0, 3.267 * mrb]),
        flags=[(out_mr, "Out of control", "C0392B")])

    # ---- Laney p' -------------------------------------------------------
    ws = wb.create_sheet("Laney p-prime")
    title(ws, "Laney p\u2032 chart — the default for support percentages",
          "FCR%, SLA met%, QA pass%, reopen%. At thousands of contacts a day an ordinary p-chart's limits are "
          "far too tight and almost every point looks out of control — until the team stops looking at the chart.", 9)
    widths(ws, [14, 14, 13, 11, 12, 10, 11, 11, 22])
    baseline_input(ws, "Laney p-prime", 9, 14, 37)
    _spc_stats(ws, [
        ("Overall proportion (p-bar)",
         '=SUMIF($T$14:$T$37,"<="&$B$3,$C$14:$C$37)/SUMIF($T$14:$T$37,"<="&$B$3,$B$14:$B$37)', "0.00%",
         "Total defectives / total opportunities — NOT the average of the daily percentages."),
        ("Average moving range of z", '=IFERROR(SUMIF($T$15:$T$37,"<="&$B$3,$G$15:$G$37)/COUNTIFS($T$15:$T$37,"<="&$B$3,$G$15:$G$37,">"&-9.9E+307),"")', "#,##0.000",
         "How much the standardised points move period to period, over the baseline window."),
        ("Sigma z (Laney adjustment)", "=MAX(1,B6/1.128)", "#,##0.000",
         "This is the whole trick. Sigma z = 1 means no overdispersion and this collapses to an ordinary p-chart. "
         "Above about 1.2 you had a real problem."),
        ("Ordinary p-chart limits would be", "=\"\u00b1\"&TEXT(3*SQRT(B5*(1-B5)/AVERAGE(B14:B37)),\"0.00%\")",
         "@", "Compare with the Laney limits in the chart. This is why the old chart cried wolf."),
    ], 9)
    band(ws, 12, "Your data — one row per day", 9)
    header(ws, 13, ["Day", "Contacts (n)", "Met / passed", "Proportion", "Sigma p", "z", "MR of z", "", "Signal"])
    ws.cell(row=13, column=8, value="").fill = HDR
    ns = [7820, 8140, 7960, 8310, 7690, 8020, 8450, 7880, 8210, 7740, 8090, 8380,
          7930, 8260, 7810, 8150, 8470, 7860, 8030, 8340, 7900, 8180, 8120, 7950]
    # Days 23-24 fall through the lower limit (~0.651 at these volumes): the
    # service level collapses while contact volume holds, which is what a p'
    # chart is for and what the blank Signal column never showed.
    ks = [5610, 5990, 5570, 6120, 5410, 5880, 6210, 5490, 6050, 5380, 5940, 6180,
          5520, 6010, 5480, 5860, 6240, 5600, 5760, 6090, 5710, 5900, 5180, 5060]
    # The centre is baselined; the per-point sigma and z still need every
    # plotted point, or the preview runs out of values before the rows do.
    pbar, _, _, sz = _laney(ns[:BASE_PTS], ks[:BASE_PTS])
    sigs = [((pbar * (1 - pbar) / n) ** 0.5) for n in ns]
    zs = [((ks[i] / float(ns[i])) - pbar) / sigs[i] for i in range(len(ns))]
    SHOWN.update({("Laney p-prime", "B5"): "%.2f%%" % (100 * pbar),
                  ("Laney p-prime", "B6"): "%.3f" % _mrbar(zs),
                  ("Laney p-prime", "B7"): "%.3f" % sz,
                  ("Laney p-prime", "B8"): "\u00b1%.2f%%" % (
                      100 * 3 * (pbar * (1 - pbar) / (sum(ns) / float(len(ns)))) ** 0.5)})
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=1, value="Day %d" % (i + 1))
        for col, v in ((2, ns[i]), (3, ks[i])):
            c = mark(ws, r, col, "in" if i else "ex")
            c.value = v
            c.number_format = "#,##0"
        p = mark(ws, r, 4, "calc"); p.value = "=IF(B%d=\"\",\"\",C%d/B%d)" % (r, r, r); p.number_format = "0.00%"
        sp = mark(ws, r, 5, "calc"); sp.value = "=IF(B%d=\"\",\"\",SQRT($B$5*(1-$B$5)/B%d))" % (r, r); sp.number_format = "0.0000"
        z = mark(ws, r, 6, "calc"); z.value = "=IF(B%d=\"\",\"\",(D%d-$B$5)/E%d)" % (r, r, r); z.number_format = "#,##0.00"
        if i:
            m = mark(ws, r, 7, "calc"); m.value = "=IF(COUNT(F%d:F%d)=2,ABS(F%d-F%d),\"\")" % (r - 1, r, r, r - 1); m.number_format = "#,##0.00"
        sg = mark(ws, r, 9, "calc")
        sg.value = "=IF(D{r}=\"\",\"\",IF(OR(D{r}>K{r},D{r}<L{r}),\"OUT OF CONTROL\",\"\"))".format(r=r)
        SHOWN[("Laney p-prime", "D%d" % r)] = "%.2f%%" % (100.0 * ks[i] / ns[i])
        SHOWN[("Laney p-prime", "E%d" % r)] = "%.4f" % sigs[i]
        SHOWN[("Laney p-prime", "F%d" % r)] = "%.2f" % zs[i]
    # limit columns live off to the right so the chart has something flat to plot
    ws.cell(row=13, column=11, value="UCL").fill = HDR; ws.cell(row=13, column=11).font = F_HDR
    ws.cell(row=13, column=12, value="LCL").fill = HDR; ws.cell(row=13, column=12).font = F_HDR
    ws.cell(row=13, column=13, value="CL").fill = HDR; ws.cell(row=13, column=13).font = F_HDR
    for i in range(N_PTS):
        r = 14 + i
        u = ws.cell(row=r, column=11, value="=IF(B%d=\"\",\"\",MIN(1,$B$5+3*E%d*$B$7))" % (r, r)); u.number_format = "0.00%"
        l = ws.cell(row=r, column=12, value="=IF(B%d=\"\",\"\",MAX(0,$B$5-3*E%d*$B$7))" % (r, r)); l.number_format = "0.00%"
        c = ws.cell(row=r, column=13, value="=IF(B%d=\"\",\"\",$B$5)" % r); c.number_format = "0.00%"
    out_p = markers(ws, "Laney p-prime", 14, 37, MARK_COL, "out of control",
                    lambda r: f'=IF($D{r}="",NA(),IF(OR($D{r}>$K{r},$D{r}<$L{r}),$D{r},NA()))')
    run_p = markers(ws, "Laney p-prime", 14, 37, MARK_COL + 1, "8 in a row one side",
                    lambda r: ('=NA()' if r < 21 else
                               f'=IF(COUNT($D{r-7}:$D{r})<8,NA(),IF(OR(MIN($D{r-7}:$D{r})>$B$5,'
                               f'MAX($D{r-7}:$D{r})<$B$5),$D{r},NA()))'))
    cats = Reference(ws, min_col=1, min_row=14, max_row=37)
    spc_chart(ws, "Laney p\u2032 — limits that survive 8,000 contacts a day", cats, [
        (Reference(ws, min_col=4, min_row=14, max_row=37), "Proportion", "1F4E79", False, True),
        (Reference(ws, min_col=13, min_row=14, max_row=37), "Centre line", "3F8F5A", False, False),
        (Reference(ws, min_col=11, min_row=14, max_row=37), "UCL", "C0392B", True, False),
        (Reference(ws, min_col=12, min_row=14, max_row=37), "LCL", "C0392B", True, False),
    ], "O4", pct=True,
        ylim=bounds([ks[i] / float(ns[i]) for i in range(N_PTS)],
                    [max(0.0, pbar - 3 * sigs[i] * sz) for i in range(N_PTS)],
                    [pbar + 3 * sigs[i] * sz for i in range(N_PTS)]),
        flags=[(out_p, "Out of control", "C0392B"),
               (run_p, "8 in a row one side — shift", "B45309")])

    # ---- Laney u' -------------------------------------------------------
    ws = wb.create_sheet("Laney u-prime")
    title(ws, "Laney u\u2032 chart — defects per unit with varying volume",
          "Defects per 100 contacts, escalations per 1,000 tickets, errors per audit. Same overdispersion problem "
          "as percentages, same fix.", 9)
    widths(ws, [14, 15, 13, 12, 12, 10, 11, 11, 22])
    baseline_input(ws, "Laney u-prime", 9, 14, 37)
    _spc_stats(ws, [
        ("Overall rate (u-bar)",
         '=SUMIF($T$14:$T$37,"<="&$B$3,$C$14:$C$37)/SUMIF($T$14:$T$37,"<="&$B$3,$B$14:$B$37)', "#,##0.0000",
         "Total defects / total units. Not the average of the daily rates."),
        ("Average moving range of z", '=IFERROR(SUMIF($T$15:$T$37,"<="&$B$3,$G$15:$G$37)/COUNTIFS($T$15:$T$37,"<="&$B$3,$G$15:$G$37,">"&-9.9E+307),"")', "#,##0.000",
         "Movement of the standardised points, over the baseline window."),
        ("Sigma z (Laney adjustment)", "=MAX(1,B6/1.128)", "#,##0.000",
         "1.0 means an ordinary u-chart was fine. Above that, it was not. It never drops below 1 — the adjustment "
         "widens limits, it never tightens them."),
        ("Rate per 100 units", "=B5*100", "#,##0.00", "The same number in the units people actually quote."),
    ], 9)
    band(ws, 12, "Your data — one row per week", 9)
    header(ws, 13, ["Week", "Contacts audited", "Defects found", "Rate (u)", "Sigma u", "z", "MR of z", "", "Signal"])
    ws.cell(row=13, column=8, value="").fill = HDR
    us = [1240, 1310, 1180, 1420, 1275, 1195, 1360, 1230, 1405, 1150, 1290, 1375,
          1215, 1330, 1185, 1265, 1440, 1205, 1300, 1355, 1170, 1285, 1225, 1390]
    # the underlying defect rate really does move week to week in an audit
    # programme — that is the overdispersion the Laney adjustment exists for,
    # so the example data has to show it rather than being tidy Poisson noise
    # Weeks 23-24 break the upper limit (~0.089): the defect rate roughly
    # triples. Overdispersion is why a plain u-chart would have cried wolf on
    # the twenty quiet weeks and had nothing left to say about these two.
    rates = [.031, .052, .036, .058, .033, .049, .041, .056, .030, .047, .038, .054,
             .032, .050, .035, .057, .040, .045, .031, .053, .037, .048, .097, .095]
    ds = [int(round(us[i] * rates[i])) for i in range(N_PTS)]
    ubar, _, _, usz = _laney(us[:BASE_PTS], ds[:BASE_PTS], poisson=True)
    usig = [((ubar / n) ** 0.5) for n in us]
    uz = [((ds[i] / float(us[i])) - ubar) / usig[i] for i in range(len(us))]
    SHOWN.update({("Laney u-prime", "B5"): "%.4f" % ubar, ("Laney u-prime", "B6"): "%.3f" % _mrbar(uz),
                  ("Laney u-prime", "B7"): "%.3f" % usz, ("Laney u-prime", "B8"): "%.2f" % (ubar * 100)})
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=1, value="Week %d" % (i + 1))
        for col, v in ((2, us[i]), (3, ds[i])):
            c = mark(ws, r, col, "in" if i else "ex"); c.value = v; c.number_format = "#,##0"
        u = mark(ws, r, 4, "calc"); u.value = "=IF(B%d=\"\",\"\",C%d/B%d)" % (r, r, r); u.number_format = "0.0000"
        su = mark(ws, r, 5, "calc"); su.value = "=IF(B%d=\"\",\"\",SQRT($B$5/B%d))" % (r, r); su.number_format = "0.0000"
        z = mark(ws, r, 6, "calc"); z.value = "=IF(B%d=\"\",\"\",(D%d-$B$5)/E%d)" % (r, r, r); z.number_format = "#,##0.00"
        if i:
            m = mark(ws, r, 7, "calc"); m.value = "=IF(COUNT(F%d:F%d)=2,ABS(F%d-F%d),\"\")" % (r - 1, r, r, r - 1); m.number_format = "#,##0.00"
        sg = mark(ws, r, 9, "calc")
        sg.value = "=IF(D{r}=\"\",\"\",IF(OR(D{r}>K{r},D{r}<L{r}),\"OUT OF CONTROL\",\"\"))".format(r=r)
        SHOWN[("Laney u-prime", "D%d" % r)] = "%.4f" % (ds[i] / float(us[i]))
        SHOWN[("Laney u-prime", "E%d" % r)] = "%.4f" % usig[i]
        SHOWN[("Laney u-prime", "F%d" % r)] = "%.2f" % uz[i]
    for col, lab in ((11, "UCL"), (12, "LCL"), (13, "CL")):
        h = ws.cell(row=13, column=col, value=lab); h.fill, h.font = HDR, F_HDR
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=11, value="=IF(B%d=\"\",\"\",$B$5+3*E%d*$B$7)" % (r, r)).number_format = "0.0000"
        ws.cell(row=r, column=12, value="=IF(B%d=\"\",\"\",MAX(0,$B$5-3*E%d*$B$7))" % (r, r)).number_format = "0.0000"
        ws.cell(row=r, column=13, value="=IF(B%d=\"\",\"\",$B$5)" % r).number_format = "0.0000"
    out_u = markers(ws, "Laney u-prime", 14, 37, MARK_COL, "out of control",
                    lambda r: f'=IF($D{r}="",NA(),IF(OR($D{r}>$K{r},$D{r}<$L{r}),$D{r},NA()))')
    run_u = markers(ws, "Laney u-prime", 14, 37, MARK_COL + 1, "8 in a row one side",
                    lambda r: ('=NA()' if r < 21 else
                               f'=IF(COUNT($D{r-7}:$D{r})<8,NA(),IF(OR(MIN($D{r-7}:$D{r})>$B$5,'
                               f'MAX($D{r-7}:$D{r})<$B$5),$D{r},NA()))'))
    cats = Reference(ws, min_col=1, min_row=14, max_row=37)
    spc_chart(ws, "Laney u\u2032 — defects per unit", cats, [
        (Reference(ws, min_col=4, min_row=14, max_row=37), "Rate", "1F4E79", False, True),
        (Reference(ws, min_col=13, min_row=14, max_row=37), "Centre line", "3F8F5A", False, False),
        (Reference(ws, min_col=11, min_row=14, max_row=37), "UCL", "C0392B", True, False),
        (Reference(ws, min_col=12, min_row=14, max_row=37), "LCL", "C0392B", True, False),
    ], "O4",
        ylim=bounds([ds[i] / float(us[i]) for i in range(N_PTS)],
                    [max(0.0, ubar - 3 * usig[i] * usz) for i in range(N_PTS)],
                    [ubar + 3 * usig[i] * usz for i in range(N_PTS)]),
        flags=[(out_u, "Out of control", "C0392B"),
               (run_u, "8 in a row one side — shift", "B45309")])

    # ---- Xbar-R ---------------------------------------------------------
    ws = wb.create_sheet("Xbar-R")
    title(ws, "Xbar-R chart — subgroups of a continuous metric",
          "Five sampled handle times per day, five audited tickets per analyst. Getting the subgrouping right is "
          "the hard part: everything inside a subgroup must share the same conditions.", 12)
    widths(ws, [12, 9, 9, 9, 9, 9, 11, 10, 11, 11, 11, 22])
    baseline_input(ws, "Xbar-R", 12, 14, 37)
    _spc_stats(ws, [
        ("Subgroup size (n)", "=COUNT(B14:F14)", "0", "Change how many observation columns you fill and this follows."),
        ("Grand average (X-double-bar)", '=IFERROR(SUMIF($T$14:$T$37,"<="&$B$3,$G$14:$G$37)/COUNTIFS($T$14:$T$37,"<="&$B$3,$G$14:$G$37,">"&-9.9E+307),"")', "#,##0.00",
         "Average of the subgroup averages over the BASELINE window."),
        ("Average range (R-bar)", '=IFERROR(SUMIF($T$14:$T$37,"<="&$B$3,$H$14:$H$37)/COUNTIFS($T$14:$T$37,"<="&$B$3,$H$14:$H$37,">"&-9.9E+307),"")', "#,##0.00",
         "Average within-subgroup spread over the baseline window."),
        ("A2 for this n", "=IFERROR(LOOKUP(B5,{2;3;4;5;6;7;8;9;10},{1.880;1.023;0.729;0.577;0.483;0.419;0.373;0.337;0.308}),\"n out of range\")",
         "#,##0.000", "Standard constant. Above n=8 most people switch to Xbar-S."),
        ("D4 for this n", "=IFERROR(LOOKUP(B5,{2;3;4;5;6;7;8;9;10},{3.267;2.574;2.282;2.114;2.004;1.924;1.864;1.816;1.777}),\"n out of range\")", "#,##0.000",
         "Sets the upper limit on the range chart: R-bar x D4. It is the only "
         "one of the three that is never zero."),
        ("D3 for this n", "=IFERROR(LOOKUP(B5,{2;3;4;5;6;7;8;9;10},{0;0;0;0;0;0.076;0.136;0.184;0.223}),\"n out of range\")", "#,##0.000", "Zero below n=7 — a range cannot be meaningfully small."),
        ("Xbar UCL / LCL", "=TEXT(B6+B8*B7,\"#,##0.00\")&\"  /  \"&TEXT(B6-B8*B7,\"#,##0.00\")", "@", "Grand average plus or minus A2 x R-bar."),
    ], 12)
    band(ws, 12, "Your data — five observations per subgroup", 12)
    header(ws, 13, ["Subgroup", "Obs 1", "Obs 2", "Obs 3", "Obs 4", "Obs 5",
                    "Average", "Range", "Xbar CL", "Xbar UCL", "Xbar LCL", "Signal"])
    import math as _m
    _avgs, _rngs = [], []
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=1, value="Day %d" % (i + 1))
        obs = XBAR_R_DATA[i]
        for j, v in enumerate(obs):
            c = mark(ws, r, 2 + j, "in" if i else "ex"); c.value = v; c.number_format = "#,##0"
        a = mark(ws, r, 7, "calc"); a.value = "=IF(COUNT(B%d:F%d)=0,\"\",AVERAGE(B%d:F%d))" % (r, r, r, r); a.number_format = "#,##0.00"
        g = mark(ws, r, 8, "calc"); g.value = "=IF(COUNT(B%d:F%d)=0,\"\",MAX(B%d:F%d)-MIN(B%d:F%d))" % (r, r, r, r, r, r); g.number_format = "#,##0.00"
        for col, f in ((9, "$B$6"), (10, "$B$6+$B$8*$B$7"), (11, "$B$6-$B$8*$B$7")):
            c = mark(ws, r, col, "calc"); c.value = "=IF(G%d=\"\",\"\",%s)" % (r, f); c.number_format = "#,##0.00"
        sg = mark(ws, r, 12, "calc")
        sg.value = "=IF(G{r}=\"\",\"\",IF(OR(G{r}>J{r},G{r}<K{r}),\"OUT OF CONTROL\",\"\"))".format(r=r)
        _avgs.append(sum(obs) / 5.0)
        _rngs.append(max(obs) - min(obs))
        SHOWN[("Xbar-R", "G%d" % r)] = "%.2f" % _avgs[-1]
        SHOWN[("Xbar-R", "H%d" % r)] = "%.2f" % _rngs[-1]
    grand = sum(_avgs[:BASE_PTS]) / float(BASE_PTS)
    rbar = sum(_rngs[:BASE_PTS]) / float(BASE_PTS)
    SHOWN.update({("Xbar-R", "B5"): "5", ("Xbar-R", "B6"): "%.2f" % grand, ("Xbar-R", "B7"): "%.2f" % rbar,
                  ("Xbar-R", "B8"): "0.577", ("Xbar-R", "B9"): "2.114", ("Xbar-R", "B10"): "0.000",
                  ("Xbar-R", "B11"): "%.2f  /  %.2f" % (grand + 0.577 * rbar, grand - 0.577 * rbar)})
    for i in range(N_PTS):
        for col, v in (("I", grand), ("J", grand + 0.577 * rbar), ("K", grand - 0.577 * rbar)):
            SHOWN[("Xbar-R", "%s%d" % (col, 14 + i))] = "%.2f" % v
    out_x = markers(ws, "Xbar-R", 14, 37, MARK_COL, "out of control",
                    lambda r: f'=IF($G{r}="",NA(),IF(OR($G{r}>$J{r},$G{r}<$K{r}),$G{r},NA()))')
    run_x = markers(ws, "Xbar-R", 14, 37, MARK_COL + 1, "8 in a row one side",
                    lambda r: ('=NA()' if r < 21 else
                               f'=IF(COUNT($G{r-7}:$G{r})<8,NA(),IF(OR(MIN($G{r-7}:$G{r})>$B$6,'
                               f'MAX($G{r-7}:$G{r})<$B$6),$G{r},NA()))'))
    out_r = markers(ws, "Xbar-R", 14, 37, MARK_COL + 2, "R out of control",
                    lambda r: f'=IF($H{r}="",NA(),IF($H{r}>$O{r},$H{r},NA()))')
    cats = Reference(ws, min_col=1, min_row=14, max_row=37)
    spc_chart(ws, "Xbar — is the average stable?", cats, [
        (Reference(ws, min_col=7, min_row=14, max_row=37), "Subgroup average", "1F4E79", False, True),
        (Reference(ws, min_col=9, min_row=14, max_row=37), "Centre line", "3F8F5A", False, False),
        (Reference(ws, min_col=10, min_row=14, max_row=37), "UCL", "C0392B", True, False),
        (Reference(ws, min_col=11, min_row=14, max_row=37), "LCL", "C0392B", True, False),
    ], "N4", ylim=bounds(_avgs, [grand - 0.577 * rbar, grand + 0.577 * rbar]),
        flags=[(out_x, "Out of control", "C0392B"),
               (run_x, "8 in a row one side — shift", "B45309")])
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=14, value="=IF(H%d=\"\",\"\",$B$7)" % r).number_format = "#,##0.00"
        ws.cell(row=r, column=15, value="=IF(H%d=\"\",\"\",$B$9*$B$7)" % r).number_format = "#,##0.00"
        ws.cell(row=r, column=16, value="=IF(H%d=\"\",\"\",$B$10*$B$7)" % r).number_format = "#,##0.00"
    spc_chart(ws, "R — is the spread stable? Read this one first", cats, [
        (Reference(ws, min_col=8, min_row=14, max_row=37), "Subgroup range", "6B4FA0", False, True),
        (Reference(ws, min_col=14, min_row=14, max_row=37), "R centre line", "3F8F5A", False, False),
        (Reference(ws, min_col=15, min_row=14, max_row=37), "R UCL", "C0392B", True, False),
        (Reference(ws, min_col=16, min_row=14, max_row=37), "R LCL", "C0392B", True, False),
    ], "N22", height=7, ylim=bounds(_rngs, [0, 2.114 * rbar]),
        flags=[(out_r, "Out of control", "C0392B")])

    # ---- EWMA -----------------------------------------------------------
    ws = wb.create_sheet("EWMA")
    title(ws, "EWMA — catch a slow slide before it becomes obvious",
          "An ordinary chart is poor at small sustained shifts, which is exactly what a decaying improvement looks "
          "like. Lambda around 0.2 detects a 1-sigma drift in a handful of periods.", 9)
    widths(ws, [14, 13, 13, 12, 12, 12, 11, 11, 22])
    _spc_stats(ws, [
        ("Lambda (weight on the newest point)", 0.2, "0.00",
         "0.05 is very smooth and slow; 0.4 behaves nearly like an ordinary chart. 0.2 is the usual choice."),
        ("L (limit width in sigmas)", 2.7, "0.0", "2.7 with lambda 0.2 gives roughly the same false-alarm rate as a 3-sigma chart."),
        ("Target / process mean", "=AVERAGE(B16:B27)", "#,##0.00",
         "Baseline window only — the first 12 periods. Averaging the whole series folds the very drift you are hunting "
         "into the centre line. Type your own target over this if you have one."),
        ("Average moving range", "=AVERAGE(C17:C27)", "#,##0.00", "Baseline window, same reason."),
        ("Sigma estimate", "=B8/1.128", "#,##0.00", "Short-term variation only — the same estimate the I chart uses."),
    ], 9)
    ws.cell(row=5, column=2).fill = IN
    ws.cell(row=6, column=2).fill = IN
    band(ws, 14, "Your data — one row per period, oldest first", 9)
    header(ws, 15, ["Period", "Value", "Moving range", "EWMA", "CL", "UCL", "LCL", "", "Signal"])
    ws.cell(row=15, column=8, value="").fill = HDR
    drift = [408, 415, 402, 431, 419, 396, 424, 410, 438, 405, 417, 429,
             418, 426, 431, 428, 437, 433, 441, 439, 448, 444, 452, 449]
    for i in range(N_PTS):
        r = 16 + i
        ws.cell(row=r, column=1, value="Day %d" % (i + 1))
        c = mark(ws, r, 2, "in" if i else "ex"); c.value = drift[i]; c.number_format = "#,##0.00"
        if i:
            m = mark(ws, r, 3, "calc"); m.value = "=IF(COUNT(B%d:B%d)=2,ABS(B%d-B%d),\"\")" % (r - 1, r, r, r - 1); m.number_format = "#,##0.00"
        e = mark(ws, r, 4, "calc")
        prev = "$B$7" if i == 0 else "D%d" % (r - 1)
        e.value = "=IF(B%d=\"\",\"\",$B$5*B%d+(1-$B$5)*%s)" % (r, r, prev)
        e.number_format = "#,##0.00"
        cl = mark(ws, r, 5, "calc"); cl.value = "=IF(B%d=\"\",\"\",$B$7)" % r; cl.number_format = "#,##0.00"
        for col, sign in ((6, "+"), (7, "-")):
            l = mark(ws, r, col, "calc")
            l.value = ("=IF(B%d=\"\",\"\",$B$7%s$B$6*$B$9*SQRT($B$5/(2-$B$5)*(1-(1-$B$5)^(2*%d))))" % (r, sign, i + 1))
            l.number_format = "#,##0.00"
        sg = mark(ws, r, 9, "calc")
        sg.value = "=IF(D{r}=\"\",\"\",IF(OR(D{r}>F{r},D{r}<G{r}),\"SIGNAL — the process has drifted\",\"\"))".format(r=r)
    _ewb = drift[:BASE_N]
    _emu, _emr = sum(_ewb) / float(BASE_N), _mrbar(_ewb)
    SHOWN.update({("EWMA", "B7"): "%.2f" % _emu, ("EWMA", "B8"): "%.2f" % _emr,
                  ("EWMA", "B9"): "%.2f" % (_emr / 1.128)})
    prev, _esig = _emu, _emr / 1.128
    for i in range(N_PTS):
        prev = 0.2 * drift[i] + 0.8 * prev
        hw = 2.7 * _esig * (0.2 / 1.8 * (1 - 0.8 ** (2 * (i + 1)))) ** 0.5
        SHOWN[("EWMA", "D%d" % (16 + i))] = "%.2f" % prev
        SHOWN[("EWMA", "E%d" % (16 + i))] = "%.2f" % _emu
        SHOWN[("EWMA", "F%d" % (16 + i))] = "%.2f" % (_emu + hw)
        SHOWN[("EWMA", "G%d" % (16 + i))] = "%.2f" % (_emu - hw)
    # Out-of-control only. An EWMA is a weighted average of everything before
    # it, so consecutive points are autocorrelated by construction and a
    # "8 in a row one side" rule fires on that autocorrelation, not on a shift.
    out_e = markers(ws, "EWMA", 16, 39, MARK_COL, "out of control",
                    lambda r: f'=IF($D{r}="",NA(),IF(OR($D{r}>$F{r},$D{r}<$G{r}),$D{r},NA()))')
    cats = Reference(ws, min_col=1, min_row=16, max_row=39)
    spc_chart(ws, "EWMA — the limits widen then settle as evidence accumulates", cats, [
        (Reference(ws, min_col=4, min_row=16, max_row=39), "EWMA", "1F4E79", False, True),
        (Reference(ws, min_col=5, min_row=16, max_row=39), "Target", "3F8F5A", False, False),
        (Reference(ws, min_col=6, min_row=16, max_row=39), "UCL", "C0392B", True, False),
        (Reference(ws, min_col=7, min_row=16, max_row=39), "LCL", "C0392B", True, False),
    ], "K4", ylim=bounds(drift, [_emu - 2.7 * _esig * (0.2 / 1.8) ** 0.5,
                                 _emu + 2.7 * _esig * (0.2 / 1.8) ** 0.5]),
        flags=[(out_e, "Signal — the process has drifted", "C0392B")])

    # ---- CUSUM ----------------------------------------------------------
    ws = wb.create_sheet("CUSUM")
    title(ws, "CUSUM — confirm a change landed on the day you deployed it",
          "Fastest detection of a step change. Two running sums: one accumulates evidence the process went up, "
          "the other that it went down. Whichever crosses h first tells you the direction.", 9)
    widths(ws, [14, 13, 13, 12, 12, 11, 11, 11, 24])
    _spc_stats(ws, [
        ("k (slack, in sigmas)", 0.5, "0.00", "Half the shift you care about. 0.5 is tuned to catch a 1-sigma shift."),
        ("h (decision interval)", 4, "0.0", "4 or 5. Lower means faster detection and more false alarms."),
        ("Target / pre-change mean", "=AVERAGE(B16:B27)", "#,##0.00",
         "Set this from the BEFORE window only. Including the after period hides the very shift you are testing for."),
        ("Average moving range", "=AVERAGE(C17:C27)", "#,##0.00", "Baseline window only, for the same reason."),
        ("Sigma estimate", "=B8/1.128", "#,##0.00",
         "The process's ordinary spread, taken from point-to-point movement rather than from a "
         "standard deviation — one bad week inflates a standard deviation and this resists that. "
         "Every CUSUM value below is counted in these units, so if this is wrong the whole chart is."),
    ], 9)
    ws.cell(row=5, column=2).fill = IN
    ws.cell(row=6, column=2).fill = IN
    band(ws, 14, "Your data — one row per period, oldest first", 9)
    header(ws, 15, ["Period", "Value", "Moving range", "Standardised", "SH (up)", "SL (down)", "h", "-h", "Signal"])
    step = [408, 415, 402, 431, 419, 396, 424, 410, 438, 405, 417, 429,
            452, 461, 448, 470, 457, 466, 449, 463, 471, 455, 468, 460]
    for i in range(N_PTS):
        r = 16 + i
        ws.cell(row=r, column=1, value="Day %d" % (i + 1))
        c = mark(ws, r, 2, "in" if i else "ex"); c.value = step[i]; c.number_format = "#,##0.00"
        if i:
            m = mark(ws, r, 3, "calc"); m.value = "=IF(COUNT(B%d:B%d)=2,ABS(B%d-B%d),\"\")" % (r - 1, r, r, r - 1); m.number_format = "#,##0.00"
        y = mark(ws, r, 4, "calc"); y.value = "=IF(B%d=\"\",\"\",(B%d-$B$7)/$B$9)" % (r, r); y.number_format = "#,##0.00"
        ph = "0" if i == 0 else "E%d" % (r - 1)
        pl = "0" if i == 0 else "F%d" % (r - 1)
        sh = mark(ws, r, 5, "calc"); sh.value = "=IF(B%d=\"\",\"\",MAX(0,D%d-$B$5+%s))" % (r, r, ph); sh.number_format = "#,##0.00"
        sl = mark(ws, r, 6, "calc"); sl.value = "=IF(B%d=\"\",\"\",MAX(0,-D%d-$B$5+%s))" % (r, r, pl); sl.number_format = "#,##0.00"
        hh = mark(ws, r, 7, "calc"); hh.value = "=IF(B%d=\"\",\"\",$B$6)" % r; hh.number_format = "#,##0.00"
        hl = mark(ws, r, 8, "calc"); hl.value = "=IF(B%d=\"\",\"\",-$B$6)" % r; hl.number_format = "#,##0.00"
        sg = mark(ws, r, 9, "calc")
        sg.value = ("=IF(E{r}=\"\",\"\",IF(E{r}>$B$6,\"SHIFT UP confirmed\","
                    "IF(F{r}>$B$6,\"SHIFT DOWN confirmed\",\"\")))").format(r=r)
    _cub = step[:BASE_N]
    _cmu, _cmr = sum(_cub) / float(BASE_N), _mrbar(_cub)
    _csig = _cmr / 1.128
    SHOWN.update({("CUSUM", "B7"): "%.2f" % _cmu, ("CUSUM", "B8"): "%.2f" % _cmr,
                  ("CUSUM", "B9"): "%.2f" % _csig})
    sh_v = sl_v = 0.0
    _cusum_shown = []
    for i in range(N_PTS):
        yv = (step[i] - _cmu) / _csig
        SHOWN[("CUSUM", "D%d" % (16 + i))] = "%.2f" % yv
        SHOWN[("CUSUM", "G%d" % (16 + i))] = "4.00"
        SHOWN[("CUSUM", "H%d" % (16 + i))] = "-4.00"
        sh_v = max(0.0, yv - 0.5 + sh_v)
        sl_v = max(0.0, -yv - 0.5 + sl_v)
        SHOWN[("CUSUM", "E%d" % (16 + i))] = "%.2f" % sh_v
        SHOWN[("CUSUM", "F%d" % (16 + i))] = "%.2f" % sl_v
        _cusum_shown.extend((sh_v, sl_v))
    out_c = markers(ws, "CUSUM", 16, 39, MARK_COL, "crossed h",
                    lambda r: f'=IF($E{r}="",NA(),IF(OR($E{r}>$B$6,$F{r}>$B$6),'
                              f'MAX($E{r},$F{r}),NA()))')
    cats = Reference(ws, min_col=1, min_row=16, max_row=39)
    spc_chart(ws, "CUSUM — the sum climbs from the day the change landed", cats, [
        (Reference(ws, min_col=5, min_row=16, max_row=39), "SH (evidence it went up)", "1F4E79", False, True),
        (Reference(ws, min_col=6, min_row=16, max_row=39), "SL (evidence it went down)", "6B4FA0", False, True),
        (Reference(ws, min_col=7, min_row=16, max_row=39),
         "Decision limit h — crossing it is the signal", "C0392B", True, False),
    ], "K4", ylim=bounds([0, 4], _cusum_shown),
        flags=[(out_c, "Crossed the decision limit", "C0392B")])

    # ---- t and g --------------------------------------------------------
    ws = wb.create_sheet("t and g (rare events)")
    title(ws, "t-chart and g-chart — for things that almost never happen",
          "Sev-1 outages, compliance misses, security incidents. Charting COUNTS of a rare event gives you a chart "
          "of zeros that tells you nothing. Chart the gap BETWEEN events instead — longer gaps mean improvement.", 8)
    widths(ws, [16, 15, 15, 13, 13, 13, 13, 26])
    _spc_stats(ws, [
        ("Events recorded", "=COUNT(B14:B37)", "0", "Ten events is a workable minimum."),
        ("Average days between (t-chart)", "=AVERAGE(B14:B37)", "#,##0.00",
         "Rising over time is the improvement you want. This is the arithmetic mean and is deliberately "
         "NOT the chart's centre line — these gaps are right-skewed, so the mean sits above the middle "
         "of the limits."),
        ("Transformed mean", "=AVERAGE(C14:C37)", "#,##0.0000",
         "Gaps between rare events are exponential, not normal. The 1/3.6 power (Nelson) makes them near-normal so ordinary limits apply."),
        ("Transformed MR-bar", "=AVERAGE(D15:D37)", "#,##0.0000",
         "MR-bar = the average moving range, the typical step from one point to the next. It is the "
         "spread the limits below are built from, and it is measured on the transformed scale, not in "
         "days — do not try to read it as a number of days."),
        ("t-chart UCL / LCL (days)", "=TEXT((B7+2.66*B8)^3.6,\"#,##0.0\")&\"  /  \"&TEXT(MAX(0,(B7-2.66*B8))^3.6,\"#,##0.0\")",
         "@", "Limits are computed on the transformed scale and then converted back to days."),
        ("Average opportunities between (g-chart)", "=AVERAGE(F14:F37)", "#,##0.00", "Use this instead of days when volume swings — 'contacts between misses' beats 'days between misses'."),
        ("g-chart UCL", "=B10+3*SQRT(B10*(B10+1))", "#,##0.00", "Geometric, so the limits are wide and the lower limit is almost always zero."),
    ], 8)
    band(ws, 12, "One row per event, oldest first", 8)
    header(ws, 13, ["Event", "Days since previous", "Transformed", "Moving range",
                    "Transformed CL", "Contacts since previous", "g CL", "Signal"])
    # The last gap breaks the UPPER limit (100.8 days), which on a t-chart is
    # the improvement signal — the sheet says rising is what you want, and a
    # reader who only ever sees a chart catch bad news learns that it is an
    # alarm rather than a detector. Rare-event charts have no frozen baseline:
    # the limits move with the data, so this had to be a gap large enough to
    # clear limits it widens itself.
    gaps = [14, 9, 22, 31, 18, 27, 41, 12, 35, 24, 48, 19, 33, 26, 52, 21, 38, 29, 44, 36, 58, 31, 47, 168]
    vol = [12800, 8400, 21600, 29800, 17400, 26100, 39900, 11200, 33500, 23100, 46800, 18300,
           31700, 25400, 50600, 20200, 36900, 27800, 43100, 34600, 56900, 29900, 45200, 161000]
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=1, value="Incident %d" % (i + 1))
        c = mark(ws, r, 2, "in" if i else "ex"); c.value = gaps[i]; c.number_format = "#,##0"
        t = mark(ws, r, 3, "calc"); t.value = "=IF(B%d=\"\",\"\",B%d^(1/3.6))" % (r, r); t.number_format = "0.0000"
        if i:
            m = mark(ws, r, 4, "calc"); m.value = "=IF(COUNT(C%d:C%d)=2,ABS(C%d-C%d),\"\")" % (r - 1, r, r, r - 1); m.number_format = "0.0000"
        cl = mark(ws, r, 5, "calc"); cl.value = "=IF(B%d=\"\",\"\",$B$7)" % r; cl.number_format = "0.0000"
        v = mark(ws, r, 6, "in" if i else "ex"); v.value = vol[i]; v.number_format = "#,##0"
        gcl = mark(ws, r, 7, "calc"); gcl.value = "=IF(F%d=\"\",\"\",$B$10)" % r; gcl.number_format = "#,##0"
        # B11 computed a g-chart UCL and nothing ever plotted it
        gu = ws.cell(row=r, column=13); gu.value = "=IF(F%d=\"\",\"\",$B$11)" % r
        gu.number_format = "#,##0"; gu.font = F_NOTE
        gm = ws.cell(row=r, column=14); gm.value = "=IF(F%d=\"\",\"\",MEDIAN($F$14:$F$37))" % r
        gm.number_format = "#,##0"; gm.font = F_NOTE
        SHOWN[("t and g (rare events)", "M%d" % r)] = ""
        SHOWN[("t and g (rare events)", "N%d" % r)] = ""
        sg = mark(ws, r, 8, "calc")
        sg.value = ("=IF(B{r}=\"\",\"\",IF(C{r}>$B$7+2.66*$B$8,\"Longer gap than expected — improvement\","
                    "IF(C{r}<$B$7-2.66*$B$8,\"Shorter gap than expected — investigate\",\"\")))").format(r=r)
        SHOWN[("t and g (rare events)", "C%d" % r)] = "%.4f" % (gaps[i] ** (1 / 3.6))
    for i in range(N_PTS):
        r = 14 + i
        ws.cell(row=r, column=10, value="=IF(B%d=\"\",\"\",($B$7+2.66*$B$8)^3.6)" % r).number_format = "#,##0.0"
        ws.cell(row=r, column=11, value="=IF(B%d=\"\",\"\",MAX(0,$B$7-2.66*$B$8)^3.6)" % r).number_format = "#,##0.0"
        # The limits are (transformed mean ± 2.66 MR)^3.6, so the centre of
        # that construction is the back-transformed mean, not the arithmetic
        # mean of the raw gaps. Gaps between rare events are right-skewed: the
        # arithmetic mean sits above the centre of its own limits, and the
        # line was drawn two days off the middle of the band it sat in.
        ws.cell(row=r, column=12, value="=IF(B%d=\"\",\"\",$B$7^3.6)" % r).number_format = "#,##0.0"
    _tt = [g ** (1 / 3.6) for g in gaps]
    _tm, _tmr = sum(_tt) / float(len(_tt)), _mrbar(_tt)
    _gm = sum(vol) / float(len(vol))
    SHOWN.update({("t and g (rare events)", "B5"): "%d" % N_PTS,
                  ("t and g (rare events)", "B6"): "%.2f" % (sum(gaps) / float(len(gaps))),
                  ("t and g (rare events)", "B7"): "%.4f" % _tm,
                  ("t and g (rare events)", "B8"): "%.4f" % _tmr,
                  ("t and g (rare events)", "B9"): "%.1f  /  %.1f" % (
                      (_tm + 2.66 * _tmr) ** 3.6, max(0.0, _tm - 2.66 * _tmr) ** 3.6),
                  ("t and g (rare events)", "B10"): "{:,.2f}".format(_gm),
                  ("t and g (rare events)", "B11"): "{:,.2f}".format(_gm + 3 * (_gm * (_gm + 1)) ** 0.5)})
    for i in range(N_PTS):
        SHOWN[("t and g (rare events)", "E%d" % (14 + i))] = "%.4f" % _tm
        SHOWN[("t and g (rare events)", "G%d" % (14 + i))] = "{:,.0f}".format(_gm)
    # The generic "8 above the centre line" rule is wrong on both of these.
    # Gaps between rare events are exponential: P(gap > mean) is about 0.368,
    # so a run rule about mean DAYS advertises 0.5^8 and delivers nothing like
    # it. Run it on the transformed column, which is what the 1/3.6 power is
    # for. Same problem on the g-chart, where opportunities-between-events are
    # geometric — there the median is the honest centre for a run rule.
    out_t = markers(ws, "t and g (rare events)", 14, 37, MARK_COL, "out of control",
                    lambda r: f'=IF($B{r}="",NA(),IF(OR($C{r}>$B$7+2.66*$B$8,'
                              f'$C{r}<$B$7-2.66*$B$8),$B{r},NA()))')
    run_t = markers(ws, "t and g (rare events)", 14, 37, MARK_COL + 1,
                    "8 in a row one side (transformed)",
                    lambda r: ('=NA()' if r < 21 else
                               f'=IF(COUNT($C{r-7}:$C{r})<8,NA(),IF(OR(MIN($C{r-7}:$C{r})>$B$7,'
                               f'MAX($C{r-7}:$C{r})<$B$7),$B{r},NA()))'))
    out_g = markers(ws, "t and g (rare events)", 14, 37, MARK_COL + 2, "above the g UCL",
                    lambda r: f'=IF($F{r}="",NA(),IF($F{r}>$M{r},$F{r},NA()))')
    run_g = markers(ws, "t and g (rare events)", 14, 37, MARK_COL + 3,
                    "8 in a row above the median",
                    lambda r: ('=NA()' if r < 21 else
                               f'=IF(COUNT($F{r-7}:$F{r})<8,NA(),IF(MIN($F{r-7}:$F{r})>$N{r},'
                               f'$F{r},NA()))'))
    cats = Reference(ws, min_col=1, min_row=14, max_row=37)
    spc_chart(ws, "t-chart — days between incidents. Rising is good.", cats, [
        (Reference(ws, min_col=2, min_row=14, max_row=37), "Days between", "1F4E79", False, True),
        (Reference(ws, min_col=12, min_row=14, max_row=37),
         "Centre line (back-transformed)", "3F8F5A", False, False),
        (Reference(ws, min_col=10, min_row=14, max_row=37), "UCL", "C0392B", True, False),
        (Reference(ws, min_col=11, min_row=14, max_row=37), "LCL", "C0392B", True, False),
    ], "N4", ylim=bounds(gaps, [0, (_tm + 2.66 * _tmr) ** 3.6]),
        flags=[(out_t, "Out of control", "C0392B"),
               (run_t, "8 in a row one side — sustained change", "B45309")])
    ws.cell(row=13, column=13, value="g UCL").font = F_NOTE
    ws.cell(row=13, column=14, value="g median").font = F_NOTE
    spc_chart(ws, "g-chart — contacts handled between incidents", cats, [
        (Reference(ws, min_col=6, min_row=14, max_row=37), "Contacts between", "6B4FA0", False, True),
        # The median is within half a percent of the mean on any realistic
        # data, so plotting both draws one line on top of the other — and it
        # was drawn in amber, which this workbook reserves for run-rule
        # warnings. The run rule still uses the median; the stats band reports
        # it; the chart shows one centre line.
        (Reference(ws, min_col=7, min_row=14, max_row=37), "Mean", "3F8F5A", False, False),
        (Reference(ws, min_col=13, min_row=14, max_row=37), "UCL", "C0392B", True, False),
    ], "Q22", height=7,
        ylim=bounds(vol, [0, sum(vol) / len(vol) + 3 * ((sum(vol) / len(vol)) ** 2) ** 0.5],),
        flags=[(out_g, "Above the UCL", "C0392B"),
               (run_g, "8 in a row above the median — sustained improvement", "B45309")])

    # ---- picker ---------------------------------------------------------
    ws = wb.create_sheet("Pick your chart", 0)
    title(ws, "Which control chart does your data need?",
          "Pick the row that matches what you have. Getting this wrong is the most common SPC mistake in support.", 4)
    widths(ws, [46, 20, 74, 2])
    header(ws, 4, ["Your data", "Use this chart", "Why, and what to watch out for", ""])
    rows = [
        ("One number per day or week (daily AHT, daily backlog, weekly SLA%)", "I-MR",
         "Fine on daily aggregates. Wrong on raw per-contact durations, which are heavily skewed. Watch for day-of-week patterns."),
        ("A percentage, with thousands of contacts a day (FCR%, SLA met%, QA pass%, reopen%)", "Laney p-prime",
         "THE DEFAULT FOR SUPPORT. An ordinary p-chart's limits are far too tight at these volumes and almost every point will look out of control, until the team stops looking at the chart entirely."),
        ("Defects per 100 contacts, with varying volume", "Laney u-prime",
         "Same overdispersion problem as percentages. Use the Laney version."),
        ("A continuous metric where you can form sensible subgroups", "Xbar-R",
         "Switch to Xbar-S when subgroups are larger than about 8. Getting the subgrouping right is the hard part."),
        ("You need to catch a slow slide before it becomes obvious", "EWMA",
         "Ordinary charts are poor at small sustained shifts — which is exactly what a decaying improvement looks like. Use lambda around 0.2."),
        ("You need to confirm a change landed on the day you deployed it", "CUSUM",
         "Fastest detection of a step change."),
        ("A rare event — sev-1 outages, compliance misses, security incidents", "t and g (rare events)",
         "Chart the time BETWEEN events. Charting counts gives you a chart of zeros and tells you nothing."),
    ]
    for i, (data, chart, why) in enumerate(rows):
        r = 5 + i
        a = ws.cell(row=r, column=1, value=data); a.alignment = Alignment(wrap_text=True, vertical="top"); a.border = THIN
        b = ws.cell(row=r, column=2, value=chart); b.font = Font(bold=True, color="FFB45309"); b.border = THIN
        b.alignment = Alignment(vertical="top")
        c = ws.cell(row=r, column=3, value=why); c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = THIN
        ws.row_dimensions[r].height = 46
    band(ws, 13, "Before you plot anything", 4)
    for i, t in enumerate([
        "Set the limits from a stable BASELINE window, then freeze them. Recomputing limits every time you add data is how a "
        "deteriorating process gets a clean chart.",
        "Twenty to twenty-five points is the working minimum. Fewer than that and the limits move every time you add a day.",
        "Limits are not targets and targets are not limits. Limits describe what the process does; a target describes what you "
        "want. Never draw a target on a control chart and call a point beyond it a special cause.",
        "A point outside the limits is a signal to investigate, not proof of a cause. Go and find out what was different that day.",
    ], start=14):
        c = ws.cell(row=i, column=1, value="\u2022  " + t)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30

    howto(wb, [
        (True, "Control charts — seven chart types, one workbook"),
        (False, "Every chart the program's selection table sends you to is built here, on your own numbers, with the "
                "control limits written as live formulas so you can see exactly where each line comes from."),
        (True, "How to use it"),
        (False, "1. Open the 'Pick your chart' tab and find the row that matches the data you have."),
        (False, "2. Go to that tab. The green cells are a worked example — 24 periods of real-shaped support "
                "data, so the chart and its limits say something the moment you open it. Overwrite them "
                "with your own numbers, oldest first. The yellow cell above is a setting you choose, not data."),
        (False, "3. The chart, the limits and the signal column update themselves. Nothing else needs touching."),
        (False, "4. Blue cells are calculated. Read the formula bar on any of them to see the arithmetic — the constants "
                "(1.128, 2.66, 3.267, A2, D3, D4) are all visible, not buried."),
        (True, "The two mistakes that kill SPC in a support organisation"),
        (False, "Using an ordinary p-chart on high-volume percentages. At 8,000 contacts a day the binomial limits around a "
                "72% FCR are about plus or minus 1.5 points, while real daily FCR swings 4 points from contact mix and "
                "staffing alone. Almost every point reads as out of control, the team stops trusting the chart, and SPC is "
                "discredited within a month. The Laney p-prime tab fixes this — check its 'sigma z' cell: if it is well "
                "above 1.0, an ordinary p-chart was never going to work on your data."),
        (False, "Recomputing the limits every week. Limits come from a baseline window and then stay put. If you recompute "
                "them continuously, a process that is slowly getting worse drags its own limits down with it and never "
                "signals."),
        (True, "Where the numbers come from"),
        (False, "Daily volumes and outcome counts: your contact platform's daily summary (Genesys, NICE, Five9, Zendesk "
                "Explore, Salesforce reports). Defect counts: your QA tool's audit export. Incident dates: the incident "
                "management system, not a spreadsheet somebody keeps by hand."),
        (True, "Reading a chart honestly"),
        (False, "A point outside a limit, eight points in a row on one side, or a run of six steadily rising or falling all "
                "mean something changed. Anything else is noise and reacting to it makes the process worse — that is "
                "tampering, and it is the single most expensive habit in operations management."),
    ])
    return wb, "27-control-charts.xlsx"



# --------------------------------------------------------------------------
# 28 — Erlang C and Erlang A. The staffing question every contact centre asks,
# and the one the calculator tab's own subtitle admits it cannot answer.
# --------------------------------------------------------------------------

def erlang():
    """Agents required to hit a service level, with and without abandonment."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Staffing")
    title(ws, "How many agents? — Erlang C, and Erlang A when people hang up",
          "Offered load alone cannot answer this. Queueing is not linear: the first agent past the "
          "load buys a great deal of service level and the tenth buys almost none.", 14)
    widths(ws, [34, 13, 3, 9, 11, 8, 11, 11, 3, 12, 11, 11, 3, 12, 3, 3])

    band(ws, 4, "YOUR NUMBERS — one planning interval", 14)
    ins = [
        ("Contacts arriving in the interval", 360, "0",
         "From your ACD, for the interval you are staffing. Erlang assumes a steady arrival rate, "
         "so staff each interval separately — a whole day is never steady."),
        ("Interval length (minutes)", 30, "0",
         "30 minutes is the WFM standard. Below 15 the arrivals are too lumpy for the maths."),
        ("Average handle time (seconds)", 300, "0",
         "Talk + hold + after-call work. Leaving ACW out is the most common staffing error, and it "
         "understates the requirement by whatever ACW actually is."),
        ("Answer target (seconds)", 20, "0", "The 20 in '80% in 20 seconds'."),
        ("Target service level", 0.80, "0%",
         "The 80. SL = service level, the share of contacts answered inside your target time — this "
         "cell and the one above are the whole promise, and both come from whoever signed it."),
        ("Shrinkage", 0.30, "0%",
         "Everything you pay for that is not on the phone: breaks, training, meetings, coaching, "
         "absence. 26-30% is typical. Measure yours rather than borrowing this number."),
        ("Average patience (seconds)", 90, "0",
         "How long a caller waits before hanging up, from your ACD's abandon-time distribution. "
         "Only the Erlang A line uses it."),
    ]
    for i, (lab, val, fmt, note) in enumerate(ins, start=5):
        ws.cell(row=i, column=1, value=lab).font = F_B
        c = mark(ws, i, 2, "in")
        c.value = val
        c.number_format = fmt
        note_cell(ws, i, 4, note, 14)

    band(ws, 12, "WHAT THAT MEANS", 14)
    outs = [
        ("Offered load (Erlangs)", "=$B$5*$B$7/($B$6*60)", "#,##0.00",
         "Arrival rate x handle time. The number of agents that would be busy if nobody ever "
         "queued. It is the floor, never the answer.", False),
        ("AGENTS REQUIRED (Erlang C)",
         '=IFERROR(INDEX($F$22:$F$121,MATCH(1,$O$22:$O$121,0)),"raise the range")',
         "0", "The smallest number of agents that meets your target.", True),
        ("Service level at that staffing",
         '=IFERROR(INDEX($J$22:$J$121,MATCH($B$14,$F$22:$F$121,0)),"")', "0.0%",
         "What you would actually deliver at that headcount. It normally overshoots the target, "
         "because agents come in whole numbers and the one that takes you over the line buys a lot "
         "of service level. Do not shave it back off — the next interval will need it.", False),
        ("Average speed of answer (seconds)",
         '=IFERROR(INDEX($K$22:$K$121,MATCH($B$14,$F$22:$F$121,0)),"")', "#,##0.0",
         "ASA = average speed of answer, in seconds — averaged over everyone, including the callers "
         "answered instantly. It is not the wait a queued caller feels, which is much longer. Report "
         "service level to customers and ASA to operations.", False),
        ("Occupancy at that staffing",
         '=IFERROR(INDEX($L$22:$L$121,MATCH($B$14,$F$22:$F$121,0)),"")', "0.0%",
         "Sustained above about 85% and attrition rises while handle time drifts up. If this is "
         "high the answer is more agents, not more discipline.", False),
        ("PAID FTE (after shrinkage)", '=IFERROR($B$14/(1-$B$10),"")', "#,##0.0",
         "What you actually roster. Shrinkage is applied to the ANSWER, never to the load — "
         "inflating the load first and then running Erlang gives a different and wrong number.",
         True),
        ("Agents if you credit abandonment (Erlang A)",
         '=IFERROR(INDEX($F$22:$F$121,MATCH(1,$P$22:$P$121,0)),"")', "0",
         "Erlang C assumes nobody ever hangs up, which is false and makes it conservative. Erlang A "
         "credits the callers who abandon and usually asks for one or two fewer.", False),
    ]
    for i, (lab, f, fmt, note, strong) in enumerate(outs, start=13):
        ws.cell(row=i, column=1, value=lab).font = (
            Font(bold=True, size=11, color="FFB45309") if strong else F_B)
        c = mark(ws, i, 2, "calc")
        c.value = f
        c.number_format = fmt
        if note:
            note_cell(ws, i, 4, note, 14)

    SHOWN.update({("Staffing", "B13"): "60.00", ("Staffing", "B14"): "68",
                  ("Staffing", "B15"): "81.5%", ("Staffing", "B16"): "10.2",
                  ("Staffing", "B17"): "88.2%", ("Staffing", "B18"): "97.1",
                  ("Staffing", "B19"): "66"})

    band(ws, 20, "THE CURVE — every agent count, so you can see the shape", 14)
    for col, lab in ((6, "Agents"), (7, "Erlang B"), (8, "Erlang C"),
                     (10, "Service level"), (11, "ASA (s)"), (12, "Occupancy"),
                     (14, "Erlang A SL")):
        h = ws.cell(row=21, column=col, value=lab)
        h.fill, h.font = HDR, F_HDR
        h.alignment = Alignment(wrap_text=True, vertical="center")

    # Erlang B recursion: B(1) = A/(1+A), B(n) = A.B(n-1) / (n + A.B(n-1)).
    # Excel does it straight down a column — no macro, no circular reference.
    for i in range(100):
        r = 22 + i
        n = i + 1
        ws.cell(row=r, column=6, value=n).font = F_NOTE
        b = ws.cell(row=r, column=7)
        b.value = ('=IFERROR($B$13/(1+$B$13),"")' if i == 0
                   else '=IFERROR($B$13*G%d/(F%d+$B$13*G%d),"")' % (r - 1, r, r - 1))
        b.number_format = "0.000000"
        c = ws.cell(row=r, column=8)
        c.value = '=IFERROR(G%d/(1-($B$13/F%d)*(1-G%d)),"")' % (r, r, r)
        c.number_format = "0.000000"
        sl = ws.cell(row=r, column=10)
        sl.value = ('=IF(F%d<=$B$13,0,IFERROR(1-H%d*EXP(-(F%d-$B$13)*$B$8/$B$7),""))'
                    % (r, r, r))
        sl.number_format = "0.0%"
        asa = ws.cell(row=r, column=11)
        asa.value = '=IF(F%d<=$B$13,"",IFERROR(H%d*$B$7/(F%d-$B$13),""))' % (r, r, r)
        asa.number_format = "#,##0.0"
        occ = ws.cell(row=r, column=12)
        # Blank below the offered load, exactly as service level and ASA already
        # are. Without the guard this row reported occupancy of 6000%, 3000%,
        # 2000% down the first sixty agent counts — arithmetically load/agents,
        # and meaningless: below the offered load the queue never clears, so
        # there is no steady state to be occupied in. Sixty rows of it read as a
        # broken sheet.
        occ.value = '=IF(F%d<=$B$13,"",IFERROR($B$13/F%d,""))' % (r, r)
        occ.number_format = "0.0%"
        # Erlang A: the waiting queue is drained by patience as well as by
        # agents, so the same staffing delivers a slightly better service level.
        ea = ws.cell(row=r, column=14)
        ea.value = ('=IF(F%d<=$B$13,0,IFERROR(1-H%d*EXP(-(F%d-$B$13)*$B$8/$B$7)'
                    '*EXP(-$B$8/$B$11),""))' % (r, r, r))
        ea.number_format = "0.0%"
        # first row that meets the target, flagged plainly so MATCH(1, ...) can
        # find it without an array formula
        f1 = ws.cell(row=r, column=15)
        f1.value = '=IF(J%d="","",IF(J%d>=$B$9,1,0))' % (r, r)
        f2 = ws.cell(row=r, column=16)
        f2.value = '=IF(N%d="","",IF(N%d>=$B$9,1,0))' % (r, r)
        for col in (15, 16):
            ws.cell(row=r, column=col).font = Font(color="FFFFFFFF", size=9)
            SHOWN[("Staffing", "%s%d" % (get_column_letter(col), r))] = ""
        for col in (7, 8, 10, 11, 12, 14):
            ws.cell(row=r, column=col).font = F_NOTE
            SHOWN[("Staffing", "%s%d" % (get_column_letter(col), r))] = ""
        SHOWN[("Staffing", "F%d" % r)] = str(n)

    # the target is the whole decision, so it belongs on the chart
    for i in range(100):
        r = 22 + i
        t = ws.cell(row=r, column=17)
        t.value = '=IF(J%d="","",$B$9)' % r
        t.number_format = "0.0%"
        t.font = Font(color="FFFFFFFF", size=9)
        SHOWN[("Staffing", "Q%d" % r)] = ""
    ws.cell(row=21, column=17, value="target").font = F_NOTE
    ws.column_dimensions["Q"].width = 4
    spc_chart(ws, "Service level against agents — the cliff, and where it flattens",
              Reference(ws, min_col=6, min_row=52, max_row=101),
              [(Reference(ws, min_col=10, min_row=52, max_row=101),
                "Erlang C (nobody abandons)", "1F4E79", False, False),
               (Reference(ws, min_col=14, min_row=52, max_row=101),
                "Erlang A (callers hang up)", "3F8F5A", False, False),
               (Reference(ws, min_col=17, min_row=52, max_row=101),
                "Your target service level", "C0392B", True, False)],
              "S4", pct=True, height=9, width=20, y_title="Service level")

    ws.merge_cells("A123:N123")
    ws.cell(row=123, column=1, value="Read the curve, not just the number. Between the offered load "
            "and your answer each agent buys a great deal of service level; past it each one buys "
            "very little. That shape is why cutting two agents from a comfortable roster is "
            "survivable and cutting two from a tight one is not — and it is the argument to take "
            "into a budget conversation.").font = F_NOTE

    howto(wb, LEGEND + [
        (True, "What this is for"),
        (False, "Answering “how many agents do we need to hit 80% in 20 seconds?”, which "
                "offered load alone cannot tell you."),
        (True, "Why the offered load is not the answer"),
        (False, "Offered load is how many agents would be busy if nobody ever waited. Real arrivals "
                "cluster, so a queue forms even when there is spare capacity on average. At 60 "
                "Erlangs of load you do not need 60 agents, you need 68 — and those eight are not "
                "slack, they are what stops the queue running away."),
        (True, "Erlang C against Erlang A"),
        (False, "Erlang C assumes every caller waits forever. Nobody does. Erlang A credits the "
                "callers who abandon and so asks for slightly fewer agents. Plan with C because it "
                "is conservative; look at A to see how much of your measured service level is "
                "actually people giving up."),
        (True, "Where it stops being true"),
        (False, "Erlang assumes a steady arrival rate, one skill, one customer per agent and no "
                "callbacks. Staff each interval separately."),
        (True, "Chat, and what to use instead"),
        (False, "Erlang C cannot represent a concurrency cap: a chat waits for an agent who is "
                "BELOW the cap, not for one who is free. It still sizes the workload, though, if "
                "you enter the AGENT VIEW of handle time above — busy time divided by chats "
                "finished, which is agent-seconds consumed per chat, so the concurrency is "
                "already divided out. Enter the case view (session duration) instead and you "
                "over-staff by exactly the concurrency factor. Treat the answer as a floor: "
                "session duration is not independent of load, so the handle time you measured at "
                "a cap of 2 is not the one you get at a cap of 3, and Erlang reports neither that "
                "nor response latency. Validate with a discrete-event simulation that models the "
                "cap — the chat engines in the WFM suites do this — and read the output as "
                "service level against concurrency cap rather than as a single headcount. With "
                "no simulator, write down that you approximated and that the tail is "
                "understated."),
        (True, "The order that matters"),
        (False, "Shrinkage is applied to the answer, not to the load. Inflating the load by "
                "shrinkage and then running Erlang produces a different and wrong number."),
    ])
    return wb, "28-erlang-staffing.xlsx"


# ------------------------------------------------------------- 29 GAGE R&R
# The continuous MSA had a document and no workbook, so the one table that
# matters — variance components, % study variation, ndc — shipped with its
# columns empty and no method anywhere in the file. "ANOVA method" is not an
# instruction. This is the arithmetic, with every step visible.

PARTS, OPS, REPS = 10, 3, 3

# Ninety readings, in minutes: ten parts down, then A/B/C × three trials across.
# A study whose parts do not span the real range flatters itself, so these span
# 5.2 to 15.9 minutes. Appraiser B reads consistently long and C consistently
# short, which is what makes reproducibility (26.1%) dominate repeatability
# (11.2%) — the finding the document reports, and the reason the fix is a
# calibration session rather than a new tool. Recalculating this grid gives
# 28.4% study variation and ndc 4 exactly, which is what 08-msa-gage-rr.md says.
_READINGS = [
    [7.01, 7.26, 7.98, 8.37, 8.43, 7.80, 7.36, 6.89, 7.05],
    [9.15, 8.68, 8.95, 9.79, 10.04, 10.42, 8.55, 7.90, 8.43],
    [12.39, 12.17, 12.90, 13.33, 13.55, 13.40, 12.53, 11.45, 11.60],
    [6.40, 6.70, 5.66, 6.99, 6.81, 7.65, 6.09, 5.53, 5.19],
    [11.02, 10.14, 11.16, 11.68, 11.40, 11.84, 10.52, 10.30, 10.20],
    [14.62, 14.93, 14.79, 15.91, 15.83, 15.44, 13.78, 13.82, 13.81],
    [7.96, 7.66, 8.25, 9.03, 9.48, 9.64, 7.73, 7.56, 8.22],
    [11.47, 11.83, 11.55, 12.90, 12.79, 12.99, 11.52, 11.83, 11.04],
    [7.33, 7.68, 7.33, 8.55, 8.73, 9.20, 7.52, 7.20, 7.97],
    [13.08, 13.37, 12.91, 14.19, 14.11, 14.28, 12.60, 12.93, 12.71],
]


def gage_rr():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Gage R&R study")
    W = 14
    title(ws, "Measurement System Analysis — continuous Gage R&R (crossed, ANOVA)",
          "How much of the variation you are about to analyse is the measurement "
          "rather than the process. The green block is a worked study — overwrite the 90 "
          "readings with your own. The yellow cells above it describe your study; "
          "everything else calculates.", W)

    band(ws, 4, "THE STUDY — what you measured, and what you will compare it against", W)
    for r, (lab, val, fmt, hint) in enumerate([
        ("What is being measured",
         "Handle time on a billing adjustment, in minutes", None,
         "Name the measurement, not the metric. 'Handle time from the CRM timestamps' "
         "and 'handle time as the analyst clocks it' are two different gages."),
        ("Parts / items in the study", PARTS, "0",
         "Ten minimum, and they must span the RANGE YOU ACTUALLY SEE. Sampling ten "
         "routine contacts is the single most common way a study flatters the gage."),
        ("Appraisers / observers", OPS, "0",
         "Three minimum. Fewer and reproducibility is a guess."),
        ("Replicates (each appraiser measures each part this many times)", REPS, "0",
         "Two minimum, three is better. Randomise the order and blind them to their "
         "own previous answer, or you are measuring memory."),
        ("Tolerance width (upper spec − lower spec), same units", 12.0, "0.0",
         "Only if you have a spec. Leave it if you do not — %study variation is the "
         "reading that matters either way. Here: an 18-minute SLA less a 6-minute floor."),
    ], start=5):
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "in")
        c.value = val
        if fmt:
            c.number_format = fmt
        note_cell(ws, r, 3, hint, W)
        note(ws, r, 2, hint)

    band(ws, 11, "THE MEASUREMENTS — one row per part, three trials per appraiser", W)
    hdr = ["Part"]
    for o in range(OPS):
        for t in range(REPS):
            hdr.append(f"{chr(65 + o)} trial {t + 1}")
    hdr += ["Part mean", "Part range", "Widest trial spread"]
    header(ws, 12, hdr)
    first, last = 13, 12 + PARTS
    for i in range(PARTS):
        r = first + i
        ws.cell(row=r, column=1, value=f"Part {i + 1}").font = F_B
        for o in range(OPS):
            for t in range(REPS):
                col = 2 + o * REPS + t
                c = mark(ws, r, col, "in")
                c.value = _READINGS[i][o * REPS + t]
                c.number_format = "0.00"
        lo, hi = get_column_letter(2), get_column_letter(1 + OPS * REPS)
        for col, f, fmt in ((11, f"=AVERAGE({lo}{r}:{hi}{r})", "0.00"),
                            (12, f"=MAX({lo}{r}:{hi}{r})-MIN({lo}{r}:{hi}{r})", "0.00"),
                            (13, "=MAX(" + ",".join(
                                f"MAX({get_column_letter(2 + o * REPS)}{r}:"
                                f"{get_column_letter(1 + (o + 1) * REPS)}{r})-"
                                f"MIN({get_column_letter(2 + o * REPS)}{r}:"
                                f"{get_column_letter(1 + (o + 1) * REPS)}{r})"
                                for o in range(OPS)) + ")", "0.00")):
            cc = mark(ws, r, col, "calc")
            cc.value, cc.number_format = f, fmt
            SHOWN[("Gage R&R study", cc.coordinate)] = ""
    note_cell(ws, last + 1, 1,
              "Widest trial spread is the repeatability signal you can read without any "
              "arithmetic: if one appraiser cannot reproduce their own answer on the same "
              "part, no amount of ANOVA will rescue the study.", W)

    # ---- ANOVA. Every sum of squares is a DEVSQ over a block you can point at.
    grid = f"B{first}:{get_column_letter(1 + OPS * REPS)}{last}"
    band(ws, last + 3, "STEP 1 — the means the ANOVA is built from", W)
    m = last + 4
    header(ws, m, ["Mean", "Value", "What it is"] + [""] * (W - 3))
    rows = [("Grand mean", f"=AVERAGE({grid})",
             "Every measurement in the study, averaged.")]
    for o in range(OPS):
        c1 = get_column_letter(2 + o * REPS)
        c2 = get_column_letter(1 + (o + 1) * REPS)
        rows.append((f"Appraiser {chr(65 + o)} mean",
                     f"=AVERAGE({c1}{first}:{c2}{last})",
                     f"All {PARTS * REPS} readings appraiser {chr(65 + o)} took. "
                     "Spread between these three is reproducibility."))
    for i, (lab, f, why) in enumerate(rows, start=m + 1):
        ws.cell(row=i, column=1, value=lab).font = F_B
        c = mark(ws, i, 2, "calc")
        c.value, c.number_format = f, "0.0000"
        SHOWN[("Gage R&R study", c.coordinate)] = ""
        note_cell(ws, i, 3, why, W)
    gm = f"$B${m + 1}"
    op_first, op_last = m + 2, m + 1 + OPS

    # cell means (part × appraiser), the block SS_int needs
    cm = m + 2 + OPS
    band(ws, cm, "STEP 2 — the part × appraiser cell means", W)
    header(ws, cm + 1, ["Part"] + [f"Appraiser {chr(65 + o)}" for o in range(OPS)]
           + [""] * (W - 1 - OPS))
    cf, cl = cm + 2, cm + 1 + PARTS
    for i in range(PARTS):
        r = cf + i
        ws.cell(row=r, column=1, value=f"Part {i + 1}").font = F_B
        for o in range(OPS):
            c1 = get_column_letter(2 + o * REPS)
            c2 = get_column_letter(1 + (o + 1) * REPS)
            c = mark(ws, r, 2 + o, "calc")
            c.value = f"=AVERAGE({c1}{first + i}:{c2}{first + i})"
            c.number_format = "0.0000"
            SHOWN[("Gage R&R study", c.coordinate)] = ""
    note_cell(ws, cl + 1, 1,
              "One number per part per appraiser. If a row is flat across the appraisers "
              "the part is easy to measure; if it is not, that part and that appraiser "
              "disagree, and the interaction term below is what prices that.", W)

    # ---- the ANOVA table itself
    a = cl + 3
    band(ws, a, "STEP 3 — ANOVA. Each sum of squares is DEVSQ over one block above", W)
    header(ws, a + 1, ["Source", "Sum of squares", "df", "Mean square",
                       "Where the sum of squares comes from"] + [""] * (W - 5))
    part_means = f"K{first}:K{last}"
    op_means = f"B{op_first}:B{op_last}"
    cell_means = f"B{cf}:{get_column_letter(1 + OPS)}{cl}"
    ss_rows = [
        ("Parts", f"={OPS * REPS}*DEVSQ({part_means})", PARTS - 1,
         f"Spread of the {PARTS} part means, times the {OPS * REPS} readings behind each."),
        ("Appraisers", f"={PARTS * REPS}*DEVSQ({op_means})", OPS - 1,
         f"Spread of the {OPS} appraiser means, times the {PARTS * REPS} readings behind each."),
        ("Part × appraiser", None, (PARTS - 1) * (OPS - 1),
         "What is left of the cell-mean spread once parts and appraisers are taken out — "
         "the appraisers disagreeing about particular parts rather than in general."),
        ("Repeatability (error)", None, PARTS * OPS * (REPS - 1),
         "What is left of the total once the cell means are taken out: the same appraiser, "
         "the same part, a different answer."),
        ("Total", f"=DEVSQ({grid})", PARTS * OPS * REPS - 1,
         "Every reading against the grand mean."),
    ]
    r0 = a + 2
    ws_ss = {}
    for i, (lab, f, df, why) in enumerate(ss_rows):
        r = r0 + i
        ws.cell(row=r, column=1, value=lab).font = F_B
        ws_ss[lab] = r
        c = mark(ws, r, 2, "calc")
        c.number_format = "0.0000"
        SHOWN[("Gage R&R study", c.coordinate)] = ""
        d = mark(ws, r, 3, "calc")
        d.value, d.number_format = df, "0"
        ms = mark(ws, r, 4, "calc")
        ms.value = f"=IF($C${r}=0,0,$B${r}/$C${r})"
        ms.number_format = "0.0000"
        SHOWN[("Gage R&R study", ms.coordinate)] = ""
        note_cell(ws, r, 5, why, W)
    ss_cells = f"{REPS}*DEVSQ({cell_means})"
    ws.cell(row=ws_ss["Parts"], column=2).value = ss_rows[0][1]
    ws.cell(row=ws_ss["Appraisers"], column=2).value = ss_rows[1][1]
    ws.cell(row=ws_ss["Part × appraiser"], column=2).value = (
        f"={ss_cells}-$B${ws_ss['Parts']}-$B${ws_ss['Appraisers']}")
    ws.cell(row=ws_ss["Repeatability (error)"], column=2).value = (
        f"=$B${ws_ss['Total']}-{ss_cells}")
    ws.cell(row=ws_ss["Total"], column=2).value = ss_rows[4][1]
    MS = {k: f"$D${v}" for k, v in ws_ss.items()}

    # ---- variance components
    v = ws_ss["Total"] + 2
    band(ws, v, "STEP 4 — variance components, and the two numbers you report", W)
    header(ws, v + 1, ["Source", "Variance component", "% Contribution",
                       "Study variation (6σ)", "% Study variation", "Verdict"]
           + [""] * (W - 6))
    # A negative component means the term is not real; AIAG pools it into error.
    rep = f"MAX(0,{MS['Repeatability (error)']})"
    inter = f"MAX(0,({MS['Part × appraiser']}-{MS['Repeatability (error)']})/{REPS})"
    oper = f"MAX(0,({MS['Appraisers']}-{MS['Part × appraiser']})/{PARTS * REPS})"
    prt = f"MAX(0,({MS['Parts']}-{MS['Part × appraiser']})/{OPS * REPS})"
    vc_rows = [
        ("Total Gage R&R", f"={rep}+{oper}+{inter}",
         "Everything the measurement contributes. This is the number you report."),
        ("Repeatability", f"={rep}",
         "Same appraiser, same part, a different answer. The instrument or the procedure — fix it with a clearer definition or a tighter method."),
        ("Reproducibility", f"={oper}+{inter}",
         "Different appraisers, same part. The people — fix it with calibration sessions and a rubric, not with a new tool."),
        ("Part-to-part", f"={prt}",
         "Real process variation. You WANT this to dominate — it is the signal."),
        ("Total variation", None,
         "Gage R&R plus part-to-part. Percentages below are shares of this."),
    ]
    vr = {}
    for i, (lab, f, why) in enumerate(vc_rows):
        r = v + 2 + i
        vr[lab] = r
        ws.cell(row=r, column=1, value=lab).font = (
            Font(size=11) if lab in ("Repeatability", "Reproducibility") else F_B)
        c = mark(ws, r, 2, "calc")
        c.value, c.number_format = f, "0.00000"
        SHOWN[("Gage R&R study", c.coordinate)] = ""
        note_cell(ws, r, 7, why, W)
    tot = vr["Total variation"]
    ws.cell(row=tot, column=2).value = f"=$B${vr['Total Gage R&R']}+$B${vr['Part-to-part']}"
    for lab, r in vr.items():
        pc = mark(ws, r, 3, "calc")
        pc.value, pc.number_format = f"=IF($B${tot}=0,\"\",$B${r}/$B${tot})", "0.0%"
        sv = mark(ws, r, 4, "calc")
        sv.value, sv.number_format = f"=6*SQRT($B${r})", "0.000"
        ps = mark(ws, r, 5, "calc")
        ps.value = f"=IF($B${tot}=0,\"\",SQRT($B${r})/SQRT($B${tot}))"
        ps.number_format = "0.0%"
        for cc in (pc, sv, ps):
            SHOWN[("Gage R&R study", cc.coordinate)] = ""
    grr = vr["Total Gage R&R"]
    vd = mark(ws, grr, 6, "calc")
    vd.value = (f'=IF($E${grr}="","",IF($E${grr}<0.1,"ACCEPTABLE — under 10%",'
                f'IF($E${grr}<=0.3,"MARGINAL — usable with stated caution",'
                f'"UNACCEPTABLE — fix it before you analyse anything")))')
    SHOWN[("Gage R&R study", vd.coordinate)] = ""
    note_cell(ws, tot + 2, 1,
              "% Study variation compares STANDARD DEVIATIONS, % Contribution compares "
              "VARIANCES, which is why the two columns differ so much. Judge the gage on "
              "% study variation: under 10% acceptable, 10–30% marginal, over 30% "
              "unacceptable. Read % contribution to see where the damage is coming from.", W)

    # ---- ndc and the tolerance reading
    n = tot + 4
    band(ws, n, "STEP 5 — how many groups this gage can actually tell apart", W)
    for i, (lab, f, fmt, why) in enumerate([
        ("Number of distinct categories (ndc)",
         f"=IF($B${grr}=0,\"\",MAX(1,INT(1.41*SQRT($B${vr['Part-to-part']})/"
         f"SQRT($B${grr}))))", "0",
         "How many groups the gage can separate across your part range. Want 5 or more. "
         "Below 5 the gage sorts into a handful of buckets, and 4 means it is telling you "
         "'high, medium-high, medium, low' and nothing finer."),
        ("% Tolerance (Gage R&R against the spec width)",
         f"=IF($B$9=0,\"\",6*SQRT($B${grr})/$B$9)", "0.0%",
         "Only meaningful if you entered a tolerance. Same bands as % study variation. "
         "A gage can be fine against the spec and useless for improvement work, because "
         "improvement needs it to resolve variation the spec does not care about."),
    ]):
        r = n + 1 + i
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "calc")
        c.value, c.number_format = f, fmt
        SHOWN[("Gage R&R study", c.coordinate)] = ""
        note_cell(ws, r, 3, why, W)

    # ---- the picture
    cat = Reference(ws, min_col=1, min_row=vr["Total Gage R&R"], max_row=vr["Part-to-part"])
    val = Reference(ws, min_col=5, min_row=vr["Total Gage R&R"], max_row=vr["Part-to-part"])
    lim = n + 4
    # "30% line" beside four cells reading 30% told the reader nothing about why
    # a number they never entered was repeated four times across their results.
    ws.cell(row=lim, column=1,
            value="30% reference line — chart plumbing, not a result. The same number "
                  "repeated once per bar is what lets the chart draw one flat line "
                  "across all four of them.").font = F_NOTE
    for i in range(4):
        c = ws.cell(row=lim, column=2 + i, value=0.30)
        c.number_format = "0%"
    ch = bar(ws, "Where your measurement error is — under 30% or the study stops here",
             cat, val, "P4", pct=True, name="% study variation",
             colours=["C0392B", "E08A3C", "B45309", "1D6F42"])
    overlay(ch, ws, Reference(ws, min_col=2, min_row=lim, max_col=5, max_row=lim),
            "30% — above this the gage is unacceptable")
    ch.y_axis.title = "% of total study variation"

    widths(ws, [46, 13, 13, 13, 13, 13, 13, 13, 13, 13, 12, 12, 14, 14])
    ws.freeze_panes = "A13"

    howto(wb, [
        (True, "Continuous Gage R&R — what this answers"),
        (False, "Before you analyse a continuous metric, find out how much of its variation is "
                "the measurement rather than the process. If the gage eats 30% of the variation, "
                "every capability number and every hypothesis test downstream is an opinion with "
                "decimal places on it."),
        (True, "Running the study"),
        (False, "Pick 10 parts that SPAN THE RANGE YOU ACTUALLY SEE — not ten typical ones. Have "
                "3 appraisers measure each part 3 times. Randomise the order, and make sure no "
                "appraiser can see their own earlier answer or anyone else's. Type the 90 "
                "readings into the yellow grid; the rest of the workbook is formulas."),
        (True, "Reading the answer"),
        (False, "% Study variation is the headline: under 10% acceptable, 10–30% marginal, over "
                "30% unacceptable. Then look at the split. Repeatability high means the method is "
                "loose — tighten the definition. Reproducibility high means the PEOPLE disagree — "
                "that is a calibration and rubric problem, and buying a new tool will not touch it."),
        (True, "ndc, and why 5"),
        (False, "Number of distinct categories is how many groups the gage can separate across "
                "your parts. Below 5 you effectively have an ordinal scale wearing a number's "
                "clothing, and the arithmetic you do on it will not mean what you think."),
        (True, "The support-specific trap"),
        (False, "In support the 'gage' is usually a timestamp definition or a human judgement, not "
                "an instrument. Two analysts timing the same call differ because they disagree "
                "about when after-call work ends, which is a definition problem showing up as "
                "measurement error. Fix the operational definition first (04-operational-"
                "definition.md), then re-run this — it is usually most of the gap."),
        (True, "Where the numbers go next"),
        (False, "The verdict and the two percentages belong in 08-msa-gage-rr.md, which is the "
                "one-page record you sign. Do not start capability analysis until this passes."),
    ])
    return wb, "29-msa-gage-rr.xlsx"


# ------------------------------------------------------------- 30 regression
# Twenty-four billing adjustments drawn at random from the 966 handled in the
# baseline month, the same month the evidence pack describes: 137 of those 966
# reopened inside seven days (14.2%) against a target of 8%, and handle time
# averaged 412 seconds. These twenty-four average 412 seconds and four of them
# reopened, so every number a reader meets on the tabs below ties back to a
# figure they have already seen in 14-root-cause-evidence-pack.md.
#
#   ref, raised after 17:00, over the $50 authority limit, agent tenure in
#   months, transfers before resolution, handle time in seconds, reopened
REG_ROWS = [
    ("BA-2201 · 17:05", 1, 0,  7, 0, 280, 0),
    ("BA-2202 · 17:20", 1, 1, 24, 1, 506, 0),
    ("BA-2203 · 08:15", 0, 1, 39, 1, 344, 0),
    ("BA-2204 · 17:45", 1, 1, 39, 2, 479, 0),
    ("BA-2205 · 18:10", 1, 1, 33, 2, 422, 1),
    ("BA-2206 · 09:05", 0, 0, 11, 3, 452, 1),
    ("BA-2207 · 09:40", 0, 0, 46, 2, 374, 0),
    ("BA-2208 · 18:35", 1, 1,  4, 0, 451, 0),
    ("BA-2209 · 10:20", 0, 0,  4, 0, 360, 0),
    ("BA-2210 · 11:05", 0, 0,  6, 0, 260, 0),
    ("BA-2211 · 11:35", 0, 1, 28, 0, 415, 0),
    ("BA-2212 · 19:00", 1, 1,  3, 2, 555, 1),
    ("BA-2213 · 12:10", 0, 0,  4, 1, 393, 0),
    ("BA-2214 · 19:25", 1, 1, 39, 0, 454, 0),
    ("BA-2215 · 13:25", 0, 0, 18, 0, 257, 0),
    ("BA-2216 · 19:50", 1, 1, 18, 1, 544, 0),
    ("BA-2217 · 20:15", 1, 1, 11, 1, 482, 0),
    ("BA-2218 · 14:00", 0, 0,  2, 1, 403, 0),
    ("BA-2219 · 20:40", 1, 0,  5, 2, 433, 0),
    ("BA-2220 · 14:35", 0, 0,  4, 2, 483, 0),
    ("BA-2221 · 15:10", 0, 0, 11, 0, 261, 0),
    ("BA-2222 · 15:50", 0, 0, 53, 1, 424, 0),
    ("BA-2223 · 16:20", 0, 0, 21, 1, 359, 0),
    ("BA-2224 · 21:10", 1, 1,  5, 1, 505, 1),
]

# The four candidate drivers, in the order they sit in the data block.
REG_X = [
    ("Raised after 17:00 (1 = yes)",
     "an adjustment raised after 17:00 misses the 02:00 posting batch, so the "
     "customer sees no credit the next morning"),
    ("Over the $50 authority limit (1 = yes)",
     "over $50 the agent cannot approve it themselves and it goes to a queue"),
    ("Agent tenure in months",
     "months since the agent finished induction, from the HR extract"),
    ("Transfers before resolution",
     "how many times the contact changed hands, counted from the CRM routing log"),
]

# What a binary logistic fit on all 966 contacts returned. Pasted here, not
# fitted here — see the honesty note on the tab.
#   term, coefficient (log-odds), standard error, p-value as the tool printed it
REG_LOGIT = [
    ("Intercept", -3.05, 0.24, "<0.0001"),
    ("Raised after 17:00 (1 = yes)", 1.16, 0.21, "<0.0001"),
    ("Over the $50 authority limit (1 = yes)", 0.83, 0.22, "0.0002"),
    ("Agent tenure in months", -0.038, 0.009, "<0.0001"),
    ("Transfers before resolution", 0.61, 0.12, "<0.0001"),
]


def _reg_data_block(ws, hdr_row, cols, values, fmts):
    """The shared 24-row example block: first row green, the rest yellow.

    `cols` is the column index of each field, `values` a row-major table. The
    first row of every data block in this workbook is the worked example, so a
    reader can see the shape before they delete it.
    """
    first = hdr_row + 1
    for i, row in enumerate(values):
        r = first + i
        for col, v, fmt in zip(cols, row, fmts):
            c = mark(ws, r, col, "ex" if i == 0 else "in")
            c.value = v
            if fmt:
                c.number_format = fmt
    return first, first + len(values) - 1


def regression():
    wb = Workbook(); wb.remove(wb.active)
    refs = [r[0] for r in REG_ROWS]

    # =================================================== 1. simple linear ===
    ws = wb.create_sheet("Simple linear")
    W = 6
    title(ws, "Simple linear regression — one driver, one outcome",
          "Fit a straight line through two columns, then find out how much of the "
          "outcome that one driver actually explains. The green rows are a worked example — "
          "replace them with your own pairs. Yellow cells above are yours to set; "
          "everything blue is a formula you can point at.", W)
    # Column A carries both the long statistic labels and the contact ref, and
    # column B both a typed answer and the x column, so both are sized for the
    # longest thing that lands in them rather than for the table alone.
    widths(ws, [42, 24, 18, 18, 16, 26, 11, 11])

    band(ws, 4, "THE QUESTION — what you are modelling, and where the rows came from", W)
    for r, (lab, val, hint) in enumerate([
        ("Outcome (y) — the thing you want to move",
         "Handle time in seconds",
         "Name the measurement, not the ambition. 'Handle time from CRM open to "
         "CRM close, in seconds' is a column you can pull; 'efficiency' is not."),
        ("Driver (x) — the one thing you are testing",
         "Transfers per contact",
         "One column, and one you could actually change. Start with the driver "
         "the process owner already believes in; you are testing their belief, "
         "not replacing it."),
        ("Where the rows come from",
         "24 of 966 March contacts",
         "Say the population and the sampling method. A convenience sample of "
         "the contacts you happened to remember will produce a slope, and the "
         "slope will be of your memory."),
    ], start=5):
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "in")
        c.value = val
        note_cell(ws, r, 3, hint, W)
        note(ws, r, 2, hint)

    band(ws, 9, "THE DATA — one row per contact, x beside y", W)
    hdr = 10
    header(ws, hdr, ["Contact ref", "Transfers before resolution (x)",
                     "Handle time in seconds (y)", "Fitted y — what the line predicts",
                     "Residual (y − fitted)",
                     "Off the line by more than 1.5 standard errors?"])
    ws.row_dimensions[hdr].height = 46
    first, last = _reg_data_block(
        ws, hdr, [1, 2, 3], [(r[0], r[4], r[5]) for r in REG_ROWS],
        [None, "0", "#,##0"])
    # The statistics block sits ten rows below the data, and the fitted-value and
    # outlier columns read three of its cells by address. Named here so a row
    # added above cannot silently repoint them at the wrong statistic.
    R_N, R_SLOPE, R_INTER, R_RSQ = last + 5, last + 6, last + 7, last + 8
    R_STEYX, R_SESL, R_T, R_DF, R_P = (last + 10, last + 11, last + 12,
                                       last + 13, last + 14)
    for i in range(len(REG_ROWS)):
        r = first + i
        for col, f, fmt in (
            (4, f'=IF($C{r}="","",$B${R_INTER}+$B${R_SLOPE}*$B{r})', "#,##0"),
            (5, f'=IF($D{r}="","",$C{r}-$D{r})', "#,##0"),
            (6, f'=IF($E{r}="","",IF(ABS($E{r})>1.5*$B${R_STEYX},'
                f'"LOOK AT THIS ONE","on the line"))', None),
        ):
            c = mark(ws, r, col, "calc")
            c.value = f
            if fmt:
                c.number_format = fmt
            if i == 0:
                c.fill = EX
    note_cell(ws, last + 1, 1,
              "The last three columns are the model marking its own homework. Fitted y "
              "is what the line says this contact should have taken; the residual is "
              "what it missed by. BA-2216 took 544 seconds on a single transfer, which "
              "the line cannot explain at all — hold on to it, because the multiple "
              "regression tab does explain it.", W)

    band(ws, last + 3, "STEP 1 — the line, and whether it is real", W)
    sh = last + 4
    header(ws, sh, ["Statistic", "Value", "What it tells you"] + [""] * (W - 3))
    xr, yr = f"$B${first}:$B${last}", f"$C${first}:$C${last}"
    stats = [
        ("Contacts in the model (n)", f"=COUNT({yr})", "0",
         "COUNT(y). Thirty is where a regression starts to be worth quoting; below "
         "about fifteen you are describing this sample and nothing beyond it."),
        ("Slope — seconds per extra transfer", f"=SLOPE({yr},{xr})", "0.00",
         "SLOPE(y, x). How much y moves for one more transfer. This is the number "
         "that goes in a business case, not R²."),
        ("Intercept — seconds at zero transfers", f"=INTERCEPT({yr},{xr})", "0.00",
         "INTERCEPT(y, x). Only meaningful if zero is inside the range you actually "
         "observed. Here it is: plenty of these contacts were never transferred."),
        ("R² — share of the spread the line explains", f"=RSQ({yr},{xr})", "0.0%",
         "RSQ(y, x). The share of the variation in y this one driver accounts for. "
         "Everything else is the drivers you did not put in the model."),
        ("Correlation r", f"=CORREL({yr},{xr})", "0.000",
         "CORREL(y, x). The square root of R², carrying the sign. Useful for "
         "direction; do not quote it as if it were a share of anything."),
        ("Standard error of the estimate", f"=STEYX({yr},{xr})", "0.00",
         "STEYX(y, x). The typical distance a point sits from the line, in seconds. "
         "This is the honest width of a prediction from this model."),
        ("Standard error of the slope", f"=INDEX(LINEST({yr},{xr},TRUE,TRUE),2,1)", "0.00",
         "INDEX(LINEST(y, x, TRUE, TRUE), 2, 1). LINEST returns a block; row 2 "
         "column 1 of it is how much the slope would wobble on another sample."),
        ("t-stat for the slope", f'=IFERROR($B${R_SLOPE}/$B${R_SESL},"")', "0.00",
         "Slope divided by its own standard error. Two or more and the slope is "
         "usually distinguishable from a flat line at this sample size."),
        ("Degrees of freedom", f"=IFERROR(COUNT({yr})-2,\"\")", "0",
         "n − 2: one for the slope, one for the intercept. Everything about "
         "significance below depends on this being honest about your row count."),
        ("p-value for the slope",
         f'=IFERROR(TDIST(ABS($B${R_T}),$B${R_DF},2),"")', "0.0000",
         "TDIST(|t|, df, 2). The chance of a slope this far from flat if transfers "
         "made no difference at all."),
    ]
    for i, (lab, f, fmt, why) in enumerate(stats):
        r = sh + 1 + i
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "calc")
        c.value, c.number_format = f, fmt
        note_cell(ws, r, 3, why, W)
    assert [sh + 1 + i for i in (0, 1, 2, 3, 5, 6, 7, 8, 9)] == [
        R_N, R_SLOPE, R_INTER, R_RSQ, R_STEYX, R_SESL, R_T, R_DF, R_P]
    eqb = sh + 1 + len(stats) + 1

    band(ws, eqb, "THE MODEL, WRITTEN OUT — the sentence you are allowed to say", W)
    for r, f in ((eqb + 1,
                  f'="Handle time in seconds  =  "&TEXT($B${R_INTER},"0.0")&"  +  "'
                  f'&TEXT($B${R_SLOPE},"0.0")&" × transfers.      One more transfer is '
                  f'worth about "&TEXT($B${R_SLOPE},"0")&" seconds, and this one driver '
                  f'explains "&TEXT($B${R_RSQ},"0%")&" of the spread in handle time "'
                  f'&"(p = "&TEXT($B${R_P},"0.000")&", n = "&TEXT($B${R_N},"0")&")."'),
                 (eqb + 2,
                  f'=IF($B${R_P}="","",IF($B${R_P}>=0.05,'
                  f'"NOT PROVEN — at this sample size the slope is not distinguishable '
                  f'from a flat line. Collect more rows before you build anything on it.",'
                  f'IF($B${R_RSQ}<0.3,"REAL BUT PARTIAL — the slope is significant (p = "'
                  f'&TEXT($B${R_P},"0.000")&") and the line still explains only "'
                  f'&TEXT($B${R_RSQ},"0%")&" of the spread. Transfers are a genuine driver '
                  f'AND most of what moves handle time is not in this model.",'
                  f'"REAL AND STRONG — significant, and the line explains "'
                  f'&TEXT($B${R_RSQ},"0%")&" of the spread. Rare in support data; check '
                  f'the sample before you celebrate.")))')):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=W)
        c = mark(ws, r, 1, "calc")
        c.value = f
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 30

    rb = eqb + 4
    band(ws, rb, "THE DECISION RULE — read this before you quote the R²", W)
    for i, txt in enumerate([
        "How high does R² have to be? Much lower than a manufacturing course will "
        "tell you. In support you are modelling human behaviour through a CRM, and "
        "0.20 to 0.40 from a single driver is a normal, useful result. Judge the "
        "model on whether the SLOPE is big enough to act on and stable enough to "
        "bet on — not on R². A model that explains 25% of handle time and hands you "
        "a lever worth 50 seconds a contact is worth more than one explaining 80% "
        "of something you cannot change.",
        "A significant coefficient is not the same as one worth acting on. "
        "Significance says the effect is probably not zero. It says nothing about "
        "size. With enough rows a 3-second effect becomes significant, and 3 seconds "
        "will not move your reopen rate. Always read the slope in the units of the "
        "problem, multiply it by the range you actually see, and compare that to the "
        "smallest change that would matter to the business.",
        "Correlation here is not causation, and one driver hides the rest. This "
        "sheet cannot tell you whether transfers CAUSE longer handle times or "
        "whether hard cases cause both. Nothing in a regression can. To claim "
        "cause you need either a mechanism you have watched at the gemba or a "
        "designed experiment (24-doe-design-matrix.xlsx). Until then this is a "
        "well-measured association, and you say so out loud.",
    ]):
        note_cell(ws, rb + 1 + i, 1, txt, W)

    # ---- scaffolding: the two endpoints the fitted line is drawn between
    ws.cell(row=hdr, column=7,
            value="fitted line endpoints — the chart reads these").font = F_NOTE
    for i, (xf, yf) in enumerate((
            (f"=IFERROR(MIN({xr}),0)",
             f'=IFERROR($B${R_INTER}+$B${R_SLOPE}*$G${first},"")'),
            (f"=IFERROR(MAX({xr}),0)",
             f'=IFERROR($B${R_INTER}+$B${R_SLOPE}*$G${first + 1},"")'))):
        ws.cell(row=first + i, column=7, value=xf).number_format = "0.00"
        ws.cell(row=first + i, column=8, value=yf).number_format = "#,##0"

    sc = ScatterChart()
    sc.title = ("Handle time against transfers — the line is the model, "
                "the scatter is what it misses")
    sc.style = 13
    sc.height, sc.width = 10, 17
    sc.scatterStyle = "lineMarker"
    sc.x_axis.title = "Transfers before resolution"
    sc.y_axis.title = "Handle time (seconds)"
    sc.x_axis.delete = False
    sc.y_axis.delete = False
    pts = Series(Reference(ws, min_col=3, min_row=first, max_row=last),
                 xvalues=Reference(ws, min_col=2, min_row=first, max_row=last),
                 title="Contacts — one dot each")
    pts.marker = Marker(symbol="circle", size=8)
    pts.marker.graphicalProperties.solidFill = "1F4E79"
    pts.graphicalProperties.line.noFill = True
    sc.series.append(pts)
    fit = Series(Reference(ws, min_col=8, min_row=first, max_row=first + 1),
                 xvalues=Reference(ws, min_col=7, min_row=first, max_row=first + 1),
                 title="Fitted line — the model's average")
    fit.marker = Marker(symbol="none")
    fit.graphicalProperties.line.solidFill = "C0392B"
    fit.graphicalProperties.line.width = 22000
    sc.series.append(fit)
    sc.legend = Legend()
    sc.legend.position = "b"
    ws.add_chart(sc, "J4")

    # ============================================== 2. multiple regression ===
    ws2 = wb.create_sheet("Multiple regression")
    W2 = 9
    title(ws2, "Multiple regression — four candidate drivers at once",
          "One driver at a time flatters whichever one you looked at first. This "
          "fits all four together, prices each one with the others held still, and "
          "shows you which of them the data cannot actually tell apart.", W2)
    widths(ws2, [44, 24, 14, 14, 15, 15, 17, 26, 40] + [9] * 14)

    band(ws2, 4, "THE MODEL — the outcome, and the smallest effect worth acting on", W2)
    for r, (lab, val, fmt, hint) in enumerate([
        ("Outcome (y) — what the model explains",
         "Handle time in seconds", None,
         "One continuous column. If your outcome is yes/no — reopened, escalated, "
         "breached — this is the wrong tab: go to Logistic regression."),
        ("Where the rows come from",
         "24 of 966 March contacts", None,
         "Population and sampling method. A model fitted on the contacts somebody "
         "escalated is a model of escalation, whatever you name the outcome."),
        ("Smallest effect worth acting on (seconds)", 20, "#,##0",
         "The smallest change in the outcome that would actually change a decision. "
         "Ask the process owner, before you see the coefficients. Twenty seconds a "
         "contact across 966 contacts a month is about five and a half hours of "
         "handling — enough to bother with; two seconds is not."),
    ], start=5):
        ws2.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws2, r, 2, "in")
        c.value = val
        if fmt:
            c.number_format = fmt
        note_cell(ws2, r, 3, hint, W2)
        note(ws2, r, 2, hint)

    band(ws2, 9, "THE DATA — the same 24 contacts, with all four drivers beside y", W2)
    h2 = 10
    header(ws2, h2, ["Contact ref"] + [x[0] for x in REG_X]
           + ["Handle time in seconds (y)", "Fitted y — what the model predicts",
              "Residual (y − fitted)"])
    ws2.row_dimensions[h2].height = 46
    f2, l2 = _reg_data_block(
        ws2, h2, [1, 2, 3, 4, 5, 6],
        [(r[0], r[1], r[2], r[3], r[4], r[5]) for r in REG_ROWS],
        [None, "0", "0", "#,##0", "0", "#,##0"])
    cb = l2 + 4                              # coefficient block band row
    hc = cb + 1                              # its header row
    # The residual degrees of freedom, which every per-coefficient p-value below
    # divides by, lands seven rows into the goodness-of-fit block. Computed here
    # because the coefficient table is written first and has to point at it.
    R_DF2 = hc + 16
    for i in range(len(REG_ROWS)):
        r = f2 + i
        for col, f, fmt in (
            (7, f'=IF($F{r}="","",$B${hc + 1}+$B${hc + 2}*$B{r}+$B${hc + 3}*$C{r}'
                f'+$B${hc + 4}*$D{r}+$B${hc + 5}*$E{r})', "#,##0"),
            (8, f'=IF($G{r}="","",$F{r}-$G{r})', "#,##0"),
        ):
            c = mark(ws2, r, col, "calc")
            c.value, c.number_format = f, fmt
            if i == 0:
                c.fill = EX
    note_cell(ws2, l2 + 1, 1,
              "Two of the four drivers are 1/0 flags, which is how you put a yes/no "
              "into a regression: the coefficient is then simply the seconds the flag "
              "adds. Fitted y and the residual are the model marking its own homework, "
              "and the residual column is what the second chart plots.", W2)

    band(ws2, cb, "STEP 1 — the coefficients, and which of them you can act on", W2)
    header(ws2, hc, ["Predictor", "Coefficient", "Std error", "t-stat", "p-value",
                     "VIF", "Effect across the observed range", "Act on it?",
                     "What it means"])
    ws2.row_dimensions[hc].height = 46
    LN = f"LINEST($F${f2}:$F${l2},$B${f2}:$E${l2},TRUE,TRUE)"
    # LINEST hands the coefficients back in reverse column order, so the first
    # predictor (column B) is the LAST of the four it returns. Getting this
    # backwards is the classic way to publish a model with the labels swapped.
    rows_c = [("Intercept", 5, None, None)] + [
        (REG_X[j][0], 4 - j, get_column_letter(2 + j), j) for j in range(4)]
    means = [
        "Predicted handle time for a contact with every driver at zero — before "
        "17:00, under the limit, a brand-new agent, no transfers. Read it as the "
        "starting point the four effects are added to, not as a prediction.",
        "Seconds the after-17:00 flag adds once the other three are held still. "
        "Compare it with the p-value beside it before you repeat it out loud.",
        "Seconds the $50 authority limit adds once the other three are held still. "
        "This is the queue, not the money.",
        "Seconds per extra month of tenure — negative means longer-serving agents "
        "are quicker. Small per month, so read the range column, not this one.",
        "Seconds each extra transfer adds once the other three are held still. "
        "Compare it with the 50.7 seconds the single-driver tab gave you.",
    ]
    for i, (lab, ln_col, xcol, j) in enumerate(rows_c):
        r = hc + 1 + i
        ws2.cell(row=r, column=1, value=lab).alignment = Alignment(
            wrap_text=True, vertical="center")
        for col, f, fmt in (
            (2, f"=IFERROR(INDEX({LN},1,{ln_col}),\"\")", "0.00"),
            (3, f"=IFERROR(INDEX({LN},2,{ln_col}),\"\")", "0.00"),
            (4, f'=IFERROR($B{r}/$C{r},"")', "0.00"),
            (5, f'=IFERROR(TDIST(ABS($D{r}),$B${R_DF2},2),"")', "0.0000"),
        ):
            c = mark(ws2, r, col, "calc")
            c.value, c.number_format = f, fmt
            c.alignment = Alignment(vertical="center")
        if xcol is not None:
            hcol = get_column_letter(10 + j * 3)
            hend = get_column_letter(12 + j * 3)
            v = mark(ws2, r, 6, "calc")
            v.value = (f"=IFERROR(1/(1-INDEX(LINEST(${xcol}${f2}:${xcol}${l2},"
                       f"${hcol}${f2}:${hend}${l2},TRUE,TRUE),3,1)),\"\")")
            v.number_format = "0.00"
            v.alignment = Alignment(vertical="center")
            e = mark(ws2, r, 7, "calc")
            e.value = (f'=IFERROR($B{r}*(MAX(${xcol}${f2}:${xcol}${l2})'
                       f'-MIN(${xcol}${f2}:${xcol}${l2})),"")')
            e.number_format = "#,##0"
            e.alignment = Alignment(vertical="center")
            a = mark(ws2, r, 8, "calc")
            a.value = (
                f'=IF($E{r}="","",IF($E{r}>=0.05,'
                f'"NOT PROVEN — cannot be separated from zero",'
                f'IF(ABS($G{r})<$B$7,"REAL BUT TOO SMALL — under the "'
                f'&TEXT($B$7,"0")&" seconds you set","ACT ON IT — real, and big '
                f'enough to matter")))')
            a.alignment = Alignment(wrap_text=True, vertical="center")
        m = mark(ws2, r, 9, "calc")
        m.value = f'=IF($B{r}="","","{means[i]}")'
        m.alignment = Alignment(wrap_text=True, vertical="top")
        # tall enough for the longest sentence in the last column at its width,
        # or the reader meets a wrapped explanation with its final line cut off
        ws2.row_dimensions[r].height = 12 * (2 + len(means[i]) // 40)
    note_cell(ws2, hc + 6, 1,
              "VIF is the one people skip, and it is why this table disagrees with "
              "the process owner. Two thirds of the after-17:00 adjustments here are "
              "also over the $50 limit, so the two columns carry much the same "
              "information and the model cannot say which of them is doing the work — "
              "it splits the credit and widens both standard errors. Under 5 is fine, "
              "5 to 10 is a warning, over 10 means drop one of the pair or combine "
              "them, and re-fit. Nine of the eleven after-17:00 contacts here are also "
              "over the limit, which is where these VIFs come from.", W2)

    sb = hc + 8
    band(ws2, sb, "STEP 2 — the model as a whole", W2)
    hs = sb + 1
    header(ws2, hs, ["Statistic", "Value", "What it tells you"] + [""] * (W2 - 3))
    fit_stats = [
        ("Contacts in the model (n)", f"=COUNT($F${f2}:$F${l2})", "0",
         "COUNT(y). Four predictors on 24 rows is already thin. The rule of thumb "
         "is ten to fifteen rows per predictor before the coefficients settle down."),
        ("Predictors (k)", "=4", "0",
         "How many x columns the model is fitting. Every one you add uses up a "
         "degree of freedom and lifts R² whether or not it means anything."),
        ("R²", f"=IFERROR(INDEX({LN},3,1),\"\")", "0.0%",
         "INDEX(LINEST(...), 3, 1). The share of the spread in handle time all four "
         "drivers explain together."),
        ("Adjusted R²", None, "0.0%",
         "R² with a penalty for each predictor: 1 − (1 − R²)(n − 1)/(n − k − 1). If "
         "this sits well below R², some of your columns are paying their way in "
         "arithmetic rather than in meaning."),
        ("Standard error of the estimate", f"=IFERROR(INDEX({LN},3,2),\"\")", "0.00",
         "INDEX(LINEST(...), 3, 2). The typical miss, in seconds. Quote a prediction "
         "as fitted ± roughly two of these, or do not quote it."),
        ("F", f"=IFERROR(INDEX({LN},4,1),\"\")", "0.00",
         "INDEX(LINEST(...), 4, 1). Tests the whole model at once: is this set of "
         "four better than no model at all?"),
        ("Degrees of freedom (residual)", f"=IFERROR(INDEX({LN},4,2),\"\")", "0",
         "INDEX(LINEST(...), 4, 2). n − k − 1. Every per-coefficient p-value on the "
         "table above reads this cell."),
        ("p-value for F", None, "0.0000",
         "FDIST(F, k, df). Small means the four together explain something. It does "
         "NOT mean each of them does — that is what the p-values above are for."),
        ("Verdict", None, None,
         "Read this with the residual chart beside it. A model can pass every "
         "number here and still be unusable if the residuals have a shape."),
    ]
    S2 = {}
    for i, (lab, f, fmt, why) in enumerate(fit_stats):
        r = hs + 1 + i
        S2[lab] = r
        ws2.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws2, r, 2, "calc")
        c.value = f
        if fmt:
            c.number_format = fmt
        note_cell(ws2, r, 3, why, W2)
    r2r, adjr, nr, kr, fr, dfr, pfr, vr = (S2["R²"], S2["Adjusted R²"],
                                           S2["Contacts in the model (n)"],
                                           S2["Predictors (k)"], S2["F"],
                                           S2["Degrees of freedom (residual)"],
                                           S2["p-value for F"], S2["Verdict"])
    # the per-coefficient p-values read the residual df cell by address
    assert dfr == R_DF2, (dfr, R_DF2)
    ws2.cell(row=adjr, column=2).value = (
        f'=IFERROR(1-(1-$B${r2r})*($B${nr}-1)/($B${nr}-$B${kr}-1),"")')
    ws2.cell(row=pfr, column=2).value = (
        f'=IFERROR(FDIST($B${fr},$B${kr},$B${dfr}),"")')
    ws2.cell(row=vr, column=2).value = (
        f'=IF($B${pfr}="","",IF($B${pfr}>=0.05,'
        f'"NO MODEL — these four together explain nothing",'
        f'IF($B${adjr}<0.3,"USABLE BUT WEAK — read the effect sizes, not the R²",'
        f'"USABLE — now go and look at the residual chart")))')

    ab = hs + 1 + len(fit_stats) + 1
    band(ws2, ab, "THE DECISION RULE — what this model does and does not license", W2)
    for i, txt in enumerate([
        "Read the residual chart before you read anything else. Residuals against "
        "fitted values should be a shapeless cloud of even width around zero. A "
        "funnel that widens to the right means the model is least reliable exactly "
        "where the numbers are biggest, and the usual fix is to model log(y) "
        "instead. A curve means you have fitted a straight line to something bent. "
        "Either way the coefficients above are not safe to quote, however small "
        "their p-values are.",
        "Adjusted R² is the number you report, not R². R² can only go up when you "
        "add a column, so a model with enough junk in it will always look better "
        "than the honest one. For support data an adjusted R² of 0.3 to 0.6 on a "
        "duration is a genuinely useful model; anything above 0.85 usually means "
        "something on the right-hand side is a repackaging of the left — check you "
        "have not put a component of the outcome in as a predictor.",
        "A tiny p-value on a tiny coefficient is not a finding. The 'Act on it?' "
        "column deliberately refuses to say yes on significance alone: it multiplies "
        "each coefficient by the range that predictor actually covers in your data, "
        "and compares that with the smallest effect you said was worth acting on. "
        "That comparison, not the p-value, is the one to take to a steering group.",
        "None of this is causation, and a coefficient is not a lever. The model "
        "cannot tell whether transfers lengthen contacts or whether hard contacts "
        "get transferred. Before you spend money on a change, either watch the "
        "mechanism at the gemba or run a designed experiment "
        "(24-doe-design-matrix.xlsx) — a regression on observational data can rank "
        "your suspects and it cannot convict one.",
    ]):
        note_cell(ws2, ab + 1 + i, 1, txt, W2)

    # ---- scaffolding: the other-three blocks each VIF regresses against, and
    #      the two endpoints of the zero line on the residual chart
    for j in range(4):
        base = 10 + j * 3
        ws2.cell(row=h2, column=base,
                 value=f"VIF helper — the other three, for {REG_X[j][0]}").font = F_NOTE
        others = [k for k in range(4) if k != j]
        for i in range(len(REG_ROWS)):
            r = f2 + i
            for n_, k in enumerate(others):
                ws2.cell(row=r, column=base + n_,
                         value=f"={get_column_letter(2 + k)}{r}").number_format = "#,##0"
    ws2.cell(row=h2, column=22, value="zero line — the chart reads these").font = F_NOTE
    for i, f in enumerate((f"=IFERROR(MIN($G${f2}:$G${l2}),0)",
                           f"=IFERROR(MAX($G${f2}:$G${l2}),0)")):
        ws2.cell(row=f2 + i, column=22, value=f).number_format = "#,##0"
        ws2.cell(row=f2 + i, column=23, value=0).number_format = "#,##0"

    rc = ScatterChart()
    rc.title = ("Residuals against fitted — a shapeless cloud is the only shape "
                "you can use")
    rc.style = 13
    rc.height, rc.width = 10, 17
    rc.scatterStyle = "lineMarker"
    rc.x_axis.title = "Fitted handle time (seconds)"
    rc.y_axis.title = "Residual (seconds)"
    rc.x_axis.delete = False
    rc.y_axis.delete = False
    rp = Series(Reference(ws2, min_col=8, min_row=f2, max_row=l2),
                xvalues=Reference(ws2, min_col=7, min_row=f2, max_row=l2),
                title="Residual for each contact")
    rp.marker = Marker(symbol="circle", size=8)
    rp.marker.graphicalProperties.solidFill = "1F4E79"
    rp.graphicalProperties.line.noFill = True
    rc.series.append(rp)
    zl = Series(Reference(ws2, min_col=23, min_row=f2, max_row=f2 + 1),
                xvalues=Reference(ws2, min_col=22, min_row=f2, max_row=f2 + 1),
                title="Zero — residuals should average out along it")
    zl.marker = Marker(symbol="none")
    zl.graphicalProperties.line.solidFill = "C0392B"
    zl.graphicalProperties.line.dashStyle = "dash"
    zl.graphicalProperties.line.width = 20000
    rc.series.append(zl)
    rc.legend = Legend()
    rc.legend.position = "b"
    ws2.add_chart(rc, "Y4")

    dv01 = DataValidation(type="whole", operator="between", formula1=0, formula2=1,
                          allow_blank=True)
    ws2.add_data_validation(dv01)
    dv01.add(f"B{f2}:C{l2}")

    # ================================================ 3. logistic scoring ===
    ws3 = wb.create_sheet("Logistic regression")
    W3 = 9
    title(ws3, "Logistic regression — the scoring sheet for a model fitted elsewhere",
          "Reopened yes/no is the outcome that actually matters, and a straight line "
          "is the wrong tool for it. Fit the model in Minitab, Python or R; paste the "
          "coefficient table here and this tab turns it into odds ratios, "
          "probabilities and a classification table you can argue about.", W3)
    widths(ws3, [46, 16, 13, 13, 13, 16, 16, 30, 16])

    band(ws3, 4, "READ THIS FIRST — what this tab does, and what it deliberately does not", W3)
    note_cell(ws3, 5, 1,
              "This sheet does NOT fit the model. Fitting a logistic regression is "
              "maximum likelihood: an iterative search that has no closed form, and "
              "no honest way to write it as a spreadsheet formula. Anything that "
              "claims to do it in plain Excel is either running a macro or lying to "
              "you. So the split is explicit — the tool fits, this sheet interprets. "
              "Everything below is arithmetic on numbers you paste in, and every step "
              "of that arithmetic is a formula you can click on.", W3)
    note_cell(ws3, 6, 1,
              "Where to get the numbers. Minitab: Stat → Regression → Binary Logistic "
              "Regression → Fit Binary Logistic Model, and tick odds ratios in the "
              "Results dialog. Python: m = smf.logit('reopened ~ after17 + over50 + "
              "tenure + transfers', data=df).fit(); m.summary(). R: glm(reopened ~ "
              "., family = binomial, data = df). Copy the coefficient column, the "
              "standard error column and the p-value column into the yellow cells "
              "below, in the same order as your predictors.", W3)

    band(ws3, 8, "STEP 1 — paste the coefficient table your tool produced", W3)
    hl = 9
    header(ws3, hl, ["Term", "Coefficient (log-odds)", "Std error", "p-value",
                     "Odds ratio", "Odds ratio 95% CI low", "Odds ratio 95% CI high",
                     "What it means"] + [""])
    ws3.row_dimensions[hl].height = 46
    lo_means = [
        '="Baseline odds of a reopen with every driver at zero: "&TEXT($E{r},"0.000")'
        '&" to 1, which is "&TEXT($E{r}/(1+$E{r}),"0.0%")&" probability."',
        '="Raising the adjustment after 17:00 multiplies the odds of a reopen by "'
        '&TEXT($E{r},"0.00")&". This is the posting-batch effect, and it is the '
        'largest single driver in the model."',
        '="Clearing the $50 authority limit multiplies the odds of a reopen by "'
        '&TEXT($E{r},"0.00")&", because the case waits in a queue for approval."',
        '="Each extra month of tenure multiplies the odds by "&TEXT($E{r},"0.000")'
        '&" — so a 24-month agent runs at "&TEXT($E{r}^24,"0.00")&" times the odds '
        'of a first-week agent."',
        '="Each extra transfer multiplies the odds of a reopen by "'
        '&TEXT($E{r},"0.00")&". Multiply by how common transfers are before you '
        'rank it against the others."',
    ]
    for i, (term, coef, se, pv) in enumerate(REG_LOGIT):
        r = hl + 1 + i
        ws3.cell(row=r, column=1, value=term).alignment = Alignment(
            wrap_text=True, vertical="center")
        for col, v, fmt in ((2, coef, "0.000"), (3, se, "0.000"), (4, pv, None)):
            c = mark(ws3, r, col, "ex" if i == 0 else "in")
            c.value = v
            c.alignment = Alignment(vertical="center")
            if fmt:
                c.number_format = fmt
        for col, f, fmt in (
            (5, f'=IFERROR(EXP($B{r}),"")', "0.000"),
            (6, f'=IFERROR(EXP($B{r}-1.96*$C{r}),"")', "0.000"),
            (7, f'=IFERROR(EXP($B{r}+1.96*$C{r}),"")', "0.000"),
            (8, lo_means[i].format(r=r), None),
        ):
            c = mark(ws3, r, col, "calc")
            c.value = f
            if fmt:
                c.number_format = fmt
            c.alignment = (Alignment(wrap_text=True, vertical="top") if col == 8
                           else Alignment(vertical="center"))
            if i == 0:
                c.fill = EX
        ws3.row_dimensions[r].height = 12 * (2 + len(lo_means[i]) // 34)
    note_cell(ws3, hl + 6, 1,
              "An odds ratio of 1.00 means the driver changes nothing; below 1 it "
              "protects. The confidence interval is the number to argue about, not "
              "the point estimate — if it straddles 1.00 the driver has not earned a "
              "place in your improvement plan, whatever the coefficient says. And a "
              "p-value printed as <0.0001 by the tool is pasted here as text on "
              "purpose: it is not a number you should be doing arithmetic on.", W3)

    band(ws3, hl + 8, "STEP 2 — the same 24 contacts, scored with those coefficients", W3)
    h3 = hl + 9
    header(ws3, h3, ["Contact ref"] + [x[0] for x in REG_X]
           + ["Reopened within 7 days (1 = yes)", "Logit", "Predicted probability",
              "Flagged at your threshold (1 = yes)"])
    ws3.row_dimensions[h3].height = 46
    f3, l3 = _reg_data_block(
        ws3, h3, [1, 2, 3, 4, 5, 6],
        [(r[0], r[1], r[2], r[3], r[4], r[6]) for r in REG_ROWS],
        [None, "0", "0", "#,##0", "0", "0"])
    # A blank row closes the data block off. Without it the block reads as
    # running on through the note and the band beneath it and into the two
    # settings below — and the finishing pass, which repaints filled-in cells
    # inside a table as worked example, would then tell the reader to delete
    # their own threshold. Those two are choices to adjust, not data to clear.
    thr = l3 + 4                              # the threshold input, set below
    b0 = hl + 1
    for i in range(len(REG_ROWS)):
        r = f3 + i
        for col, f, fmt in (
            (7, f'=IF($B{r}="","",$B${b0}+$B${b0 + 1}*$B{r}+$B${b0 + 2}*$C{r}'
                f'+$B${b0 + 3}*$D{r}+$B${b0 + 4}*$E{r})', "0.000"),
            (8, f'=IFERROR(1/(1+EXP(-$G{r})),"")', "0.0%"),
            (9, f'=IF($H{r}="","",IF($H{r}>=$B${thr},1,0))', "0"),
        ):
            c = mark(ws3, r, col, "calc")
            c.value, c.number_format = f, fmt
            if i == 0:
                c.fill = EX
    note_cell(ws3, l3 + 2, 1,
              "The logit is the straight-line part — intercept plus each coefficient "
              "times its column. The probability is that logit squashed into 0-1 by "
              "1/(1+EXP(−logit)), which is the whole reason this is not a linear "
              "regression: a line would happily predict a 130% chance of reopening.", W3)

    band(ws3, thr - 1, "STEP 3 — the threshold you choose, and what it costs you", W3)
    for r, (lab, val, fmt, hint) in enumerate([
        ("Flag a contact at this probability or above", 0.20, "0%",
         "Not 0.5. Roughly one billing adjustment in seven reopens, so at 0.5 the "
         "model flags nobody and still beats the base rate by predicting 'no' every "
         "time. Set the threshold from what a flag COSTS you: a quality check on a "
         "flagged contact takes about four minutes, so you can afford to be wrong "
         "often."),
        ("AUC the tool reported", 0.74, "0.00",
         "Straight off the tool's output — Minitab calls it the concordance or the "
         "area under the ROC curve. It is threshold-free: 0.5 is a coin toss, 0.7 "
         "is useful for targeting, 0.8 is strong for support data. It says nothing "
         "about whether the probabilities themselves are calibrated."),
    ], start=thr):
        ws3.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws3, r, 2, "in")
        c.value, c.number_format = val, fmt
        note_cell(ws3, r, 3, hint, W3)
        note(ws3, r, 2, hint)

    ct = thr + 3
    band(ws3, ct, "STEP 4 — the classification table at that threshold", W3)
    header(ws3, ct + 1, ["What the model said", "Actually reopened", "Did not reopen",
                         "Total"] + [""] * (W3 - 4))
    act, pred = f"$F${f3}:$F${l3}", f"$I${f3}:$I${l3}"
    for i, (lab, a) in enumerate([("Flagged — predicted reopen", 1),
                                  ("Not flagged", 0)]):
        r = ct + 2 + i
        ws3.cell(row=r, column=1, value=lab).font = F_B
        for col, want in ((2, 1), (3, 0)):
            c = mark(ws3, r, col, "calc")
            c.value = f"=SUMPRODUCT(({pred}={a})*({act}={want}))"
            c.number_format = "#,##0"
        c = mark(ws3, r, 4, "calc")
        c.value, c.number_format = f"=SUM($B{r}:$C{r})", "#,##0"
    r = ct + 4
    ws3.cell(row=r, column=1, value="Total").font = F_B
    for col in (2, 3, 4):
        c = mark(ws3, r, col, "calc")
        L = get_column_letter(col)
        c.value, c.number_format = f"=SUM(${L}${ct + 2}:${L}${ct + 3})", "#,##0"
    tp, fp = f"$B${ct + 2}", f"$C${ct + 2}"
    fn, tn = f"$B${ct + 3}", f"$C${ct + 3}"
    note_cell(ws3, ct + 5, 1,
              "Top left is a contact you flagged that did reopen; top right is a "
              "false alarm; bottom left is the one you missed. There is no threshold "
              "that empties both off-diagonal cells, and choosing which of them to "
              "fill is a business decision, not a statistical one.", W3)

    mb = ct + 7
    band(ws3, mb, "STEP 5 — what that table says about the model", W3)
    header(ws3, mb + 1, ["Measure", "Value", "What it tells you"] + [""] * (W3 - 3))
    for i, (lab, f, fmt, why) in enumerate([
        ("Sensitivity — reopens you caught", f'=IFERROR({tp}/({tp}+{fn}),"")', "0.0%",
         "Of the contacts that did reopen, the share the model flagged. This is the "
         "number the improvement is actually paid on: a reopen you never flagged is "
         "a reopen you never prevented."),
        ("Specificity — quiet contacts left alone",
         f'=IFERROR({tn}/({tn}+{fp}),"")', "0.0%",
         "Of the contacts that did not reopen, the share the model correctly left "
         "alone. Falls as you lower the threshold; that is the trade."),
        ("Precision — flags that were right", f'=IFERROR({tp}/({tp}+{fp}),"")', "0.0%",
         "Of the contacts you flagged, the share that really reopened. At a 14% base "
         "rate this is always the ugly number, and it is the one that decides "
         "whether the team trusts the flag."),
        ("Accuracy — rows called correctly",
         f'=IFERROR(({tp}+{tn})/({tp}+{fp}+{fn}+{tn}),"")', "0.0%",
         "The share of all rows called right. Almost useless on its own: predicting "
         "'never reopens' for every row scores whatever the base rate below leaves "
         "it, and prevents nothing. Quote it only beside that base rate."),
        ("Base rate — share that actually reopened",
         f'=IFERROR(({tp}+{fn})/({tp}+{fp}+{fn}+{tn}),"")', "0.0%",
         "What you would score by always guessing 'no'. Accuracy has to beat this "
         "by a wide margin before it means anything at all."),
        ("Verdict at this threshold", None, None,
         "Sensitivity is what the model buys you; precision is what it costs the "
         "team. Move the threshold in step 3 and watch both move."),
    ]):
        r = mb + 2 + i
        ws3.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws3, r, 2, "calc")
        c.value = f
        if fmt:
            c.number_format = fmt
        note_cell(ws3, r, 3, why, W3)
    ws3.cell(row=mb + 7, column=2).value = (
        f'=IF({tp}+{fn}=0,"NO REOPENS IN THE SAMPLE",'
        f'IF({tp}/({tp}+{fn})<0.5,"MISSES MORE THAN HALF — lower the threshold",'
        f'IF({tp}/({tp}+{fp})<0.2,"TOO MANY FALSE ALARMS — raise the threshold",'
        f'"USABLE FOR TARGETING — check it against the cost of a flag")))')

    lb = mb + 9
    band(ws3, lb, "THE DECISION RULE — before this leaves the room", W3)
    for i, txt in enumerate([
        "Never fit a straight line to a yes/no outcome. Linear regression on a 0/1 "
        "column predicts probabilities above 1 and below 0, and gets the error "
        "structure wrong on top of it. If your outcome is reopened, escalated, "
        "breached or churned, it is this tab or nothing.",
        "Report odds ratios with their confidence intervals, never raw log-odds. "
        "'Each transfer multiplies the odds of a reopen by 1.84 (95% CI 1.46 to "
        "2.33)' is a sentence the Improve phase can act on. '0.61' is not. And an "
        "interval that includes 1.00 means you have not shown anything, however "
        "confident the point estimate looks.",
        "The biggest odds ratio is not the biggest opportunity. Multiply each odds "
        "ratio by how COMMON that driver is. After-17:00 has the largest ratio here "
        "and applies to 46% of contacts; a rare driver with an odds ratio of 5 that "
        "touches 2% of volume moves almost nothing. Rank by ratio times prevalence "
        "and you will usually reorder the list.",
        "AUC and accuracy are not the same question, and neither is causation. AUC "
        "asks whether the model RANKS reopens above non-reopens; accuracy asks how "
        "many rows it called right at one threshold; and neither says the drivers "
        "cause the reopens. Before you change the 17:00 cut-off, confirm the "
        "posting-batch mechanism at the gemba or pilot the change against a control "
        "group (16-pilot-protocol.md) — the model chose your suspect, the pilot "
        "convicts it.",
    ]):
        note_cell(ws3, lb + 1 + i, 1, txt, W3)
    dvp = DataValidation(type="decimal", operator="between", formula1=0, formula2=1,
                         allow_blank=True)
    ws3.add_data_validation(dvp)
    dvp.add(f"B{thr}")
    dv01b = DataValidation(type="whole", operator="between", formula1=0, formula2=1,
                           allow_blank=True)
    ws3.add_data_validation(dv01b)
    dv01b.add(f"B{f3}:C{l3}")
    dv01b.add(f"F{f3}:F{l3}")

    howto(wb, LEGEND + [
        (True, "What this workbook is for"),
        (False, "Two of the tools in the library — linear/multiple regression and logistic "
                "regression — had nowhere to write anything down. This is that place. It "
                "carries one worked case all the way through: billing adjustments, a 7-day "
                "reopen rate of 14.2% against a target of 8%, 966 contacts in the baseline "
                "month, mean handle time 412 seconds, and four drivers the team had already "
                "named. The same 24 sampled contacts appear on all three tabs, so you can "
                "paste your own extract once and reuse the columns."),
        (True, "Which tab, and in what order"),
        (False, "Simple linear first, always — one driver against the outcome, plotted, so "
                "you see curvature, clusters and outliers before any coefficient hides them. "
                "Then Multiple regression, which fits all four drivers together and shows "
                "which of them the data cannot tell apart. Logistic regression last, and "
                "only when the outcome is yes/no."),
        (True, "What the logistic tab does not do"),
        (False, "It does not fit the model. Maximum likelihood is an iterative search with no "
                "closed form, so no plain spreadsheet formula can do it honestly. Fit it in "
                "Minitab, statsmodels or R, paste the coefficient table into the yellow "
                "cells, and this tab does the interpretation: odds ratios with confidence "
                "intervals, a predicted probability per contact, and a classification table "
                "at whatever threshold you choose."),
        (True, "How good is good enough, for support data"),
        (False, "Lower than a manufacturing course will tell you. A single driver explaining "
                "20-40% of a duration is a normal, useful result; an adjusted R² of 0.3 to "
                "0.6 across several drivers is a good model. Anything above 0.85 usually "
                "means a component of the outcome has crept in as a predictor. Judge a model "
                "on the size of the effects it hands you and on the residual plot, not on "
                "R²."),
        (True, "The trap that catches everybody"),
        (False, "A significant coefficient is not a coefficient worth acting on. Significance "
                "only says the effect is probably not zero. The 'Act on it?' column "
                "multiplies each coefficient by the range that driver actually covers in "
                "your data and compares it with the smallest effect you said would matter — "
                "and it refuses to say yes on a p-value alone."),
        (True, "And the one that ends projects"),
        (False, "Correlation here is not causation. No regression on observed data can tell "
                "you whether transfers lengthen contacts or whether hard contacts get "
                "transferred. Regression ranks your suspects. A designed experiment "
                "(24-doe-design-matrix.xlsx) or a controlled pilot (16-pilot-protocol.md) "
                "convicts one. Write the association down here, then go and test it."),
        (True, "Where the numbers go next"),
        (False, "The coefficients, their intervals and the residual verdict belong in "
                "14-root-cause-evidence-pack.md, next to the mechanism you observed. The "
                "driver you decide to act on becomes a candidate solution in "
                "15-solution-selection-matrix.xlsx."),
    ])
    return wb, "30-regression.xlsx"


# --------------------------------------------------------------- 31 multi-vari
#
# A balanced nested sample: two sites, two billing teams in each, three agents on
# each team, two consecutive contacts from each agent. Balance is not decoration
# — the mean squares below only carry the variance components they are supposed
# to when every cell holds the same number of observations, which is why the
# sheet refuses to let a reader quietly delete a row.
#
# The handle times are the same billing-adjustment case the rest of the pack
# runs on: grand mean 429 seconds against the 412 the baseline document quotes
# for the whole queue, because adjustments are the slow end of it.
MV_DATA = [
    ("Manchester", "Billing A", "R. Okonjo",  (312, 377)),
    ("Manchester", "Billing A", "P. Adeyemi", (478, 415)),
    ("Manchester", "Billing A", "S. Novak",   (356, 410)),
    ("Manchester", "Billing B", "T. Iqbal",   (527, 462)),
    ("Manchester", "Billing B", "M. Chen",    (341, 398)),
    ("Manchester", "Billing B", "L. Haddad",  (401, 459)),
    ("Kraków",     "Billing C", "J. Barros",  (438, 375)),
    ("Kraków",     "Billing C", "D. Wexler",  (552, 487)),
    ("Kraków",     "Billing C", "A. Ferreira", (329, 388)),
    ("Kraków",     "Billing D", "K. Nowak",   (495, 433)),
    ("Kraków",     "Billing D", "E. Duarte",  (368, 424)),
    ("Kraków",     "Billing D", "N. Osei",    (571, 505)),
]


def multi_vari():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Multi-vari study")
    W = 8
    title(ws, "Multi-vari study — which family of variation owns the problem",
          "Before you improve anything, find out where the variation lives: between sites, "
          "between teams, between agents, or inside one agent from one contact to the next. "
          "Improving the wrong family is the most expensive mistake in the Measure phase. "
          "The green block is a worked example — replace it with your own extract.", W)
    # Column A carries site names and the long statistic labels; B carries team
    # names, agent names and a column of sums of squares. Both are sized for the
    # longest thing that lands in them rather than for one block.
    widths(ws, [34, 20, 18, 15, 17, 14, 50, 12])

    band(ws, 4, "THE QUESTION — what you are partitioning, and how the contacts were drawn", W)
    MIN_ROW = 8
    for r, (lab, val, hint) in enumerate([
        ("What you are measuring",
         "Handle time in seconds, CRM open to CRM close",
         "One measurement, with its start and stop written down. A multi-vari study on "
         "'productivity' partitions an opinion."),
        ("The families you sampled, outer to inner",
         "Site → team → agent → contact",
         "Each family has to sit INSIDE the one before it. Agents belong to one team, teams "
         "to one site. If a family cuts across the others — shift, say, or contact type — it "
         "is not nested and this sheet will mis-attribute it."),
        ("How the contacts were drawn",
         "Two consecutive contacts per agent, same week",
         "Consecutive and close together on purpose: the innermost family is meant to capture "
         "contact-to-contact variation, not the drift between March and June."),
        ("Smallest difference worth chasing (seconds)",
         30,
         "The gap you would actually change something for. Set it before you look at the "
         "answer, or you will set it to whatever the answer turned out to be."),
    ], start=5):
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "in")
        c.value = val
        if r == MIN_ROW:
            c.number_format = "#,##0"
        note_cell(ws, r, 3, hint, W)
        note(ws, r, 2, hint)

    band(ws, 10, "THE DATA — one row per contact, nested: site, then team, then agent", W)
    HDR = 11
    header(ws, HDR, ["Site", "Team", "Agent", "Contact ref", "Handle time (seconds)",
                     "", "", ""])
    FIRST = HDR + 1
    r = FIRST
    for site, team, agent, obs in MV_DATA:
        for o in obs:
            for col, v in ((1, site), (2, team), (3, agent),
                           (4, f"BA-{3000 + r - FIRST + 1}"), (5, o)):
                c = mark(ws, r, col, "ex")
                c.value = v
                if col == 5:
                    c.number_format = "#,##0"
            r += 1
    LAST = r - 1
    note_cell(ws, LAST + 1, 1,
              "Twenty-four contacts: 2 sites × 2 teams × 3 agents × 2 contacts each. Keep it "
              "balanced. The arithmetic below splits the variation using the mean square of "
              "each family, and a mean square only means what it is supposed to when every "
              "cell holds the same number of contacts. Add an agent and you add two rows, not "
              "one; drop a contact and the partition below is no longer a partition.", W)

    # ------------------------------------------------ the averages, bottom up
    AB = LAST + 3
    band(ws, AB, "STEP 1 — the averages, from the bottom up", W)
    AH = AB + 1
    header(ws, AH, ["Level", "Name", "Mean handle time (s)", "Contacts behind it",
                    "Gap to its parent (s)", "Spread inside (s)", "", ""])
    A0 = AH + 1                       # first agent row
    T0 = A0 + 12                      # first team row
    S0 = T0 + 4                       # first site row
    GR = S0 + 2                       # the grand mean
    for i, (site, team, agent, _) in enumerate(MV_DATA):
        d1, d2 = FIRST + 2 * i, FIRST + 2 * i + 1
        ws.cell(row=A0 + i, column=1, value="Agent")
        ws.cell(row=A0 + i, column=2, value=f"{agent} · {team}")
        for col, f, fmt in (
            (3, f"=AVERAGE(E{d1}:E{d2})", "#,##0"),
            (4, f"=COUNT(E{d1}:E{d2})", "0"),
            (5, f"=C{A0 + i}-C{T0 + i // 3}", "+#,##0;−#,##0;0"),
            # Signed, because the sum of squares below squares it. The reader
            # sees how far apart one agent's two contacts were; the partition
            # needs the same number with its sign kept.
            (6, f"=E{d1}-E{d2}", "+#,##0;−#,##0;0"),
        ):
            c = mark(ws, A0 + i, col, "calc")
            c.value, c.number_format = f, fmt
    for j, (site, team) in enumerate([(MV_DATA[k * 3][0], MV_DATA[k * 3][1])
                                      for k in range(4)]):
        d1, d2 = FIRST + 6 * j, FIRST + 6 * j + 5
        ws.cell(row=T0 + j, column=1, value="Team")
        ws.cell(row=T0 + j, column=2, value=f"{team} · {site}")
        for col, f, fmt in ((3, f"=AVERAGE(E{d1}:E{d2})", "#,##0"),
                            (4, f"=COUNT(E{d1}:E{d2})", "0"),
                            (5, f"=C{T0 + j}-C{S0 + j // 2}", "+#,##0;−#,##0;0")):
            c = mark(ws, T0 + j, col, "calc")
            c.value, c.number_format = f, fmt
    for k, site in enumerate([MV_DATA[0][0], MV_DATA[6][0]]):
        d1, d2 = FIRST + 12 * k, FIRST + 12 * k + 11
        ws.cell(row=S0 + k, column=1, value="Site")
        ws.cell(row=S0 + k, column=2, value=site)
        for col, f, fmt in ((3, f"=AVERAGE(E{d1}:E{d2})", "#,##0"),
                            (4, f"=COUNT(E{d1}:E{d2})", "0"),
                            (5, f"=C{S0 + k}-C{GR}", "+#,##0;−#,##0;0")):
            c = mark(ws, S0 + k, col, "calc")
            c.value, c.number_format = f, fmt
    ws.cell(row=GR, column=1, value="All contacts").font = F_B
    ws.cell(row=GR, column=2, value="Grand mean").font = F_B
    for col, f, fmt in ((3, f"=AVERAGE(E{FIRST}:E{LAST})", "#,##0"),
                        (4, f"=COUNT(E{FIRST}:E{LAST})", "0")):
        c = mark(ws, GR, col, "calc")
        c.value, c.number_format = f, fmt
    note_cell(ws, GR + 1, 1,
              "Read the GAP column, not the means. Every gap is measured against the level "
              "that contains it — an agent against their own team, a team against its own "
              "site — so the four families never claim the same second twice. SPREAD INSIDE "
              "is the difference between one agent's two contacts, and it is the only column "
              "here that owes nothing to who or where they are.", W)

    # ------------------------------------------------------ the partition
    VB = GR + 3
    band(ws, VB, "STEP 2 — which family owns the variation", W)
    VH = VB + 1
    header(ws, VH, ["Family", "Sum of squares", "Degrees of freedom", "Mean square",
                    "Variance component", "Share of total", "What this family means", ""])
    ws.row_dimensions[VH].height = 42
    V0 = VH + 1                                             # site row
    VT, VA, VE, VTOT = V0 + 1, V0 + 2, V0 + 3, V0 + 4
    for i, (fam, ss, df, comp, why) in enumerate([
        ("Between sites", f"=12*SUMSQ(E{S0}:E{S0 + 1})", 1,
         f"=MAX(0,(D{V0}-D{VT})/12)",
         "Manchester against Kraków, once the teams and agents inside them are accounted "
         "for. A big share here means the two sites are running different processes."),
        ("Between teams, within a site", f"=6*SUMSQ(E{T0}:E{T0 + 3})", 2,
         f"=MAX(0,(D{VT}-D{VA})/6)",
         "One billing team against another in the same building — different team leader, "
         "different coaching, same everything else."),
        ("Between agents, within a team", f"=2*SUMSQ(E{A0}:E{A0 + 11})", 8,
         f"=MAX(0,(D{VA}-D{VE})/2)",
         "One agent against their own team-mates. A big share here is a training, tooling "
         "or method problem, and it is the family a site-level project cannot touch."),
        ("Within one agent, contact to contact", f"=SUMSQ(F{A0}:F{A0 + 11})/2", 12,
         f"=D{VE}",
         "The same agent, two contacts, one week. Whatever is left after everything you can "
         "name — contact mix, system waits, luck. It sets the floor on what any of the "
         "families above can be measured against."),
    ]):
        r = V0 + i
        ws.cell(row=r, column=1, value=fam).font = F_B
        for col, v, fmt in ((2, ss, "#,##0.0"), (3, df, "0"), (4, f"=B{r}/C{r}", "#,##0.0"),
                            (5, comp, "#,##0.0"),
                            (6, f"=IFERROR(E{r}/SUM($E${V0}:$E${VE}),\"\")", "0.0%")):
            c = mark(ws, r, col, "calc")
            c.value, c.number_format = v, fmt
        note_cell(ws, r, 7, why, W)
    ws.cell(row=VTOT, column=1, value="TOTAL").font = F_B
    for col, v, fmt in ((2, f"=DEVSQ(E{FIRST}:E{LAST})", "#,##0.0"), (3, 23, "0"),
                        (6, f"=IFERROR(SUM(F{V0}:F{VE}),\"\")", "0.0%")):
        c = mark(ws, VTOT, col, "calc")
        c.value, c.number_format = v, fmt
    note_cell(ws, VTOT, 7,
              "The four sums of squares add up to this one exactly. If they ever do not, the "
              "design has stopped being balanced.", W)
    note_cell(ws, VTOT + 1, 1,
              "A VARIANCE COMPONENT can come out negative, and one has here: the teams "
              "component is showing zero because the arithmetic produced a number below zero "
              "and there is no such thing as negative variance. It does not mean the teams "
              "are identical. It means this sample cannot tell teams apart from the agents "
              "sitting inside them — the agent-to-agent spread is wide enough to swallow any "
              "team effect. Report it as zero, say why, and if the team question matters, go "
              "back with more agents per team rather than more contacts per agent.", W)
    note_cell(ws, VTOT + 2, 1,
              "SUM OF SQUARES says how much of the raw spread each family accounts for; the "
              "VARIANCE COMPONENT is what is left after removing the variation the families "
              "underneath already explain, and it is the one to quote. They rank differently "
              "on purpose. Share of total is computed on the components.", W)

    # ----------------------------------------------------------- the verdict
    DB = VTOT + 4
    band(ws, DB, "STEP 3 — what it tells you to do next", W)
    for i, (lab, f, fmt, hint) in enumerate([
        ("Look here first",
         f'=IFERROR(INDEX($A${V0}:$A${VE},MATCH(MAX($F${V0}:$F${VE}),$F${V0}:$F${VE},0)),"")',
         None,
         "The family holding the largest share. Scope your project to this family — an "
         "improvement aimed anywhere else is competing with variation it cannot reach."),
        ("Biggest gap between two agents (s)",
         f"=IFERROR(MAX(C{A0}:C{A0 + 11})-MIN(C{A0}:C{A0 + 11}),\"\")", "#,##0",
         "Fastest agent against slowest, in seconds of handle time. This is the number to "
         "take to the process owner, because it is one they can picture."),
        ("Is that gap worth chasing?",
         f'=IF(B{DB + 2}="","",IF(B{DB + 2}>=$B${MIN_ROW},'
         f'"YES — larger than the difference you said would matter",'
         f'"NO — smaller than the difference you said would matter"))', None,
         "Measured against the threshold you set at the top, before you saw the data."),
        ("Contacts per agent needed to halve the noise",
         f"=IFERROR(ROUNDUP(4*E{VE}/(($B${MIN_ROW}/2)^2),0),\"\")", "#,##0",
         "If the answer above is too close to call, this is roughly how many contacts per "
         "agent the next study needs to resolve half your threshold. It grows with the "
         "square, which is why precision gets expensive quickly."),
    ]):
        r = DB + 1 + i
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "calc")
        c.value = f
        if fmt:
            c.number_format = fmt
        note_cell(ws, r, 3, hint, W)

    bar(ws, "Where the variation lives — share of the total",
        Reference(ws, min_col=1, min_row=V0, max_row=VE),
        Reference(ws, min_col=6, min_row=V0, max_row=VE),
        "J4", horizontal=True, pct=True, name="Share of total variation",
        colours=["6B4FA0", "3F8F5A", "C0392B", "B45309"])
    agch = bar(ws, "Mean handle time by agent — the multi-vari picture",
               Reference(ws, min_col=2, min_row=A0, max_row=A0 + 11),
               Reference(ws, min_col=3, min_row=A0, max_row=A0 + 11),
               "J20", name="Mean handle time (s)")
    # Left to itself Excel starts this axis just under the fastest agent, and a
    # 344-against-538 gap then draws as though one agent were four times the
    # other. A bar says "this much, measured from nothing" whether or not the
    # axis agrees, and handle time genuinely starts at nothing — so the bound is
    # a fact about the measurement rather than a framing of this example, which
    # is the only kind of axis this pack fixes.
    agch.y_axis.scaling.min = 0

    dv = DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E{FIRST}:E{LAST}")
    dv.add(f"B{MIN_ROW}")

    howto(wb, LEGEND + [
        (True, "What this workbook is for"),
        (False, "A multi-vari study is the cheapest thing you can run in Measure, and the tool "
                "library had nowhere to record one. It answers a single question before you "
                "spend anything: of all the variation in your metric, how much sits between "
                "sites, how much between teams, how much between agents, and how much inside "
                "one agent from one contact to the next. You cannot improve a family you have "
                "not found, and every hour spent improving the wrong one is gone."),
        (True, "Why the design has to be balanced"),
        (False, "Two sites, two teams in each, three agents on each team, two contacts from "
                "each agent. The same number everywhere. The partition works by comparing "
                "mean squares, and a mean square carries the variance components it is "
                "supposed to only when the cells are equal in size. An unbalanced extract "
                "still produces numbers; they are just not the numbers on the label."),
        (True, "Nested, not crossed"),
        (False, "Each family must sit inside the one before it. Agent R. Okonjo belongs to "
                "Billing A and Billing A belongs to Manchester, so 'agent' is nested in "
                "'team' and 'team' in 'site'. Shift and contact type are NOT nested — every "
                "shift happens at every site — and putting one of them in this sheet would "
                "hand its variation to whichever column you happened to type it in. For a "
                "crossed factor, use 24-doe-design-matrix.xlsx or stratify and run "
                "13-hypothesis-test-log.xlsx."),
        (True, "What the worked example found"),
        (False, "Sixty-three per cent of the variation is between agents on the same team, and "
                "another thirty-two is inside one agent from one contact to the next. Sites "
                "hold four and a half per cent, teams none that this sample can see. Read "
                "that as an instruction: the fastest agent averages 344 seconds and the "
                "slowest 538, on the same queue, at the same site, under the same team "
                "leader. Whatever the difference between those two people is, it is worth "
                "more than any site-level programme in the plan."),
        (True, "The trap"),
        (False, "A family with a large share is where the variation IS, not where the CAUSE "
                "is. 'Between agents' does not mean the agents are the problem — in the "
                "worked case it turned out to be which of them had been shown the deferred-"
                "close rule. The study tells you which door to open. What is behind it is "
                "still a root cause investigation."),
        (True, "Where the numbers go next"),
        (False, "The family shares and the agent-to-agent gap belong in "
                "14-root-cause-evidence-pack.md as the evidence for where you scoped. If the "
                "answer was 'within one agent', the next step is usually "
                "29-msa-gage-rr.xlsx — measurement error lands in exactly that family and "
                "looks identical to it. If it was 'between agents', 30-regression.xlsx will "
                "tell you which of their measurable differences tracks the gap."),
    ])
    return wb, "31-multi-vari.xlsx"


# -------------------------------------------------------------- 32 system hop
#
# Six billing adjustments, observed at the desk, every move between systems
# timed. Four of the six run the same four-hop loop; one detours through the
# knowledge base; the seconds differ because real observations do.
CRM, BILL, PAY, KB = ("Case tool", "Billing platform",
                      "Payment gateway", "Knowledge base")
HOP_LOOP = [(CRM, BILL, "Find the adjustment on the account", "Account number"),
            (BILL, PAY, "Check whether the refund has actually posted", "Transaction ID"),
            (PAY, BILL, "Record the posting reference against the adjustment",
             "Posting reference"),
            (BILL, CRM, "Write the case note and close", "Adjustment amount")]
HOP_ROWS = [
    ("BA-2201", [(HOP_LOOP[0], 34), (HOP_LOOP[1], 41), (HOP_LOOP[2], 22), (HOP_LOOP[3], 27)]),
    ("BA-2202", [(HOP_LOOP[0], 31), (HOP_LOOP[1], 38), (HOP_LOOP[2], 19), (HOP_LOOP[3], 24)]),
    ("BA-2203", [((CRM, KB, "Check whether this adjustment type needs a manager", ""), 46),
                 ((KB, CRM, "Back to the case", ""), 11),
                 (HOP_LOOP[0], 36), (HOP_LOOP[3], 29)]),
    ("BA-2204", [(HOP_LOOP[0], 33), (HOP_LOOP[1], 44), (HOP_LOOP[2], 21), (HOP_LOOP[3], 26)]),
    ("BA-2205", [(HOP_LOOP[0], 29), (HOP_LOOP[1], 52), (HOP_LOOP[2], 24), (HOP_LOOP[3], 31)]),
    ("BA-2206", [(HOP_LOOP[0], 35), (HOP_LOOP[1], 39), (HOP_LOOP[2], 20), (HOP_LOOP[3], 25)]),
]
HOP_PAIRS = [(CRM, BILL), (BILL, PAY), (PAY, BILL), (BILL, CRM), (CRM, KB), (KB, CRM)]


def system_hop():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Hop log")
    W = 8
    title(ws, "System-hop log — what the swivel chair costs",
          "One row per move between systems, timed at the desk. Handle time hides this: the "
          "contact takes seven minutes and nobody can say how much of it was the agent "
          "waiting for a second screen to load and typing the account number into it again. "
          "The green block is a worked example — replace it with your own observations.", W)
    widths(ws, [30, 20, 22, 22, 34, 12, 14, 26])

    band(ws, 4, "THE QUESTION — what you watched, and what counts", W)
    Q0 = 5
    R_AHT, R_VOL, R_COST = Q0 + 2, Q0 + 3, Q0 + 4
    for r, (lab, val, fmt, hint) in enumerate([
        ("Process observed", "Billing adjustment, end to end", None,
         "One contact type. Hop counts are not comparable across contact types, and averaging "
         "them together produces a number that describes nothing anybody does."),
        ("What counts as a hop", "Any move to a different application window", None,
         "Write the rule down before you start counting, because the borderline cases decide "
         "the answer. Tabs inside the same application: your call — but make the same call "
         "every time, and record which way you called it."),
        ("Mean handle time for this contact type (seconds)", 412, "#,##0",
         "From the platform, over the same period you observed. It is the denominator for "
         "everything below, so take it from a report rather than from the six contacts you "
         "happened to watch."),
        ("Contacts of this type per year", 11592, "#,##0",
         "The population the hops apply to. Adjustments here, not the whole billing queue — "
         "annualising against a population you did not observe is how a swivel-chair saving "
         "becomes a number Finance rejects."),
        ("Loaded cost per agent hour", 38, '"$"#,##0.00',
         "Salary, employer costs, benefits and paid non-productive time, divided by paid "
         "hours. Your Finance partner has this figure; do not build one."),
    ], start=Q0):
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "in")
        c.value = val
        if fmt:
            c.number_format = fmt
        note_cell(ws, r, 3, hint, W)
        note(ws, r, 2, hint)

    LB = R_COST + 2
    band(ws, LB, "THE LOG — one row per move, in the order the agent made them", W)
    LH = LB + 1
    header(ws, LH, ["Contact ref", "Step", "From system", "To system",
                    "Why the agent moved", "Seconds", "Data re-keyed?",
                    "Which fields were re-keyed"])
    LF = LH + 1
    r = LF
    for ref, hops in HOP_ROWS:
        for step, ((frm, to, why, fields), secs) in enumerate(hops, start=1):
            for col, v in ((1, ref), (2, step), (3, frm), (4, to), (5, why),
                           (6, secs), (7, "Y" if fields else "N"), (8, fields or "—")):
                c = mark(ws, r, col, "ex")
                c.value = v
                if col == 6:
                    c.number_format = "#,##0"
            r += 1
    LL = r - 1
    note_cell(ws, LL + 1, 1,
              "Time the hop, not the task. The seconds here are the ones between leaving one "
              "screen and being able to work in the next — the load, the second login, the "
              "search that finds the account again. The work the agent does once they arrive "
              "is not a hop and belongs in the handle time above. DATA RE-KEYED means the "
              "agent typed something into the second system that the first one already knew, "
              "which is both the waste and the defect: every re-key is a chance to mistype an "
              "account number.", W)

    # ------------------------------------------------------ contact by contact
    CB = LL + 3
    band(ws, CB, "STEP 1 — how much of each contact went on moving between systems", W)
    CH = CB + 1
    header(ws, CH, ["Level", "Contact ref", "Seconds in hops", "Share of handle time",
                    "Hops", "Hops that re-keyed data", "", ""])
    C0 = CH + 1
    for i, (ref, _) in enumerate(HOP_ROWS):
        r = C0 + i
        ws.cell(row=r, column=1, value="Contact")
        c = mark(ws, r, 2, "ex")
        c.value = ref
        for col, f, fmt in (
            (3, f"=SUMIF($A${LF}:$A${LL},$B{r},$F${LF}:$F${LL})", "#,##0"),
            (4, f'=IFERROR(C{r}/$B${R_AHT},"")', "0.0%"),
            (5, f"=COUNTIF($A${LF}:$A${LL},$B{r})", "0"),
            (6, f'=COUNTIFS($A${LF}:$A${LL},$B{r},$G${LF}:$G${LL},"Y")', "0"),
        ):
            c = mark(ws, r, col, "calc")
            c.value, c.number_format = f, fmt
    CL = C0 + len(HOP_ROWS) - 1
    note_cell(ws, CL + 1, 1,
              "Look at the spread before you quote the average. If one contact carries most of "
              "the hopping you have found an exception to investigate, not a process to "
              "redesign — and a business case built on the mean would be wrong in the "
              "direction that gets it withdrawn.", W)

    # -------------------------------------------------------------- the totals
    TB = CL + 3
    band(ws, TB, "STEP 2 — what the hopping costs across a year", W)
    T0 = TB + 1
    R_N, R_HOPS, R_HPC, R_SPC = T0, T0 + 1, T0 + 2, T0 + 3
    R_SHARE, R_REKEY, R_HRS, R_CASH = T0 + 4, T0 + 5, T0 + 6, T0 + 7
    for r, (lab, f, fmt, hint) in enumerate([
        ("Contacts observed", f"=COUNTA($B${C0}:$B${CL})", "0",
         "Six is enough to size a problem and nowhere near enough to price one. Treat "
         "everything below as an order of magnitude until you have watched thirty."),
        ("Hops logged", f"=SUM($E${C0}:$E${CL})", "0",
         "Every move, across every contact you watched."),
        ("Hops per contact", f'=IFERROR(B{R_HOPS}/B{R_N},"")', "0.0",
         "The number people remember. It is also the one that makes a process owner "
         "uncomfortable enough to look at the log."),
        ("Seconds per contact spent hopping", f'=IFERROR(SUM($C${C0}:$C${CL})/B{R_N},"")', "0",
         "Total hop time divided by contacts observed."),
        ("Share of handle time", f'=IFERROR(B{R_SPC}/B{R_AHT},"")', "0.0%",
         "Against the platform's mean handle time above, not against the contacts you "
         "watched — an observed agent works faster, and the share would flatter you."),
        ("Share of hops that re-key data", f'=IFERROR(SUM($F${C0}:$F${CL})/B{R_HOPS},"")',
         "0.0%",
         "The part of the waste that is also a defect risk. A hop that carries data across by "
         "hand can carry it across wrong."),
        ("Agent hours a year spent hopping",
         f'=IFERROR(B{R_SPC}*B{R_VOL}/3600,"")', "#,##0",
         "Seconds per contact times contacts a year. This is capacity, not cash."),
        ("What that time costs a year", f'=IFERROR(B{R_HRS}*B{R_COST},"")', '"$"#,##0',
         "Hours at the loaded rate. It becomes a saving only when something is harvested — "
         "01-project-charter.md will not accept it without a named mechanism."),
    ], start=T0):
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "calc")
        c.value, c.number_format = f, fmt
        note_cell(ws, r, 3, hint, W)
    note_cell(ws, R_CASH + 1, 1,
              "The tool library says to expect 15-30% of handle time here. Treat that as a "
              "rule of thumb to test, not a published benchmark to be judged against: it is "
              "the range the method usually turns up, and it is not evidence about your "
              "queue. What decides whether there is a project is the cost on the line above "
              "and how many of these hops somebody who knows the systems believes are "
              "avoidable. A queue at a fifth of handle time where every hop is unavoidable "
              "has nothing in it; one at eight per cent where half the hops exist only "
              "because two systems were never joined up has a great deal.", W)

    # --------------------------------------------------------- which hop hurts
    PB = R_CASH + 3
    band(ws, PB, "STEP 3 — which hop to attack", W)
    PH = PB + 1
    header(ws, PH, ["From system", "To system", "The pair", "Hops", "Total seconds",
                    "Seconds per contact", "Share of hop time", "Any re-keying?"])
    ws.row_dimensions[PH].height = 30
    P0 = PH + 1
    for i, (frm, to) in enumerate(HOP_PAIRS):
        r = P0 + i
        for col, v in ((1, frm), (2, to)):
            c = mark(ws, r, col, "ex")
            c.value = v
        for col, f, fmt in (
            # The pair is spelled out in one cell because a chart needs a single
            # label per bar, and a reader deserves to see the same thing the
            # chart is showing rather than a column that only makes sense to it.
            (3, f'=IF($A{r}="","",$A{r}&" → "&$B{r})', None),
            (4, f"=COUNTIFS($C${LF}:$C${LL},$A{r},$D${LF}:$D${LL},$B{r})", "0"),
            (5, f"=SUMIFS($F${LF}:$F${LL},$C${LF}:$C${LL},$A{r},$D${LF}:$D${LL},$B{r})",
             "#,##0"),
            (6, f'=IFERROR(E{r}/$B${R_N},"")', "0.0"),
            (7, f'=IFERROR(E{r}/SUM($E${P0}:$E${P0 + len(HOP_PAIRS) - 1}),"")', "0.0%"),
            (8, f'=IF(COUNTIFS($C${LF}:$C${LL},$A{r},$D${LF}:$D${LL},$B{r},'
                f'$G${LF}:$G${LL},"Y")>0,"yes","no")', None),
        ):
            c = mark(ws, r, col, "calc")
            c.value = f
            if fmt:
                c.number_format = fmt
    PL = P0 + len(HOP_PAIRS) - 1
    note_cell(ws, PL + 1, 1,
              "One row per direction, because they are different problems: leaving the case "
              "tool for the billing platform is a lookup, and coming back is a write-up. Add "
              "a pair by inserting a row inside this block — the formulas above and below it "
              "will follow. Sort the block by total seconds if you want the chart ranked; it "
              "plots the order you typed.", W)

    VB = PL + 3
    band(ws, VB, "STEP 4 — the verdict", W)
    for i, (lab, f, fmt, hint) in enumerate([
        ("The pair costing the most time",
         f'=IFERROR(INDEX($C${P0}:$C${PL},MATCH(MAX($E${P0}:$E${PL}),$E${P0}:$E${PL},0)),"")',
         None,
         "Where an integration, a deep link or a copied field would buy the most. It is not "
         "always the slowest single hop — a fast hop made on every contact beats a slow one "
         "made on a tenth of them."),
        ("What that one pair costs a year",
         f'=IFERROR(MAX($E${P0}:$E${PL})/B{R_N}*B{R_VOL}/3600*B{R_COST},"")', '"$"#,##0',
         "Seconds per contact on that pair, annualised at the loaded rate. Take this to the "
         "system owner, not the contact-centre manager — it is their backlog it belongs in."),
        ("Does that pair re-key data?",
         f'=IFERROR(INDEX($H${P0}:$H${PL},MATCH(MAX($E${P0}:$E${PL}),$E${P0}:$E${PL},0)),"")',
         None,
         "If yes, the case is stronger than the time alone: you are also arguing about "
         "accuracy, and a mistyped account number costs far more than the forty seconds it "
         "took to type."),
    ]):
        r = VB + 1 + i
        ws.cell(row=r, column=1, value=lab).font = F_B
        c = mark(ws, r, 2, "calc")
        c.value = f
        if fmt:
            c.number_format = fmt
        note_cell(ws, r, 3, hint, W)
    note_cell(ws, VB + 4, 1,
              "Removing a hop does not return all of its seconds. Some of that time is the "
              "agent thinking, and thinking moves to whatever is on the screen next. Halve the "
              "figure above before you put it in a charter, say in the charter that you "
              "halved it, and then prove the real number in a pilot (16-pilot-protocol.md) "
              "against a control group.", W)

    pch = bar(ws, "Where the hop time goes — total seconds by system pair",
              Reference(ws, min_col=3, min_row=P0, max_row=PL),
              Reference(ws, min_col=5, min_row=P0, max_row=PL),
              "J4", horizontal=True, name="Total seconds observed")
    pch.width = 19          # a pair name is two system names and an arrow
    cch = bar(ws, "Seconds spent hopping, contact by contact",
              Reference(ws, min_col=2, min_row=C0, max_row=CL),
              Reference(ws, min_col=3, min_row=C0, max_row=CL),
              "J20", name="Seconds in system hops")
    # Same reason as the multi-vari agent chart: left alone the axis starts just
    # under the quickest contact, and 112 seconds against 136 draws as a sliver
    # against a tower. Seconds start at zero whatever the data says.
    cch.y_axis.scaling.min = 0

    dvs = DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=True)
    ws.add_data_validation(dvs)
    dvs.add(f"F{LF}:F{LL}")
    dvyn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dvyn)
    dvyn.add(f"G{LF}:G{LL}")

    howto(wb, LEGEND + [
        (True, "What this workbook is for"),
        (False, "The tool library tells you to map swivel-chair waste with a system-hop "
                "diagram and then gave you nowhere to record one. This is that place. It "
                "turns 'the agents are always jumping between screens', which nobody can act "
                "on, into a number of seconds per contact, a cost per year, and a named pair "
                "of systems with an owner."),
        (True, "How to collect it"),
        (False, "Sit beside the agent with a stopwatch and this sheet, and watch whole "
                "contacts end to end — never a sampled minute. Log every move as it happens; "
                "reconstructing it afterwards produces the loop people believe they run "
                "rather than the one they ran. Watch at least ten contacts and more than one "
                "agent: the multi-vari study in 31-multi-vari.xlsx exists because agent-to-"
                "agent variation is usually the largest family there is, and hopping is no "
                "different."),
        (True, "Tell the agent what you are doing"),
        (False, "Say plainly that you are timing the systems and not the person, and show "
                "them the finished sheet. An observation that feels like an audit produces a "
                "performance, and you will have carefully timed the performance."),
        (True, "What the worked example found"),
        (False, "Six billing adjustments, 24 hops, about 123 seconds a contact — near enough "
                "thirty per cent of a 412-second handle time — and 22 of the 24 hops re-typed "
                "something the previous system already held. The heaviest pair is the billing "
                "platform to the payment gateway, which is a confirmation the agent should "
                "never have had to go and fetch."),
        (True, "The trap"),
        (False, "Hop time is not free time, and it is not a saving. Removing every hop would "
                "not hand you the hours back: some of that time is the agent thinking, and "
                "thinking follows them to the next screen. Take the figure, halve it, say out "
                "loud that you halved it, and let a pilot tell you the real number. A "
                "swivel-chair business case that claims the full amount is the classic way to "
                "lose a Finance partner for the rest of the programme."),
        (True, "Where the numbers go next"),
        (False, "The hop time per contact is a wait in 10-value-stream-map.xlsx and drags "
                "process cycle efficiency down with it. Every re-keyed field is a failure mode "
                "for 12-fmea.xlsx, with a detection score that is usually terrible because "
                "nothing checks a hand-typed account number. The named pair of systems becomes "
                "a candidate in 15-solution-selection-matrix.xlsx, where an integration will "
                "score badly on effort and well on durability — which is exactly the argument "
                "worth having."),
    ])
    return wb, "32-system-hop.xlsx"


BUILDERS = [five_whys, fishbone, stakeholder, kano, doe, pareto, flow, control_charts,
            erlang, gage_rr, regression, multi_vari, system_hop]


def main() -> int:
    previews = {}
    for fn in BUILDERS:
        wb, name = fn()
        polish_workbook(wb)
        path = TEMPLATES / name
        save_workbook(wb, path)
        # regenerate from the saved file so the preview describes what ships,
        # with every formula's real result rather than a value remembered by
        # hand in SHOWN — which only ever covered the cells someone thought to
        # add to it, and left the rest as empty blue boxes.
        saved = load_workbook(path)
        shown = dict(SHOWN)
        try:
            shown.update(shown_from(saved, chartsvg.values_for(path)))
        except Exception as exc:                                 # noqa: BLE001
            print(f"    (no recalculation: {exc}; falling back to SHOWN)")
        previews[name] = workbook_html(saved, shown)
        print(f"  built {name:34s} {path.stat().st_size:>7,} bytes  "
              f"{len(previews[name]):>7,} chars of preview")
    (ROOT / "tools" / "previews.json").write_text(json.dumps(previews), encoding="utf-8")
    print(f"\n  {len(previews)} templates written, previews cached for sync_html.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
