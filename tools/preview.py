#!/usr/bin/env python3
"""Generate a template's preview HTML from the workbook itself.

The original previews were hand-authored, which is why they could drift from the
files they describe. Anything built by tools/build_templates.py gets its preview
generated here instead, from the same workbook the user downloads.

openpyxl stores no cached formula results, so a formula cell needs a display
value supplied in `shown` — {(sheet, "B4"): "28,800"}. Everything else is read
straight off the cell and formatted from its number format.
"""
from __future__ import annotations

import html as H

from openpyxl.utils import get_column_letter, range_boundaries

# A recalculation can legitimately produce these — NA() is how a chart is told
# to leave a gap — and printing "#N/A" into a preview cell helps nobody.
ERRORS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

# Fill colour -> preview class. These are the same colours the legend explains.
FILL_CLASS = {
    "FF151B24": "x-title",
    "FFEEF1F6": "x-band",
    "FF333C49": "x-hdr",
    "FFFFF9E3": "x-fill",
    "FFF2F7FF": "x-calc",
    "FFECFAEF": "x-ex",
    "FFFDECEC": "x-bad",
}


def fmt_value(v, number_format: str | None) -> str:
    """Render a value the way a cell with this number format would show it."""
    if v is None:
        return ""
    if isinstance(v, str):
        return "" if v.strip() in ERRORS else v
    nf = (number_format or "").lower()
    try:
        if "%" in nf:
            digits = 1 if ".0" in nf else 0
            return f"{float(v) * 100:.{digits}f}%"
        if "$" in nf or "£" in nf:
            return "$" + f"{float(v):,.0f}" if "0.00" not in nf else "$" + f"{float(v):,.2f}"
        if isinstance(v, float):
            if "0.00" in nf:
                return f"{v:,.2f}"
            if "#,##0" in nf:
                return f"{v:,.0f}"
            return f"{v:g}"
        if isinstance(v, int):
            return f"{v:,}" if "#,##0" in nf else str(v)
    except (TypeError, ValueError):
        pass
    return str(v)


def _fmt(cell):
    """Render a cell value the way the sheet would show it."""
    return fmt_value(cell.value, cell.number_format)


def shown_from(wb, cells: dict) -> dict:
    """{(sheet, coord): display text} for every formula cell, from real results.

    The preview leaves a formula cell blank unless something supplies its
    displayed value, and that something used to be a dict maintained by hand.
    Anything added after the fact showed as an empty blue box with a little
    "fx" on it — a declared column with nothing in it, which is the exact
    defect the EXAMPLE layer fails a workbook for. These come from a real
    recalculation instead, so a new formula cannot ship blank.
    """
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    out[(ws.title, c.coordinate)] = fmt_value(
                        cells.get((ws.title, c.row, c.column)), c.number_format)
    return out


def _edge_style(cell) -> str:
    """Diagonals and heavy rules, which is how the fishbone is actually drawn.

    The Ishikawa diagram's bones are a merged block carrying a single diagonal
    border, and its spine is a thick bottom rule. Without these the preview
    renders the diagram as a grid of empty boxes and the shape disappears.
    """
    b = getattr(cell, "border", None)
    if b is None:
        return ""
    bits = []
    diag = getattr(b, "diagonal", None)
    if diag is not None and diag.style:
        line = "#333c49"
        if getattr(b, "diagonalDown", False):
            corner = "to bottom right"
        elif getattr(b, "diagonalUp", False):
            corner = "to top right"
        else:
            corner = None
        if corner:
            bits.append(
                "background-image:linear-gradient(%s,transparent calc(50%% - 1px),"
                "%s calc(50%% - 1px),%s calc(50%% + 1px),transparent calc(50%% + 1px))"
                % (corner, line, line))
    bottom = getattr(b, "bottom", None)
    if bottom is not None and bottom.style in ("thick", "medium"):
        bits.append("border-bottom:%dpx solid #151b24" % (3 if bottom.style == "thick" else 2))
    return ";".join(bits)


def scaffold_from(ws) -> int | None:
    """First column of the chart-feeding scaffolding, if the sheet has any.

    Every builder parks its helper columns to the right of the content: the
    marker series a chart needs, a sorted copy of a block, the noise floor. In
    Excel nobody sees them — they are off the edge and you scroll past. The
    preview laid them straight into the table beside the data, so a reader met a
    wall of unheaded numbers (36.55, 77.27, 0, repeated down every row) with
    nothing anywhere saying what they were.

    No guidance check could have caught it: every one of them inspects columns
    that have a header, and having no header is exactly what makes these a
    problem. The precondition of the check was the signature of the defect.

    So they come out of the preview. Anything to the right of the last headed
    column that carries no header of its own is scaffolding, and the reader is
    told it exists rather than shown it raw.
    """
    headed = 0
    for row in ws.iter_rows():
        for c in row:
            try:
                dark = (c.fill and c.fill.patternType
                        and str(c.fill.fgColor.rgb) == "FF333C49")
            except Exception:                                    # noqa: BLE001
                dark = False
            if dark and c.value not in (None, ""):
                headed = max(headed, c.column)
    if not headed or headed >= ws.max_column:
        return None
    has = any(c.value not in (None, "")
              for row in ws.iter_rows(min_col=headed + 1) for c in row)
    return headed + 1 if has else None


def sheet_html(ws, shown: dict) -> str:
    """One sheet as the preview's table markup."""
    hide_from = scaffold_from(ws)
    # merged ranges: the anchor spans, the rest are skipped. Both directions —
    # emitting colspan alone silently drops every vertically merged block and
    # shifts the rest of the grid left.
    spans, covered = {}, set()
    for rng in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(str(rng))
        spans[(r1, c1)] = (c2 - c1 + 1, r2 - r1 + 1)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    covered.add((r, c))

    max_r, max_c = ws.max_row, ws.max_column
    if hide_from:
        max_c = hide_from - 1
    out = ['<table class="xgrid">']
    for r in range(1, max_r + 1):
        cells = []
        c = 1
        while c <= max_c:
            if (r, c) in covered:
                c += 1
                continue
            cell = ws.cell(row=r, column=c)
            span, rspan = spans.get((r, c), (1, 1))

            cls = ""
            try:
                if cell.fill and cell.fill.patternType and cell.fill.fgColor.rgb:
                    cls = FILL_CLASS.get(str(cell.fill.fgColor.rgb), "")
            except Exception:
                pass
            # A section band is styled as a short upper-case label. Once it also
            # carries the plain-English key to the columns below it, that
            # styling turns a full sentence into a shout — the explanation
            # becomes harder to read than the jargon it explains.
            if cls == "x-band" and isinstance(cell.value, str) and len(cell.value) > 80:
                cls = "x-band x-bandlong"
            if not cls and cell.font and cell.font.bold:
                cls = "x-b"

            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if is_formula:
                cls = (cls + " x-f").strip() if cls else "x-calc x-f"
                text = shown.get((ws.title, cell.coordinate), "")
                title = ' title="%s"' % H.escape(cell.value, quote=True)
            else:
                text = H.escape(_fmt(cell), quote=True)
                title = ""

            attrs = ""
            if span > 1:
                attrs += ' colspan="%d"' % span
            if rspan > 1:
                attrs += ' rowspan="%d"' % rspan
            if cls:
                attrs += ' class="%s"' % cls
            edge = _edge_style(cell)
            if edge:
                attrs += ' style="%s"' % H.escape(edge, quote=True)
            attrs += title
            cells.append("<td%s>%s</td>" % (attrs, text))
            c += span
        out.append("<tr>" + "".join(cells) + "</tr>")
    out.append("</table>")
    if hide_from:
        out.append(
            '<p class="xscaf">Columns %s onward are hidden here: they are working '
            'cells that feed this sheet&rsquo;s chart &mdash; sorted copies, marker '
            'series, control-limit constants &mdash; not anything you fill in. They '
            'are all in the .xlsx if you want them.</p>'
            % get_column_letter(hide_from))
    return "".join(out)


def workbook_html(wb, shown: dict | None = None) -> str:
    """Tab strip plus one table per sheet, matching the existing previews."""
    shown = shown or {}
    tabs = "".join(
        '<button class="xtab%s" data-xs="%d">%s</button>'
        % (" on" if i == 0 else "", i, H.escape(ws.title))
        for i, ws in enumerate(wb.worksheets)
    )
    sheets = "".join(
        '<div class="xsheet%s" data-xsheet="%d">%s</div>'
        % (" on" if i == 0 else "", i, sheet_html(ws, shown))
        for i, ws in enumerate(wb.worksheets)
    )
    return '<div class="xtabs">%s</div>%s' % (tabs, sheets)
