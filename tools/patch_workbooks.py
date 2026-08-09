#!/usr/bin/env python3
"""Canonical formula set for the Excel templates.

This file is the single source of truth for every *calculated* cell in
templates/*.xlsx. Running it is idempotent: it rewrites each listed cell to the
formula below, leaving all formatting, validation and conditional formatting
untouched.

Why this exists: each workbook is mirrored in three other places (the base64
copy inside the HTML, the preview table's formula tooltips, and docs/index.html).
Editing a formula by hand in a spreadsheet is how those four copies drifted apart.
Change a formula here, then run tools/sync_html.py.

    python3 tools/patch_workbooks.py
    python3 tools/sync_html.py
    python3 tools/verify.py

Requires: openpyxl
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
from chart_specs import add_charts  # noqa: E402
from worked_examples import add_examples  # noqa: E402
from xlpolish import polish_workbook, save_workbook  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "templates"

# Built by tools/build_templates.py, which owns them end to end. Read from the
# previews that builder writes rather than guessed from the file name: the guess
# was `^2[0-9]-`, which silently stopped covering the pack the moment a
# thirtieth template existed, and re-saving a generated workbook here is
# destructive — openpyxl's reader keeps only the first chart type in a combo.
# The regex stays as the fallback for a checkout with no previews.json.
_BUILT_RE = re.compile(r"^(2[0-9]|[3-9][0-9])-")


def _is_built(name: str) -> bool:
    if name in _BUILT_NAMES:
        return True
    return bool(_BUILT_RE.match(name))


def _built_names() -> set:
    import json
    p = Path(__file__).resolve().parent / "previews.json"
    try:
        return set(json.loads(p.read_text(encoding="utf-8")))
    except Exception:                                            # noqa: BLE001
        return set()


_BUILT_NAMES = _built_names()

# Fill colours used across the workbooks; the preview renderer keys off these.
YELLOW_INPUT = "FFFFF9E3"

# ---------------------------------------------------------------------------
# Formula patches: (workbook, sheet, cell, formula)
# Ranges are expressed with {r} and expanded over `rows`.
# ---------------------------------------------------------------------------

SINGLE: list[tuple[str, str, str, str]] = [
    # -- 19 calculators, sheet 1: guard every division and the two NORMSINV
    #    calls. Zero defects makes NORMSINV(1) a #NUM!, and a perfect process
    #    is a plausible input.
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "B9", '=IFERROR(B7/B5,"")'),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "B11", '=IFERROR(B7/(B5*B6),"")'),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "B12", '=IFERROR(B7/(B5*B6)*1000000,"")'),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "B13", '=IFERROR(EXP(-B7/B5),"")'),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "B14", '=IFERROR(NORMSINV(1-B7/(B5*B6)),"")'),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "B15", '=IFERROR(NORMSINV(1-B7/(B5*B6))+1.5,"")'),

    # -- 19 calculators, sheet 2: the verdict read B12 (chance agreement) instead
    #    of B13 (kappa). The shipped example is the one case where both land in
    #    the same band, which is why it survived review.
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "B11", '=IFERROR((B5+B8)/SUM(B5:B8),"")'),
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "B12",
     '=IFERROR(((B5+B6)/SUM(B5:B8))*((B5+B7)/SUM(B5:B8))+((B7+B8)/SUM(B5:B8))*((B6+B8)/SUM(B5:B8)),"")'),
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "B13",
     '=IFERROR((B11-B12)/(1-B12),"")'),
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "B14",
     # The bands the pack runs on, from 07-msa-attribute-agreement.md: good
     # above 0.80, marginal 0.60-0.80, unacceptable below 0.60. This carried
     # the looser AIAG-style 0.75/0.40 cuts, so the workbook called a kappa of
     # 0.41 MARGINAL where the template halts all use of the data — and the
     # same split sat in the curriculum table, the tool card, the on-page
     # calculator, the glossary and the worked example.
     '=IF(B13="","",IF(B13>0.9,"EXCELLENT — fit for purpose",'
     'IF(B13>0.8,"GOOD — usable, fix the weakest rubric items",'
     'IF(B13>0.6,"MARGINAL — aggregate analysis only, never individual performance management",'
     '"UNACCEPTABLE — halt all use of this data until fixed"))))'),

    # -- 19 calculators, sheet 3: the verdict read B15, which sits inside the
    #    merged note A15:D15 and is therefore always empty -> every input
    #    returned "NOT CAPABLE". Ppu lives in B13.
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "B11", '=IFERROR((B5-B6)/B7,"")'),
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "B12", '=IFERROR(1-NORMSDIST((B5-B6)/B7),"")'),
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "B13", '=IFERROR((B5-B6)/(3*B7),"")'),
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "B17",
     '=IF(B13="","",IF(B13>=1.33,"CAPABLE",IF(B13>=1,"MARGINAL — customers notice",'
     '"NOT CAPABLE — you are producing breaches")))'),

    # -- 19 calculators, sheets 4-7: divide-by-zero guards.
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "B9", '=IFERROR(B5/B6,"")'),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "B10", '=IFERROR(B5/B6*24,"")'),
    ("19-black-belt-calculators.xlsx", "5 Process efficiency", "B9", '=IFERROR(B6*60-B5,"")'),
    ("19-black-belt-calculators.xlsx", "5 Process efficiency", "B10", '=IFERROR(B5/(B6*60),"")'),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "B11", '=IFERROR(B5*B6/3600/B7,"")'),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "B12",
     '=IFERROR(B5*B6/3600/B7/(1-B8)-B5*B6/3600/B7,"")'),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "B13", '=IFERROR(B5*B6/3600/B7/(1-B8),"")'),
    # Occupancy converts handle time into AVAILABLE time, not paid time. This
    # chain stopped there and called the result "Paid hours freed", pricing the
    # saving as though nobody were paid for a break, a coaching session or a
    # holiday. Shrinkage is the step that was missing, and leaving it out
    # understates the benefit by 1/(1-shrinkage) — about 47% at 32%. The same
    # stop-at-occupancy error was in the page's formula card and in the
    # business-case wizard; all three carry the full chain now.
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "A10", "Shrinkage"),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "B10", 0.32),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "A12", "On-phone hours freed"),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "B12", '=IFERROR(B5*B6/3600/B8,"")'),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "A13", "Paid hours freed"),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "B13",
     '=IFERROR(B5*B6/3600/B8/(1-B10),"")'),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "B14",
     '=IFERROR(B5*B6/3600/B8/(1-B10)*B7,"")'),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "B15",
     '=IFERROR(B5*B6/3600/B8/(1-B10)*B7*B9,"")'),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "A16", "People-equivalent (FTE)"),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "B16",
     '=IFERROR(B5*B6/3600/B8/(1-B10)/1760,"")'),

    # -- 19 calculators, sheet 9: NPV hardcoded years 1-3 while Simple ROI used
    #    B6*B7 for any B7, so the two disagreed above three years; year 1 was
    #    also unguarded, so B7=0 still banked a full year. The closed-form
    #    annuity is correct for any number of years and matches the HTML card,
    #    which loops y=1..yr.
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "B10", '=IF($B$7>=1,B6/(1+B8)^1,0)'),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "B11", '=IF($B$7>=2,B6/(1+B8)^2,0)'),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "B12", '=IF($B$7>=3,B6/(1+B8)^3,0)'),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "B14", '=IFERROR((B6*B7-B5)/B5,"")'),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "B15",
     '=IFERROR(-B5+IF(B8=0,B6*B7,B6*(1-(1+B8)^-B7)/B8),"")'),

    # -- 10 VSM: the column header says "% of lead time" but the denominator was
    #    E32 (total waiting time), overstating every share. Lead time is E33.
    #    E33/E34/E35 now reference the subtotals instead of re-summing.
    ("10-value-stream-map.xlsx", "Value stream", "E33", "=E31+E32"),
    ("10-value-stream-map.xlsx", "Value stream", "E34", "=(E31+E32)/60"),
    ("10-value-stream-map.xlsx", "Value stream", "E35", '=IFERROR(E31/E33,"")'),
    # PRODUCT() over an empty range returns 0, not an error, so IFERROR never
    # fired and a blank template reported RTY = 0%.
    ("10-value-stream-map.xlsx", "Value stream", "E36", '=IF(COUNT(G10:G28)=0,"",PRODUCT(G10:G28))'),

    # -- 12 FMEA: risk reduction divided the mean of the remediated rows by the
    #    mean of *all* scored rows. AVERAGEIF restricts both sides to the rows
    #    that actually have an after-score.
    ("12-fmea.xlsx", "FMEA", "D44",
     '=IFERROR(1-AVERAGE(R11:R35)/AVERAGEIF(R11:R35,">0",J11:J35),"")'),

    # -- 15 solution selection: the old form ran VALUE() over the whole range, so
    #    one pasted free-text value made the entire summary #VALUE!. Wrapping it
    #    in IFERROR does not help - IFERROR is not reliably element-wise inside
    #    SUMPRODUCT. Comparing the leading character as text cannot raise an
    #    error at all, and the trailing space keeps a blank cell from matching.
    ("15-solution-selection-matrix.xlsx", "Solution selection", "E37",
     '=SUMPRODUCT((M14:M32="Yes")*((LEFT(L14:L32&" ",1)="1")+(LEFT(L14:L32&" ",1)="2")'
     '+(LEFT(L14:L32&" ",1)="3")))'),
    ("15-solution-selection-matrix.xlsx", "Solution selection", "E38",
     '=SUMPRODUCT((M14:M32="Yes")*((LEFT(L14:L32&" ",1)="5")+(LEFT(L14:L32&" ",1)="6")))'),

    # -- 17 control plan: same treatment, and E32 now references E30/E29 rather
    #    than duplicating the whole SUMPRODUCT.
    ("17-control-plan.xlsx", "Control plan", "E30",
     '=SUMPRODUCT((A10:A27<>"")*((LEFT(P10:P27&" ",1)="1")+(LEFT(P10:P27&" ",1)="2")'
     '+(LEFT(P10:P27&" ",1)="3")))'),
    ("17-control-plan.xlsx", "Control plan", "E31",
     '=SUMPRODUCT((A10:A27<>"")*((LEFT(P10:P27&" ",1)="5")+(LEFT(P10:P27&" ",1)="6")))'),
    ("17-control-plan.xlsx", "Control plan", "E32", '=IFERROR(E30/E29,"")'),

    # -- 05 data collection: a target rate at or above 100% produced a negative
    #    variance term and a confidently wrong sample size. Blank it instead.
    ("05-data-collection-plan.xlsx", "Sample size calculator", "C13",
     '=IF(OR(C5<=0,C5>=1,C6<=0,C5+C6>=1,C7<=0,C7>=1,C8<=0,C8>=1),"",'
     'ROUNDUP((NORMSINV(1-C7/2)+NORMSINV(C8))^2*(C5*(1-C5)+(C5+C6)*(1-C5-C6))/C6^2,0))'),
    ("05-data-collection-plan.xlsx", "Sample size calculator", "C14", '=IF(C13="","",2*C13)'),
    ("05-data-collection-plan.xlsx", "Sample size calculator", "C11", '=IFERROR(NORMSINV(1-C7/2),"")'),
    ("05-data-collection-plan.xlsx", "Sample size calculator", "C12", '=IFERROR(NORMSINV(C8),"")'),
]

# Repeated row formulas: (workbook, sheet, cell_template, formula_template, rows)
REPEATED: list[tuple[str, str, str, str, range]] = [
    # -- 13 hypothesis log: the blank-guard checked the p-value and the effect
    #    size but not the practical threshold. A blank threshold fell through
    #    VALUE("") -> IFERROR -> 0, so |effect| >= 0 was always true and every
    #    significant result was auto-labelled "real and matters" - the exact
    #    discipline the template exists to enforce. Alpha now reads $B$8 instead
    #    of being hardcoded to 0.05.
    ("13-hypothesis-test-log.xlsx", "Test log", "N{r}",
     '=IF(OR(J{r}="",K{r}="",M{r}=""),"",IF(J{r}>$B$8,"NOT SIGNIFICANT",'
     'IF(ABS(IFERROR(VALUE(SUBSTITUTE(SUBSTITUTE(K{r},"pts",""),"%","")),0))'
     '>=ABS(IFERROR(VALUE(SUBSTITUTE(SUBSTITUTE(M{r},"pts",""),"%","")),0)),'
     '"YES — real and matters","NO — significant but too small")))',
     range(11, 34)),

    # -- 10 VSM waiting states: "% of lead time" must divide by lead time (E33),
    #    not total waiting time (E32).
    ("10-value-stream-map.xlsx", "Value stream", "C{r}", '=IFERROR(B{r}/$E$33,"")', range(44, 50)),
]

# Literal value / label corrections: (workbook, sheet, cell, value)
VALUES: list[tuple[str, str, str, object]] = [
    # The matrix ranked this project's causes against five CTQs that appear in
    # no CTQ tree — and the rule it breaks is printed on its own How-to sheet:
    # "Weights come from the CTQ tree, which came from customers. If the team
    # invents the weights in the room, you have built an opinion aggregator."
    # One of the five, "Not asked to repeat myself", paraphrases the tree row
    # 03-voc-ctq-tree.md forbids weighting, because it would put the queue's
    # 266,000 contacts back into a matrix ranking causes for 11,592.
    #
    # Re-headed on the three the tree actually weights, and re-scored, because
    # the columns no longer mean what they meant. The scores also now use the
    # 9/6/3/1 scale the How-to sheet asks for — the old ones used 8, 7, 5, 4
    # and 2, which are not on it.
    # Says why two of the five columns are empty, so the blank is read as a
    # decision rather than an omission — and answers the question the
    # unlabelled input at G10 would otherwise leave open.
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "A12",
     "Columns 4 and 5 are spare, and this example leaves them empty on purpose. "
     "The CTQ tree behind this project weights three requirements, so three are "
     "filled in. A fourth column with no tree row behind it has no weight anyone "
     "can defend, and the handle-time row the tree explicitly refuses to weight is "
     "the one most teams would have put there."),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "C10", "7-day reopen rate, in-scope billing adjustments (OD-BIL-004-ADJ)"),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "C11", 10),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "D10", "Share of in-scope adjustments confirmed posted before the case is closed"),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "D11", 9),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "E10", "Share of contacts with a commitment date logged"),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "E11", 7),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F10", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F11", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G10", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G11", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F15", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G15", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F16", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G16", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F17", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G17", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F18", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G18", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F19", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G19", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F20", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G20", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F21", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G21", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "F22", None),
    ("11-cause-effect-xy-matrix.xlsx", "X-Y matrix", "G22", None),
    # The collection plan gave the baseline window as a calendar quarter while
    # 09-baseline-document.md gives it as 2026-01-05 to 2026-03-29, "12 whole
    # weeks", and the charter says in terms that the 61,400 is "over a 12-week
    # window at ~5,100/week ... not one quarter". The dates decide it: that
    # span is 84 days, Monday to Sunday, and 61,400 over 12 weeks is 5,117 a
    # week and annualises to 266,067 — the pack's 266,000. Read as the 90-day
    # quarter this cell stated, the same 61,400 annualises to 249,011, which
    # appears nowhere.
    ("05-data-collection-plan.xlsx", "Collection plan", "G10", "2026-01-05 to 2026-03-29"),
    # The calculator's own D8 note tells the reader that pricing a reopen at
    # cost-to-serve is "the defect its charter records", and that the rate and
    # the volume must be measured on the SAME population — then B5 and B8
    # committed both errors, on the sheet the note is attached to. 480,000 is
    # the whole billing queue; 14.2% and 8.0% are measured on in-scope
    # adjustments. Same numbers as the charter now: 11,592 x 6.2 points at
    # $38.60, realised at 0.85.
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "B5", 11592),
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "B8", 38.6),
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D5",
     "Units per year in the population the rates below are measured on — not the "
     "wider queue it sits inside. The worked example uses 11,592 in-scope billing "
     "adjustments, not the 266,000 contacts of the queue around them."),
    # A column heading that went through an HTML escaper on its way into the
    # workbook and never came back out. Excel shows the entity, and so do the
    # preview and the Word export, because all three read the cell. The phrase
    # check in qa_wordtables unescapes before comparing, so it structurally
    # could not see this; a sweep of every cell in all 22 workbooks for a raw
    # entity found this one and no others.
    ("10-value-stream-map.xlsx", "Value stream", "G9", "% complete & accurate"),
    # The worked example used the bare number 2, which is not a member of the
    # cell's own dropdown list ("2 Design it out"). 17-control-plan already uses
    # the string form.
    ("15-solution-selection-matrix.xlsx", "Solution selection", "L14", "2 Design it out"),
    # Say out loud that the ratio is computed on remediated rows only.
    ("12-fmea.xlsx", "FMEA", "A44", "Risk reduction (on remediated rows, shipped only)"),
    # The sheet stated no rule for when the New columns may be filled, and its
    # own worked example used two: one row re-scored a control that was built
    # but not live, another left an unshipped row at its original scores. The
    # first put 19.6 points into the headline, 57.7% reading as 77.3%.
    ("12-fmea.xlsx", "FMEA", "E44",
     "Average New RPN against average original RPN, over the rows that have a "
     "New score. RE-SCORE ONLY WHAT HAS SHIPPED. A control that is built, "
     "signed off or awaiting release has not changed the failure rate a customer "
     "meets, so leave the row at its original scores until it is live — an FMEA "
     "that books the improvement at design time reports a risk reduction the "
     "process has not had. Re-score on evidence, the way the benefit is "
     "validated at closure."),
    ("13-hypothesis-test-log.xlsx", "Test log", "A8", "Significance level (alpha)"),
    ("13-hypothesis-test-log.xlsx", "Test log", "B8", 0.05),
    ("13-hypothesis-test-log.xlsx", "Test log", "C8",
     "Used by the 'Above threshold?' column. 0.05 is standard; drop to 0.01 when acting on "
     "the result is expensive. WHICH TEST GOES IN THE TEST COLUMN — the log never named one, "
     "and eight tools in the library send you here. Two means, normal-ish: two-sample t-test. "
     "Two means, skewed (most support duration data): Mann-Whitney U. Before and after on the "
     "SAME units: paired t-test, or Wilcoxon signed-rank if skewed. Three or more groups: "
     "one-way ANOVA, or Kruskal-Wallis if skewed. Rates or proportions: one- and two-proportion "
     "tests. Two categorical variables: chi-square test of independence. Comparing spread "
     "rather than centre: Levene's test. Check the shape first on the 'Shape and spread' tab "
     "of 25-pareto-and-distribution.xlsx — it decides which half of that list you are in."),
    # The ROI tab's worked example carried a realised benefit of $156,672, and
    # that number reconciles with nothing. The same workbook computes a realised
    # benefit two tabs earlier — tab 8 lands on $172,012.80 from its own inputs
    # — so the one workbook stated two answers to its own question, $15,340.80
    # apart, with nothing to say which was meant. B6 is and stays a yellow input:
    # a generic calculator must let you type a benefit that came from anywhere.
    # What it may not do is ship a worked example that contradicts the tab next
    # to it, so the example is now tab 8's own output and the note says so.
    # Sheet 9 carries sheet 8's realised benefit as a literal, so it moves
    # with it: 11,592 x 6.2 points x $38.60 x 0.85.
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "B6", 23580.68),
    # And say what the numbers below it now show. On the corrected benefit this
    # example does not pay back inside the model, which is the honest result
    # for the programme's worked project and worth more as an example than a
    # comfortable one would be.
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D6",
     "Validated and realised — not gross, not the pilot rate. Signed by Finance after "
     "90 days of control data. The worked example carries tab 8's own realised benefit "
     "so this workbook agrees with itself; replace it with yours, from whichever tab or "
     "source produced it. Note what it produces below: against $193,000 of investment "
     "this project never pays back inside three years, and the NPV stays negative. That "
     "is the correct answer for it — its charter records that it does not clear the "
     "$50,000 Finance floor — and a worked example that came out comfortable would be "
     "teaching you the wrong thing."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D6",
     "Validated and realised — not gross, not the pilot rate. Signed by Finance after 90 days "
     "of control data. The worked example carries tab 8's own realised benefit so this "
     "workbook agrees with itself; replace it with yours, from whichever tab or source "
     "produced it."),
    # The cost basis is where this programme's own worked project went wrong, and
    # this tab was quietly instructing the same mistake: D6 invites you to reduce
    # a REOPEN rate, and D8 priced whatever you removed at the cost of an average
    # contact. A reopen is not an average contact — it carries the investigation
    # and the redo — and the two rates are 5.7x apart in the pack's own example.
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D8",
     "From Finance: total support cost divided by total contacts, with an agreed definition "
     "of 'total support cost'. CAREFUL when what you are removing is a REOPEN, an escalation "
     "or a repeat rather than a first contact: cost to serve an average contact is the wrong "
     "price for rework, because rework carries the investigation and the redo on top of the "
     "handling, and Finance usually books it far higher. This programme's worked project uses "
     "$6.80 to serve a contact and $38.60 for a reopened one — 5.7x apart, and pricing reopens "
     "at $6.80 is the defect its charter records. Ask Finance which of the two applies to the "
     "thing you are actually removing, and make sure the rate above and the volume in B5 are "
     "measured on the SAME population: a rate from one and a volume from a wider one is how a "
     "benefit case comes out ten times its arithmetic maximum."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "A17",
     "Under 18 months payback is generally an easy sell. Net present value is calculated across all the years "
     "you model; the three rows above show the first three. And set year-one expectations honestly: a Black Belt "
     "completes only one project while learning, and a first cohort picks the least well-selected projects."),
    ("05-data-collection-plan.xlsx", "Sample size calculator", "A16",
     "Notice how fast this grows as the effect you want to detect gets smaller — halving the detectable effect "
     "roughly quadruples the sample. The sample size stays blank if the inputs are impossible, for example when "
     "the current rate plus the change you want to detect reaches 100%."),
]


# ---------------------------------------------------------------------------
# Explanations for standalone label:value rows: (workbook, sheet, cell, text)
#
# On the web page every acronym is clickable. In a downloaded workbook nothing
# is, so a row that reads "DPO | 0.0507" and nothing else told the reader
# neither what it is nor what to do about it. These are the missing right-hand
# notes: for a yellow input, where the number comes from and the trap; for a
# blue calculated cell, what it means and how to read it — never how it is
# computed, because the formula is already visible.
#
# The cell is named explicitly rather than assumed to be column C, because each
# sheet already has a note column and the note belongs in it: column D on the
# calculator tabs, which is 72 wide and where every input note already sits.
# Column C on those tabs is 18 wide, and on "9 ROI and payback" it holds the
# breakeven series the chart plots — writing there would have broken the chart.
# ---------------------------------------------------------------------------

NOTES: list[tuple[str, str, str, str]] = [
    # -- 17 control plan: the durability bar chart's six categories. The plan
    #    lives or dies on which level its controls sit at, and the levels were
    #    named but never defined anywhere in the workbook.
    ("17-control-plan.xlsx", "Control plan", "C39",
     "Strongest of all: the failure cannot happen because the step is gone. Retiring the manual "
     "adjustment queue counts here; a better checklist does not."),
    ("17-control-plan.xlsx", "Control plan", "C40",
     "The system will not permit the failure — a ticket that cannot be closed while an adjustment "
     "is still in flight. Survives attrition, busy weeks and everyone forgetting."),
    ("17-control-plan.xlsx", "Control plan", "C41",
     "Mistake-proofing: the wrong action is blocked or corrected as it is attempted. Still durable, "
     "but it can be switched off, so name who owns the setting."),
    ("17-control-plan.xlsx", "Control plan", "C42",
     "The failure still happens and something catches it without anyone looking. Useful, but you "
     "are now managing defects rather than preventing them."),
    ("17-control-plan.xlsx", "Control plan", "C43",
     "Depends on a person remembering. Decays with every new hire and every busy week — count "
     "these honestly rather than relying on them."),
    ("17-control-plan.xlsx", "Control plan", "C44",
     "A briefing, a poster, a line in the wiki. Half-life measured in weeks. If most of your "
     "controls sit here, the process will drift back after handover."),

    # -- 19 calculators, tab 1: five different names for the same defect count.
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D9",
     "DPU = defects per unit: how many things go wrong on the average contact. Above 1.0 the "
     "typical contact is carrying more than one defect."),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D10",
     "The denominator everything below is measured against. Change the opportunities-per-unit "
     "figure and every rate on this tab moves with it, which is why you fix it once and leave it."),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D11",
     "DPO = defects per opportunity: of all the chances to get something wrong, the share that "
     "actually went wrong. This is what the sigma scale is built on."),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D12",
     "DPMO = defects per million opportunities, the same figure scaled up so it stays readable. "
     "This is the one to quote — published benchmarks are all in DPMO."),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D13",
     "The share of contacts that get through with no defect at all — a customer's chance of a "
     "clean experience, which is usually a more sobering number than the sigma level."),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D14",
     "Your process restated in standard deviations, with no shift added. Use this one when you "
     "compare yourself against your own history."),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D15",
     "The headline number, on the convention the industry quotes. It is 1.5 higher than the Z "
     "above by construction, so say which of the two you are reporting and never mix them."),
    ("19-black-belt-calculators.xlsx", "1 Sigma level", "D25",
     "Where you land on the published ladder. Read the band you are actually in, not the one the "
     "rounded sigma level lets you claim."),

    # -- 19 calculators, tab 2: kappa. Row 14 (Verdict) is skipped: B14 is
    #    merged across B:D, so there is no free cell on that row.
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "D10",
     "The size of the study. Below about 50 items kappa bounces around too much to act on, and "
     "below 30 it is not worth calculating."),
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "D11",
     "Raw agreement — the share of items the two scored the same way. It flatters you badly: two "
     "analysts who pass everything agree 95% of the time while measuring nothing."),
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "D12",
     "What the two would have matched on by luck alone, given how often each says pass. The higher "
     "your pass rate the higher this is, which is exactly why raw agreement misleads."),
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "D13",
     "Agreement with the luck stripped out: 0 is no better than guessing, 1 is perfect. This is the "
     "number to report, and it is always lower than the raw figure above."),
    ("19-black-belt-calculators.xlsx", "2 QA agreement (kappa)", "D20",
     "Your kappa against the bars everyone uses. Below 0.6 the scores are not a measurement system "
     "yet — fix the rubric and re-calibrate before anything is built on them."),

    # -- 19 calculators, tab 3: capability. Row 17 (Verdict) is skipped: B17 is
    #    merged across B:D.
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "D9",
     "SLA = service level agreement. How far the average sits below what you promised, in hours. "
     "Negative means the average contact already breaches and no amount of tuning will save it."),
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "D10",
     "The width of your process, in the units of your promise. Compare it with the headroom above: "
     "if it is the larger of the two, you are manufacturing breaches."),
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "D11",
     "Your headroom counted in standard deviations. Around 3 or more is a process that breaches "
     "rarely; around 1 is one that breaches most weeks."),
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "D12",
     "The breach rate this model expects. Treat it as a floor — support durations have a long "
     "right tail, so the real rate is normally higher than this, sometimes much higher."),
    ("19-black-belt-calculators.xlsx", "3 SLA capability", "D13",
     "Ppu = one-sided process performance: headroom divided by spread. 1.33 is the usual bar for "
     "'capable'; below 1.0 customers notice. It says nothing about whether the promise was sane."),

    # -- 19 calculators, tab 4: Little's Law.
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D9",
     "How long the average customer waits, given the backlog you hold and the rate you close at. "
     "Headcount and effort are not in this calculation, and cannot be argued into it."),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D10",
     "The same wait in hours, which is usually the unit your SLA and your customers speak in."),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D11",
     "The most open work you can hold and still deliver the target wait. This is your work-in-"
     "progress limit — enforce it in the tool, because a limit in a document is a suggestion."),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D12",
     "The one-off clear-down that has to happen before the cap can hold. Plan it as a separate "
     "push with a start and an end date, or the limit gets abandoned in week one."),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D17",
     "What customers experience today, drawn against the promise below."),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D18",
     "The wait you want to commit to. The distance between the two bars is backlog, not effort."),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D21",
     "Today's open work against the ceiling below. Everything above the ceiling is wait the "
     "customer feels and the team cannot work off by trying harder."),
    ("19-black-belt-calculators.xlsx", "4 Backlog and lead time", "D22",
     "The ceiling that delivers your target. Hold more than this and the promise fails as "
     "arithmetic, whatever the team does."),

    # -- 19 calculators, tab 5: PCE.
    ("19-black-belt-calculators.xlsx", "5 Process efficiency", "D8",
     "The customer's entire wait, restated in the same unit as the work so the two can be compared "
     "honestly. Everything here that is not touch time is queue."),
    ("19-black-belt-calculators.xlsx", "5 Process efficiency", "D9",
     "Queue time: the part of the wait when nobody was working on the issue. In support this is "
     "nearly all of it, and it is where the improvement lives."),
    ("19-black-belt-calculators.xlsx", "5 Process efficiency", "D10",
     "PCE = process cycle efficiency, the share of the customer's wait that was actual work. "
     "1-8% is normal in support. Below about 5%, making agents faster cannot move it — you have to "
     "remove a waiting state."),
    ("19-black-belt-calculators.xlsx", "5 Process efficiency", "D15",
     "The touch time, drawn to scale against the wait. It is usually a sliver, and the sliver is "
     "the argument."),
    ("19-black-belt-calculators.xlsx", "5 Process efficiency", "D16",
     "The queue time. Every hour of it is a handoff, an approval or a reply being waited on — and "
     "each one is a candidate for removal."),

    # -- 19 calculators, tab 6: staffing without queueing.
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D10",
     "One Erlang is one agent busy for the whole hour. This is the work that arrives, before any "
     "allowance for queueing, breaks or occupancy. It is the floor, never the answer."),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D11",
     "Bodies logged in and taking contacts, once you accept that nobody handles contacts every "
     "minute they are available."),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D12",
     "Headcount that exists purely to cover breaks, meetings, training and absence. Show it "
     "separately — it is the line Finance always challenges, and it is defensible when named."),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D13",
     "FTE = full-time equivalent: what actually goes on the payroll. This model has no queueing in "
     "it, so if you are staffing to a service-level promise use 28-erlang-staffing.xlsx instead."),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D16",
     "The work itself. Everything stacked above this is overhead you should be able to name."),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D17",
     "The gap between the work arriving and the people needed to absorb it without running agents "
     "flat out all day."),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D18",
     "Paid time that is never on a contact. Usually the largest single block, and the one most "
     "often left out of a business case."),
    ("19-black-belt-calculators.xlsx", "6 Staffing", "D19",
     "The total you are asking for. Presenting it as three named parts is how the request survives "
     "a budget conversation."),

    # -- 19 calculators, tab 7: handle-time benefit.
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D11",
     "AHT = average handle time. The raw hours of talk, hold and after-call work your change "
     "removes across a year."),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D10",
     "Share of PAID time that is not available for contacts at all: breaks, training, coaching, "
     "meetings, sickness, holiday. From workforce management, and agree the list — 25% and 40% "
     "are both common and the difference is entirely definitional. Occupancy above gets you to "
     "available hours; this gets you the rest of the way to payroll."),
    # This note used to say freeing an hour of handle time "frees rather more
    # than an hour of payroll", which is the error the row itself made:
    # occupancy converts handle time into AVAILABLE time, and payroll is one
    # step further out.
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D12",
     "More than the handle-hours above, because agents are never occupied every minute. This is "
     "AVAILABLE time, not payroll — occupancy is measured against the hours an agent is logged "
     "in and ready, and paid time carries shrinkage on top."),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D13",
     "What payroll actually carries, and therefore what the saving is worth. Stopping at the row "
     "above prices it as though nobody were paid for a break, and understates the case by "
     "1 ÷ (1 − shrinkage) — about 47% at 32%."),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D16",
     "The same saving expressed as whole people, at 1,760 paid hours a year. This is the unit an "
     "executive actually hears, and it is on paid hours so it compares to a headcount line."),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D14",
     "The value before anything is discounted for what decays at rollout. Never quote this figure "
     "on its own."),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D15",
     "The number to take to Finance — and still only real if the freed capacity is harvested. Read "
     "the trap below before you claim it."),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D20",
     "The claim before realisation is applied."),
    ("19-black-belt-calculators.xlsx", "7 Benefit — AHT", "D21",
     "What belongs in the business case. Showing the discount rather than burying it is what stops "
     "the conversation happening later, in front of the steering committee."),

    # -- 19 calculators, tab 8: avoided contacts.
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D11",
     "The gap you are closing, in percentage POINTS, not as a percentage change. 14.2% down to 8% "
     "is 6.2 points — describing it as a 44% reduction is how these numbers lose credibility."),
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D12",
     "Contacts that simply never happen. Unlike a handle-time saving this needs no harvesting "
     "argument, because the work is not there to be redeployed."),
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D13",
     "Avoided contacts at your fully-loaded cost per contact, before realisation is applied."),
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D14",
     "The defensible number. Realisation is higher here than on handle-time work because an "
     "avoided contact is unambiguously avoided."),
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D20",
     "The claim before realisation is applied."),
    ("19-black-belt-calculators.xlsx", "8 Benefit — avoided contacts", "D21",
     "What belongs in the business case. A wide gap between the two bars means your target is "
     "running ahead of the causes you have actually verified."),

    # -- 19 calculators, tab 9: the three numbers Finance asks for.
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D10",
     "Benefit arriving a year from now is worth less than cash today, so it is discounted. This is "
     "year one in today's money."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D11",
     "Year two, discounted twice. It shows zero if you modelled fewer years than this."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D12",
     "Year three, discounted three times. It shows zero if you modelled fewer years than this."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D13",
     "How long before the benefit has repaid what you spent. Under 18 months is generally an easy "
     "sell; past three years the assumptions matter more than the answer."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D14",
     "ROI = return on investment: what you got back less what you spent, as a share of what you "
     "spent. It ignores the time value of money, so it always flatters — quote it beside the net "
     "present value below, never instead of it."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D15",
     "NPV = net present value: every year's benefit brought back to today's money, minus the "
     "investment. Above zero the project beat the cost of the money. This is the one Finance "
     "trusts."),
    ("19-black-belt-calculators.xlsx", "9 ROI and payback", "D22",
     "The hole you start in — everything spent before any benefit lands. The line climbs from here "
     "and crosses zero at payback."),
]


# Ranges where a fill has to go because it collides with the legend.
#
# The countermeasure hierarchy shades its six levels as a traffic light, which
# is the right idea — the whole point of the hierarchy is to see at a glance
# that most of your controls sit at the weak end. The problem is the palette it
# reached for: the greens on levels 1-3 are FFECFAEF and the amber on level 4 is
# FFFFF9E3, byte-identical to the two colours every legend in this pack reserves
# for "a worked example, delete it when you start" and "you fill this in". So a
# fixed definition of the hierarchy read as a cell to type into, three rows that
# must never be deleted read as disposable, and the colour checks agreed with
# the reader's misreading because they key on exactly those bytes.
#
# Fill in this pack means one thing: what you do with the cell. The durability
# signal moves to the font instead, where it cannot be confused with either —
# and column D already says "Permanent / Moderate / Weak" in words, so nothing
# is lost if a reader ignores colour entirely.
PLAIN: list[tuple[str, str, str]] = [
    ("15-solution-selection-matrix.xlsx", "Countermeasure hierarchy", "A5:D10"),
]

# (workbook, sheet, row) -> font colour carrying the traffic light.
TRAFFIC: dict[tuple[str, str, int], str] = {
    ("15-solution-selection-matrix.xlsx", "Countermeasure hierarchy", 5): "FF1D6F42",
    ("15-solution-selection-matrix.xlsx", "Countermeasure hierarchy", 6): "FF1D6F42",
    ("15-solution-selection-matrix.xlsx", "Countermeasure hierarchy", 7): "FF1D6F42",
    ("15-solution-selection-matrix.xlsx", "Countermeasure hierarchy", 8): "FFB45309",
    ("15-solution-selection-matrix.xlsx", "Countermeasure hierarchy", 9): "FFB91C1C",
    ("15-solution-selection-matrix.xlsx", "Countermeasure hierarchy", 10): "FFB91C1C",
}


def patch_plain(wb, wbname: str) -> int:
    """Clear fills that collide with the legend, and re-signal in the font."""
    from openpyxl.styles import PatternFill
    none_fill = PatternFill(fill_type=None)
    n = 0
    for name, sheet, ref in PLAIN:
        if name != wbname:
            continue
        for row in wb[sheet][ref]:
            for c in row:
                if c.fill and c.fill.patternType:
                    c.fill = none_fill
                    n += 1
    for (name, sheet, r), rgb in TRAFFIC.items():
        if name != wbname:
            continue
        ws = wb[sheet]
        for col in range(2, 5):                  # the words, not the level number
            c = ws.cell(row=r, column=col)
            if getattr(c.font.color, "rgb", None) == rgb and c.font.bold:
                continue                          # already done; do not churn
            c.font = Font(name=c.font.name, size=c.font.sz, bold=True, color=rgb)
            n += 1
    return n


# The per-sheet legend on the seven hand-authored workbooks. It said "Row 1
# example" and "delete it", which was true of none of them — the worked example
# runs to nine rows on the value stream map and twenty-four on a control chart,
# and deleting it on a seeded sheet leaves an empty chart.
LEGEND_FIX: list[tuple[str, str, str, object]] = [
    # The calculators explain their colours in prose — "change the YELLOW cells,
    # everything else is a formula" — which is true and is not the key every
    # other workbook in the pack carries. A reader who has learnt to look for
    # one will not find it here. Same words, same place as the rest.
    ("19-black-belt-calculators.xlsx", "Start here", "B21", "What the colours mean"),
    ("19-black-belt-calculators.xlsx", "Start here", "B22", "Yellow cells"),
    ("19-black-belt-calculators.xlsx", "Start here", "B23", "You fill these in."),
    ("19-black-belt-calculators.xlsx", "Start here", "B24", "Blue cells"),
    ("19-black-belt-calculators.xlsx", "Start here", "B25",
     "Calculated for you. Do not type over them \u2014 they contain formulas."),
    ("12-fmea.xlsx", "How to use this", "B6",
     "1.  Open the FMEA tab. The green rows are a worked example — read them, then "
     "overwrite them with your own failure modes."),
]


def patch_notes(wb, wbname: str) -> int:
    """Write the explanations above, styled as the notes already on the sheet.

    Wrapping is decided by the column, not applied blindly. Wrapping text into a
    26-wide column and then sizing the row for a 90-character line clips it —
    the reader sees a truncated sentence and no scroll bar. A narrow column is
    left unwrapped so the text overflows across the empty cells beside it, which
    is what a plain note in column C has always done on these sheets.
    """
    n = 0
    for name, sheet, cell, text in NOTES:
        if name != wbname:
            continue
        ws = wb[sheet]
        c = ws[cell]
        if c.__class__.__name__ == "MergedCell":
            raise SystemExit(f"{wbname} {sheet}!{cell} is inside a merged range")
        # Refuse to clobber something that is not a note — a formula, a number,
        # a heading. But a cell already holding PROSE is a note this registry
        # owns, and refusing that made the registry able to add a note and
        # never to correct one: the AHT tab's D12 said freeing an hour of
        # handle time "frees rather more than an hour of payroll", which is the
        # error the row itself made, and it could not be edited without the
        # build exiting. The same shape as write_chart_notes, which could not
        # replace a note either.
        held = c.value
        if held not in (None, "") and held != text:
            editable = isinstance(held, str) and len(held) > 40 and not held.startswith("=")
            if not editable:
                raise SystemExit(
                    f"{wbname} {sheet}!{cell} already holds {held!r} — that is not a "
                    f"note, so NOTES will not overwrite it")
        if c.value != text:
            c.value = text
            n += 1
        # Only assign a style that is not already there. _style_new_cells runs
        # first and styles these same cells; overwriting its font and alignment
        # unconditionally made the pair oscillate, minting one fresh cellXf on
        # every build. The workbook grew a style per run forever, and nobody
        # could see it because the document timestamps churned the bytes anyway.
        if (c.font.italic, c.font.sz,
                getattr(c.font.color, "rgb", None)) != (True, 9.0, "FF6B7280"):
            c.font = Font(italic=True, size=9, color="FF6B7280")
        width = ws.column_dimensions[c.column_letter].width or 8.43
        wrap, vert = (True, "top") if width >= 40 else (False, "center")
        if (bool(c.alignment.wrap_text), c.alignment.vertical) != (wrap, vert):
            c.alignment = Alignment(wrap_text=wrap, vertical=vert)
        if wrap:
            lines = 1 + int(len(text) / (width * 0.95))
            ws.row_dimensions[c.row].height = max(
                ws.row_dimensions[c.row].height or 0, 13 * lines)
    return n


def patch_formulas(verbose: bool = True) -> int:
    """Apply every patch. Returns the number of cells changed."""
    by_file: dict[str, list] = {}
    for wbname, sheet, cell, formula in SINGLE:
        by_file.setdefault(wbname, []).append((sheet, cell, formula))
    for wbname, sheet, celltpl, ftpl, rows in REPEATED:
        for r in rows:
            by_file.setdefault(wbname, []).append(
                (sheet, celltpl.format(r=r), ftpl.format(r=r))
            )
    for wbname, sheet, cell, value in VALUES:
        by_file.setdefault(wbname, []).append((sheet, cell, value))
    for wbname, *_ in NOTES:
        by_file.setdefault(wbname, [])
    for wbname, *_ in PLAIN:
        by_file.setdefault(wbname, [])
    for wbname, sheet, cell, value in LEGEND_FIX:
        by_file.setdefault(wbname, []).append((sheet, cell, value))

    changed = 0
    for wbname in sorted(by_file):
        path = TEMPLATES / wbname
        if not path.exists():
            raise SystemExit(f"missing workbook: {path}")
        wb = load_workbook(path)
        local = 0
        for sheet, cell, value in by_file[wbname]:
            ws = wb[sheet]
            if ws[cell].value != value:
                ws[cell] = value
                local += 1
        _style_new_cells(wb)
        local += patch_notes(wb, wbname)
        local += patch_plain(wb, wbname)
        local += _add_validations(wb, wbname)
        local += add_examples(wb, wbname)
        local += add_charts(wb, wbname)
        local += polish_workbook(wb)
        # openpyxl's zip output is not byte-stable, so saving an unchanged
        # workbook would churn its base64 copy in the HTML on every run.
        if local:
            save_workbook(wb, path)
        changed += local
        if verbose:
            state = f"{local:3d} cell(s) rewritten" if local else "  unchanged"
            print(f"  {wbname:46s} {state}")

    # Workbooks with no calculated cells still need the finishing pass — the
    # X-Y matrix is pure scoring and would otherwise never be polished at all.
    for path in sorted(TEMPLATES.glob("*.xlsx")):
        if path.name in by_file:
            continue
        if _is_built(path.name):
            # build_templates.py already polishes these at build time. Loading
            # and re-saving here would be destructive: openpyxl's reader keeps
            # only the first chart type in a combo, so every reference line laid
            # over a bar chart would be dropped on the way back out.
            continue
        wb = load_workbook(path)
        local = add_examples(wb, path.name)
        local += add_charts(wb, path.name)
        local += polish_workbook(wb)
        if local:
            save_workbook(wb, path)
            changed += local
            if verbose:
                print(f"  {path.name:46s} {local:3d} layout fix(es)")
    return changed


def _style_new_cells(wb) -> None:
    """Style the alpha input added to the hypothesis-test log."""
    if "Test log" not in wb.sheetnames:
        return
    ws = wb["Test log"]
    if ws["A8"].value != "Significance level (alpha)":
        return
    ws["A8"].font = Font(bold=True, size=10)
    ws["B8"].fill = PatternFill("solid", fgColor=YELLOW_INPUT)
    ws["B8"].number_format = "0.00"
    ws["B8"].alignment = Alignment(horizontal="center")
    ws["C8"].font = Font(size=9, italic=True, color="FF6B7280")


def _add_validations(wb, wbname: str) -> int:
    """Bound the inputs that previously accepted impossible values.

    Returns the number of validations added, so the caller can skip the save
    when nothing changed.
    """
    specs: list[tuple[str, str, str, str, str, str]] = []
    if wbname == "19-black-belt-calculators.xlsx":
        specs = [
            ("9 ROI and payback", "B7", "whole", "1", "10",
             "Years to model must be a whole number from 1 to 10."),
            ("9 ROI and payback", "B8", "decimal", "0", "1",
             "Discount rate is a share, for example 0.10 for 10%."),
            ("3 SLA capability", "B7", "decimal", "0.0000001", "100000",
             "Standard deviation must be greater than zero."),
            ("6 Staffing", "B7", "decimal", "0.05", "1", "Target occupancy is a share between 0.05 and 1."),
            ("6 Staffing", "B8", "decimal", "0", "0.95", "Shrinkage is a share below 0.95."),
            ("7 Benefit — AHT", "B8", "decimal", "0.05", "1", "Occupancy is a share between 0.05 and 1."),
            ("1 Sigma level", "B6", "whole", "1", "50", "Opportunities per unit must be at least 1."),
        ]
    elif wbname == "05-data-collection-plan.xlsx":
        specs = [
            ("Sample size calculator", "C5", "decimal", "0.001", "0.999", "Current rate is a share between 0 and 1."),
            ("Sample size calculator", "C6", "decimal", "0.001", "0.999",
             "Change to detect is a share, for example 0.03 for 3 percentage points."),
            ("Sample size calculator", "C7", "decimal", "0.001", "0.5", "Alpha is a share, typically 0.05 or 0.01."),
            ("Sample size calculator", "C8", "decimal", "0.5", "0.999", "Power is a share, typically 0.80 or 0.90."),
        ]
    elif wbname == "13-hypothesis-test-log.xlsx":
        specs = [("Test log", "B8", "decimal", "0.001", "0.5",
                  "Alpha is a share, typically 0.05 or 0.01.")]

    added = 0
    for sheet, cell, dtype, lo, hi, msg in specs:
        ws = wb[sheet]
        # openpyxl normalises a single-cell sqref to "C5", but writing one via
        # dv.add() can leave "C5:C5" - accept either so re-running is a no-op.
        existing = {str(dv.sqref) for dv in ws.data_validations.dataValidation}
        if cell in existing or f"{cell}:{cell}" in existing:
            continue
        dv = DataValidation(type=dtype, operator="between", formula1=lo, formula2=hi,
                            allow_blank=True, showErrorMessage=True)
        dv.error = msg
        dv.errorTitle = "Check this number"
        ws.add_data_validation(dv)
        dv.add(ws[cell])
        added += 1
    return added


if __name__ == "__main__":
    print("Patching workbooks in", TEMPLATES)
    n = patch_formulas()
    print(f"Done: {n} cell(s) rewritten.")
    sys.exit(0)
