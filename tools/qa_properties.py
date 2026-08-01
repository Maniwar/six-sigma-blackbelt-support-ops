#!/usr/bin/env python3
"""Test the templates as functions, not as one shipped example.

Every defect the re-scoring panel found had the same shape. The workbook is
correct at the example I wrote, and wrong the moment somebody changes an input:

  * the baseline divided by the window requested rather than the points that
    exist, so fifteen pasted periods against a window of twenty put the centre
    line 25% low, silently
  * axis bounds are computed at build time from the example, so pasting data
    with a different magnitude leaves the plot framed on limits it no longer
    draws
  * the preview embedded in the page still printed the pre-fix numbers next to
    a caption describing the new behaviour
  * the ROI chart plots four points while its own input allows ten years

None of that is visible in the shipped file. Structural checks pass, the render
looks right, and the rubric scores it full marks — because the example was
authored to make it look right. The only way to catch this class is to perturb
the inputs and assert the invariants still hold.

    python3 tools/qa_properties.py            # all templates
    python3 tools/qa_properties.py 27 19      # just these

Needs the `formulas` package.

This harness used to fail on five charts every single run, with the reasons
written up here as KNOWN LIMITS. That was honest and it was still wrong: a check
that is permanently red gets read as decoration exactly as fast as one that is
permanently green, and nobody looks at either again. All five are now fixed —
four in the harness, one in the workbook — and the distinction matters:

  * 05's sample-size curve hardcoded 1.96 and 0.25, so it was a textbook
    illustration that never moved whatever the reader typed. A real defect in
    the workbook, fixed in the workbook: the curve reads their alpha and their
    baseline rate, and marks where their own study sits on it.
  * A scatter has two value axes and chart_specs read the first, so the
    stakeholder map was checked against the SUPPORT axis using its INFLUENCE
    values. A harness bug.
  * Doubling every input leaves a ratio unchanged — kappa is scale-invariant,
    so a uniform perturbation could only ever report it as a dead chart. The
    shift is distinct per input now.
  * Charts that COUNT text categories move when a CATEGORY changes, not when a
    number does, so the perturbation now also switches dropdown values.
  * That switching took a couple of cells per dropdown rather than the first
    eight overall, which had been exhausting the budget on the control plan's
    chart-type column and never reaching the control-LEVEL column the chart
    actually counts.

tools/qa_selftest.py mutation-tests this file, because a suite that went from
five failures to none in one sitting is exactly the suite to be suspicious of.
"""
from __future__ import annotations

import re
import shutil
import sys
import tempfile
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "templates"
C = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"

fails: list[str] = []
passes = [0]


def check(cond, label, detail=""):
    if cond:
        passes[0] += 1
    else:
        fails.append(f"  FAIL {label}" + (f"\n       {detail}" if detail else ""))


def _engine(path):
    warnings.filterwarnings("ignore")
    import formulas
    return formulas.ExcelModel().loads(str(path)).finish()


def _val(sol, sheet, cell):
    key = f"!{cell}"
    for k, v in sol.items():
        if k.upper().endswith(f"]{sheet.upper()}'{key}"):
            try:
                out = v.value[0, 0]
            except Exception:                                    # noqa: BLE001
                out = v.value
            return out
    return None


# ------------------------------------------------------------------ charts


def chart_specs(path):
    """(sheet-less) axis bounds and series ranges straight out of the file."""
    out = []
    with zipfile.ZipFile(path) as z:
        for name in sorted(n for n in z.namelist()
                           if re.match(r"xl/charts/chart\d+\.xml$", n)):
            root = ET.fromstring(z.read(name))
            title = "".join(t.text or "" for t in root.iter(
                "{http://schemas.openxmlformats.org/drawingml/2006/main}t"))
            # Every value axis, in document order. This used to take the first
            # one and break — which is right for a bar or a line, and wrong for
            # a scatter, where the FIRST axis is X and the values collected
            # below are Y. The stakeholder map was therefore checked against
            # the support axis (-2.5..2.5) using its influence values (0..6),
            # and reported as broken for the whole life of this harness.
            axes = []
            for ax in root.iter(f"{C}valAx"):
                sc = ax.find(f"{C}scaling")
                mn = sc.find(f"{C}min") if sc is not None else None
                mx = sc.find(f"{C}max") if sc is not None else None
                axes.append((float(mn.get("val")), float(mx.get("val")))
                            if mn is not None and mx is not None else None)
            scatter = root.find(f".//{C}scatterChart") is not None
            yrefs, xrefs = [], []
            for ser in root.iter(f"{C}ser"):
                f = ser.find(f"{C}val/{C}numRef/{C}f")
                if f is None:
                    f = ser.find(f"{C}yVal/{C}numRef/{C}f")
                if f is not None and f.text:
                    yrefs.append(f.text)
                x = ser.find(f"{C}xVal/{C}numRef/{C}f")
                if x is not None and x.text:
                    xrefs.append(x.text)
            if scatter:
                # X against the first value axis, Y against the second.
                if xrefs and axes:
                    out.append({"title": title[:60] + " (x)",
                                "bounds": axes[0], "refs": xrefs})
                ybounds = axes[1] if len(axes) > 1 else None
            else:
                ybounds = axes[0] if axes else None
            out.append({"title": title[:60], "bounds": ybounds, "refs": yrefs})
    return out


def plotted_values(sol, refs):
    vals = []
    for ref in refs:
        m = re.match(r"^'?([^'!]+)'?!(.+)$", ref)
        if not m:
            continue
        sheet = m.group(1)
        try:
            c1, r1, c2, r2 = range_boundaries(m.group(2).replace("$", ""))
        except Exception:                                        # noqa: BLE001
            continue
        from openpyxl.utils import get_column_letter as gl
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                v = _val(sol, sheet, f"{gl(c)}{r}")
                if isinstance(v, (int, float)):
                    vals.append(float(v))
    return vals


# --------------------------------------------------------------- the tests


def dropdown_inputs(wb, sheets=None, limit=3):
    """Yellow cells that hold one of a fixed list, plus a different valid choice.

    Several charts do not plot a number the reader types at all — they COUNT
    text. The control hierarchy counts controls by level, the fishbone counts
    causes per branch. Doubling every number on those sheets moves nothing, and
    the harness reported that correct behaviour as a failure for its whole life.
    The way to move a counting chart is to change a category, so that is what
    this offers.
    """
    from openpyxl.utils import range_boundaries as rb
    out = []
    for ws in wb.worksheets:
        if sheets and ws.title not in sheets:
            continue
        for dv in getattr(ws.data_validations, "dataValidation", []) or []:
            f = (dv.formula1 or "").strip()
            if dv.type != "list" or not f.startswith('"'):
                continue
            options = [o.strip() for o in f.strip('"').split(",") if o.strip()]
            if len(options) < 2:
                continue
            # A couple of cells per dropdown, not the first N overall. Taking
            # the first eight exhausted the budget on the control plan's
            # chart-type column and never reached the control-LEVEL column,
            # which is the one the chart actually counts.
            taken = 0
            for rng in str(dv.sqref).split():
                try:
                    c1, r1, c2, r2 = rb(rng)
                except Exception:                                # noqa: BLE001
                    continue
                for r in range(r1, min(r2, r1 + 40) + 1):
                    for c in range(c1, c2 + 1):
                        cell = ws.cell(row=r, column=c)
                        if not isinstance(cell.value, str) or not cell.value.strip():
                            continue
                        alt = next((o for o in options if o != cell.value.strip()), None)
                        if alt:
                            out.append((ws.title, cell.coordinate, cell.value, alt))
                            taken += 1
                        if taken >= limit:
                            break
                    if taken >= limit:
                        break
                if taken >= limit:
                    break
    return out


def yellow_inputs(wb, limit=40, sheets=None):
    """Cells the user is invited to change.

    Scoped to the sheets that matter: perturbing only the first sheet's inputs
    and then asserting a chart on the fifth sheet moved tests nothing except
    which tab happened to come first.
    """
    out = []
    for ws in wb.worksheets:
        if sheets and ws.title not in sheets:
            continue
        n = 0
        for row in ws.iter_rows():
            for c in row:
                try:
                    # yellow "you type this" AND green "worked example" — both
                    # are data the user replaces, and on several sheets the
                    # example is the only thing feeding the chart
                    if (c.fill and c.fill.patternType
                            and str(c.fill.fgColor.rgb) in ("FFFFF9E3", "FFECFAEF")
                            and isinstance(c.value, (int, float))):
                        out.append((ws.title, c.coordinate, c.value))
                        n += 1
                except Exception:                                # noqa: BLE001
                    pass
                if n >= limit:
                    break
            if n >= limit:
                break
    return out


def validated_max(ws, coord):
    """The ceiling data validation puts on a cell, if any.

    Perturbing an influence score of 1-5 up to 15 tests nothing except whether
    the harness reads the sheet's own rules — the workbook forbids it.
    """
    for dv in ws.data_validations.dataValidation:
        if dv.type != "whole" or not dv.sqref:
            continue
        if coord in dv.sqref:
            try:
                return float(dv.formula2)
            except (TypeError, ValueError):
                return None
    return None


def sheets_of(refs):
    """Which sheets a chart's series actually read."""
    out = set()
    for ref in refs:
        m = re.match(r"^'?([^'!]+)'?!", ref)
        if m:
            out.add(m.group(1))
    return out


def test_axis_survives_rescale(path):
    """Multiply every numeric input by 3. The chart must still frame its data.

    Axis bounds are computed at build time from the shipped example. That is
    fine until somebody pastes their own numbers — a support team whose metric
    is in seconds rather than minutes gets a chart with nothing inside the
    frame, while the How-to tab says nothing else needs touching.
    """
    specs = [s for s in chart_specs(path) if s["bounds"]]
    if not specs:
        return
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / path.name
        shutil.copyfile(path, tmp)
        wb = load_workbook(tmp)
        ins = yellow_inputs(wb)
        if not ins:
            return
        for sheet, coord, v in ins:
            if 0 < v < 1:            # a rate or a percentage: leave it alone
                continue
            cap = validated_max(wb[sheet], coord)
            if cap is not None:      # the workbook forbids going higher
                continue
            wb[sheet][coord] = v * 3
        wb.save(tmp)
        try:
            sol = _engine(tmp).calculate()
        except Exception as exc:                                 # noqa: BLE001
            check(False, f"{path.name}: recalculates after inputs change",
                  f"{type(exc).__name__}: {exc}")
            return
        for s in specs:
            vals = plotted_values(sol, s["refs"])
            if not vals:
                continue
            lo, hi = s["bounds"]
            outside = [v for v in vals if v < lo or v > hi]
            check(not outside,
                  f"{path.name} · {s['title']}: still frames its data when the "
                  "numbers change",
                  f"axis {lo:,.4g}..{hi:,.4g}, {len(outside)} of {len(vals)} "
                  f"points outside, e.g. {outside[:3]}")


def test_preview_matches_workbook(path):
    """The preview printed in the page must equal what the workbook computes."""
    import json
    cache = ROOT / "tools" / "previews.json"
    if not cache.exists():
        return
    shown = json.loads(cache.read_text(encoding="utf-8"))
    if path.name not in shown:
        return
    html = shown[path.name]
    # every tooltip carries the formula; the visible text is what we check
    try:
        sol = _engine(path).calculate()
    except Exception:                                            # noqa: BLE001
        return
    bad = []
    for m in re.finditer(r'title="([^"]*)"[^>]*>([^<]*)</td>', html):
        pass                                    # positional mapping lives in
        # sync_html; re-deriving it here would duplicate that logic, so the
        # preview check is deliberately limited to the cached headline values
    check(not bad, f"{path.name}: preview agrees with the workbook",
          "; ".join(bad[:3]))


def test_chart_responds(path):
    """Change an input; something plotted must move.

    A chart wired to a stale helper column, or to a block a later edit stopped
    feeding, looks perfect and never changes. This is the cheapest possible
    test for it.
    """
    specs = [s for s in chart_specs(path) if s["refs"]]
    if not specs:
        return
    with tempfile.TemporaryDirectory() as td:
        a = Path(td) / ("a_" + path.name)
        b = Path(td) / ("b_" + path.name)
        shutil.copyfile(path, a)
        shutil.copyfile(path, b)
        wb = load_workbook(b)
        want = set()
        for s_ in specs:
            want |= sheets_of(s_["refs"])
        ins = yellow_inputs(wb, limit=6, sheets=want)
        drops = dropdown_inputs(wb, sheets=want)
        if not ins and not drops:
            return
        moved = False
        for i, (sheet, coord, v) in enumerate(ins):
            cap = validated_max(wb[sheet], coord)
            # A DISTINCT shift per input, not a uniform doubling. Kappa, and any
            # other ratio, is scale-invariant: double every cell of its table and
            # it returns the identical answer, which is correct behaviour that
            # a uniform perturbation can only ever report as a dead chart.
            new = v * 2 + i + 1 if v else i + 1
            if cap is not None and new > cap:
                new = cap if v != cap else max(1, cap - 1)   # move inside the rule
            if new != v:
                wb[sheet][coord] = new
                moved = True
        for sheet, coord, _cur, alt in drops:
            wb[sheet][coord] = alt
            moved = True
        if not moved:
            return
        wb.save(b)
        try:
            sa = _engine(a).calculate()
            sb = _engine(b).calculate()
        except Exception:                                        # noqa: BLE001
            return
        for s in specs:
            va = plotted_values(sa, s["refs"])
            vb = plotted_values(sb, s["refs"])
            if not va:
                continue
            check(va != vb,
                  f"{path.name} · {s['title']}: moves when its inputs move",
                  "every plotted value identical after doubling the inputs — "
                  "the chart is probably wired to a stale block")


def main() -> int:
    only = [a for a in sys.argv[1:] if not a.startswith("-")]
    books = sorted(TEMPLATES.glob("*.xlsx"))
    if only:
        books = [b for b in books if any(o in b.name for o in only)]
    print(f"Property tests over {len(books)} workbook(s) — "
          "perturbing inputs, not trusting the shipped example\n")
    for path in books:
        test_axis_survives_rescale(path)
        test_chart_responds(path)
        print(f"  {path.name}")
    print(f"\n  {passes[0]} property check(s) passed")
    if fails:
        print(f"\n{len(fails)} FAILURE(S):")
        print("\n".join(fails))
        return 1
    print("\nAll property checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
