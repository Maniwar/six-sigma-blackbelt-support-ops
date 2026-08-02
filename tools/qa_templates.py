#!/usr/bin/env python3
"""Quality audit for every workbook in templates/.

A template is only useful if someone can open it, understand what to type, and
trust what comes back. That is three separate failure modes, so this checks
three separate layers:

  CHARTS   every chart resolves to real data, is titled, and is readable
  GUIDED   every cell you are asked to fill in says where the number comes from
  EXAMPLE  the guidance that exists actually says something — no empty columns,
           no opaque codes, no one sentence pasted onto every row
  NUMERIC  the whole workbook recalculates without a single error value

Run it after any change to tools/build_templates.py or tools/patch_workbooks.py:

    python3 tools/qa_templates.py            # all three layers
    python3 tools/qa_templates.py --fast     # skip NUMERIC (it is the slow one)

Exit status is non-zero if anything failed, so it drops straight into CI.
"""
from __future__ import annotations

import math
import re
import sys
import warnings
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "templates"

# Colours the builders use. Yellow is "you fill this in", so a yellow cell that
# carries no explanation is the single most common way a template goes unused.
FILL_INPUT = "FFFFF9E3"
FILL_CALC = "FFF2F7FF"

# Workbooks where a chart would be noise rather than signal: a 5 Whys chain and
# a RACI grid are structures, not distributions. Everything else must chart.
NO_CHART_OK = {
    "20-five-whys-tree.xlsx": "a causal chain is a structure, not a distribution",
}

ERR_VALUES = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

# NA() is the only construction that reliably leaves a gap in a plotted series;
# "" plots as zero and draws a row of markers along the axis floor. Marker and
# horizon columns are therefore FULL of #N/A by design.
#
# This used to be a hand-written book -> sheet -> column-span map, and the map
# was the problem: two workbooks that use the same idiom were never added to
# it, so the harness ran red on 37 deliberate cells for months. A permanently
# red gate is read by nobody, which costs exactly as much as no gate.
#
# The formula already says which it is. A deliberate gap CALLS NA() — the
# author wrote it. An accidental #N/A propagates out of a lookup that missed
# (INDEX/MATCH, VLOOKUP) and never contains the word. So the exemption is a
# property of the cell rather than a list somebody has to remember to update:
#
#   =IF(I5="",NA(),SUM(...))          deliberate: gap this point
#   =INDEX(A:A,MATCH(x,B:B,0))        accidental: the match failed
#
# A cell whose formula is a bare reference to a marker cell inherits #N/A
# without calling it, so one hop of propagation is followed too.
RE_CALLS_NA = re.compile(r"\bNA\s*\(\s*\)")
RE_REF = re.compile(r"(?<![A-Za-z0-9_!$])(\$?[A-Z]{1,3}\$?[0-9]{1,7})(?![0-9(])")

fails: list[str] = []
warns: list[str] = []


def fail(book: str, layer: str, msg: str) -> None:
    fails.append(f"  FAIL [{layer}] {book}: {msg}")


def warn(book: str, layer: str, msg: str) -> None:
    warns.append(f"  warn [{layer}] {book}: {msg}")


# ---------------------------------------------------------------- helpers


def merged_shadow(ws) -> set[tuple[int, int]]:
    """Cells swallowed by a merge. A formula reading one of these is always empty."""
    out = set()
    for rng in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(str(rng))
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    out.add((r, c))
    return out


REF_RE = re.compile(r"^'?(?P<sheet>[^'!]+)'?!\$?(?P<c1>[A-Z]+)\$?(?P<r1>\d+)"
                    r"(?::\$?(?P<c2>[A-Z]+)\$?(?P<r2>\d+))?$")


def parse_ref(ref: str):
    """('Sheet1'!$B$4:$B$27) -> (sheet, min_col, min_row, max_col, max_row)."""
    m = REF_RE.match((ref or "").strip())
    if not m:
        return None
    from openpyxl.utils import column_index_from_string as ci
    c1, r1 = ci(m.group("c1")), int(m.group("r1"))
    c2 = ci(m.group("c2")) if m.group("c2") else c1
    r2 = int(m.group("r2")) if m.group("r2") else r1
    return m.group("sheet"), min(c1, c2), min(r1, r2), max(c1, c2), max(r1, r2)


def series_refs(ser):
    """(values_ref, categories_ref) for a chart series, as written in the XML."""
    val = None
    if ser.val is not None and ser.val.numRef is not None:
        val = ser.val.numRef.f
    elif getattr(ser, "yVal", None) is not None and ser.yVal.numRef is not None:
        val = ser.yVal.numRef.f
    cat = None
    if ser.cat is not None:
        if ser.cat.numRef is not None:
            cat = ser.cat.numRef.f
        elif ser.cat.strRef is not None:
            cat = ser.cat.strRef.f
    elif getattr(ser, "xVal", None) is not None:
        if ser.xVal.numRef is not None:
            cat = ser.xVal.numRef.f
        elif ser.xVal.strRef is not None:
            cat = ser.xVal.strRef.f
    return val, cat


def chart_title_text(ch) -> str:
    """openpyxl nests the title several layers down; dig it out."""
    t = getattr(ch, "title", None)
    if t is None:
        return ""
    rich = getattr(getattr(t, "tx", None), "rich", None)
    if rich is None:
        return str(t) if isinstance(t, str) else ""
    return "".join(r.t or "" for p in rich.p for r in (p.r or []))


# ---------------------------------------------------------------- layers


def audit_charts(path: Path, wb) -> int:
    """Every chart must plot real cells, be titled, and not sit on top of the data."""
    book = path.name
    total = 0
    for ws in wb.worksheets:
        shadow = merged_shadow(ws)
        for ch in getattr(ws, "_charts", []):
            total += 1
            label = f"{ws.title!r} chart {total}"
            if not chart_title_text(ch).strip():
                fail(book, "CHARTS", f"{label} has no title — a chart with no title is a decoration")
            if not ch.series:
                fail(book, "CHARTS", f"{label} has no data series at all")
                continue
            for si, ser in enumerate(ch.series, start=1):
                vref, cref = series_refs(ser)
                if not vref:
                    fail(book, "CHARTS", f"{label} series {si} has no value reference")
                    continue
                p = parse_ref(vref)
                if not p:
                    fail(book, "CHARTS", f"{label} series {si} value ref is unparseable: {vref}")
                    continue
                sheet, c1, r1, c2, r2 = p
                if sheet not in wb.sheetnames:
                    fail(book, "CHARTS", f"{label} series {si} points at missing sheet {sheet!r}")
                    continue
                src = wb[sheet]
                cells = [src.cell(row=r, column=c)
                         for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)]
                if len(cells) < 2:
                    fail(book, "CHARTS", f"{label} series {si} plots {len(cells)} point(s) — not a chart")
                live = [c for c in cells
                        if c.value is not None and (r1, c1) not in shadow]
                if not live:
                    fail(book, "CHARTS",
                         f"{label} series {si} range {vref} is entirely empty — this renders blank in Excel")
                hidden = [c.coordinate for c in cells if (c.row, c.column) in merged_shadow(src)]
                if hidden:
                    fail(book, "CHARTS",
                         f"{label} series {si} reads merged-shadow cells {hidden[:3]} which are always empty")
                if cref:
                    pc = parse_ref(cref)
                    if not pc:
                        fail(book, "CHARTS", f"{label} series {si} category ref unparseable: {cref}")
                    else:
                        _, cc1, cr1, cc2, cr2 = pc
                        n_cat = (cr2 - cr1 + 1) * (cc2 - cc1 + 1)
                        n_val = (r2 - r1 + 1) * (c2 - c1 + 1)
                        if n_cat != n_val:
                            fail(book, "CHARTS",
                                 f"{label} series {si} has {n_cat} categories for {n_val} values — "
                                 "Excel will silently drop the overhang")
                else:
                    warn(book, "CHARTS", f"{label} series {si} has no category axis — points get numbered 1..n")
            if ch.__class__.__name__ == "BarChart":
                for si, ser in enumerate(ch.series, start=1):
                    if ser.invertIfNegative is not False:
                        fail(book, "CHARTS",
                             f"{label} series {si} leaves invertIfNegative unset — any negative "
                             "value renders as a blank bar")
            # A series with no name shows as "Series1" the moment anything
            # turns the legend on — which laying a reference line over a bar
            # chart does.
            if getattr(ch, "legend", None) is not None:
                for si, ser in enumerate(ch.series, start=1):
                    tx = getattr(ser, "tx", None)
                    nm = ""
                    if tx is not None:
                        nm = (getattr(tx, "v", None) or
                              (tx.strRef.f if getattr(tx, "strRef", None) else "") or "")
                    if not nm:
                        fail(book, "CHARTS",
                             f"{label} series {si} has no name but the chart has a legend — "
                             "Excel will label it 'Series1'")
            # Does the chart sit on top of content? A chart anchored over the
            # block it reads from hides the very table it is explaining, and
            # nothing in the file records the collision — you only see it when
            # you open the sheet.
            anc = getattr(ch, "anchor", None)
            frm = getattr(anc, "_from", None)
            if frm is not None:
                c0, r0 = frm.col, frm.row              # zero-based
                # roughly 0.65 cm per column, 0.5 cm per row at default sizes
                c1 = c0 + int((ch.width or 15) / 0.65)
                r1 = r0 + int((ch.height or 7.5) / 0.5)
                covered = []
                for rr in range(r0 + 1, min(r1 + 1, ws.max_row) + 1):
                    for cc in range(c0 + 1, min(c1 + 1, ws.max_column) + 1):
                        cell = ws.cell(row=rr, column=cc)
                        if cell.value not in (None, "") and cell.__class__.__name__ != "MergedCell":
                            covered.append(cell.coordinate)
                if len(covered) > 4:
                    fail(book, "CHARTS",
                         f"{label} is anchored over {len(covered)} populated cells "
                         f"(e.g. {covered[:4]}) — it hides the data it explains")
            if ch.height and ch.height < 5:
                warn(book, "CHARTS", f"{label} is only {ch.height}cm tall — labels will collide")
    # Per-SHEET, not just per-workbook. A workbook passed the "has charts" rule
    # on the strength of one chart while another of its tabs shipped completely
    # empty — every statistic blank and a verdict reading "paste your data".
    for ws in wb.worksheets:
        low = ws.title.lower()
        if low.startswith(("how to", "start here", "pick your", "category prompts",
                           "countermeasure", "scoring guide", "which chart",
                           "six-trap", "raci", "fishbone diagram", "5 whys")):
            continue
        # Not "few numbers" — a calculator tab legitimately has three inputs
        # and that is the whole tab. The signature is formulas with NOTHING
        # driving them: every statistic blank and a verdict reading "paste your
        # data in", which is how the distribution tab shipped.
        def _drives(c):
            if isinstance(c.value, (int, float)):
                return True
            try:                                   # a green worked-example cell
                return (c.fill and c.fill.patternType
                        and str(c.fill.fgColor.rgb) == "FFECFAEF"
                        and c.value not in (None, ""))
            except Exception:                      # noqa: BLE001
                return False
        nums = sum(1 for row in ws.iter_rows() for c in row if _drives(c))
        forms = sum(1 for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("="))
        if forms >= 5 and nums < 2:
            fail(book, "CHARTS",
                 f"sheet {ws.title!r} has {forms} formulas and {nums} numbers to drive them — "
                 "it ships blank, so nothing on it demonstrates the point of the tab")

    if total == 0 and book not in NO_CHART_OK:
        fail(book, "CHARTS", "no charts at all — the numbers are there but nobody can see the shape")
    return total


GUIDE_TABS = ("how to use", "start here", "read me", "readme", "pick your chart")


def _column_header(ws, col: int, row: int) -> str:
    """The dark banded header above this cell, if it is in a data table.

    A table column explains itself once, in its header — demanding a note on all
    230 cells of an entry grid is not a quality bar, it is noise. A standalone
    input (a label in column A, a value in column B) has no header, and that is
    the case that genuinely needs a per-cell explanation.
    """
    for r in range(row - 1, 0, -1):
        try:
            rgb = ""
            c = ws.cell(row=r, column=col)
            if c.fill and c.fill.patternType:
                rgb = str(c.fill.fgColor.rgb)
            if rgb == "FF333C49":                # the table's header row
                return str(c.value or "") or "\x00"
            if rgb == "FFEEF1F6":                # a section band: different block
                return ""
            # a band can be merged, so it may only be filled in column A
            a = ws.cell(row=r, column=1)
            if a.fill and a.fill.patternType and str(a.fill.fgColor.rgb) == "FFEEF1F6":
                return ""
        except Exception:                                        # noqa: BLE001
            pass
    return ""


FILL_EXAMPLE = "FFECFAEF"


# Deliberately allowed to cross a sentence and a cell boundary. The legend puts
# "Green cells" in one cell and "Replace it with your own data when you start"
# in the next, so a pattern that stops at a full stop can never match the one
# place every workbook states this.
RE_REPLACE_GREEN = re.compile(
    r"green[\s\S]{0,240}?\b(overwrite|replace|your own)|"
    r"\b(overwrite|replace)[\s\S]{0,240}?green", re.I)


def _check_example_has_somewhere_to_go(path: Path, wb) -> None:
    """A sheet that is all example has to say the example is what you overwrite.

    The whole-workbook input count cannot see this: it asks whether the FILE has
    a yellow cell anywhere, so a sheet could lose every one it had and the
    workbook would still pass on the strength of a different tab. That is how
    825 cells of entry grid turned green on the control charts, the Gage R&R
    study and the regression tabs while their instructions still read "overwrite
    the yellow column with your own numbers" — a colour that by then existed
    nowhere on those sheets.

    The rule is not "every sheet needs a yellow cell". Some sheets are entirely
    a worked example on purpose: 24 periods of control-chart data, 90 Gage R&R
    readings, 40 handle times whose whole job is to show you a skewed
    distribution. Splitting those into one green demonstration row and a yellow
    remainder tells the reader the other rows are theirs to invent, which they
    are not. What must hold is that the sheet and its instructions agree — if
    the entry area is green, the workbook has to tell you to overwrite the green
    rather than send you looking for yellow that is not there.
    """
    book = path.name
    # Every string cell, with no length floor: the legend's own label is the
    # eleven characters "Green cells", and a filter for prose dropped exactly
    # the cell that defines the colour.
    prose = " ".join(
        str(c.value) for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if isinstance(c.value, str))
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in GUIDE_TABS):
            continue
        green = yellow = 0
        for row in ws.iter_rows():
            for c in row:
                try:
                    rgb = str(c.fill.fgColor.rgb) if c.fill and c.fill.patternType else ""
                except Exception:                                # noqa: BLE001
                    continue
                if rgb == FILL_EXAMPLE:
                    green += 1
                elif rgb == FILL_INPUT:
                    yellow += 1
        if green and not yellow and not RE_REPLACE_GREEN.search(prose):
            fail(book, "GUIDED",
                 f"{ws.title!r} is {green} example cell(s) and no input cell, and "
                 "nothing in the workbook tells the reader to overwrite the green — "
                 "so the sheet shows them an answer and no way to tell what is theirs")


def audit_guided(path: Path, wb) -> None:
    """Every yellow cell has to say where its number comes from."""
    book = path.name
    if not any(any(s.lower().startswith(g) for g in GUIDE_TABS) for s in wb.sheetnames):
        fail(book, "GUIDED", "no instructions tab — nothing tells a first-time user what to do")
    naked, unlabelled, inputs = [], [], 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                try:
                    rgb = str(c.fill.fgColor.rgb) if c.fill and c.fill.patternType else ""
                except Exception:                                # noqa: BLE001
                    rgb = ""
                if rgb != FILL_INPUT:
                    continue
                inputs += 1
                head = _column_header(ws, c.column, c.row)
                if head:
                    # a table cell: the header carries the explanation, so the
                    # only real failure is a header that says nothing
                    # "#" is a fine header for a row-number column; only a
                    # genuinely blank one leaves the user guessing
                    if head == "\x00" or not head.strip():
                        unlabelled.append(f"{ws.title}!{c.coordinate}")
                    continue
                # standalone input: needs a comment, or prose beside it
                has_note = c.comment is not None
                # prose beside it, or the label that names it, or a note in the
                # row underneath the block
                probes = [(c.row, c.column + 1), (c.row, c.column + 2),
                          (c.row, max(1, c.column - 1)), (c.row + 1, 1), (c.row + 2, 1)]
                for pr, pc in probes:
                    if has_note:
                        break
                    v = ws.cell(row=pr, column=pc).value
                    has_note = isinstance(v, str) and len(v) > 8
                if not has_note:
                    naked.append(f"{ws.title}!{c.coordinate}")
    if inputs == 0:
        warn(book, "GUIDED", "no yellow input cells — is anything meant to be filled in?")
    _check_example_has_somewhere_to_go(path, wb)
    if naked:
        fail(book, "GUIDED",
             f"{len(naked)} standalone input(s) with no explanation of where the number "
             f"comes from: {naked[:5]}")
    if unlabelled:
        fail(book, "GUIDED", f"{len(unlabelled)} input(s) under a blank column header: {unlabelled[:5]}")
    # A label appearing twice on one sheet is the signature of a ghost block:
    # a helper table that was moved, leaving the old copy behind, because the
    # builders only ever write cells and never clear the ground they left.
    for ws in wb.worksheets:
        # Only BOLD text counts. Ordinary content legitimately repeats — two
        # solutions can address the same root cause, two metrics can use the
        # same chart type — but a heading appearing twice means a block was
        # moved and its old copy was left behind.
        labels = {}
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and len(v) > 24 and not v.startswith("=")):
                    continue
                if not (c.font and c.font.bold):
                    continue
                labels.setdefault(v, []).append(c.coordinate)
        ghosts = {v: at for v, at in labels.items() if len(at) > 1}
        if ghosts:
            v, at = next(iter(ghosts.items()))
            fail(book, "GUIDED",
                 f"{len(ghosts)} label(s) appear more than once on {ws.title!r} — "
                 f"a moved helper block left its old copy behind: {v[:40]!r} at {at[:3]}")

    if not any(ws.sheet_view.showGridLines is False for ws in wb.worksheets):
        warn(book, "GUIDED", "gridlines left on everywhere — reads like a spreadsheet, not a tool")


def _blocks(ws):
    """Every banded data table on the sheet: (header_row, first, last, cols).

    A block runs from its dark header row down to the first row that is blank
    across all of its columns.
    """
    shadow = merged_shadow(ws)
    out = []
    for row in ws.iter_rows():
        hdr = [c for c in row
               if (c.row, c.column) not in shadow
               and c.fill and c.fill.patternType
               and str(c.fill.fgColor.rgb) == "FF333C49"]
        if len(hdr) < 2:
            continue
        cols = {c.column: str(c.value or "").strip() for c in hdr}
        r = row[0].row + 1
        last = r - 1
        while r <= ws.max_row:
            vals = [ws.cell(row=r, column=k).value for k in cols]
            if all(v in (None, "") for v in vals):
                break
            last = r
            r += 1
        if last >= row[0].row + 1:
            out.append((row[0].row, row[0].row + 1, last, cols))
    return out


def audit_example(path: Path, wb) -> None:
    """The shipped example has to be worth reading, not merely present.

    GUIDED asks whether an explanation EXISTS. Every defect a reader has
    actually reported was one that exists and says nothing: a Notes column
    headed and then left empty on all eight runs, a header reading "AB", the
    same sentence pasted onto six effect rows. Presence is not substance, so
    this layer asks the three questions presence cannot.
    """
    book = path.name
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in GUIDE_TABS):
            continue
        for hrow, first, last, cols in _blocks(ws):
            n = last - first + 1
            if n < 3:
                continue
            for col, head in cols.items():
                cells = [ws.cell(row=r, column=col) for r in range(first, last + 1)]
                filled = [c for c in cells if c.value not in (None, "")]
                typeable = any(c.fill and c.fill.patternType
                               and str(c.fill.fgColor.rgb) == FILL_INPUT for c in cells)
                # (1) a column the example declares and then never demonstrates.
                #
                # This used to exempt any column shaded yellow, on the reasoning
                # that a cell you are asked to fill in may legitimately be
                # blank. That reasoning is backwards, and it blinded the check
                # to the columns it matters most for: the reader is asked to
                # write a paragraph into "What you found" or "Why does it
                # wait?" and never shown one filled-in example of what a good
                # answer looks like. Every workbook here ships a worked example
                # row for exactly this purpose; a yellow column has to appear in
                # it. Only the REST of the column stays blank.
                if head and not filled:
                    fail(book, "EXAMPLE",
                         f"{ws.title}!{get_column_letter(col)}{first}:{last} — the "
                         f"example declares a {head!r} column and leaves it empty on "
                         f"all {n} rows; "
                         + ("the reader is asked to fill it in and never shown one "
                            "filled in" if typeable else
                            "the reader is told it matters and never shown what "
                            "goes in it"))
                # (2) a header that names nothing
                if head and len(head) <= 3 and head not in ("#", "No", "Qty", "Wk",
                                                            "Day", "RPN", "n", "SL"):
                    legend = any(
                        isinstance(ws.cell(row=r, column=c).value, str)
                        and len(str(ws.cell(row=r, column=c).value)) > 25
                        for r in (hrow - 1, hrow + 1)
                        for c in range(1, max(cols) + 1))
                    if not legend:
                        fail(book, "EXAMPLE",
                             f"{ws.title}!{get_column_letter(col)}{hrow} — column "
                             f"headed {head!r} with no legend anywhere near it; a "
                             "code is only readable to whoever wrote it")
        # (3) the same sentence on every row tells the reader nothing about
        #     their own row. Deliberately NOT scoped to banded blocks: the six
        #     identical DOE effect notes sat under a section band, not a table
        #     header, which is exactly how the first version of this check
        #     missed the defect it was written for.
        percol: dict[int, dict[str, list[int]]] = {}
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and len(v) > 30 and not v.startswith("="):
                    percol.setdefault(c.column, {}).setdefault(v, []).append(c.row)
        for col, seen in percol.items():
            dup = sorted(((v, rs) for v, rs in seen.items()
                          if len(rs) >= 3 and not _mirrored(ws, col, rs)),
                         key=lambda x: -len(x[1]))
            if dup:
                v, rs = dup[0]
                fail(book, "EXAMPLE",
                     f"{ws.title}!{get_column_letter(col)} rows {rs[:4]} — the same "
                     f"note is repeated on {len(rs)} rows ({v[:44]!r}…); boilerplate "
                     "that cannot be about any particular row is not guidance")


def _mirrored(ws, col, rows) -> bool:
    """Does another column repeat in exactly the same step as this one?

    Repetition is not automatically boilerplate. An observation log records the
    same event on every occasion it happened — the system-hop sheet says "find
    the adjustment on the account" once per contact because every contact made
    that hop — and there the repetition IS the finding.

    What separates the two is whether the rest of the row repeats with it. On a
    log, "why the agent moved" varies in lockstep with "from system" and "to
    system": same hop, same reason. A sentence pasted onto five rows repeats
    while the rows underneath it stay different, which is the defect this check
    was written for and the shape the mutant reproduces.

    The mirroring column has to vary somewhere across the span, or a column
    holding one word all the way down would excuse anything beside it.
    """
    lo, hi = min(rows), max(rows)
    if hi - lo + 1 <= len(rows):          # a solid run mirrors nothing useful
        return False
    for other in range(1, ws.max_column + 1):
        if other == col:
            continue
        here = {ws.cell(row=r, column=other).value for r in rows}
        if len(here) != 1 or next(iter(here)) in (None, ""):
            continue
        span = {ws.cell(row=r, column=other).value for r in range(lo, hi + 1)}
        if len(span) > 1:
            return True
    return False


from xlpolish import matches as _matches      # noqa: E402


def audit_jargon(path: Path, wb) -> None:
    """A term of art in a column header has to be explained on the sheet.

    The page makes every one of these clickable. A downloaded workbook cannot,
    so the reader meets "MR of z", "A2 for this n" and "OBS 1" with nothing to
    click and no way to guess — and every guidance check we had asked whether a
    header EXISTS, never whether a person outside the discipline could read it.
    The opaque-header rule above only fires at three characters or fewer, so
    "SIGMA U" and "Effect 95% CI" sailed through it.

    The glosses live in xlpolish, which is what writes them, so the check and
    the fix cannot disagree about what counts as jargon.
    """
    from xlpolish import GLOSSES
    book = path.name
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in GUIDE_TABS):
            continue
        # Everything readable on the sheet, so a gloss anywhere counts.
        prose = " ".join(
            str(c.value) for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and len(c.value) > 25).lower()
        for hrow, first, last, cols in _blocks(ws):
            for col, head in sorted(cols.items()):
                if not (isinstance(head, str) and head.strip()):
                    continue
                lab = head.strip().lower()
                for term in GLOSSES:
                    # THE matcher, not a third copy of it. This kept its own,
                    # and when xlpolish stopped letting a key prefix-match
                    # across an "of" this one carried on pairing "Share of hop
                    # time" with the Pareto's "Share" — then failed the sheet
                    # for not carrying a gloss that no longer belonged on it.
                    if not _matches(head, term):
                        continue
                    # Ask for the canonical gloss verbatim. Guessing at the
                    # shape of an explanation ("term =", "term /") reported the
                    # Xbar-R sheet as unexplained while the explanation sat two
                    # rows above the header, in a wording the guess did not
                    # anticipate. The fix writes GLOSSES[term]; the check looks
                    # for GLOSSES[term]; neither can drift from the other.
                    if GLOSSES[term].lower() in prose:
                        continue
                    fail(book, "JARGON",
                         f"{ws.title}!{get_column_letter(col)}{hrow} — column headed "
                         f"{head.strip()!r} uses {term!r} and nothing on the sheet says "
                         "what it means; on the page it would be clickable, in Excel "
                         "it is not")
                    break


ROWLABEL_COLS = range(3, 10)          # C..I — every column a note could sit in
RE_LITERAL = re.compile(r'"([^"]*)"')


def _speaks_for_itself(cell) -> bool:
    """Does this value read as a sentence, so a note beside it would repeat it?

    Two signals have to agree, because either alone is wrong.

    The number format decides whether the cell produces a NUMBER. Text alone is
    not enough: the Erlang staffing cell is
    ``=IFERROR(INDEX(...),"raise the range")`` — it returns a headcount, and the
    only long string in it is the error message. Judging by the literals would
    have exempted a bare number on the strength of text the reader never sees.

    The literal decides whether that text is an explanation or a label. A
    verdict cell returns "EXCELLENT — fit for purpose", which tells the reader
    what to do; a factor cell holds "Authority limit", which is a name and needs
    a note like any other. Twenty-five characters is where one becomes the
    other.
    """
    fmt = cell.number_format or "General"
    if any(ch in fmt for ch in "0#"):                # produces a number
        return False
    v = cell.value
    lits = RE_LITERAL.findall(v) if isinstance(v, str) and v.startswith("=") else \
        ([v] if isinstance(v, str) else [])
    return max((len(s) for s in lits), default=0) >= 25


def audit_formula_colour(path: Path, wb) -> None:
    """A formula cell must be blue — never yellow, never green.

    Each colour has exactly one job, and these two are instructions: yellow says
    type here, green says this is the example, replace it. Both are wrong on a
    cell the workbook computes, and following either destroys the calculation.
    55 cells across six workbooks carried one of them, including the Share and
    Rank on the Pareto's own worked-example row.

    Nothing caught it because every colour check so far asked what a cell's fill
    means for guidance, never whether the fill contradicts the cell's contents.
    """
    book = path.name
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                try:
                    rgb = str(c.fill.fgColor.rgb) if c.fill and c.fill.patternType else ""
                except Exception:                                # noqa: BLE001
                    continue
                if rgb in (FILL_INPUT, FILL_EXAMPLE):
                    which = "yellow (type here)" if rgb == FILL_INPUT else "green (replace this)"
                    bad.append(f"{ws.title}!{c.coordinate} is {which}")
    if bad:
        fail(book, "GUIDED",
             f"{len(bad)} formula cell(s) painted as something the reader should "
             f"type over: {bad[:4]}")


def audit_note_width(path: Path, wb) -> None:
    """A paragraph needs a paragraph's width, or it wraps into a column of soup.

    Measured in lines, not characters. A flat 200-character threshold caught the
    hypothesis log's 776-character guidance and missed the worse case: the six
    hierarchy definitions on the control plan, 130-170 characters each in a
    column 26 wide, each wrapping six or seven lines deep so the block read as
    shredded text. Length alone says nothing about readability — the same
    sentence is fine in a wide column and unreadable in a narrow one.

    Every guidance layer until now asked whether an explanation EXISTS. None
    asked whether it is legible where it sits.
    """
    book = path.name
    bad = []
    for ws in wb.worksheets:
        anchors = {str(m).split(":")[0] for m in ws.merged_cells.ranges}
        shadow = merged_shadow(ws)
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and len(v) >= 60 and not v.startswith("=")):
                    continue
                if c.coordinate in anchors or (c.row, c.column) in shadow:
                    continue
                w = ws.column_dimensions[c.column_letter].width or 8.43
                lines = len(v) / max(4.0, w * 0.95)
                # Only where there is somewhere to widen INTO. A note in the last
                # column of a populated table row has neighbours; wrapping deep
                # there is the table's shape, not a defect anyone can fix here.
                free = all(ws.cell(row=c.row, column=k).value in (None, "")
                           for k in range(c.column + 1, ws.max_column + 1))
                if lines > 4 and free:
                    bad.append(f"{ws.title}!{c.coordinate} ({len(v)} chars wrapping "
                               f"~{lines:.0f} lines in a {w:.0f}-wide column)")
    if bad:
        fail(book, "EXAMPLE",
             f"{len(bad)} note(s) wrapped into a column too narrow to read them: {bad[:3]}")


def audit_rowlabel(path: Path, wb) -> None:
    """A label in column A beside a filled value in B has to say what it is.

    This is the JARGON check turned ninety degrees. That one asks whether a
    COLUMN header is readable to someone outside the discipline; it walks the
    banded tables, so it only ever sees cells that sit under a header. A
    standalone row — "DPO | 0.0507" and nothing else — has no header by
    construction, so the entire class was invisible to it. Same defect, other
    axis, which is why the header check never fired on any of the 77.

    GUIDED does not cover it either: it inspects yellow cells only, so every
    CALCULATED row could print a bare number forever. Those are the rows that
    most need a sentence — the reader did not type the value and has nothing
    telling them what it is saying.
    """
    book = path.name
    bare: list[str] = []
    for ws in wb.worksheets:
        if any(ws.title.lower().startswith(g) for g in GUIDE_TABS):
            continue
        # Rows that belong to a banded table whose column A carries a real name.
        grid_rows = {r for hrow, first, last, cols in _blocks(ws)
                     if cols.get(1, "").strip()
                     for r in range(first, last + 1)}
        for r in range(1, ws.max_row + 1):
            lab = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2)
            if not (isinstance(lab, str) and lab.strip()):
                continue
            try:
                rgb = str(val.fill.fgColor.rgb) if val.fill and val.fill.patternType else ""
            except Exception:                                    # noqa: BLE001
                rgb = ""
            if rgb not in (FILL_INPUT, FILL_CALC) or val.value in (None, ""):
                continue
            # A row of a data grid, where column A is a real column with a name
            # of its own — "Part", "Period", "Category" — and the value is a
            # number the reader types. The header explains all 24 rows at once;
            # a note on each would be noise, and JARGON already requires that
            # header to be readable.
            #
            # Two things this exemption must NOT do, both of which it did on the
            # first attempt and a mutant caught:
            #
            # Membership comes from the block, not from walking up the sheet
            # looking for a dark fill. That walk has no lower bound, so a row
            # three below the end of a table still found the table's header and
            # inherited an explanation from a table it is not in.
            #
            # And it is limited to LITERAL values. A calculated cell under the
            # same header is a named quantity, not a repeated measurement —
            # "Count", "Mean", "p90" sit under a "Statistic" header and each one
            # needs its own sentence. Exempting on the header alone silently
            # dropped every stats block in the pack.
            if (r in grid_rows and not (isinstance(val.value, str)
                                        and val.value.startswith("="))):
                continue
            note = " ".join(
                str(ws.cell(row=r, column=c).value) for c in ROWLABEL_COLS
                if isinstance(ws.cell(row=r, column=c).value, str))
            if len(note.strip()) >= 12 or _speaks_for_itself(val):
                continue
            bare.append(f"{ws.title}!A{r} {lab.strip()[:32]!r}")
    if bare:
        fail(book, "ROWLABEL",
             f"{len(bare)} label:value row(s) with a filled value and no explanation "
             f"beside it — the reader sees a name and a number: {bare[:5]}")


def _na_ok(wb, where: str, depth: int = 1) -> bool:
    """Is this cell #N/A because somebody wrote NA(), or because it broke?"""
    sheet, _, addr = where.rpartition("!")
    # the recalculated keys come back with the sheet name upper-cased
    ws = next((w for w in wb.worksheets if w.title.upper() == sheet.upper()), None)
    if ws is None:
        return False
    try:
        f = ws[addr].value
    except (ValueError, KeyError):
        return False
    if not isinstance(f, str) or not f.startswith("="):
        return False
    if RE_CALLS_NA.search(f):
        return True
    if depth <= 0:
        return False
    # A helper column that is just `=AD5` inherits the gap without calling it.
    # One hop only: past that the chain is long enough to be hiding something.
    refs = RE_REF.findall(f)
    return bool(refs) and all(
        _na_ok(wb, f"{ws.title}!{r.replace('$', '')}", depth - 1) for r in refs)


def audit_numeric(path: Path) -> None:
    """Recalculate the whole workbook and refuse a single error value."""
    book = path.name
    try:
        import formulas
    except ImportError:
        warn(book, "NUMERIC", "the `formulas` package is not installed — skipped")
        return
    warnings.filterwarnings("ignore")
    try:
        xl = formulas.ExcelModel().loads(str(path)).finish()
        sol = xl.calculate()
    except Exception as exc:                                   # noqa: BLE001
        fail(book, "NUMERIC", f"workbook will not recalculate: {type(exc).__name__}: {exc}")
        return
    from openpyxl import load_workbook as _lw
    wb = _lw(path)
    bad = {}
    for key, cell in sol.items():
        if "'!" not in key or "!_XLFN" in key.upper():
            continue
        try:
            v = cell.value[0, 0]
        except Exception:
            continue
        if isinstance(v, str) and v.strip() in ERR_VALUES:
            where = key.split("]")[-1].replace("'", "")
            if v.strip() == "#N/A" and _na_ok(wb, where):
                continue
            bad.setdefault(v.strip(), []).append(where)
    for err, where in sorted(bad.items()):
        fail(book, "NUMERIC", f"{len(where)} cell(s) evaluate to {err}: {where[:5]}")


# ---------------------------------------------------------------- main


SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"


def render(books: list[Path], outdir: Path) -> list[Path]:
    """Render each workbook to page PNGs so the charts can actually be looked at.

    Structural checks prove a chart references live cells. They cannot tell you
    the axis runs 0-500 for data that lives between 396 and 442, so every point
    is squashed into the top fifth of the plot. Only looking at it tells you
    that, so the audit renders through LibreOffice and leaves PNGs behind.
    """
    import shutil
    import subprocess

    soffice = SOFFICE if Path(SOFFICE).exists() else shutil.which("soffice")
    if not soffice:
        print("  (no LibreOffice — skipping the visual layer)")
        return []
    if not shutil.which("pdftoppm"):
        print("  (no pdftoppm — `brew install poppler` — skipping the visual layer)")
        return []
    outdir.mkdir(parents=True, exist_ok=True)
    pngs = []
    for book in books:
        subprocess.run([soffice, "--headless", "--norestore", "--convert-to",
                        "pdf:calc_pdf_Export", "--outdir", str(outdir), str(book)],
                       check=False, capture_output=True, stdin=subprocess.DEVNULL)
        pdf = outdir / (book.stem + ".pdf")
        if not pdf.exists():
            fail(book.name, "VISUAL", "LibreOffice could not open it — the file may be corrupt")
            continue
        subprocess.run(["pdftoppm", "-r", "100", "-png", str(pdf), str(outdir / book.stem)],
                       check=False, capture_output=True)
        got = sorted(outdir.glob(book.stem + "-*.png"))
        pngs += got
        print(f"  rendered {book.name:36s} {len(got)} page(s)")
    return pngs


def audit_rubric(path: Path, wb, guidance: int) -> list[dict]:
    """Grade every chart against CHART-QA.md and fail the ones below the bar."""
    from chart_rubric import charts_in, grade, ships
    out = []
    for ch in charts_in(path):
        g = grade(path.name, ch, wb)
        g["guidance"] = guidance
        g["ships"] = ships(g, guidance)
        out.append(g)
        if not g["ships"]:
            fail(path.name, "RUBRIC",
                 f"{g['title'][:46] or '(untitled)'} — " + "; ".join(g["notes"] or ["below the bar"]))
    return out


# --------------------------------------------------------------- markdown
# Eleven of the twenty-eight templates are markdown documents, and until now not
# one check had ever opened one: every layer above globs *.xlsx. That is the
# whole reason 08-msa-gage-rr shipped an ANOVA results table with its
# "% Contribution" and "% Study variation" columns empty on all five rows and no
# method anywhere in the file — the identical defect the EXAMPLE layer catches
# in a workbook, sitting in a file the audit did not glob.


MD_MARKER = "<!-- guidance -->"   # md_guidance.py's idempotency marker
RE_MD_ROW = re.compile(r"^\s*\|(.+)\|\s*$")
# The dash is required. `[\s:|-]+` also matches `| | | |`, so a row that was
# empty in EVERY column read as a second header rule and was dropped from the
# body — the check could see a half-empty table but not a wholly empty one,
# which is the case it exists for.
RE_MD_RULE = re.compile(r"^\s*\|[\s:|-]*-[\s:|-]*\|\s*$")

# Words that mean a number has to be worked out rather than looked up. A
# document that asks for one owes the reader the arithmetic.
RE_ARITH = re.compile(
    r"\b(kappa|sigma level|dpmo|ndc|%\s*(study|contribution|grr)|p-value|"
    r"confidence interval|sample size|power|npv|payback|roi|"
    r"variance component|std ?dev|standard deviation|cpk|ppk|takt|"
    r"cycle efficiency|yield)\b", re.I)

# How that arithmetic may be discharged: a formula, a named method with a
# reference, or a pointer at the workbook that does it.
RE_METHOD = re.compile(
    r"(=\s*[A-Za-z(]|[÷×]|\bformula\b|\bcalculat|\bhow to work (it|this) out\b|"
    r"\.xlsx\b|\bworkbook\b|\bcalculator\b|\bMinitab\b|\bANOVA table\b)", re.I)


def md_tables(text: str):
    """Every pipe table as (header cells, [body rows]), with its line numbers."""
    lines = text.splitlines()
    out, i = [], 0
    while i < len(lines):
        m = RE_MD_ROW.match(lines[i])
        if m and i + 1 < len(lines) and RE_MD_RULE.match(lines[i + 1]):
            head = [c.strip() for c in m.group(1).split("|")]
            body, j = [], i + 2
            while j < len(lines) and RE_MD_ROW.match(lines[j]) \
                    and not RE_MD_RULE.match(lines[j]):
                body.append(([c.strip() for c in
                              RE_MD_ROW.match(lines[j]).group(1).split("|")], j + 1))
                j += 1
            out.append((head, body, i + 1))
            i = j
        else:
            i += 1
    return out


def audit_markdown(path: Path) -> None:
    """A document template has to be as fillable-in as a workbook."""
    book = path.name
    text = path.read_text(encoding="utf-8")

    # Guidance written inside an HTML comment reaches nobody. A markdown viewer
    # hides it, and the page escaped it into visible body text, so four real
    # rules — "no cause and no solution in the problem statement", "exactly one
    # A per row", the seven-step SIPOC limit — were simultaneously invisible in
    # the file and printed raw on the page, angle brackets and all.
    #
    # The one comment that stays is the marker md_guidance.py writes to know it
    # has already run. It is machinery, not prose, and it is named here so that
    # nothing else can hide behind the same exemption.
    for m in re.finditer(r"<!--(.*?)-->", text, re.S):
        if m.group(0) == MD_MARKER:
            continue
        line = text[:m.start()].count("\n") + 1
        fail(book, "MARKDOWN",
             f"line {line}: guidance is inside an HTML comment and reaches no "
             f"reader — {' '.join(m.group(1).split())[:60]!r}. Write it as text, "
             "or delete it")

    for head, body, line in md_tables(text):
        if not body:
            continue
        # A sign-off block is blank on purpose: the worked example must not
        # forge a signature, and a date filled in for you is worse than useless.
        # The exemption is scoped to tables that actually have a signature
        # column, so a "Date" in a data table is still expected to carry one.
        signing = any(h.strip(" *").lower() in ("signature", "signed") for h in head)
        for col, name in enumerate(head):
            if not name or name.startswith("**Total"):
                continue
            if signing and name.strip(" *").lower() in (
                    "signature", "signed", "date", "initials"):
                continue
            # A Total row does not count as filling a column. The Gage R&R
            # results table carried "100%" on its total line and nothing on any
            # of the four component rows above it, which is the column being
            # empty in every way that matters to a reader.
            got = [row[col].strip(" *") for row, _ in body
                   if col < len(row) and not row[0].strip(" *").lower().startswith("total")]
            if got and not any(got):
                fail(book, "MARKDOWN",
                     f"line {line}: column {name!r} is declared and then left "
                     f"empty on all {len(got)} rows — a worked example that "
                     "fills none of a column has not shown the reader anything")

    # A share column whose total row claims 100% has to actually sum to 100.
    # The stratified baseline listed shares of 38 / 18 / 34 / 14 under a total
    # row asserting 100%, because each had been rounded up on its own. Every
    # other number in that table reconciled exactly, which is what makes this
    # worth checking by machine: the one figure a reader would spot-check is the
    # one a careful author is least likely to re-derive.
    for head, body, line in md_tables(text):
        for col in range(len(head)):
            vals, total = [], None
            for row, _ in body:
                if col >= len(row):
                    continue
                first = row[0].strip(" *").lower()
                # The number has to BE the cell, not merely appear in it. A
                # first pass matched anywhere and read the 8 out of "None — 2
                # reopens below the 8% target" as that stratum's share.
                m = re.match(r"\*?\*?(-?\d+(?:\.\d+)?)\s*%", row[col].strip())
                if not m:
                    continue
                v = float(m.group(1))
                # "Total Gage R&R" is a COMPONENT, not the total of its table —
                # the total is the last such row, not the first.
                if first.startswith(("total", "all ", "check")):
                    total = v
                else:
                    vals.append(v)
            if total is None or abs(total - 100.0) > 0.5 or len(vals) < 3:
                continue
            s = sum(vals)
            # Percentages of a STANDARD DEVIATION combine in quadrature, not by
            # addition: Gage R&R's 11.2 / 26.1 / 95.9 are right and sum to 133.
            rss = math.sqrt(sum(v * v for v in vals))
            if abs(s - 100.0) > 1.01 and abs(rss - 100.0) > 1.01:
                fail(book, "MARKDOWN",
                     f"line {line}: column {head[col]!r} is totalled at 100% and its "
                     f"{len(vals)} rows sum to {s:g}% (and {rss:.1f}% in quadrature) "
                     "— the shares do not add up either way")

    # A number the reader has to derive, with nothing saying how.
    asked = sorted({m.group(0).lower() for m in RE_ARITH.finditer(text)})
    if asked and not RE_METHOD.search(text):
        fail(book, "MARKDOWN",
             f"asks for {', '.join(asked[:4])} but never says how to work any "
             "of it out — no formula, no method, no workbook to send them to")


def main() -> int:
    fast = "--fast" in sys.argv
    visual = "--visual" in sys.argv
    rubric = "--rubric" in sys.argv
    only = [a for a in sys.argv[1:] if not a.startswith("-")]
    books = sorted(TEMPLATES.glob("*.xlsx"))
    if only:
        books = [b for b in books if any(o in b.name for o in only)]
    print(f"Auditing {len(books)} workbook(s) in templates/\n")
    charts = 0
    graded: list[dict] = []
    for path in books:
        wb = load_workbook(path)
        n = audit_charts(path, wb)
        before = len(fails)
        audit_guided(path, wb)
        guidance = 2 if len(fails) == before else 1
        audit_example(path, wb)
        audit_jargon(path, wb)
        audit_rowlabel(path, wb)
        audit_formula_colour(path, wb)
        audit_note_width(path, wb)
        charts += n
        if rubric:
            graded += audit_rubric(path, wb, guidance)
        note = f"  ({NO_CHART_OK[path.name]})" if n == 0 and path.name in NO_CHART_OK else ""
        print(f"  {path.name:36s} sheets={len(wb.worksheets):2d}  charts={n:2d}{note}")
        if not fast:
            audit_numeric(path)
    docs = sorted(TEMPLATES.glob("*.md"))
    if only:
        docs = [d for d in docs if any(o in d.name for o in only)]
    for path in docs:
        audit_markdown(path)
    print(f"\n  {charts} charts across {len(books)} workbooks, "
          f"{len(docs)} markdown template(s) audited")
    if rubric and graded:
        ok = sum(1 for g in graded if g["ships"])
        print(f"\n  SCORECARD — {ok}/{len(graded)} charts clear the bar "
              "(gates pass, worth>=3, example=3, guidance=2)\n")
        print(f"  {'chart':50s} {'form':>6s} {'read':>5s} {'worth':>6s} {'exmpl':>6s} {'guide':>6s}")
        for g in sorted(graded, key=lambda x: (x["ships"], x["book"])):
            mark = "  " if g["ships"] else "->"
            print(f"{mark}{(g['title'] or '(untitled)')[:48]:50s} "
                  f"{'ok' if g['right_chart'] else 'FAIL':>6s} "
                  f"{'ok' if g['readable'] else 'FAIL':>5s} "
                  f"{g['worth']}/4{'':>3s} {g['example']}/3{'':>3s} {g['guidance']}/2")
    if visual:
        # --visual takes no value of its own: the positional arguments are
        # workbook filters, and consuming one here rendered into a directory
        # named after the filter.
        out = ROOT / ".qa-render"
        for a in sys.argv[1:]:
            if a.startswith("--out="):
                out = Path(a.split("=", 1)[1])
        print(f"\nRendering to {out} …")
        pngs = render(books, out)
        print(f"  {len(pngs)} page image(s) — open them and look at every chart")
    if warns:
        print(f"\n{len(warns)} warning(s):")
        print("\n".join(warns))
    if fails:
        print(f"\n{len(fails)} FAILURE(S):")
        print("\n".join(fails))
        return 1
    print("\nAll checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
