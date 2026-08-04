#!/usr/bin/env python3
"""The finishing pass every workbook gets, whoever built it.

Both builders (tools/build_templates.py for the newer templates,
tools/patch_workbooks.py for the originals) call polish_workbook() last. It
fixes the things that are invisible while you are writing formulas and glaring
the moment somebody opens, scrolls or prints the file:

  * charts that straddle a page break and print as two useless halves
  * axes that vanish in Excel because openpyxl leaves `delete` unset
  * 24 category labels crushed into an axis six centimetres wide
  * headers that scroll off the top the moment you reach row 30

None of it changes a single number. It is the difference between a workbook
somebody uses and one they close again.
"""
from __future__ import annotations

import re
from copy import copy

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.properties import PageSetupProperties

# openpyxl writes <c:delete val="1"/> unless told otherwise on some paths, and
# Excel then renders a chart with no axes at all.
def _fix_axes(ch) -> None:
    for ax in (getattr(ch, "x_axis", None), getattr(ch, "y_axis", None)):
        if ax is None:
            continue
        ax.delete = False
        # keep tick labels next to the axis rather than jumping to the far side
        # when a series goes negative (CUSUM's lower arm does)
        ax.tickLblPos = "low" if ax is getattr(ch, "x_axis", None) else "nextTo"


# Roughly how many characters of category label a 15cm axis carries before they
# start colliding, allowing for the rotation Excel applies on its own.
AXIS_CHARS = 90


def _thin_category_labels(ch, labels: list[str]) -> None:
    """Twenty-four labels on a 15cm axis is an unreadable smear.

    How many fit depends on how WIDE they are, not just how many there are:
    twenty-five test numbers sit comfortably where twenty-four "Day 12"s do not.
    Counting alone thinned the hypothesis log to one label in three on an axis
    with room for every one of them.
    """
    ax = getattr(ch, "x_axis", None)
    if ax is None or not labels:
        return
    # On a ranked bar chart the category label IS the content — thinning it
    # hid five of eight causes on the X-Y matrix. Only thin a long time axis.
    #
    # This used to read `ch.type == "bar"`, but on a BarChart `.type` is the bar
    # DIRECTION, so it only ever matched the horizontal ones. Every vertical
    # column chart — the ranked FMEA, the Pareto — kept being thinned, which is
    # the same defect the line above says was fixed. Excel rotates labels that
    # do not fit; it cannot invent one that was skipped.
    if ch.tagname == "barChart":
        skip = 1
    else:
        need = sum(len(t) + 1 for t in labels)
        skip = max(1, -(-need // AXIS_CHARS))
    ax.tickLblSkip = skip
    ax.tickMarkSkip = skip


def _cat_labels(ch, wb=None) -> list[str]:
    """The category labels the widest series actually plots.

    Reads the cells, rather than measuring the range the chart reserves. The
    value stream map books nineteen rows against nine filled ones; sizing the
    label thinning off nineteen made Excel show one step in three on a chart
    with room for all nine — a defect in the file the user downloads, not just
    in how we preview it. A formula cell counts as one label of average width,
    since its result is not knowable until Excel opens the file.
    """
    best: list[str] = []
    for ser in getattr(ch, "series", []) or []:
        ref = None
        for holder in (getattr(ser, "cat", None), getattr(ser, "xVal", None)):
            if holder is None:
                continue
            ref = getattr(getattr(holder, "numRef", None), "f", None) or \
                getattr(getattr(holder, "strRef", None), "f", None)
            if ref:
                break
        if not ref or "!" not in ref:
            continue
        sheet, addr = ref.split("!", 1)
        try:
            c1, r1, c2, r2 = range_boundaries(addr.replace("$", ""))
        except Exception:                                        # noqa: BLE001
            continue
        sheet = sheet.strip("'")
        if wb is None or sheet not in wb.sheetnames:
            got = ["______"] * ((r2 - r1 + 1) * (c2 - c1 + 1))
        else:
            ws = wb[sheet]
            got = []
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    v = ws.cell(row=r, column=c).value
                    if v in (None, ""):
                        continue
                    got.append("______" if isinstance(v, str) and v.startswith("=")
                               else str(v))
        if len(got) > len(best):
            best = got
    return best


def _header_row(ws) -> int | None:
    """The dark banded header, so panes can freeze just below it."""
    for row in ws.iter_rows(min_row=1, max_row=40):
        for c in row:
            try:
                if c.fill and c.fill.patternType and str(c.fill.fgColor.rgb) == "FF333C49":
                    return c.row
            except Exception:                                    # noqa: BLE001
                pass
    return None


# Plain English for every term of art this pack puts in a column header. The
# page makes these clickable; a downloaded workbook cannot, so the reader meets
# "MR of z" and "A2 for this n" with nothing to click and no way to guess. Every
# guidance check we had inspected whether a header EXISTS, never whether a
# person outside the discipline could read it.
GLOSSES = {
    "OBS": "OBS 1..n are your individual measurements in one subgroup — one row per period",
    "CL": "CL = centre line, the average the chart is drawn around",
    "UCL": "UCL / LCL = upper and lower control limits, ±3 sigma from the centre line",
    "LCL": "UCL / LCL = upper and lower control limits, ±3 sigma from the centre line",
    "MR": "MR = moving range, the gap between one point and the one before it",
    "R-bar": "R-bar = the average range within your subgroups",
    "X-double-bar": "X-double-bar = the average of the subgroup averages",
    "A2": "A2, D3, D4 = standard constants that depend only on your subgroup size",
    "D3": "A2, D3, D4 = standard constants that depend only on your subgroup size",
    "D4": "A2, D3, D4 = standard constants that depend only on your subgroup size",
    "u-bar": "u-bar = defects per unit across the whole baseline, not the average of daily rates",
    "p-bar": "p-bar = the overall proportion across the whole baseline",
    "sigma z": "sigma z = the Laney adjustment; 1.0 means a plain chart was fine, above that it was not",
    "z": "z = your rate restated in standard deviations, so periods of different size compare",
    "p-value": "p-value = the chance of seeing a difference this big if nothing had really changed",
    "CI": "CI = confidence interval, the range the true value plausibly sits in",
    "Effect size": "Effect size = how big the difference is, in units a manager can act on",
    "ASA": "ASA = average speed of answer, in seconds",
    "SL": "SL = service level, the share of contacts answered inside your target time",
    "Occupancy": "Occupancy = the share of logged-in time an agent is actually handling contacts",
    "Erlang A": "Erlang A = the staffing model that allows for callers hanging up; Erlang C assumes nobody does",
    "WFM": "WFM = Workforce Management, the team that owns schedules and headcount",
    "Champion": "Champion = the executive who owns the problem and clears the obstacles",
    "Black Belt": "Black Belt = the person running the project day to day",
    "AB": "AB, AC, BC = interaction columns: the product of two factor columns, +1 or -1",
    "AC": "AB, AC, BC = interaction columns: the product of two factor columns, +1 or -1",
    "BC": "AB, AC, BC = interaction columns: the product of two factor columns, +1 or -1",
    "Coefficient": "Coefficient = how far the outcome moves for a one-unit change in that driver, with the other drivers held still",
    "Std error": "Std error = how much that number would wobble if you drew another sample the same size",
    "t-stat": "t-stat = the coefficient divided by its own standard error; roughly 2 or more is the usual signal",
    "VIF": "VIF = variance inflation factor, how much two drivers move together; under 5 is fine, over 10 means drop one",
    "Residual": "Residual = what the model missed on that row: the value that actually happened minus the fitted one",
    "Fitted y": "Fitted y = what the model predicts for that row, before you compare it with what actually happened",
    "R²": "R-squared = the share of the spread in the outcome the model explains; the rest is everything you did not measure",
    "Adjusted R²": "Adjusted R-squared = R-squared with a penalty for each extra driver, so adding columns cannot flatter the model",
    "Odds ratio": "Odds ratio = how many times the odds of the outcome multiply when that driver goes up by one",
    "AUC": "AUC = how well the model ranks a case that happened above one that did not; 0.5 is a coin toss, 0.7 upwards is useful",
    "Logit": "Logit = the log-odds the model works in, before it is turned back into a probability",
    "Predicted probability": "Predicted probability = the model's chance that this particular row ends in the outcome",
    # --- 29-msa-gage-rr. Seventeen headers on this sheet carried no key at all,
    # and it is the sheet where a confident wrong reading does most damage: the
    # verdict it produces decides whether the project's baseline can be trusted.
    # Every figure below is read from the shipped worked example.
    "trial": "trial = one reading. Each appraiser measures every part three times, so a row holds "
             "nine numbers for one unchanged item and any difference between them is the "
             "measurement, not the process. On Part 1 appraiser A recorded 7.01, 7.26 and 7.98 "
             "minutes on the same call (B13:D13) — a 0.97-minute swing from one person measuring "
             "one thing three times",
    "Widest trial spread": "Widest trial spread = for that part, the largest range any single "
             "appraiser produced across their own three readings — repeatability, readable "
             "without arithmetic. Part 1 shows 0.97 minutes (M13), from appraiser A's 7.01 to "
             "7.98; B spread 0.63 and C 0.47, so A sets the row. The part's whole range across "
             "everyone is 1.54 (L13), so most of it is one person disagreeing with himself",
    "Appraiser": "Appraiser A / B / C = that person's average of their three readings for that "
             "part, one number per person per part. Part 1 reads 7.42, 8.20 and 7.10 minutes "
             "(B33:D33) for an unchanged call. Read a row across for disagreement about one "
             "part, and a column down for someone consistently high or low",
    "Sum of squares": "Sum of squares = total squared distance from the mean, before it is "
             "divided by anything. Two of them can never be compared directly, because each "
             "carries a different number of degrees of freedom: parts contribute 644.15 over 9 "
             "df and the repeat readings 6.53 over 60 (B47, B50). Divide each by its own df and "
             "you get the mean squares, 71.57 against 0.109, which may be compared",
    "df": "df = degrees of freedom, the number of independent comparisons behind that row — 9 for "
             "ten parts, 2 for three appraisers, 18 for their interaction, 60 for the repeats "
             "(C47:C50). It is the divisor that turns a sum of squares into a mean square",
    "Variance component": "Variance component = that source's share of the variation as a "
             "VARIANCE, in minutes squared, not minutes. Repeatability is 0.109 and "
             "reproducibility 0.589 (B56, B57) — a ratio of 5.4 on the squared scale, which is "
             "2.3 once you take square roots and get back to minutes, 0.33 against 0.77. Quote "
             "the minutes to anyone who has to act on it",
    "% Contribution": "% Contribution = each row's share of the total VARIANCE, so it sits on the "
             "squared scale, and only the unindented rows add to 100%. Measurement takes 8.1% "
             "and the parts 91.9% (C55, C58); the two indented rows, 1.3% and 6.8%, are the "
             "halves of that 8.1% and are already counted inside it",
    "Study variation (6\u03c3)": "Study variation (6\u03c3) = six standard deviations of that "
             "component, back in minutes — the span that covers virtually everything it "
             "produces. Measurement spans 5.01 minutes against the parts' 16.91 (D55, D58). Six "
             "sigma here is the convention for quoting a width, not a capability target",
    "% Study variation": "% Study variation = each row's share of the total STANDARD DEVIATION, "
             "not of the variance — which is why it does not match the % Contribution beside it: "
             "28.4% against 8.1% (E55, C55), the first being the square root of the second. This "
             "is the column the verdict is read on: under 10% acceptable, 10-30% marginal, over "
             "30% unusable. 28.4% is why this study reads MARGINAL",
    # --- 12-fmea. Six of these are three-letter columns holding a 1-10 score
    # with nothing on the sheet saying which end is bad. Figures below are read
    # from the shipped worked example.
    "SEV": "SEV = severity, 1-10, how badly the CUSTOMER is hurt when this failure happens — "
           "not how often, and not how likely you are to catch it. It is the one score an "
           "action cannot lower: you can stop the batch failing, but if it fails the customer "
           "is still late. The batch-failure row scores 8 and stays 8 after the fix (D14, O14). "
           "Anything at 9 or 10 is acted on whatever its RPN says",
    "OCC": "OCC = occurrence, 1-10, how often the cause actually happens. This is the score a "
           "prevention control moves. Alerting on the nightly batch took its row from 4 to 2 "
           "(F14, P14) because the cause stopped recurring, not because it became easier to see",
    "DET": "DET = detection, 1-10, and the scale runs BACKWARDS: 1 means you catch it before the "
           "customer does, 10 means the customer tells you. The deferred-close row scored 9 — "
           "'only detected if the customer complains again' — and fell to 3 once the webhook "
           "confirmed the posting (I11, Q11). In support this column is usually the worst of "
           "the three, and the cheapest to improve",
    "RPN": "RPN = risk priority number, severity x occurrence x detection, so 1 to 1000. It is a "
           "sorting tool, not a measurement: 504 against 336 says look at the first one first "
           "(J11, J16) and nothing more. Two rows can share an RPN with completely different "
           "shapes, which is why severity 9 or 10 is acted on regardless of where it ranks",
    "Detection control": "Detection control = what would catch this failure if it happened today, "
           "and how late. 'Only detected if the customer complains again' scores DET 9; "
           "'detected immediately on bounce-back' scores 3 (H11, H13). Write what actually "
           "happens now, not what the process document says should",
    "Effect on the customer": "Effect on the customer = what the person on the other end "
           "experiences, in their words, which is what SEV is scored against. 'Customer "
           "discovers it themselves and contacts us again' scores 7; 'margin leakage, caught in "
           "reconciliation' scores 4, because the customer never feels it (C11, C15)",
    "New SEV": "New SEV / New OCC / New DET = the same three scores AFTER the action, and only "
           "once it has SHIPPED. A control that is built, signed off or awaiting release has "
           "not changed what a customer meets, so the row keeps its original scores until it is "
           "live — the 'pending release' row sits at 6 / 9 / 8, unchanged, and the 'in backlog' "
           "row at 4 / 6 / 7. Re-scoring at design time reported a 77.3% risk reduction where "
           "the shipped evidence supports 57.7%",
    "Moving range": "Moving range = the gap between this point and the one before it, always "
           "positive. It is where an I-MR chart gets its limits: the average of these gaps, "
           "19.53 here, divided by 1.128 gives a sigma of 17.31 (B7, B8), so a point-to-point "
           "measure sets the limits rather than the spread of the whole series. That is the "
           "point — a slow drift cannot widen the limits and hide itself",
    "Centre line": "Centre line = the average the chart is drawn around, frozen from the baseline "
           "window and then held still. On the I-MR sheet it is 415.05 seconds from the first "
           "twenty points (B6). Held still is what makes it useful: recomputing it as new data "
           "arrives lets the line follow a drift, and a chart that follows the process it is "
           "watching cannot signal that the process moved",
    "Sigma p": "Sigma p = the standard deviation expected for THAT day's sample size, so a "
           "quiet day gets wider limits than a busy one. Day 1 has 7,820 contacts and a sigma p "
           "of 0.00507 (E14) from the overall 72.05% (B5). On a p-prime chart this is then "
           "multiplied by sigma z, 4.61 here (B7), which is the Laney correction for a rate "
           "that moves more between days than binomial sampling alone can explain",
    "Sigma u": "Sigma u = the standard deviation expected for that week's exposure, so a week "
           "with more audited contacts gets tighter limits. Week 1 audited 1,240 and gets "
           "0.00592 (E14) from the overall rate of 0.0434 defects per contact (B5); the Laney "
           "sigma z of 2.63 (B7) then widens it for overdispersion. A plain u-chart would have "
           "cried wolf on the quiet weeks and had nothing left to say about the loud ones",
    "Share": "Share = that category's count as a percentage of the total, before any sorting. "
           "The largest billing category runs 412 of 966 contacts, 42.7% (B5, C5). Read it "
           "beside the cumulative column: share tells you how big one bar is, cumulative tells "
           "you how few bars you need",
    "Cumulative": "Cumulative = the running total of share down the SORTED list, which is the "
           "line a Pareto chart draws. The top category alone reaches 42.7% and the top two "
           "reach 66.8% (J5, J6), so two of the categories carry two-thirds of the volume. "
           "Where the line crosses 80% is the set worth working on",
    "Share of hop time": "Share of hop time = that system pair's seconds as a percentage of all "
           "the time spent moving between systems — a share of the SWIVEL, not of the contact. "
           "Case tool to Billing platform takes 198 of the 737 logged seconds, 26.9% (E62, "
           "G62), so a third of the swivelling happens on one pair. It is the column that says "
           "which integration to build first",
    "Share of handle time": "Share of handle time = the swivel measured against the whole "
           "contact, which is the number to take to anyone who owns the handle-time target. "
           "Where 'share of hop time' ranks the pairs against each other, this one says what "
           "the hopping costs overall — the same seconds against the 412-second baseline "
           "rather than against the other hops",
    "Share of total": "Share of total = that family's variance as a percentage of the total "
           "variance, which is how a multi-vari study apportions the spread between sites, "
           "teams, agents and the noise within one agent. The shares are of VARIANCE, so they "
           "add to 100% while the standard deviations they come from do not. A family clamped "
           "to zero has not been shown to be identical — only that this sample cannot separate "
           "it from the family inside it",
    "Touch time (min)": "Touch time = minutes anyone is actually working on the contact at that "
           "step, not the minutes it sits at that step. The whole map totals 26 minutes of touch "
           "inside 2,426 minutes of lead time (E31, E33) — 1.07% (E35). That ratio is the point "
           "of the map: the waste is almost never in the working",
    "% of lead time": "% of lead time = that wait as a percentage of the whole door-to-door lead "
           "time, 2,426 minutes here (E33), not of the total waiting. Measuring each wait against "
           "the other waits makes every one look large and tells you nothing about which to "
           "remove",
    "Weighted total": "Weighted total = that cause's scores multiplied by the CTQ weights along "
           "the top and added up. With weights of 10, 7, 8, 5 and 9 (C11:G11) the closing-before-"
           "posting row totals 270 (H15). It is a ranking device, not a measurement — the gap "
           "between 270 and 261 is inside anyone's scoring noise, so treat the top few as a set "
           "to investigate rather than an order to work through",
    "Rank": "Rank = position by weighted total, 1 being highest. Ties are not broken, so two "
           "rows can share a rank. Read it with the totals beside it: rank 1 at 270 and rank 2 at "
           "261 (H15, H16) is a photo finish, and a rank alone hides that",
    "Weighted score": "Weighted score = each criterion's 1-5 score times the weight above it, "
           "added up, so it lands back on the 1-5 scale. The weights here run 0.25, 0.20, 0.15 "
           "and four at 0.10 (C10:I10) and must sum to 1, or the scores are not comparable. The "
           "deferred-close option scores 4.30 against 3.60 for the next one (J14, J15)",
    "Low risk of side-effects": "Low risk of side-effects = 1-5, and the label points the same "
           "way as every other column: 5 is GOOD, meaning few side-effects. A criterion phrased "
           "as a negative would invert the arithmetic against its neighbours, which is why it is "
           "written as 'low risk' rather than 'risk'",
    "Degrees of freedom": "Degrees of freedom = the number of independent comparisons that "
           "family supports: 1 between two sites, 2 between the three teams inside them (C63, "
           "C64). It is what a sum of squares is divided by to get a mean square, and it is why "
           "a family measured across two levels can never be pinned down as tightly as one "
           "measured across ten",
    "Erlang B": "Erlang B = the chance a contact arrives to find every agent busy AND no queue "
           "to wait in, so it is lost. At the recommended 67 agents against an offered load of "
           "60 erlangs it is 3.9%. Support queues almost never behave this way — it is here "
           "because Erlang C is built from it",
    "Erlang C": "Erlang C = the chance a contact arrives to find every agent busy and has to "
           "WAIT. At 67 agents that is 28.2%, which is what produces a service level of 82.3%: "
           "roughly three contacts in ten queue at all, and most of those are answered inside "
           "the target. It assumes nobody ever hangs up, which is false and makes it "
           "conservative — Erlang A is the version that credits abandonment",
    "Score": "Score = the 1-10 anchor band, written out so two people scoring the same failure "
           "mode land in the same place. Severity runs from 1-2 'the customer does not notice' "
           "to 9-10 'regulatory breach, safety issue, data exposure or mass impact'. Score "
           "against the anchor, not against how the last one felt",
    "Unit of analysis": "Unit of analysis = the thing one row of your data IS — a contact, an "
           "issue, a customer, an agent-day. Getting it wrong is the most expensive mistake on "
           "this sheet: a rate counted per contact and a rate counted per issue are different "
           "numbers with the same name, which is how one measurement ends up meaning two things",
    "Effect across the observed range": "Effect across the observed range = the coefficient "
           "multiplied by the full span of that driver in the data you actually have. A "
           "coefficient answers 'per one unit'; this answers 'over the range that exists', which "
           "is the size a process owner can act on. A large per-unit effect on a driver that "
           "barely varies moves nothing in practice, and only this column shows that",
    "Sensitivity": "Sensitivity = of the cases that really happened, the share the model flagged",
    "Specificity": "Specificity = of the cases that did not happen, the share the model correctly left alone",
}
_GL_LOW = {k.lower(): v for k, v in GLOSSES.items()}


def matches(label: str, term: str) -> bool:
    """Does this header label carry that glossary term?

    ONE copy of this rule. verify.py kept its own, they drifted the moment the
    "of" guard landed here, and the coverage check then reported three headers
    as explained by a gloss that no longer reached them.

    "X of Y" is a different quantity from "X", so a key may not prefix-match
    across an "of" — the Pareto's "Share" was otherwise explaining "Share of
    handle time" in terms of a category's count over 966 contacts. Suffix and
    interior matches are wanted: "OCC" has to reach "New OCC", "trial" has to
    reach "A trial 1".
    """
    lab, t = label.lower(), term.lower()
    return (lab == t or lab.endswith(" " + t) or (" " + t + " ") in lab
            or (lab.startswith(t + " ") and not lab.startswith(t + " of ")))


def glossed(label: str) -> bool:
    """Would a key line explain this header?"""
    return any(matches(label, t) for t in GLOSSES)


RE_FIRST_SENTENCE = re.compile(r"(?<=[a-z0-9%)\]])\.\s+(?=[A-Z])")


def gloss_short(gloss: str) -> str:
    """The definition without its worked example.

    Every gloss is written "TERM = what it is. Then the example, with this
    sheet's own numbers." — so the first sentence is already a complete
    definition and the split needs no rewriting. That matters: the examples are
    the thing that was asked for and none of them is cut, they just stop being
    the first thing a reader meets.
    """
    m = RE_FIRST_SENTENCE.search(gloss)
    return gloss[:m.start() + 1] if m else gloss


def write_full_glossary(wb) -> int:
    """The full definitions, with their worked examples, on the how-to tab.

    The key above a header row has to be short enough to read before the table
    it explains. On the FMEA that meant seven full glosses — 2,470 characters —
    sitting exactly where the reader arrives, ahead of any data. Above the
    header now carries the definition; the example goes here, which is where
    somebody who wants the depth already goes.
    """
    from openpyxl.styles import Alignment, Font

    sheet = next((w for w in wb.worksheets
                  if w.title.lower().startswith(("how to use", "read me"))), None)
    if sheet is None:
        return 0
    used = []
    for ws in wb.worksheets:
        if ws is sheet:
            continue
        for hrow, cols in _header_rows(ws).items():
            for label in cols.values():
                for term in GLOSSES:
                    if matches(label, term) and term not in used:
                        used.append(term)
    used = [t for t in used if gloss_short(GLOSSES[t]) != GLOSSES[t]]
    if not used:
        return 0
    HEAD = "WHAT THE COLUMNS MEAN — the short version sits above each table; this is the same thing with the worked example"
    start = next((r for r in range(1, sheet.max_row + 2)
                  if str(sheet.cell(row=r, column=1).value or "").startswith(
                      "WHAT THE COLUMNS MEAN")), None)
    if start is None:
        start = sheet.max_row + 2
    rows = [HEAD] + [GLOSSES[t] for t in used]
    n = 0
    for i, text in enumerate(rows):
        c = sheet.cell(row=start + i, column=1)
        if c.__class__.__name__ == "MergedCell" or c.value == text:
            continue
        c.value = text
        c.font = Font(bold=True, size=10) if i == 0 else Font(italic=True, size=9,
                                                              color="FF4B5563")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        # Merge across, or the text wraps into whatever column A happens to be.
        # On 32-system-hop that column is 2 characters wide, so a 347-character
        # definition became 87 stacked lines — caught by the check that exists
        # for exactly this, on the first run after the block was added.
        taken = {(r, cc) for rng in sheet.merged_cells.ranges
                 for r in range(rng.min_row, rng.max_row + 1)
                 for cc in range(rng.min_col, rng.max_col + 1)}
        span = 8
        if not any((start + i, cc) in taken for cc in range(1, span + 1)):
            sheet.merge_cells(start_row=start + i, start_column=1,
                              end_row=start + i, end_column=span)
        width = sum(sheet.column_dimensions[get_column_letter(cc)].width or 9
                    for cc in range(1, span + 1))
        sheet.row_dimensions[start + i].height = max(
            14, 11 * (1 + len(text) // max(40, int(width))))
        n += 1
    # A run that used to be longer must not leave orphans behind it.
    r = start + len(rows)
    while str(sheet.cell(row=r, column=1).value or "").strip().startswith(
            tuple(t.split(" =")[0] for t in GLOSSES)):
        sheet.cell(row=r, column=1).value = None
        r += 1
        n += 1
    return n


def _header_rows(ws):
    """(row, {col: label}) for every dark banded header row on the sheet."""
    rows = {}
    for row in ws.iter_rows():
        for c in row:
            try:
                dark = (c.fill and c.fill.patternType
                        and str(c.fill.fgColor.rgb) == "FF333C49")
            except Exception:                                    # noqa: BLE001
                dark = False
            if dark and isinstance(c.value, str) and c.value.strip():
                rows.setdefault(c.row, {})[c.column] = c.value.strip()
    return rows


def explain_headers(wb) -> int:
    """Write a plain-English key under any header row that uses jargon.

    Placed in the first blank row above the header, never by inserting one — an
    inserted row shifts every formula and chart range on the sheet.
    """
    from openpyxl.styles import Alignment, Font
    n = 0
    for ws in wb.worksheets:
        if ws.title.lower().startswith(("how to use", "read me", "legend")):
            continue
        for hrow, cols in _header_rows(ws).items():
            seen, wanted = set(), []
            for label in cols.values():
                for term, gloss in GLOSSES.items():
                    t = term.lower()
                    lab = label.lower()
                    # "X of Y" is a different quantity from "X", so a key must
                    # not prefix-match across an "of". The Pareto's "Share"
                    # otherwise explained "Share of handle time" on the
                    # system-hop sheet and "Share of total" on the multi-vari
                    # one, in terms of a category's count over 966 contacts.
                    # Suffix and interior matches are safe and wanted: "OCC"
                    # has to reach "New OCC", "trial" has to reach "A trial 1".
                    hit = matches(label, term)
                    if hit and gloss not in seen:
                        seen.add(gloss)
                        wanted.append(gloss_short(gloss))
            if not wanted:
                continue
            # One term per LINE. Joined with " · " these ran together into a
            # single orange paragraph — the FMEA header carries seven of them
            # and rendered as roughly 1,400 characters of unbroken prose above
            # the table, which is less readable than the jargon it explains.
            key = "Key:\n" + "\n".join("• " + w for w in wanted)
            above = hrow - 1
            if above < 1:
                continue
            width = max(cols) if cols else 6
            # The band above a header is not always merged from column A — the
            # DOE's legend spans B..G, so anchoring on column 1 found an empty
            # cell, then refused to write because the row was merged.
            anchor = next(
                (ws.cell(row=above, column=c) for c in range(1, width + 1)
                 if isinstance(ws.cell(row=above, column=c).value, str)
                 and ws.cell(row=above, column=c).value.strip()),
                ws.cell(row=above, column=1))
            existing = [ws.cell(row=above, column=c).value for c in range(1, width + 1)]
            has_text = any(v not in (None, "") for v in existing)

            if has_text:
                # Almost always the section band that introduces the table —
                # exactly where a reader looks for what the columns mean. Append
                # to it rather than inserting a row, because inserting shifts
                # every formula and chart range on the sheet.
                if not isinstance(anchor.value, str):
                    continue
                # REPLACE any key already there, do not skip. Refusing to touch a
                # row that had one froze every key line at whatever GLOSSES said
                # the first time the workbook was patched — and the nine
                # workbooks that are patched in place rather than rebuilt never
                # saw another word of it. Seven new FMEA glosses landed in the
                # dict, the build reported success, and the sheet went on
                # showing the one sentence it was written with.
                base = anchor.value
                # Match "Key:" and NOT "Key: " with a trailing space. The key
                # was reformatted to one bullet per line, so it now begins
                # "Key:\n" — the space in the old pattern stopped matching, the
                # previous key was never stripped, and every build appended
                # another copy. Four stacked keys shipped on eleven sheets
                # before anyone scrolled far enough to see it.
                cut = base.find("Key:")
                if cut != -1:
                    base = base[:cut]
                fresh = base.rstrip().rstrip("·").rstrip() + "  ·  " + key
                if anchor.value == fresh:
                    # Already right. Fall through and openpyxl appends another
                    # Alignment to the style table for a value that did not
                    # move — xl/styles.xml grew on every single run, which is
                    # the defect the byte-reproducible build was meant to end
                    # and which nothing caught because the only check looked at
                    # timestamps rather than running the pass twice.
                    continue
                anchor.value = fresh
            else:
                if any((above, c) in {(r, cc) for rng in ws.merged_cells.ranges
                                      for r in range(rng.min_row, rng.max_row + 1)
                                      for cc in range(rng.min_col, rng.max_col + 1)}
                       for c in range(1, width + 1)):
                    continue
                ws.merge_cells(start_row=above, start_column=1,
                               end_row=above, end_column=width)
                anchor = ws.cell(row=above, column=1, value=key)
                anchor.font = Font(italic=True, size=9, color="FF6B7280")
            anchor.alignment = Alignment(wrap_text=True, vertical="center")
            # Height has to allow for the explicit line breaks as well as the
            # wrapping, or a seven-term key is drawn one row tall with the rest
            # hidden under the header below it.
            lines = sum(1 + len(seg) // 110 for seg in str(anchor.value).split("\n"))
            ws.row_dimensions[above].height = max(14, 12 * lines)
            n += 1
    return n


def stamp(wb):
    """Freeze the document timestamps so a rebuild that changes nothing is empty.

    openpyxl writes the current time into docProps/core.xml on every save. Every
    worksheet XML can be byte-identical and all 19 workbooks still come out with
    different bytes, which changes the base64 embedded in the page, which makes
    the HTML differ too. So `git diff` after a rebuild always showed 20 modified
    files and could never tell anyone whether the rebuild actually did anything
    — the same blindness as a check that cannot fail, in the one place a
    reviewer looks first.

    The date is the project's, fixed, and deliberately not derived from
    anything: a value that moves is the entire problem.
    """
    from datetime import datetime
    fixed = datetime(2026, 1, 1)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.properties.creator = "six-sigma-blackbelt-support-ops"
    wb.properties.lastModifiedBy = "six-sigma-blackbelt-support-ops"
    return wb


ZIP_EPOCH = (2026, 1, 1, 0, 0, 0)


def save_workbook(wb, path) -> None:
    """Save, then rewrite the archive with fixed entry timestamps.

    Freezing the document properties is only half of it, and setting them before
    the save does not survive it — openpyxl stamps dcterms:modified with the
    clock as it writes, so that one has to be corrected here, afterwards. An
    .xlsx is also a zip, and a zip records a modification time per entry, which
    Python's zipfile likewise takes from the clock. All three move independently;
    all three have to be pinned before "the build produced no diff" means what it
    says.
    """
    import re as _re
    import zipfile
    wb.save(path)
    src = zipfile.ZipFile(path)
    items = [(i, src.read(i.filename)) for i in src.infolist()]
    src.close()
    tmp = path.with_name(path.name + ".rezip")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for info, data in items:
            if info.filename == "docProps/core.xml":
                data = _re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2026-01-01T00:00:00Z\g<2>", data)
            zi = zipfile.ZipInfo(info.filename, date_time=ZIP_EPOCH)
            zi.compress_type = info.compress_type
            zi.external_attr = info.external_attr
            zi.internal_attr = info.internal_attr
            zi.create_system = info.create_system
            out.writestr(zi, data)
    tmp.replace(path)


FILL_IN = "FFFFF9E3"        # "you fill this in"
FILL_EX = "FFECFAEF"        # "a worked example — replace it with your own"
FILL_CALC = "FFF2F7FF"      # "calculated — do not type over it"
FILL_BAND = "FFEEF1F6"      # a section band: the table above it has ended
FILL_HDR = "FF333C49"


def _boundary(ws, row: int, span) -> bool:
    """Does this row end the table above it?"""
    for c in span:
        cell = ws.cell(row=row, column=c)
        try:
            if cell.fill and cell.fill.patternType and \
                    str(cell.fill.fgColor.rgb) in (FILL_BAND, FILL_HDR):
                return True
        except Exception:                                        # noqa: BLE001
            pass
    # a note merged across the width of the block
    for rng in ws.merged_cells.ranges:
        if rng.min_row == row and (rng.max_col - rng.min_col + 1) >= max(2, len(span) - 1):
            return True
    return False


NOTE_FLOOR = 60          # shorter than this is a label, not a paragraph
MAX_LINES = 4            # deeper than this and the column is the problem


def _wrapped_lines(text: str, width: float) -> float:
    """Roughly how many lines this text takes in a column of that width."""
    return len(text) / max(4.0, width * 0.95)


def widen_long_notes(wb) -> int:
    """Give a paragraph somewhere to go, instead of a 13-wide column.

    Measured in LINES, not characters. A first attempt used a flat 200-character
    threshold and missed the worst case in the pack: the six hierarchy
    definitions on the control plan are 130-170 characters each in a column 26
    wide, so each one wrapped six or seven lines deep and the block read as a
    column of shredded text. Length alone says nothing — 170 characters is a
    sentence in a wide column and a paragraph in a narrow one.

    Only merges when every cell to the right of it on that row is empty, so it
    cannot swallow a neighbour, and sizes the row from the width it ends up
    with rather than guessing.
    """
    from openpyxl.styles import Alignment
    n = 0
    for ws in wb.worksheets:
        anchors = {str(m).split(":")[0] for m in ws.merged_cells.ranges}
        shadow = {(r, c) for rng in ws.merged_cells.ranges
                  for r in range(rng.min_row, rng.max_row + 1)
                  for c in range(rng.min_col, rng.max_col + 1)}
        last = ws.max_column
        for row in list(ws.iter_rows()):
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and len(v) >= NOTE_FLOOR
                        and not v.startswith("=")):
                    continue
                if cell.coordinate in anchors or (cell.row, cell.column) in shadow:
                    continue
                if _wrapped_lines(
                        v, ws.column_dimensions[cell.column_letter].width or 8.43) <= MAX_LINES:
                    continue
                if any(ws.cell(row=cell.row, column=k).value not in (None, "")
                       for k in range(cell.column + 1, last + 1)):
                    continue
                span = sum(ws.column_dimensions[get_column_letter(k)].width or 8.43
                           for k in range(cell.column, last + 1))
                ws.merge_cells(start_row=cell.row, start_column=cell.column,
                               end_row=cell.row, end_column=last)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                lines = 1 + int(_wrapped_lines(v, span))
                ws.row_dimensions[cell.row].height = max(
                    ws.row_dimensions[cell.row].height or 0, 13 * lines)
                n += 1
    return n


def _typed(cell) -> bool:
    """Did a person put this here? A formula did not, and neither did nothing."""
    v = cell.value
    return v not in (None, "") and not (isinstance(v, str) and v.startswith("="))


def _index_columns(ws, hrow: int, span) -> set:
    """Columns that are just a pre-printed row number: 1, 2, 3 down the block.

    Every entry grid in this pack numbers its rows in advance so the reader can
    refer to "row 14" in a meeting. Those numbers are furniture, not data, and a
    row that holds nothing else is blank however filled it looks.
    """
    out = set()
    for c in span:
        seen = []
        for r in range(hrow + 1, min(hrow + 40, ws.max_row) + 1):
            v = ws.cell(row=r, column=c).value
            if v in (None, ""):
                break
            if not isinstance(v, int) or isinstance(v, bool):
                seen = []
                break
            seen.append(v)
        if len(seen) >= 3 and seen == list(range(seen[0], seen[0] + len(seen))):
            out.add(c)
    return out


def recolour_formulas(wb) -> int:
    """A cell holding a formula is blue, wherever it sits.

    The legend gives each colour one job: yellow is yours to type, green is the
    worked example you replace, blue is calculated and must not be typed over.
    55 formula cells across six workbooks were painted green or yellow — the
    Pareto's own Share and Rank on the example row among them — which tells the
    reader to overwrite a formula and lose the calculation.

    Runs after mark_examples on purpose. Example colouring only ever repaints
    literals, so it cannot create this, but the hand-authored sheets already had
    it and nothing looked.
    """
    from openpyxl.styles import PatternFill
    blue = PatternFill("solid", fgColor=FILL_CALC)
    wrong = {FILL_IN, FILL_EX}
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                try:
                    rgb = str(c.fill.fgColor.rgb) if c.fill and c.fill.patternType else ""
                except Exception:                                # noqa: BLE001
                    continue
                if rgb in wrong:
                    c.fill = blue
                    n += 1
    return n


def mark_examples(wb) -> int:
    """Demonstration data inside a table is a worked example, so colour it one.

    The legend promises yellow means "you fill these in" and green means "a
    worked example so you can see the expected format; delete it when you
    start". 104 rows across seven sheets were demonstration data painted
    yellow, which contradicts both at once: the cell is not something you fill
    in, it is already filled.

    Scoped to banded tables on purpose. A standalone yellow input holding a
    value is a sensible DEFAULT to adjust — alpha at 0.05, a subgroup size of
    five — not an example to delete, and those must stay yellow.

    Scoped to the CONTIGUOUS pre-filled rows at the top of the block, and it
    stops at the first row that has nothing in it. Those rows are one worked
    example and the reader reads them as one: the Pareto ships five billing
    defect categories with counts and shares, and colouring the first green and
    the next four yellow says the last four are yours to invent, which is not
    what they are.

    The two wrong answers before this one are worth keeping, because they are
    opposite errors. The first painted every yellow cell in the block whatever
    lay below, which ran past the end of the logistic tab's data and turned its
    threshold and AUC settings into what looks like reference data. The second
    over-corrected to the first row alone, which split every seeded example down
    the middle. The boundary is the data, not a row count: an example ends where
    the blank rows begin.
    """
    from openpyxl.styles import PatternFill
    green = PatternFill("solid", fgColor=FILL_EX)
    n = 0
    for ws in wb.worksheets:
        if ws.title.lower().startswith(("how to use", "read me", "legend")):
            continue
        shadow = {(r, c) for rng in ws.merged_cells.ranges
                  for r in range(rng.min_row, rng.max_row + 1)
                  for c in range(rng.min_col, rng.max_col + 1)
                  if (r, c) != (rng.min_row, rng.min_col)}
        for hrow, cols in _header_rows(ws).items():
            span = sorted(cols)
            index_cols = _index_columns(ws, hrow, span)
            body = [c for c in span if c not in index_cols]
            r = hrow + 1
            while r <= ws.max_row:
                row_cells = [ws.cell(row=r, column=c) for c in span]
                # A row is part of the example only if somebody TYPED something
                # into it. Two things otherwise make an empty row look full:
                # the index column, pre-numbered down the whole block so the
                # reader can say "row 14" in a meeting, and the computed columns,
                # which carry a formula on every row whether or not it has data.
                # Counting either made the hypothesis log's rows 16-33 a worked
                # example on the strength of the numbers 6 to 23 and a column of
                # IF(...,"") that returns blank.
                if not any(_typed(ws.cell(row=r, column=c)) for c in (body or span)):
                    break
                # A table also ends at the next section band, the next header,
                # or a merged note spanning it. Walking only to the first blank
                # row ran straight past the end of the logistic tab's data and
                # recoloured the threshold and AUC cells below it — settings the
                # reader chooses, relabelled as an example to delete.
                if _boundary(ws, r, span):
                    break
                for cell in row_cells:
                    if (cell.row, cell.column) in shadow:
                        continue
                    try:
                        yellow = (cell.fill and cell.fill.patternType
                                  and str(cell.fill.fgColor.rgb) == FILL_IN)
                    except Exception:                            # noqa: BLE001
                        yellow = False
                    literal = (cell.value not in (None, "") and
                               not (isinstance(cell.value, str)
                                    and cell.value.startswith("=")))
                    if yellow and literal:
                        cell.fill = green
                        n += 1
                r += 1
    return n


def _anchor_ref(ch) -> str | None:
    """Where a chart is anchored, as "K3", whichever form openpyxl is using.

    A chart built in this session carries a plain string. One read back from a
    file carries a OneCellAnchor/TwoCellAnchor object with zero-based col/row,
    so comparing it against a string never matched and every chart looked moved
    on every run.
    """
    a = getattr(ch, "anchor", None)
    if isinstance(a, str):
        return a
    frm = getattr(a, "_from", None)
    if frm is None:
        return None
    return f"{get_column_letter(frm.col + 1)}{frm.row + 1}"



def write_chart_notes(wb) -> int:
    """Put each chart's note and action into the workbook, under the chart.

    They existed in the preview only, as a <figcaption>. This pack's whole
    premise is that the files get downloaded, and the same argument that put
    the glossary keys into the sheets applies here: on the page a caption is
    right there under the picture, in Excel there was nothing at all.

    Placed below the chart in the chart's own column, so scrolling right to
    look at it brings the note with it. A chart is about sixteen rows tall at
    the default 8cm, and the landing cell is checked for content before
    anything is written, so a second chart stacked underneath is never
    overwritten — the note walks down until it finds free space.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    from chart_notes import ACTIONS, NOTES

    n = 0
    for ws in wb.worksheets:
        taken = {(r, c) for rng in ws.merged_cells.ranges
                 for r in range(rng.min_row, rng.max_row + 1)
                 for c in range(rng.min_col, rng.max_col + 1)}
        for ch in getattr(ws, "_charts", []):
            title = _chart_title(ch)
            note = NOTES.get(title)
            if not note:
                continue
            # A chart's anchor is a plain STRING ("AH3") until the workbook has
            # been saved and reloaded, and a TwoCellAnchor object afterwards.
            # Reading only the object form skipped every freshly built chart —
            # so this wrote notes into the nine workbooks that are patched in
            # place and none of the thirteen that are generated, while the
            # build reported success either way.
            anc = getattr(ch, "anchor", None)
            frm = getattr(anc, "_from", None)
            if frm is not None:
                col, top = frm.col + 1, frm.row + 1   # anchors are 0-based
            elif isinstance(anc, str):
                from openpyxl.utils.cell import coordinate_to_tuple
                top, col = coordinate_to_tuple(anc)
            else:
                continue
            # 8cm of chart is roughly sixteen default rows; clear it, then a gap.
            row = top + int((ch.height or 7.5) / 0.5) + 2
            text = f"HOW TO READ IT.  {note}\n\nSO:  {ACTIONS.get(title, '')}".rstrip()
            # Idempotent, and it has to be: the landing cell is only free the
            # FIRST time, so walking down to the next empty one wrote a second
            # copy on every rebuild — 42 notes became 76 in one round trip.
            # Look for this chart's note anywhere on the sheet before placing it.
            if any(c.value == text for r in ws.iter_rows() for c in r
                   if isinstance(c.value, str)):
                continue
            while True:
                cell = ws.cell(row=row, column=col)
                if (row, col) in taken or cell.__class__.__name__ == "MergedCell":
                    row += 1
                    continue
                if cell.value in (None, "") or cell.value == text:
                    break
                row += 1
            if cell.value == text:
                continue
            cell.value = text
            cell.font = Font(italic=True, size=9, color="FF4B5563")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            span = 9
            if not any((row, c) in taken for c in range(col, col + span)):
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=col + span - 1)
                taken.update((row, c) for c in range(col, col + span))
            width = sum(ws.column_dimensions[get_column_letter(c)].width or 9
                        for c in range(col, col + span))
            ws.row_dimensions[row].height = max(
                28, 11 * sum(1 + len(seg) // max(30, int(width))
                             for seg in text.split("\n")))
            n += 1
    return n


def _chart_title(ch) -> str:
    """The chart's title text.

    A title reads differently before and after a save. openpyxl wraps an
    assigned string in a Title object, so the build-time chart has no plain
    string to return and the rich-text walk is the only route — which is why
    reading it one way wrote the notes into every SAVED workbook and none of
    the freshly built ones.
    """
    t = getattr(ch, "title", None)
    if t is None:
        return ""
    if isinstance(t, str):
        return t.strip()
    rich = getattr(getattr(t, "tx", None), "rich", None)
    if rich is not None:
        got = "".join(r.t or "" for para in rich.p for r in (para.r or [])).strip()
        if got:
            return got
    strref = getattr(getattr(t, "tx", None), "strRef", None)
    return (getattr(strref, "f", "") or "").strip()


# ---------------------------------------------------------------- colour key
#
# The one definition of what the three fills mean. build_templates builds its
# LEGEND from this, and reconcile_legend() below repairs a workbook whose key
# has drifted from its sheets, so there is no second copy to fall out of step.
LEGEND_KEY = [
    ("Yellow cells", "FFFFF9E3", "You fill these in."),
    ("Blue cells", "FFF2F7FF",
     "Calculated for you. Do not type over them \u2014 they contain formulas."),
    ("Green cells", "FFECFAEF",
     "A worked example, so you can see the expected format and the charts say "
     "something the moment you open the file. Replace it with your own data "
     "when you start."),
]


def fills_used(wb) -> set:
    """Every solid fill colour anywhere in the workbook."""
    seen = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = getattr(c, "fill", None)
                if f is None or f.fill_type != "solid":
                    continue
                rgb = getattr(getattr(f, "start_color", None), "rgb", None)
                if isinstance(rgb, str):
                    seen.add(rgb.upper())
    return seen


def reconcile_legend(wb) -> int:
    """Add any colour the workbook uses and its key does not mention.

    The key was derived at the end of each builder, which is one step too
    early: this function's own caller adds fills, so the DOE matrix and the
    Gage R&R study shipped using green cells their key never mentioned. In
    memory they had none; the saved file had them. Deriving it here, where
    every fill exists, is the only place the answer is the final one.

    Only adds. Removing a line would need to be right about a colour being
    absent from a workbook this pass has not finished with, and the cost of
    being wrong is a reader looking for cells nobody explained.
    """
    if "How to use this" not in wb.sheetnames:
        return 0
    ws = wb["How to use this"]
    used = fills_used(wb)
    heads = {name for name, _, _ in LEGEND_KEY}

    # Where each colour's heading sits, if it is there at all.
    at = {}
    for r in range(1, min(ws.max_row, 40) + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip() in heads:
            at[v.strip()] = r
    if not at:
        return 0                      # no key on this sheet; not ours to invent

    # Copy the look from a heading and an explanation that are already there.
    any_head = ws.cell(row=min(at.values()), column=2)
    any_body = ws.cell(row=min(at.values()) + 1, column=2)
    added = 0
    for i, (name, colour, text) in enumerate(LEGEND_KEY):
        if name in at or colour not in used:
            continue
        # After the previous colour's explanation, or at the top of the key.
        prev = [LEGEND_KEY[j][0] for j in range(i) if LEGEND_KEY[j][0] in at]
        row = (at[prev[-1]] + 2) if prev else min(at.values())
        # insert_rows moves cells and leaves merged ranges and row heights
        # where they were, so everything below the insertion came unmerged
        # from its neighbour. On the Gage R&R sheet that dropped four notes
        # into column A, which is two characters wide, and the longest of
        # them wrapped to ninety-two lines. Move them by hand.
        below = [(r.min_row, r.min_col, r.max_row, r.max_col)
                 for r in list(ws.merged_cells.ranges) if r.min_row >= row]
        heights = {r: d.height for r, d in ws.row_dimensions.items()
                   if r >= row and d.height}
        for r0, c0, r1, c1 in below:
            ws.unmerge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)
        ws.insert_rows(row, 2)
        for r0, c0, r1, c1 in below:
            ws.merge_cells(start_row=r0 + 2, start_column=c0,
                           end_row=r1 + 2, end_column=c1)
        for r in sorted(heights, reverse=True):
            ws.row_dimensions[r + 2].height = heights[r]
        h = ws.cell(row=row, column=2, value=name)
        h.font = copy(any_head.font)
        h.alignment = copy(any_head.alignment)
        b = ws.cell(row=row + 1, column=2, value=text)
        b.font = copy(any_body.font)
        b.alignment = copy(any_body.alignment)
        ws.row_dimensions[row + 1].height = max(15, 15 * (len(text) // 100 + 1))
        at = {k: (v + 2 if v >= row else v) for k, v in at.items()}
        at[name] = row
        added += 1
    return added


def sheet_fills(ws) -> set:
    """Solid fill colours on one sheet."""
    seen = set()
    for row in ws.iter_rows():
        for c in row:
            f = getattr(c, "fill", None)
            if f is None or f.fill_type != "solid":
                continue
            rgb = getattr(getattr(f, "start_color", None), "rgb", None)
            if isinstance(rgb, str):
                seen.add(rgb.upper())
    return seen


def reconcile_sheet_key(ws) -> int:
    """The colour key that sits under 'HOW TO USE THIS SHEET' on a data sheet.

    A different shape from the one on a 'How to use this' tab: the heading is
    in column A and its explanation shares the row in column B, and it
    describes THAT SHEET rather than the workbook. It was written by a fixed
    list of cell addresses asserting all three colours, so it was right only
    for as long as nobody changed a sheet underneath it.

    Rows are rewritten in place, never inserted or deleted. The data on these
    sheets starts a couple of rows below and several hundred cell addresses in
    patch_workbooks point at it, so shifting a row to tidy a legend would move
    every one of them. A colour the sheet does not use empties its row and the
    rest close up above it.

    The wording already on the sheet is kept. These sheets say the worked
    example is "realistic" and the tabs say it makes the charts say something
    straight away; both are true of their own sheet and neither is this
    function's to rewrite.
    """
    names = [n for n, _, _ in LEGEND_KEY]
    rows = {}
    for r in range(1, min(ws.max_row, 14) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() in names:
            rows[v.strip()] = r
    if not rows:
        return 0
    span = sorted(rows.values())
    used = sheet_fills(ws)
    want = [(n, t) for n, c, t in LEGEND_KEY if c in used]
    # More colours than the block has rows: adding one means inserting a row,
    # which is the thing this must not do. Leave it and let verify say so.
    if len(want) > len(span):
        return 0
    keep = {n: ws.cell(row=r, column=2).value for n, r in rows.items()}
    changed = 0
    for i, r in enumerate(span):
        head, body = (want[i][0], keep.get(want[i][0]) or want[i][1]) \
            if i < len(want) else (None, None)
        if ws.cell(row=r, column=1).value != head:
            ws.cell(row=r, column=1).value = head
            changed += 1
        if ws.cell(row=r, column=2).value != body:
            ws.cell(row=r, column=2).value = body
            changed += 1
    return changed


def polish_workbook(wb, landscape: bool = True) -> int:
    """Print setup, frozen headers and chart legibility, for every sheet.

    Returns how many sheets it actually changed, so a caller that only saves on
    a real change (patch_workbooks.py, to keep the embedded base64 stable) can
    tell whether this pass did anything.

    Runs at every save site in the build, which is why stamp() lives here: the
    document timestamps have to be frozen on the way out or the bytes move
    whether or not the content did.
    """
    stamp(wb)
    changed = (explain_headers(wb) + mark_examples(wb) + recolour_formulas(wb)
               + widen_long_notes(wb) + write_chart_notes(wb)
               + write_full_glossary(wb))
    for ws in wb.worksheets:
        before = (ws.page_setup.fitToWidth, ws.page_setup.orientation, ws.freeze_panes)

        # --- printing: one sheet wide, so a chart is never cut in half ---
        ws.page_setup.orientation = "landscape" if landscape else "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0          # as many pages tall as it needs
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5

        # --- reading: keep the header visible past row 30 ---
        if ws.freeze_panes is None:
            hr = _header_row(ws)
            if hr and hr < 30:
                ws.freeze_panes = f"A{hr + 1}"

        # --- charts: never on top of the data they explain ---
        # Thirteen charts were anchored over populated cells, several over the
        # very block they read from. Fixing them one at a time is how the last
        # three rounds of this went, so it is done here for the class: every
        # chart is pushed clear of the rightmost populated column on its sheet,
        # and stacked vertically in anchor order so two charts never collide.
        charts = list(getattr(ws, "_charts", []))
        if charts:
            last_col = 0
            for row in ws.iter_rows():
                for c in row:
                    if c.value not in (None, ""):
                        last_col = max(last_col, c.column)
            free = get_column_letter(last_col + 2)
            row_at = 3
            for ch in charts:
                target = f"{free}{row_at}"
                # Only a real move counts. This incremented unconditionally, so
                # every workbook with a chart reported a change on every run and
                # re-saved forever — which was invisible while the timestamps
                # churned the bytes anyway, and is the whole reason "did the
                # build change anything?" had no answer.
                if _anchor_ref(ch) != target:
                    changed += 1
                ch.anchor = target
                row_at += max(16, int((ch.height or 7.5) / 0.5) + 3)


        for ch in getattr(ws, "_charts", []):
            # Charts skip hidden cells by default. Several templates keep their
            # reference columns hidden — a quadrant divider, a noise floor —
            # and those series silently vanished from the plot.
            ch.plotVisOnly = False
            _fix_axes(ch)
            _thin_category_labels(ch, _cat_labels(ch, wb))
            # openpyxl never round-trips this one, so setting it must not count
            # as a change or every run would re-save and churn the base64 copies
            ch.roundedCorners = False

        if before != (ws.page_setup.fitToWidth, ws.page_setup.orientation, ws.freeze_panes):
            changed += 1
    # Last, so both see every fill this pass has put in.
    changed += reconcile_legend(wb)
    for ws in wb.worksheets:
        changed += reconcile_sheet_key(ws)
    return changed
